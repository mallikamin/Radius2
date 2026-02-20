import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.load_workbook('C:/Users/Malik/Downloads/Service Areas_updated.xlsx')
ws = wb['Sheet1']

# Find where Tech Services starts
tech_row = None
for row in ws.iter_rows(min_row=1, max_col=3):
    for cell in row:
        if cell.value and 'Tech Services' in str(cell.value):
            tech_row = cell.row
            break
    if tech_row:
        break

print(f"Tech Services starts at row: {tech_row}")

# Clear everything from tech_row onward
for row in ws.iter_rows(min_row=tech_row, max_row=ws.max_row, max_col=10):
    for cell in row:
        cell.value = None

# Styles
header_font = Font(name='Calibri', bold=True, size=11)
sub_header_font = Font(name='Calibri', bold=True, size=10, color='333333')
normal_font = Font(name='Calibri', size=10)
gold_fill = PatternFill(start_color='FFF8E7', end_color='FFF8E7', fill_type='solid')

r = tech_row

# Main category header
ws.cell(row=r, column=1, value=8).font = header_font
ws.cell(row=r, column=2, value='Tech Services').font = header_font

# Structured service lines
services = [
    # --- 1. Digital Strategy & Enterprise Architecture ---
    ('DIGITAL STRATEGY & ENTERPRISE ARCHITECTURE', True),
    ('Target operating model design and enterprise system blueprints', False),
    ('Technology consolidation strategy and modernization roadmaps', False),
    ('Business-to-technical architecture translation', False),
    ('System integration planning and vendor-neutral technology selection', False),
    ('Data architecture design (relational, document, geospatial)', False),

    # --- 2. Custom Platform Engineering ---
    ('CUSTOM PLATFORM ENGINEERING (FULL-STACK)', True),
    ('FastAPI + React enterprise application development', False),
    ('Secure REST API design and implementation (200+ endpoints per system)', False),
    ('Multi-role portals with RBAC and organizational hierarchy-based access', False),
    ('Production-grade responsive UI with real-time data binding', False),
    ('Database design, optimization, and migration (PostgreSQL, SQLAlchemy, Alembic)', False),
    ('File and media management systems (MinIO S3-compatible object storage)', False),
    ('PDF/Excel report generation engines with dynamic templates', False),
    ('Real-time notification systems (in-app, event-driven)', False),

    # --- 3. AI, Analytics & Intelligence Systems ---
    ('AI, ANALYTICS & INTELLIGENCE SYSTEMS', True),
    ('Predictive analytics engines and forecasting models (94% accuracy achieved)', False),
    ('Real-time voice transcription and processing (Whisper, GPU-accelerated)', False),
    ('AI agent orchestration frameworks (multi-agent parallel execution)', False),
    ('Sentiment analysis and NLP pipelines', False),
    ('Commodity and financial signal analysis (Gold, Silver, WTI crude)', False),
    ('Decision-support models with scenario analysis and what-if modeling', False),
    ('Geospatial intelligence and visualization', False),
    ('KPI instrumentation and executive dashboard design', False),

    # --- 4. Business Systems ---
    ('BUSINESS SYSTEMS & DOMAIN PLATFORMS', True),
    ('CRM lifecycle management: sales, receivables, collections, buybacks', False),
    ('Restaurant POS: order orchestration, KDS, menu builder, multi-location', False),
    ('Financial reporting, P&L monitoring, and revenue forecasting', False),
    ('Inventory, transaction, and installment management', False),
    ('Task management with delegation, collaboration, and org-hierarchy visibility', False),
    ('Vector mapping and spatial design tools (annotation, export, PDF overlay)', False),
    ('B2B client portals and collaboration platforms', False),
    ('Portfolio management and investment tracking', False),

    # --- 5. Data Infrastructure & Platform Operations ---
    ('DATA INFRASTRUCTURE & PLATFORM OPERATIONS', True),
    ('PostgreSQL performance tuning, indexing strategy, and query optimization', False),
    ('ETL pipeline engineering and data transformation', False),
    ('Redis caching layer and Celery distributed task orchestration', False),
    ('Real-time data pipelines (500K+ records processed at scale)', False),
    ('Data migration from legacy systems with validated balances', False),
    ('Backup automation, recovery procedures, and data integrity verification', False),

    # --- 6. Infrastructure & DevOps ---
    ('INFRASTRUCTURE & DEVOPS', True),
    ('Docker Compose multi-service orchestration with health checks', False),
    ('nginx reverse proxy with SSL/TLS termination and domain routing', False),
    ('LAN, VPS (DigitalOcean), and cloud deployment', False),
    ('Multi-application co-hosting on single server (shared network architecture)', False),
    ('Domain management, DNS configuration, and certificate automation', False),
    ('CI/CD pipelines, release management, and deployment runbooks', False),
    ('Environment parity (dev, staging, production) with container isolation', False),

    # --- 7. Security & Quality Assurance ---
    ('SECURITY & QUALITY ASSURANCE', True),
    ('JWT authentication and secure session management', False),
    ('Role-based access control (RBAC) with multi-level organizational hierarchy', False),
    ('OWASP-aware hardening (IDOR, XSS, CORS, SQL injection prevention)', False),
    ('Architecture reviews and milestone-gated release process', False),
    ('End-to-end integration testing and runtime verification', False),
    ('API contract validation and cross-boundary failure detection', False),

    # --- 8. AI-Augmented Delivery ---
    ('AI-AUGMENTED DELIVERY METHODOLOGY', True),
    ('Proprietary AI org: 23 specialized agents across 10 departments (TARS)', False),
    ('Parallel engineering: multiple agents writing, reviewing, and testing simultaneously', False),
    ('Automated architecture review and 54-point audit checklist per release', False),
    ('Session continuity and crash-recovery across long engagements', False),
    ('Delivery timeline compression (10x vs conventional development)', False),
]

for text, is_header in services:
    r += 1
    cell = ws.cell(row=r, column=3, value=text)
    if is_header:
        cell.font = sub_header_font
        cell.fill = gold_fill
    else:
        cell.font = normal_font

# Set column widths
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 80

wb.save('C:/Users/Malik/Downloads/Service Areas_updated.xlsx')
print(f"Done. Tech Services section: rows {tech_row} to {r} ({r - tech_row} sub-items)")

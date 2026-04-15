"""
Radius CRM v3 - Complete Backend
Customers, Brokers, Projects, Inventory, Transactions
"""
from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Query, Form, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy import create_engine, Column, String, Text, DateTime, Numeric, ForeignKey, Integer, Date, Boolean, text, or_, and_, case, literal, union_all
from sqlalchemy.dialects.postgresql import UUID, JSONB
from sqlalchemy.orm import sessionmaker, Session, declarative_base
from sqlalchemy.sql import func
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, timedelta, datetime
from dateutil.relativedelta import relativedelta
from jose import JWTError, jwt
from passlib.context import CryptContext
import os
import uuid
import csv
import io
import time
import logging
import httpx
from collections import defaultdict

logger = logging.getLogger(__name__)
_interaction_target_indexes_ready = False
CUSTOMER_SYNC_APPROVER_REP_IDS = ["REP-0014", "REP-0015", "REP-0008"]  # Iram Riaz, Imran Younas, Syed Faisal

# ============================================
# DATABASE SETUP
# ============================================
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://sitara:sitara123@localhost:5432/sitara_crm")
engine = create_engine(
    DATABASE_URL,
    pool_size=20,
    max_overflow=10,
    pool_pre_ping=True,
    pool_recycle=3600
)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def find_entity(db, model, id_field_name, value):
    """Safely lookup by UUID id or string entity_id (handles non-UUID strings)"""
    try:
        uuid.UUID(str(value))
        return db.query(model).filter(
            (model.id == value) | (getattr(model, id_field_name) == value)
        ).first()
    except ValueError:
        return db.query(model).filter(getattr(model, id_field_name) == value).first()

# ============================================
# MODELS
# ============================================
class Customer(Base):
    __tablename__ = "customers"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    customer_id = Column(String(20), unique=True, nullable=False)
    name = Column(String(255), nullable=False)
    mobile = Column(String(20), unique=True, nullable=False)
    address = Column(Text)
    cnic = Column(String(20))
    email = Column(String(255))
    additional_mobiles = Column(JSONB, default=list)
    source = Column(String(100))
    source_other = Column(Text)
    occupation = Column(String(100))
    occupation_other = Column(Text)
    interested_project_id = Column(UUID(as_uuid=True), ForeignKey("projects.id"))
    interested_project_other = Column(Text)
    area = Column(Text)
    city = Column(Text)
    country_code = Column(String(5), default="+92")
    notes = Column(Text)
    assigned_rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    temperature = Column(String(10))  # hot | mild | cold
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class Broker(Base):
    __tablename__ = "brokers"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    broker_id = Column(String(20), unique=True, nullable=False)
    name = Column(String(255), nullable=False)
    mobile = Column(String(20), unique=True, nullable=False)
    cnic = Column(String(20))
    email = Column(String(255))
    company = Column(String(255))
    address = Column(Text)
    commission_rate = Column(Numeric(5, 2), default=2.00)
    bank_name = Column(String(100))
    bank_account = Column(String(50))
    bank_iban = Column(String(50))
    notes = Column(Text)
    status = Column(String(20), default="active")
    linked_customer_id = Column(UUID(as_uuid=True), ForeignKey("customers.id"))
    assigned_rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class Project(Base):
    __tablename__ = "projects"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    project_id = Column(String(20), unique=True, nullable=False)
    name = Column(String(255), nullable=False)
    location = Column(String(255))
    description = Column(Text)
    status = Column(String(20), default="active")
    # Vector-specific fields (optional - for linked projects)
    map_pdf_base64 = Column(Text, nullable=True)
    map_name = Column(String(255), nullable=True)
    map_size = Column(JSONB, nullable=True)
    vector_metadata = Column(JSONB, nullable=True)
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class CompanyRep(Base):
    __tablename__ = "company_reps"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    rep_id = Column(String(20), unique=True, nullable=False)
    name = Column(String(255), nullable=False)
    mobile = Column(String(20))
    email = Column(String(255))
    password_hash = Column(String(255))  # Hashed password
    role = Column(String(50), default="user")  # admin, manager, cco, user, viewer, creator
    rep_type = Column(String(20))  # direct, indirect, both, or NULL (no isolation)
    title = Column(String(100))  # CEO, CFO, COO, CCO, Director Land, etc.
    reports_to = Column(String(20))  # rep_id of manager (org hierarchy)
    status = Column(String(20), default="active")
    created_at = Column(DateTime, server_default=func.now())

class Inventory(Base):
    __tablename__ = "inventory"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    inventory_id = Column(String(20), unique=True, nullable=False)
    project_id = Column(UUID(as_uuid=True), ForeignKey("projects.id"), nullable=False)
    unit_number = Column(String(50), nullable=False)
    unit_type = Column(String(50), default="plot")
    block = Column(String(50))
    area_marla = Column(Numeric(10, 2), nullable=False)
    rate_per_marla = Column(Numeric(15, 2), nullable=False)
    status = Column(String(20), default="available")
    factor_details = Column(Text)
    notes = Column(Text)
    # Vector-specific fields (plot coordinates for map rendering)
    plot_coordinates = Column(JSONB, nullable=True)  # {x, y, width, height}
    plot_offset = Column(JSONB, nullable=True)  # {offset_x, offset_y, rotation}
    is_manual_plot = Column(String(10), default="false")  # "true" or "false" as string for compatibility
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class Transaction(Base):
    __tablename__ = "transactions"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    transaction_id = Column(String(20), unique=True, nullable=False)
    customer_id = Column(UUID(as_uuid=True), ForeignKey("customers.id"), nullable=False)
    broker_id = Column(UUID(as_uuid=True), ForeignKey("brokers.id"))
    project_id = Column(UUID(as_uuid=True), ForeignKey("projects.id"), nullable=False)
    inventory_id = Column(UUID(as_uuid=True), ForeignKey("inventory.id"), nullable=False)
    company_rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    broker_commission_rate = Column(Numeric(5, 2), default=2.00)
    unit_number = Column(String(50), nullable=False)
    block = Column(String(50))
    area_marla = Column(Numeric(10, 2), nullable=False)
    rate_per_marla = Column(Numeric(15, 2), nullable=False)
    total_value = Column(Numeric(15, 2), nullable=False)
    installment_cycle = Column(String(20), default="bi-annual")
    num_installments = Column(Integer, default=4)
    first_due_date = Column(Date, nullable=False)
    status = Column(String(20), default="active")
    notes = Column(Text)
    booking_date = Column(Date, default=date.today)
    payment_plan_version_id = Column(UUID(as_uuid=True), ForeignKey("project_payment_plan_versions.id", ondelete="SET NULL"))
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class Installment(Base):
    __tablename__ = "installments"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    transaction_id = Column(UUID(as_uuid=True), ForeignKey("transactions.id"), nullable=False)
    installment_number = Column(Integer, nullable=False)
    due_date = Column(Date, nullable=False)
    amount = Column(Numeric(15, 2), nullable=False)
    amount_paid = Column(Numeric(15, 2), default=0)
    status = Column(String(20), default="pending")
    notes = Column(Text)
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class Interaction(Base):
    __tablename__ = "interactions"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    interaction_id = Column(String(20), unique=True, nullable=False)
    company_rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"), nullable=False)
    customer_id = Column(UUID(as_uuid=True), ForeignKey("customers.id"))
    broker_id = Column(UUID(as_uuid=True), ForeignKey("brokers.id"))
    interaction_type = Column(String(20), nullable=False)  # call, message, whatsapp
    status = Column(String(50))
    notes = Column(Text)
    next_follow_up = Column(Date)
    lead_id = Column(UUID(as_uuid=True), ForeignKey("leads.id", ondelete="SET NULL"))
    created_at = Column(DateTime, server_default=func.now())

class Campaign(Base):
    __tablename__ = "campaigns"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    campaign_id = Column(String(20), unique=True, nullable=False)
    name = Column(String(255), nullable=False)
    source = Column(String(50))  # facebook, google, instagram, etc
    start_date = Column(Date)
    end_date = Column(Date)
    budget = Column(Numeric(15, 2))
    status = Column(String(20), default="active")
    notes = Column(Text)
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class Lead(Base):
    __tablename__ = "leads"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    lead_id = Column(String(20), unique=True, nullable=False)
    campaign_id = Column(UUID(as_uuid=True), ForeignKey("campaigns.id"))
    assigned_rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    name = Column(String(255), nullable=False)
    mobile = Column(String(20))
    email = Column(String(255))
    source_details = Column(Text)
    status = Column(String(50), default="new")  # new, contacted, qualified, converted, lost
    lead_type = Column(String(20), default="prospect")  # prospect, customer, broker
    converted_customer_id = Column(UUID(as_uuid=True), ForeignKey("customers.id"))
    converted_broker_id = Column(UUID(as_uuid=True), ForeignKey("brokers.id"))
    notes = Column(Text)
    pipeline_stage = Column(String(100), default="New")
    last_contacted_at = Column(DateTime)
    is_stale = Column(Boolean, default=False)
    additional_mobiles = Column(JSONB, default=list)
    source = Column(String(100))
    source_other = Column(Text)
    occupation = Column(String(100))
    occupation_other = Column(Text)
    interested_project_id = Column(UUID(as_uuid=True), ForeignKey("projects.id"))
    interested_project_other = Column(Text)
    area = Column(Text)
    city = Column(Text)
    country_code = Column(String(5), default="+92")
    lead_metadata = Column("metadata", JSONB, default=dict)
    temperature = Column(String(10))  # hot | mild | cold
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class Receipt(Base):
    __tablename__ = "receipts"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    receipt_id = Column(String(20), unique=True, nullable=False)
    customer_id = Column(UUID(as_uuid=True), ForeignKey("customers.id"), nullable=False)
    transaction_id = Column(UUID(as_uuid=True), ForeignKey("transactions.id"))
    amount = Column(Numeric(15, 2), nullable=False)
    payment_method = Column(String(50))  # cash, cheque, bank_transfer, online
    reference_number = Column(String(100))  # cheque number, transaction id etc
    payment_date = Column(Date, default=date.today)
    notes = Column(Text)
    created_by_rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    created_at = Column(DateTime, server_default=func.now())

class ReceiptAllocation(Base):
    __tablename__ = "receipt_allocations"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    receipt_id = Column(UUID(as_uuid=True), ForeignKey("receipts.id"), nullable=False)
    installment_id = Column(UUID(as_uuid=True), ForeignKey("installments.id"), nullable=False)
    amount = Column(Numeric(15, 2), nullable=False)
    created_at = Column(DateTime, server_default=func.now())

class EOICollection(Base):
    __tablename__ = "eoi_collections"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    eoi_id = Column(String(20), unique=True, nullable=False)
    project_id = Column(UUID(as_uuid=True), ForeignKey("projects.id"), nullable=False)
    customer_id = Column(UUID(as_uuid=True), ForeignKey("customers.id"))
    broker_id = Column(UUID(as_uuid=True), ForeignKey("brokers.id"))
    party_name = Column(String(255), nullable=False)
    party_mobile = Column(String(20))
    party_cnic = Column(String(20))
    broker_name = Column(String(255))
    amount = Column(Numeric(15, 2), nullable=False)
    marlas = Column(Numeric(8, 2))
    unit_number = Column(String(50))
    inventory_id = Column(UUID(as_uuid=True), ForeignKey("inventory.id"))
    payment_method = Column(String(50))
    reference_number = Column(String(100))
    payment_received = Column(Boolean, default=False, server_default="false")
    payment_received_at = Column(DateTime)
    eoi_date = Column(Date, default=date.today, nullable=False)
    notes = Column(Text)
    status = Column(String(20), default="active", nullable=False)  # active|converted|cancelled|refunded
    converted_transaction_id = Column(UUID(as_uuid=True), ForeignKey("transactions.id"))
    created_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class ZakatBeneficiary(Base):
    __tablename__ = "zakat_beneficiaries"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    beneficiary_id = Column(String(20), unique=True, nullable=False)
    name = Column(String(255), nullable=False)
    cnic = Column(String(20))
    mobile = Column(String(20))
    address = Column(Text)
    notes = Column(Text)
    status = Column(String(20), default="active")
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class ZakatRecord(Base):
    __tablename__ = "zakat_records"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    zakat_id = Column(String(20), unique=True, nullable=False)
    beneficiary_id = Column(UUID(as_uuid=True), ForeignKey("zakat_beneficiaries.id"))
    beneficiary_name = Column(String(255), nullable=False)
    beneficiary_cnic = Column(String(20))
    beneficiary_mobile = Column(String(20))
    beneficiary_address = Column(Text)
    amount = Column(Numeric(15, 2), nullable=False)
    category = Column(String(50), nullable=False)
    purpose = Column(Text)
    approval_reference = Column(Text)
    approved_by = Column(String(255))
    payment_method = Column(String(50))
    reference_number = Column(String(100))
    receipt_number = Column(String(100))
    disbursed_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    disbursement_date = Column(Date)
    status = Column(String(20), default="pending")
    approval_status = Column(String(20), default="pending")  # pending|approved|rejected
    approved_amount = Column(Numeric(15, 2))
    approved_by_rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    approved_at = Column(DateTime)
    approval_decision = Column(String(20))  # full|partial|rejected
    approval_notes = Column(Text)
    case_status = Column(String(20), default="pending")  # pending|open|closed
    disbursement_approval_status = Column(String(20), default="pending")  # pending|approved|postponed
    disbursement_approved_by_rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    disbursement_approved_at = Column(DateTime)
    disbursement_approval_notes = Column(Text)
    notes = Column(Text)
    payment_id = Column(UUID(as_uuid=True), ForeignKey("payments.id"))
    created_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class ZakatDisbursement(Base):
    __tablename__ = "zakat_disbursements"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    zakat_record_id = Column(UUID(as_uuid=True), ForeignKey("zakat_records.id"), nullable=False)
    amount = Column(Numeric(15, 2), nullable=False)
    disbursed_by_rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    payment_id = Column(UUID(as_uuid=True), ForeignKey("payments.id"))
    payment_method = Column(String(50))
    reference_number = Column(String(100))
    receipt_number = Column(String(100))
    notes = Column(Text)
    disbursement_date = Column(Date)
    disbursed_at = Column(DateTime, server_default=func.now())
    created_at = Column(DateTime, server_default=func.now())

class Creditor(Base):
    __tablename__ = "creditors"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    creditor_id = Column(String(20), unique=True, nullable=False)
    name = Column(String(255), nullable=False)
    mobile = Column(String(20))
    company = Column(String(255))
    bank_name = Column(String(100))
    bank_account = Column(String(50))
    bank_iban = Column(String(50))
    notes = Column(Text)
    status = Column(String(20), default="active")
    created_at = Column(DateTime, server_default=func.now())

class Payment(Base):
    __tablename__ = "payments"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    payment_id = Column(String(20), unique=True, nullable=False)
    payment_type = Column(String(50), nullable=False)  # broker_commission, rep_incentive, creditor, other
    payee_type = Column(String(50))  # broker, company_rep, creditor
    broker_id = Column(UUID(as_uuid=True), ForeignKey("brokers.id"))
    company_rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    creditor_id = Column(UUID(as_uuid=True), ForeignKey("creditors.id"))
    transaction_id = Column(UUID(as_uuid=True), ForeignKey("transactions.id"))
    amount = Column(Numeric(15, 2), nullable=False)
    payment_method = Column(String(50))  # cash, cheque, bank_transfer
    reference_number = Column(String(100))
    payment_date = Column(Date, default=date.today)
    notes = Column(Text)
    approved_by_rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    status = Column(String(20), default="completed")  # pending, completed, cancelled
    created_at = Column(DateTime, server_default=func.now())

class PaymentAllocation(Base):
    __tablename__ = "payment_allocations"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    payment_id = Column(UUID(as_uuid=True), ForeignKey("payments.id"), nullable=False)
    transaction_id = Column(UUID(as_uuid=True), ForeignKey("transactions.id"), nullable=False)
    amount = Column(Numeric(15, 2), nullable=False)
    created_at = Column(DateTime, server_default=func.now())

class MediaFile(Base):
    __tablename__ = "media_files"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    file_id = Column(String(20), unique=True, nullable=False)
    entity_type = Column(String(50), nullable=False)  # transaction, interaction, project, receipt, payment
    entity_id = Column(UUID(as_uuid=True), nullable=False)
    file_name = Column(String(255), nullable=False)
    file_path = Column(String(500), nullable=False)
    file_type = Column(String(50))  # pdf, excel, image, audio, video, other
    file_size = Column(Integer)
    uploaded_by_rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    description = Column(Text)
    created_at = Column(DateTime, server_default=func.now())

# ============================================
# DELETION REQUESTS (Soft-delete approval workflow)
# ============================================
class DeletionRequest(Base):
    __tablename__ = "deletion_requests"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    request_id = Column(String(20), unique=True, nullable=False)
    entity_type = Column(String(50), nullable=False)
    entity_id = Column(UUID(as_uuid=True), nullable=False)
    entity_name = Column(String(255))
    requested_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    requested_at = Column(DateTime, server_default=func.now())
    reason = Column(Text)
    status = Column(String(20), default="pending")
    reviewed_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    reviewed_at = Column(DateTime)
    rejection_reason = Column(Text)

# ============================================
# NOTIFICATION & PIPELINE MODELS
# ============================================
class Notification(Base):
    __tablename__ = "notifications"
    id = Column(Integer, primary_key=True)
    notification_id = Column(String(20), unique=True)
    user_rep_id = Column(String(20), nullable=False)
    title = Column(String(200), nullable=False)
    message = Column(Text)
    type = Column(String(50), default="info")
    category = Column(String(50))
    entity_type = Column(String(50))
    entity_id = Column(String(50))
    is_read = Column(Boolean, default=False)
    data = Column(JSONB, default=dict)
    created_at = Column(DateTime, server_default=func.now())
    read_at = Column(DateTime)

class SearchLog(Base):
    __tablename__ = "search_log"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    searcher_rep_id = Column(String(20), nullable=False)
    search_query = Column(String(255), nullable=False)
    search_type = Column(String(50))
    matched_entity_type = Column(String(50))
    matched_entity_id = Column(String(50))
    matched_entity_name = Column(String(255))
    owner_rep_id = Column(String(20))
    created_at = Column(DateTime, server_default=func.now())

class LeadAssignmentRequest(Base):
    __tablename__ = "lead_assignment_requests"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    request_id = Column(String(20), unique=True)
    lead_id = Column(UUID(as_uuid=True), ForeignKey("leads.id"), nullable=False)
    requested_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id", ondelete="SET NULL"))
    reason = Column(Text)
    status = Column(String(20), default="pending")
    reviewed_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    reviewed_at = Column(DateTime)
    created_at = Column(DateTime, server_default=func.now())

class PipelineStage(Base):
    __tablename__ = "pipeline_stages"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    name = Column(String(100), nullable=False, unique=True)
    display_order = Column(Integer, nullable=False)
    color = Column(String(20), default="#6B7280")
    is_terminal = Column(Boolean, default=False)
    created_at = Column(DateTime, server_default=func.now())

class LookupValue(Base):
    __tablename__ = "lookup_values"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    category = Column(String(50), nullable=False)
    label = Column(String(255), nullable=False)
    sort_order = Column(Integer, nullable=False, default=99)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, server_default=func.now())

# ============================================
# VECTOR MODELS
# ============================================
class VectorProject(Base):
    __tablename__ = "vector_projects"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    name = Column(String(255), nullable=False)
    map_name = Column(String(255), nullable=True)
    map_pdf_base64 = Column(Text, nullable=True)
    map_size = Column(JSONB, nullable=True)  # {width, height}
    linked_project_id = Column(UUID(as_uuid=True), ForeignKey("projects.id"), nullable=True)  # Optional link to Radius project
    vector_metadata = Column(JSONB, nullable=True)  # Vector-specific settings
    system_branches = Column(JSONB, nullable=True)  # System-generated branches (sales, inventory) as JSON
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class VectorAnnotation(Base):
    __tablename__ = "vector_annotations"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    project_id = Column(UUID(as_uuid=True), ForeignKey("vector_projects.id", ondelete="CASCADE"), nullable=False)
    annotation_id = Column(String(100), nullable=False)  # Changed from INTEGER to VARCHAR to support timestamp-based IDs
    note = Column(Text, nullable=True)
    category = Column(String(100), nullable=True)
    color = Column(String(50), nullable=True)
    font_size = Column(Integer, nullable=True)
    rotation = Column(Integer, default=0)
    plot_ids = Column(JSONB, nullable=True)  # Array of plot IDs
    plot_nums = Column(JSONB, nullable=True)  # Array of plot numbers
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class VectorShape(Base):
    __tablename__ = "vector_shapes"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    project_id = Column(UUID(as_uuid=True), ForeignKey("vector_projects.id", ondelete="CASCADE"), nullable=False)
    shape_id = Column(String(100), nullable=False)
    type = Column(String(50), nullable=True)
    x = Column(Integer, nullable=True)
    y = Column(Integer, nullable=True)
    width = Column(Integer, nullable=True)
    height = Column(Integer, nullable=True)
    color = Column(String(50), nullable=True)
    data = Column(JSONB, nullable=True)
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class VectorLabel(Base):
    __tablename__ = "vector_labels"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    project_id = Column(UUID(as_uuid=True), ForeignKey("vector_projects.id", ondelete="CASCADE"), nullable=False)
    label_id = Column(String(100), nullable=False)
    text = Column(Text, nullable=True)
    x = Column(Integer, nullable=True)
    y = Column(Integer, nullable=True)
    size = Column(Integer, nullable=True)
    color = Column(String(50), nullable=True)
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class VectorLegend(Base):
    __tablename__ = "vector_legend"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    project_id = Column(UUID(as_uuid=True), ForeignKey("vector_projects.id", ondelete="CASCADE"), nullable=False, unique=True)
    visible = Column(String(10), default="true")  # "true" or "false" as string
    minimized = Column(String(10), default="false")  # "true" or "false" as string
    position = Column(String(50), nullable=True)
    manual_entries = Column(JSONB, nullable=True)
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class VectorBranch(Base):
    __tablename__ = "vector_branches"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    project_id = Column(UUID(as_uuid=True), ForeignKey("vector_projects.id", ondelete="CASCADE"), nullable=False)
    name = Column(String(255), nullable=False)
    timestamp = Column(DateTime, nullable=True)
    anno_count = Column(Integer, nullable=True)
    data = Column(JSONB, nullable=True)  # {annos, labels, shapes, plotOffsets, plotRotations}
    created_at = Column(DateTime, server_default=func.now())

class VectorCreatorNote(Base):
    __tablename__ = "vector_creator_notes"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    project_id = Column(UUID(as_uuid=True), ForeignKey("vector_projects.id", ondelete="CASCADE"), nullable=False)
    timestamp = Column(DateTime, server_default=func.now())
    note = Column(Text, nullable=False)
    created_at = Column(DateTime, server_default=func.now())

class VectorChangeLog(Base):
    __tablename__ = "vector_change_log"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    project_id = Column(UUID(as_uuid=True), ForeignKey("vector_projects.id", ondelete="CASCADE"), nullable=False)
    timestamp = Column(DateTime, server_default=func.now())
    action = Column(Text, nullable=True)
    details = Column(JSONB, nullable=True)
    created_at = Column(DateTime, server_default=func.now())

class VectorProjectBackup(Base):
    __tablename__ = "vector_project_backups"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    project_id = Column(UUID(as_uuid=True), ForeignKey("vector_projects.id", ondelete="CASCADE"), nullable=False)
    backup_data = Column(JSONB, nullable=False)
    created_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id", ondelete="SET NULL"), nullable=True)
    created_at = Column(DateTime, server_default=func.now())

class VectorBackupSettings(Base):
    __tablename__ = "vector_backup_settings"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    frequency = Column(String(100), default="0 2 * * *")  # Default: Daily at 2 AM (cron format)
    enabled = Column(String(10), default="true")  # "true" or "false" as string
    last_run = Column(DateTime, nullable=True)
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class VectorReconciliation(Base):
    __tablename__ = "vector_reconciliation"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    project_id = Column(UUID(as_uuid=True), ForeignKey("vector_projects.id", ondelete="CASCADE"), nullable=False)
    linked_project_id = Column(UUID(as_uuid=True), ForeignKey("projects.id", ondelete="CASCADE"), nullable=True)
    sync_status = Column(String(50), nullable=True)  # linked, synced, needs-reconciliation
    discrepancies = Column(JSONB, nullable=True)  # Store reconciliation discrepancies
    last_sync_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

# ============================================
# TASK & VOICE MODELS
# ============================================
class Task(Base):
    __tablename__ = "tasks"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    task_id = Column(String(20), unique=True)
    title = Column(String(300), nullable=False)
    description = Column(Text)
    task_type = Column(String(30), default="general")
    department = Column(String(30))
    priority = Column(String(10), default="medium")
    status = Column(String(30), default="pending")
    assignee_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    created_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"), nullable=False)
    delegated_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    parent_task_id = Column(UUID(as_uuid=True), ForeignKey("tasks.id"))
    due_date = Column(Date)
    completed_at = Column(DateTime)
    completion_notes = Column(Text)
    pending_assignment = Column(Boolean, default=False)
    original_assignee_text = Column(String(200))
    linked_inventory_id = Column(UUID(as_uuid=True), ForeignKey("inventory.id"))
    linked_transaction_id = Column(UUID(as_uuid=True), ForeignKey("transactions.id"))
    linked_customer_id = Column(UUID(as_uuid=True), ForeignKey("customers.id"))
    linked_project_id = Column(UUID(as_uuid=True), ForeignKey("projects.id"))
    collaborator_ids = Column(JSONB, default=list)  # Additional assignee UUIDs
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now())

class TaskComment(Base):
    __tablename__ = "task_comments"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    task_id = Column(UUID(as_uuid=True), ForeignKey("tasks.id", ondelete="CASCADE"), nullable=False)
    micro_task_id = Column(UUID(as_uuid=True), ForeignKey("micro_tasks.id", ondelete="CASCADE"))
    author_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"), nullable=False)
    content = Column(Text, nullable=False)
    created_at = Column(DateTime, server_default=func.now())

class TaskActivity(Base):
    __tablename__ = "task_activities"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    task_id = Column(UUID(as_uuid=True), ForeignKey("tasks.id", ondelete="CASCADE"), nullable=False)
    actor_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"), nullable=False)
    action = Column(String(50), nullable=False)
    old_value = Column(String(200))
    new_value = Column(String(200))
    details = Column(JSONB, default=dict)
    created_at = Column(DateTime, server_default=func.now())

class MicroTask(Base):
    __tablename__ = "micro_tasks"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    task_id = Column(UUID(as_uuid=True), ForeignKey("tasks.id", ondelete="CASCADE"), nullable=False)
    title = Column(String(300), nullable=False)
    is_completed = Column(Boolean, nullable=False, server_default=text("false"))
    assignee_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id", ondelete="SET NULL"))
    due_date = Column(Date)
    sort_order = Column(Integer, nullable=False, server_default=text("0"))
    created_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"), nullable=False)
    completed_at = Column(DateTime(timezone=True))
    completed_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id"))
    created_at = Column(DateTime(timezone=True), server_default=func.now())
    updated_at = Column(DateTime(timezone=True), server_default=func.now())

class QueryHistory(Base):
    __tablename__ = "query_history"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    query_text = Column(Text, nullable=False)
    intent = Column(String(30))
    domain = Column(String(30))
    confidence = Column(Numeric(3, 2), default=0.0)
    sql_query = Column(Text)
    response_text = Column(Text)
    success = Column(Boolean, default=False)
    usage_count = Column(Integer, default=1)
    is_learned = Column(Boolean, default=False)
    feedback_score = Column(Numeric(3, 2), default=0.0)
    error_message = Column(Text)
    processing_time_ms = Column(Numeric(10, 2))
    corrected_intent = Column(String(30))
    corrected_domain = Column(String(30))
    corrected_response = Column(Text)
    is_approved = Column(Boolean, default=False)
    user_rep_id = Column(String(20))
    created_at = Column(DateTime, server_default=func.now())
    last_used_at = Column(DateTime, server_default=func.now())

class QueryFeedback(Base):
    __tablename__ = "query_feedback"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    query_history_id = Column(UUID(as_uuid=True), ForeignKey("query_history.id"), nullable=False)
    feedback_type = Column(String(20), nullable=False)
    rating = Column(Integer)
    correct_intent = Column(String(30))
    correct_domain = Column(String(30))
    correct_response = Column(Text)
    user_comment = Column(Text)
    created_at = Column(DateTime, server_default=func.now())

# ============================================
# PAYMENT PLANS, TARGETS & TEMPERATURE MODELS
# ============================================

class ProjectPaymentPlan(Base):
    __tablename__ = "project_payment_plans"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    plan_id = Column(String(20), unique=True, nullable=False)           # PPL-XXXX
    project_id = Column(UUID(as_uuid=True), ForeignKey("projects.id", ondelete="CASCADE"), nullable=False)
    name = Column(String(255), nullable=False)
    description = Column(Text)
    is_default = Column(Boolean, default=False, nullable=False)
    is_locked = Column(Boolean, default=False, nullable=False)
    locked_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id", ondelete="SET NULL"))
    locked_at = Column(DateTime(timezone=True))
    lock_reason = Column(Text)
    status = Column(String(20), default="active", nullable=False)       # active | archived
    created_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id", ondelete="SET NULL"))
    created_at = Column(DateTime(timezone=True), server_default=func.now())
    updated_at = Column(DateTime(timezone=True), server_default=func.now())

class ProjectPaymentPlanVersion(Base):
    __tablename__ = "project_payment_plan_versions"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    plan_id = Column(UUID(as_uuid=True), ForeignKey("project_payment_plans.id", ondelete="CASCADE"), nullable=False)
    version_number = Column(Integer, default=1, nullable=False)
    is_active = Column(Boolean, default=True, nullable=False)
    installments = Column(JSONB, default=[], nullable=False)            # [{number, label, percentage, month_offset}]
    num_installments = Column(Integer, default=4, nullable=False)
    installment_cycle = Column(String(20), default="bi-annual")         # hint field
    notes = Column(Text)
    created_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id", ondelete="SET NULL"))
    created_at = Column(DateTime(timezone=True), server_default=func.now())

class ProjectInventoryPaymentOverride(Base):
    __tablename__ = "project_inventory_payment_overrides"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    inventory_id = Column(UUID(as_uuid=True), ForeignKey("inventory.id", ondelete="CASCADE"), unique=True, nullable=False)
    override_type = Column(String(20), default="plan_version", nullable=False)  # plan_version | custom
    plan_version_id = Column(UUID(as_uuid=True), ForeignKey("project_payment_plan_versions.id", ondelete="SET NULL"))
    custom_installments = Column(JSONB)                                 # same format as version installments
    custom_num_installments = Column(Integer)
    custom_installment_cycle = Column(String(20))
    notes = Column(Text)
    created_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id", ondelete="SET NULL"))
    created_at = Column(DateTime(timezone=True), server_default=func.now())
    updated_at = Column(DateTime(timezone=True), server_default=func.now())

class MonthlyRepTarget(Base):
    __tablename__ = "monthly_rep_targets"
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    rep_id = Column(UUID(as_uuid=True), ForeignKey("company_reps.id", ondelete="CASCADE"), nullable=False)
    target_month = Column(Integer, nullable=False)
    target_year = Column(Integer, nullable=False)
    revenue_target = Column(Numeric(15, 2), default=0, nullable=False)
    transaction_target = Column(Integer, default=0, nullable=False)
    lead_target = Column(Integer, default=0, nullable=False)
    notes = Column(Text)
    created_by = Column(UUID(as_uuid=True), ForeignKey("company_reps.id", ondelete="SET NULL"))
    created_at = Column(DateTime(timezone=True), server_default=func.now())
    updated_at = Column(DateTime(timezone=True), server_default=func.now())

# ============================================
# HELPER FUNCTIONS
# ============================================
def sync_lead_id_sequence(db):
    """Sync lead_id_seq to MAX existing lead_id. Prevents UniqueViolation after bulk imports."""
    try:
        result = db.execute(text(
            "SELECT setval('lead_id_seq', COALESCE((SELECT MAX(CAST(SUBSTRING(lead_id FROM 6) AS INTEGER)) FROM leads), 1))"
        ))
        new_val = result.scalar()
        print(f"[sequence-sync] lead_id_seq synced to {new_val}")
        return new_val
    except Exception as e:
        print(f"[sequence-sync] Failed to sync lead_id_seq: {e}")
        return None

def normalize_mobile(mobile):
    """Normalize Pakistan mobile number for duplicate detection.
    All variants collapse to canonical 03XXXXXXXXX form:
      03001239856 → 03001239856
      3001239856  → 03001239856
      +923001239856 → 03001239856
      923001239856  → 03001239856
      +92 300-123-9856 → 03001239856
    """
    if not mobile:
        return None
    # Strip everything except digits and leading +
    m = mobile.strip()
    has_plus = m.startswith("+")
    m = "".join(c for c in m if c.isdigit())
    if not m:
        return None
    # +92... or 92... prefix → strip to local
    if has_plus and m.startswith("92"):
        m = m[2:]
    elif m.startswith("92") and len(m) > 10:
        m = m[2:]
    # Ensure leading 0 for Pakistani mobiles (3xx → 03xx)
    if len(m) == 10 and m[0] == "3":
        m = "0" + m
    return m

def create_notification(db, user_rep_id, notif_type, title, message, category=None, entity_type=None, entity_id=None, data=None):
    """Helper to create a notification record"""
    n = Notification(
        user_rep_id=user_rep_id,
        type=notif_type,
        title=title,
        message=message,
        category=category or notif_type,
        entity_type=entity_type,
        entity_id=str(entity_id) if entity_id else None,
        data=data or {}
    )
    db.add(n)
    return n

def _is_customer_sync_approver(user: "CompanyRep") -> bool:
    if not user:
        return False
    if user.role in ["admin", "cco", "director", "coo"]:
        return True
    return user.rep_id in CUSTOMER_SYNC_APPROVER_REP_IDS

def _ensure_lead_sync_requests_table(db: Session):
    """Create lead_sync_requests table lazily if migration hasn't been applied yet."""
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS lead_sync_requests (
            id UUID PRIMARY KEY DEFAULT uuid_generate_v4(),
            request_id VARCHAR(20) UNIQUE NOT NULL,
            lead_id UUID NOT NULL REFERENCES leads(id) ON DELETE CASCADE,
            requested_by UUID NOT NULL REFERENCES company_reps(id) ON DELETE CASCADE,
            reason TEXT,
            status VARCHAR(20) NOT NULL DEFAULT 'pending',
            reviewed_by UUID REFERENCES company_reps(id) ON DELETE SET NULL,
            reviewed_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_lead_sync_requests_lead_id ON lead_sync_requests(lead_id);
        CREATE INDEX IF NOT EXISTS idx_lead_sync_requests_status ON lead_sync_requests(status);
    """))
    db.commit()

def _sync_lead_to_customer_record(lead: "Lead", db: Session):
    """Idempotently sync a lead to customer DB and return (linked_existing, entity_id, customer_uuid_str)."""
    existing = None
    if lead.mobile:
        normalized = normalize_mobile(lead.mobile)
        if normalized:
            patterns = [normalized]
            if normalized.startswith("0"):
                patterns.append("+92" + normalized[1:])
                patterns.append("92" + normalized[1:])
            existing = db.query(Customer).filter(or_(*[Customer.mobile.ilike(p) for p in patterns])).first()

    linked_existing = False
    if existing:
        linked_existing = True
        if not existing.assigned_rep_id and lead.assigned_rep_id:
            existing.assigned_rep_id = lead.assigned_rep_id
        lead.converted_customer_id = existing.id
        entity_id = existing.customer_id
        customer_uuid = str(existing.id)
    else:
        c = Customer(
            name=lead.name,
            mobile=lead.mobile or f"lead-{lead.lead_id}",
            email=lead.email,
            assigned_rep_id=lead.assigned_rep_id
        )
        db.add(c); db.flush()
        lead.converted_customer_id = c.id
        entity_id = c.customer_id
        customer_uuid = str(c.id)

    lead.lead_type = "customer"
    db.query(Interaction).filter(
        Interaction.lead_id == lead.id,
        Interaction.customer_id.is_(None)
    ).update({"customer_id": lead.converted_customer_id}, synchronize_session="fetch")

    return linked_existing, entity_id, customer_uuid

def notify_duplicate_attempt(db, duplicates, attempted_by_user, action_type, campaign_name=None):
    """Send notifications when duplicate lead addition is attempted.
    action_type: 'single' = notify original rep + admin/CCO, 'bulk' = admin/CCO only.
    """
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    campaign_info = f" to campaign '{campaign_name}'" if campaign_name else ""

    # For single add: notify original lead owner
    if action_type == "single":
        notified_reps = set()
        for dup in duplicates:
            if dup["type"] == "lead" and dup.get("assigned_rep_id"):
                rep_id = dup["assigned_rep_id"]
                if rep_id not in notified_reps and rep_id != attempted_by_user.rep_id:
                    notified_reps.add(rep_id)
                    create_notification(
                        db, rep_id, "duplicate_attempt",
                        "Duplicate Lead Attempt",
                        f"{attempted_by_user.name} tried adding your lead {dup['name']} ({dup['entity_id']}){campaign_info} on {timestamp}",
                        category="lead", entity_type="lead", entity_id=dup["entity_id"],
                        data={"attempted_by": attempted_by_user.rep_id, "attempted_by_name": attempted_by_user.name,
                              "mobile": dup["mobile"], "campaign": campaign_name}
                    )

    # Notify all admin/CCO users
    admins = db.query(CompanyRep).filter(
        CompanyRep.role.in_(["admin", "cco"]), CompanyRep.status == "active"
    ).all()
    for admin in admins:
        if admin.rep_id == attempted_by_user.rep_id:
            continue
        if action_type == "single":
            dup = duplicates[0]
            create_notification(
                db, admin.rep_id, "duplicate_attempt",
                "Duplicate Lead Blocked",
                f"{attempted_by_user.name} tried adding duplicate lead (matches {dup['type']} {dup['entity_id']} — {dup['name']}){campaign_info} on {timestamp}",
                category="admin", entity_type=dup["type"], entity_id=dup["entity_id"],
                data={"attempted_by": attempted_by_user.rep_id, "attempted_by_name": attempted_by_user.name,
                      "mobile": dup["mobile"], "campaign": campaign_name}
            )
        else:  # bulk
            create_notification(
                db, admin.rep_id, "duplicate_attempt",
                "Bulk Import — Duplicates Found",
                f"{attempted_by_user.name} bulk imported leads with {len(duplicates)} duplicate(s) detected{campaign_info} on {timestamp}",
                category="admin", entity_type="lead", entity_id=None,
                data={"attempted_by": attempted_by_user.rep_id, "attempted_by_name": attempted_by_user.name,
                      "duplicate_count": len(duplicates), "campaign": campaign_name}
            )

def _generate_eoi_code(db: Session) -> str:
    """Generate EOI-XXXXX with collision retry safety."""
    count = db.query(func.count(EOICollection.id)).scalar() or 0
    candidate = f"EOI-{str(count + 1).zfill(5)}"
    while db.query(EOICollection).filter(EOICollection.eoi_id == candidate).first():
        count += 1
        candidate = f"EOI-{str(count + 1).zfill(5)}"
    return candidate

def _resolve_eoi_project(db: Session, project_id: str):
    try:
        pid = uuid.UUID(str(project_id))
        return db.query(Project).filter(
            (Project.id == pid) | (Project.project_id == project_id)
        ).first()
    except ValueError:
        return db.query(Project).filter(Project.project_id == project_id).first()

def _resolve_eoi_broker(db: Session, broker_id: Optional[str]):
    if not broker_id:
        return None
    try:
        bid = uuid.UUID(str(broker_id))
        return db.query(Broker).filter(
            (Broker.id == bid) | (Broker.broker_id == broker_id)
        ).first()
    except ValueError:
        return db.query(Broker).filter(Broker.broker_id == broker_id).first()

def _resolve_eoi_customer(db: Session, customer_id: Optional[str]):
    if not customer_id:
        return None
    try:
        cid = uuid.UUID(str(customer_id))
        return db.query(Customer).filter(
            (Customer.id == cid) | (Customer.customer_id == customer_id) | (Customer.mobile == customer_id)
        ).first()
    except ValueError:
        return db.query(Customer).filter(
            (Customer.customer_id == customer_id) | (Customer.mobile == customer_id)
        ).first()

def _resolve_eoi_inventory(db: Session, inventory_id: Optional[str]):
    if not inventory_id:
        return None
    try:
        iid = uuid.UUID(str(inventory_id))
        return db.query(Inventory).filter(
            (Inventory.id == iid) | (Inventory.inventory_id == inventory_id)
        ).first()
    except ValueError:
        return db.query(Inventory).filter(Inventory.inventory_id == inventory_id).first()

def _eoi_row_dict(eoi: EOICollection, db: Session):
    project = db.query(Project).filter(Project.id == eoi.project_id).first() if eoi.project_id else None
    customer = db.query(Customer).filter(Customer.id == eoi.customer_id).first() if eoi.customer_id else None
    broker = db.query(Broker).filter(Broker.id == eoi.broker_id).first() if eoi.broker_id else None
    inv = db.query(Inventory).filter(Inventory.id == eoi.inventory_id).first() if eoi.inventory_id else None
    creator = db.query(CompanyRep).filter(CompanyRep.id == eoi.created_by).first() if eoi.created_by else None
    converted_txn = db.query(Transaction).filter(Transaction.id == eoi.converted_transaction_id).first() if eoi.converted_transaction_id else None
    return {
        "id": str(eoi.id),
        "eoi_id": eoi.eoi_id,
        "project_id": project.project_id if project else None,
        "project_uuid": str(eoi.project_id) if eoi.project_id else None,
        "project_name": project.name if project else None,
        "customer_id": customer.customer_id if customer else None,
        "customer_uuid": str(eoi.customer_id) if eoi.customer_id else None,
        "broker_id": broker.broker_id if broker else None,
        "broker_uuid": str(eoi.broker_id) if eoi.broker_id else None,
        "party_name": eoi.party_name,
        "party_mobile": eoi.party_mobile,
        "party_cnic": eoi.party_cnic,
        "broker_name": broker.name if broker else (eoi.broker_name or "Direct"),
        "amount": float(eoi.amount or 0),
        "marlas": float(eoi.marlas) if eoi.marlas is not None else None,
        "unit_number": eoi.unit_number or (inv.unit_number if inv else None),
        "inventory_id": inv.inventory_id if inv else None,
        "inventory_uuid": str(eoi.inventory_id) if eoi.inventory_id else None,
        "payment_method": eoi.payment_method,
        "reference_number": eoi.reference_number,
        "payment_received": bool(eoi.payment_received) if eoi.payment_received is not None else False,
        "payment_received_at": str(eoi.payment_received_at) if eoi.payment_received_at else None,
        "eoi_date": str(eoi.eoi_date) if eoi.eoi_date else None,
        "notes": eoi.notes,
        "status": eoi.status,
        "converted_transaction_id": converted_txn.transaction_id if converted_txn else None,
        "converted_transaction_uuid": str(eoi.converted_transaction_id) if eoi.converted_transaction_id else None,
        "created_by_rep_id": creator.rep_id if creator else None,
        "created_by_name": creator.name if creator else None,
        "created_at": str(eoi.created_at) if eoi.created_at else None,
        "updated_at": str(eoi.updated_at) if eoi.updated_at else None,
    }

def _resolve_eoi_record(db: Session, eoi_ref: str):
    try:
        eid = uuid.UUID(str(eoi_ref))
        return db.query(EOICollection).filter(
            (EOICollection.id == eid) | (EOICollection.eoi_id == eoi_ref)
        ).first()
    except ValueError:
        return db.query(EOICollection).filter(EOICollection.eoi_id == eoi_ref).first()

def check_duplicate_mobile(db, mobile, exclude_lead_id=None, exclude_customer_id=None, exclude_broker_id=None, include_converted=False):
    """Check if mobile exists in customers, leads, or brokers tables. Returns list of matches."""
    if not mobile:
        return []
    normalized = normalize_mobile(mobile)
    if not normalized:
        return []
    # Use last 7 digits for fuzzy SQL match (avoids dash/format mismatch in stored values)
    suffix = normalized[-7:]
    # Strip non-digits in SQL for reliable matching: regexp_replace(mobile, '[^0-9]', '', 'g')
    stripped_cust = func.regexp_replace(Customer.mobile, '[^0-9]', '', 'g')
    stripped_lead = func.regexp_replace(Lead.mobile, '[^0-9]', '', 'g')
    stripped_broker = func.regexp_replace(Broker.mobile, '[^0-9]', '', 'g')
    matches = []
    # Check customers via SQL
    cust_q = db.query(Customer, CompanyRep).outerjoin(
        CompanyRep, Customer.assigned_rep_id == CompanyRep.id
    ).filter(Customer.mobile.isnot(None), stripped_cust.like(f"%{suffix}%"))
    if exclude_customer_id:
        cust_q = cust_q.filter(Customer.id != exclude_customer_id)
    for c, rep in cust_q.all():
        if normalize_mobile(c.mobile) == normalized:
            matches.append({
                "type": "customer", "id": str(c.id), "entity_id": c.customer_id,
                "name": c.name, "mobile": c.mobile,
                "assigned_rep": rep.name if rep else None,
                "assigned_rep_id": rep.rep_id if rep else None
            })
    # Check leads via SQL
    lead_q = db.query(Lead, CompanyRep).outerjoin(
        CompanyRep, Lead.assigned_rep_id == CompanyRep.id
    ).filter(Lead.mobile.isnot(None), stripped_lead.like(f"%{suffix}%"))
    if not include_converted:
        lead_q = lead_q.filter(Lead.status != "converted")
    if exclude_lead_id:
        lead_q = lead_q.filter(Lead.id != exclude_lead_id)
    for l, rep in lead_q.all():
        if normalize_mobile(l.mobile) == normalized:
            matches.append({
                "type": "lead", "id": str(l.id), "entity_id": l.lead_id,
                "name": l.name, "mobile": l.mobile,
                "assigned_rep": rep.name if rep else None,
                "assigned_rep_id": rep.rep_id if rep else None
            })
    # Check customer additional_mobiles via SQL JSONB
    try:
        add_mob_cust = db.execute(text("""
            SELECT c.id, c.customer_id, c.name, c.mobile, am.elem
            FROM customers c, jsonb_array_elements_text(c.additional_mobiles) AS am(elem)
            WHERE c.additional_mobiles IS NOT NULL AND c.additional_mobiles != '[]'::jsonb
              AND regexp_replace(am.elem, '[^0-9]', '', 'g') LIKE :suffix_pattern
        """), {"suffix_pattern": f"%{suffix}%"})
        for row in add_mob_cust:
            if exclude_customer_id and str(row.id) == str(exclude_customer_id):
                continue
            if normalize_mobile(row.elem) == normalized:
                already = any(m["id"] == str(row.id) and m["type"] == "customer" for m in matches)
                if not already:
                    matches.append({
                        "type": "customer", "id": str(row.id), "entity_id": row.customer_id,
                        "name": row.name, "mobile": row.elem, "primary_mobile": row.mobile
                    })
    except Exception:
        pass  # additional_mobiles column may not exist yet on older DBs

    # Check lead additional_mobiles via SQL JSONB
    try:
        add_mob_lead = db.execute(text("""
            SELECT l.id, l.lead_id, l.name, l.mobile, am.elem, r.name as rep_name, r.rep_id as rep_rep_id
            FROM leads l
            LEFT JOIN company_reps r ON l.assigned_rep_id = r.id
            CROSS JOIN jsonb_array_elements_text(l.additional_mobiles) AS am(elem)
            WHERE l.status != 'converted'
              AND l.additional_mobiles IS NOT NULL AND l.additional_mobiles != '[]'::jsonb
              AND regexp_replace(am.elem, '[^0-9]', '', 'g') LIKE :suffix_pattern
        """), {"suffix_pattern": f"%{suffix}%"})
        for row in add_mob_lead:
            if exclude_lead_id and str(row.id) == str(exclude_lead_id):
                continue
            if normalize_mobile(row.elem) == normalized:
                already = any(m["id"] == str(row.id) and m["type"] == "lead" for m in matches)
                if not already:
                    matches.append({
                        "type": "lead", "id": str(row.id), "entity_id": row.lead_id,
                        "name": row.name, "mobile": row.elem, "primary_mobile": row.mobile,
                        "assigned_rep": row.rep_name, "assigned_rep_id": row.rep_rep_id
                    })
    except Exception:
        pass

    # Check brokers via SQL
    broker_q = db.query(Broker, CompanyRep).outerjoin(
        CompanyRep, Broker.assigned_rep_id == CompanyRep.id
    ).filter(Broker.mobile.isnot(None), stripped_broker.like(f"%{suffix}%"))
    if exclude_broker_id:
        broker_q = broker_q.filter(Broker.id != exclude_broker_id)
    for b, rep in broker_q.all():
        if normalize_mobile(b.mobile) == normalized:
            matches.append({
                "type": "broker", "id": str(b.id), "entity_id": b.broker_id,
                "name": b.name, "mobile": b.mobile,
                "assigned_rep": rep.name if rep else None,
                "assigned_rep_id": rep.rep_id if rep else None
            })
    return matches

def get_rep_isolation_filter(current_user, db):
    """Returns isolation context for the current user. Used by GET list endpoints
    to filter data by rep ownership based on rep_type.

    Returns dict with:
        isolated: bool - True if filtering should apply
        rep_type: str - 'direct', 'indirect', 'both'
        rep_uuid: UUID - current user's UUID
        team_rep_uuids: list[UUID] - UUIDs of user + their direct reports (for managers)
    """
    role = current_user.role
    rep_type = getattr(current_user, 'rep_type', None)

    # Admin, CCO: no isolation
    if role in ("admin", "cco"):
        return {"isolated": False}

    # rep_type NULL or 'none' = non-sales / legacy user, no isolation
    if not rep_type or rep_type == "none":
        return {"isolated": False}

    # Build team: self + direct reports for managers
    team = [current_user.id]
    if role == "manager":
        direct_reports = db.query(CompanyRep).filter(
            CompanyRep.reports_to == current_user.rep_id,
            CompanyRep.status == "active"
        ).all()
        team.extend([r.id for r in direct_reports])

    return {
        "isolated": True,
        "rep_type": rep_type,
        "rep_uuid": current_user.id,
        "team_rep_uuids": team,
    }

# ============================================
# AUTHENTICATION SETUP
# ============================================
# SECRET_KEY for JWT tokens - MUST be set via SECRET_KEY env var in production
# Generate: python -c "import secrets; print(secrets.token_urlsafe(32))"
SECRET_KEY = os.getenv("SECRET_KEY", "orbit-local-dev-key")
if SECRET_KEY in ("your-secret-key-change-in-production-min-32-chars", "changeme", ""):
    logger.critical("INSECURE SECRET_KEY detected! Set SECRET_KEY env var to a strong random value.")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("TOKEN_EXPIRE_MINUTES", "480"))  # 8 hours default

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

def verify_password(plain_password, hashed_password):
    try:
        # Ensure password is a string
        if not isinstance(plain_password, str):
            plain_password = str(plain_password)
        # Truncate password to 72 bytes if necessary (bcrypt limit)
        password_bytes = plain_password.encode('utf-8')
        if len(password_bytes) > 72:
            plain_password = password_bytes[:72].decode('utf-8', errors='ignore')
        return pwd_context.verify(plain_password, hashed_password)
    except Exception as e:
        # Log the error for debugging but still return False for security
        import logging
        logging.error(f"Password verification error: {str(e)}")
        return False

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def sync_vector_branches_from_orbit(project_id, db: Session):
    """Helper function to sync Vector branches when ORBIT data changes"""
    import time as time_module

    # Find Vector project linked to this ORBIT project
    vector_project = db.query(VectorProject).filter(
        VectorProject.linked_project_id == project_id
    ).first()

    if not vector_project:
        return  # No linked Vector project

    # Get Vector plots for matching
    plots = vector_project.vector_metadata.get('plots', []) if vector_project.vector_metadata else []

    # Build plot map
    plot_map = {}
    for p in plots:
        plot_num = p.get('n', '')
        if plot_num:
            plot_map[plot_num] = p.get('id')
            plot_map[plot_num.upper()] = p.get('id')
            plot_map[plot_num.strip().upper()] = p.get('id')

    # Get current inventory and transactions
    inventory_items = db.query(Inventory).filter(Inventory.project_id == project_id).all()
    transactions = db.query(Transaction).filter(Transaction.project_id == project_id).all()

    # Build sold units set
    sold_units = set()
    for txn in transactions:
        if txn.unit_number:
            sold_units.add(txn.unit_number.upper())
    for item in inventory_items:
        if item.status and item.status.lower() == 'sold' and item.unit_number:
            sold_units.add(item.unit_number.upper())

    # Build annotation arrays
    sold_plot_ids = []
    sold_plot_nums = []
    available_plot_ids = []
    available_plot_nums = []

    for item in inventory_items:
        unit_num = item.unit_number
        if not unit_num:
            continue
        plot_id = plot_map.get(unit_num) or plot_map.get(unit_num.upper())
        if plot_id:
            if unit_num.upper() in sold_units:
                sold_plot_ids.append(plot_id)
                sold_plot_nums.append(unit_num)
            else:
                available_plot_ids.append(plot_id)
                available_plot_nums.append(unit_num)

    # Generate annotations
    current_time = int(time_module.time() * 1000)

    sales_annotation = None
    if sold_plot_ids:
        sales_annotation = {
            "id": f"auto_sold_{current_time}",
            "note": "SOLD",
            "cat": "SOLD",
            "color": "#ef4444",
            "plotIds": sold_plot_ids,
            "plotNums": sold_plot_nums,
            "fontSize": 12,
            "rotation": 0,
            "isAutoGenerated": True
        }

    inventory_annotation = None
    if available_plot_ids:
        inventory_annotation = {
            "id": f"auto_inventory_{current_time}",
            "note": "INVENTORY",
            "cat": "INVENTORY",
            "color": "#22c55e",
            "plotIds": available_plot_ids,
            "plotNums": available_plot_nums,
            "fontSize": 12,
            "rotation": 0,
            "isAutoGenerated": True
        }

    # Update system_branches
    vector_project.system_branches = {
        "sales": {
            "annotations": [sales_annotation] if sales_annotation else [],
            "lastSyncAt": datetime.utcnow().isoformat(),
            "transactionCount": len(transactions),
            "plotCount": len(sold_plot_ids)
        },
        "inventory": {
            "annotations": [inventory_annotation] if inventory_annotation else [],
            "lastSyncAt": datetime.utcnow().isoformat(),
            "unitCount": len(inventory_items),
            "plotCount": len(available_plot_ids)
        }
    }
    vector_project.updated_at = func.now()

    # Update reconciliation
    recon = db.query(VectorReconciliation).filter(
        VectorReconciliation.project_id == vector_project.id
    ).first()
    if recon:
        recon.last_sync_at = datetime.utcnow()
        recon.sync_status = "synced"

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    try:
        token = credentials.credentials
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            user_id: str = payload.get("sub")
            if user_id is None:
                raise HTTPException(status_code=401, detail="Invalid authentication credentials")
        except JWTError as e:
            print(f"JWT Error in get_current_user: {str(e)}")
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
        
        try:
            user = db.query(CompanyRep).filter((CompanyRep.id == user_id) | (CompanyRep.rep_id == user_id)).first()
            if user is None:
                print(f"User not found for ID: {user_id}")
                raise HTTPException(status_code=401, detail="User not found")
            if user.status != "active":
                print(f"User {user_id} is not active, status: {user.status}")
                raise HTTPException(status_code=401, detail="User not active")
            return user
        except Exception as e:
            print(f"Database error in get_current_user: {str(e)}")
            import traceback
            print(traceback.format_exc())
            raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except HTTPException:
        raise
    except Exception as e:
        print(f"Unexpected error in get_current_user: {str(e)}")
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Authentication error: {str(e)}")

def require_role(allowed_roles: List[str]):
    def role_checker(current_user: CompanyRep = Depends(get_current_user)):
        if current_user.role not in allowed_roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return current_user
    return role_checker

# ============================================
# VECTOR UTILITY FUNCTIONS
# ============================================
def normalize_plot_number(plot_num: str) -> str:
    """Normalize plot/unit number for matching"""
    if not plot_num:
        return ""
    # Remove spaces, convert to uppercase
    normalized = plot_num.strip().upper()
    # Remove common separators and replace with standard format
    normalized = normalized.replace("-", "").replace("_", "").replace(".", "").replace(" ", "")
    return normalized

def extract_numeric_parts(text: str) -> List[int]:
    """Extract numeric parts from text"""
    import re
    return [int(x) for x in re.findall(r'\d+', text)]

def levenshtein_distance(s1: str, s2: str) -> int:
    """Calculate Levenshtein distance between two strings"""
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)
    if len(s2) == 0:
        return len(s1)
    
    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row
    return previous_row[-1]

def calculate_similarity(str1: str, str2: str) -> float:
    """Calculate similarity score between two strings (0-1)"""
    if not str1 or not str2:
        return 0.0
    
    # Normalize both strings
    norm1 = normalize_plot_number(str1)
    norm2 = normalize_plot_number(str2)
    
    # Exact match after normalization
    if norm1 == norm2:
        return 1.0
    
    # Extract numeric parts
    nums1 = extract_numeric_parts(norm1)
    nums2 = extract_numeric_parts(norm2)
    
    # If numeric parts match, high similarity
    if nums1 and nums2 and nums1 == nums2:
        return 0.9
    
    # Calculate Levenshtein distance
    max_len = max(len(norm1), len(norm2))
    if max_len == 0:
        return 1.0
    
    distance = levenshtein_distance(norm1, norm2)
    similarity = 1.0 - (distance / max_len)
    
    # Boost similarity if numeric parts are similar
    if nums1 and nums2:
        num_similarity = len(set(nums1) & set(nums2)) / max(len(set(nums1)), len(set(nums2)), 1)
        similarity = max(similarity, num_similarity * 0.8)
    
    return max(0.0, min(1.0, similarity))

def smart_match_plot_numbers(vector_plots: List[str], radius_units: List[str], threshold: float = 0.5) -> List[dict]:
    """
    Smart match plot numbers from Vector to unit numbers from Radius Inventory.
    Returns list of matches with confidence scores.
    
    Args:
        vector_plots: List of plot numbers from Vector
        radius_units: List of unit numbers from Radius Inventory
        threshold: Minimum confidence score (0-1) for a match
    
    Returns:
        List of dicts with keys: vector_plot, radius_unit, confidence, auto_match
    """
    matches = []
    used_radius_units = set()
    
    for v_plot in vector_plots:
        best_match = None
        best_score = 0.0
        
        for r_unit in radius_units:
            if r_unit in used_radius_units:
                continue
            
            score = calculate_similarity(v_plot, r_unit)
            if score > best_score and score >= threshold:
                best_score = score
                best_match = r_unit
        
        if best_match:
            matches.append({
                "vector_plot": v_plot,
                "radius_unit": best_match,
                "confidence": round(best_score, 3),
                "auto_match": best_score >= 0.8  # High confidence = auto-match
            })
            used_radius_units.add(best_match)
        else:
            # No match found
            matches.append({
                "vector_plot": v_plot,
                "radius_unit": None,
                "confidence": 0.0,
                "auto_match": False
            })
    
    return matches

# ============================================
# FASTAPI APP
# ============================================
app = FastAPI(title="Radius CRM", version="3.0.0")
# CORS Configuration - Update with your office server domain
CORS_ORIGINS = os.getenv("CORS_ORIGINS", "http://localhost:5180,http://localhost:5174,http://localhost:5173,http://localhost:3000").split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================
# RATE LIMITER & API USAGE TRACKING
# ============================================
class RateLimiter:
    """In-memory sliding window rate limiter per user."""
    def __init__(self):
        self._requests = defaultdict(list)  # user_key -> [timestamps]

    def check(self, user_key: str, max_requests: int, window_seconds: int) -> bool:
        """Returns True if request is allowed, False if rate-limited."""
        now = time.time()
        cutoff = now - window_seconds
        # Prune old entries
        self._requests[user_key] = [t for t in self._requests[user_key] if t > cutoff]
        if len(self._requests[user_key]) >= max_requests:
            return False
        self._requests[user_key].append(now)
        return True

    def get_usage(self, user_key: str, window_seconds: int) -> int:
        """Get number of requests in current window."""
        now = time.time()
        cutoff = now - window_seconds
        self._requests[user_key] = [t for t in self._requests[user_key] if t > cutoff]
        return len(self._requests[user_key])

    def cleanup(self, max_age_seconds: int = 86400):
        """Remove entries older than max_age_seconds to prevent memory leak."""
        now = time.time()
        cutoff = now - max_age_seconds
        empty_keys = []
        for key in self._requests:
            self._requests[key] = [t for t in self._requests[key] if t > cutoff]
            if not self._requests[key]:
                empty_keys.append(key)
        for key in empty_keys:
            del self._requests[key]

# Rate limiters for different endpoint categories
voice_limiter = RateLimiter()   # Voice/AI queries
task_limiter = RateLimiter()    # Task CRUD operations
api_limiter = RateLimiter()     # General API catch-all

# Usage limits (configurable via env vars)
VOICE_RATE_LIMIT = int(os.getenv("VOICE_RATE_LIMIT", "30"))         # queries per window
VOICE_RATE_WINDOW = int(os.getenv("VOICE_RATE_WINDOW", "60"))       # window in seconds
VOICE_DAILY_LIMIT = int(os.getenv("VOICE_DAILY_LIMIT", "500"))      # per user per day
TASK_RATE_LIMIT = int(os.getenv("TASK_RATE_LIMIT", "60"))           # task ops per window
TASK_RATE_WINDOW = int(os.getenv("TASK_RATE_WINDOW", "60"))         # window in seconds

# API usage counters (daily, in-memory — resets on restart, persisted via query_history)
_daily_usage = defaultdict(lambda: {"voice_queries": 0, "task_ops": 0, "date": None})

def _track_daily_usage(user_key: str, category: str):
    """Track daily API usage per user."""
    today = date.today().isoformat()
    if _daily_usage[user_key]["date"] != today:
        _daily_usage[user_key] = {"voice_queries": 0, "task_ops": 0, "date": today}
    _daily_usage[user_key][category] = _daily_usage[user_key].get(category, 0) + 1

def _check_daily_limit(user_key: str, category: str, limit: int) -> bool:
    """Returns True if within daily limit."""
    today = date.today().isoformat()
    if _daily_usage[user_key]["date"] != today:
        return True
    return _daily_usage[user_key].get(category, 0) < limit

def require_rate_limit(limiter: RateLimiter, max_req: int, window: int, category: str = "api"):
    """Dependency that enforces rate limiting."""
    def _check(current_user=Depends(get_current_user)):
        user_key = current_user.rep_id or str(current_user.id)
        if not limiter.check(user_key, max_req, window):
            raise HTTPException(
                status_code=429,
                detail=f"Rate limit exceeded. Max {max_req} requests per {window}s. Please wait."
            )
        _track_daily_usage(user_key, category)
        return current_user
    return _check

@app.get("/")
def root():
    return {"status": "running", "app": "Radius CRM v3"}

login_limiter = RateLimiter()
LOGIN_RATE_LIMIT = int(os.getenv("LOGIN_RATE_LIMIT", "5"))       # attempts per window
LOGIN_RATE_WINDOW = int(os.getenv("LOGIN_RATE_WINDOW", "60"))    # window in seconds

@app.post("/api/auth/login")
def login(username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    """Login endpoint - returns JWT token"""
    # Rate limit login attempts per username
    if not login_limiter.check(username.lower(), LOGIN_RATE_LIMIT, LOGIN_RATE_WINDOW):
        raise HTTPException(status_code=429, detail=f"Too many login attempts. Try again in {LOGIN_RATE_WINDOW} seconds.")
    try:
        # Find user by rep_id, email, or mobile
        user = db.query(CompanyRep).filter(
            (CompanyRep.rep_id == username) | 
            (CompanyRep.email == username) | 
            (CompanyRep.mobile == username)
        ).first()
        
        if not user:
            print(f"Login attempt: User not found for username: {username}")
            raise HTTPException(status_code=401, detail="Invalid username or password")
        
        if user.status != "active":
            print(f"Login attempt: User {username} is not active, status: {user.status}")
            raise HTTPException(status_code=401, detail="User account is not active")
        
        # Check if password is set, if not allow first-time login with any password
        if not user.password_hash:
            # Set password on first login
            user.password_hash = get_password_hash(password)
            db.commit()
        elif not verify_password(password, user.password_hash):
            print(f"Login attempt: Invalid password for user: {username}")
            raise HTTPException(status_code=401, detail="Invalid username or password")
        
        # Create access token
        try:
            access_token = create_access_token(data={"sub": str(user.id), "role": user.role})
        except Exception as e:
            print(f"Error creating access token: {str(e)}")
            import traceback
            print(traceback.format_exc())
            raise HTTPException(status_code=500, detail="Error creating access token")
        
        try:
            return {
                "access_token": access_token,
                "token_type": "bearer",
                "user": {
                    "id": str(user.id),
                    "rep_id": user.rep_id,
                    "name": user.name,
                    "email": user.email if hasattr(user, 'email') else None,
                    "role": user.role if hasattr(user, 'role') else 'user',
                    "rep_type": user.rep_type if hasattr(user, 'rep_type') else None
                }
            }
        except Exception as e:
            print(f"Error building login response: {str(e)}")
            import traceback
            print(traceback.format_exc())
            raise HTTPException(status_code=500, detail=f"Error building response: {str(e)}")
    except HTTPException:
        raise
    except Exception as e:
        print(f"Unexpected error in login: {str(e)}")
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Login error: {str(e)}")

@app.get("/api/auth/me")
def get_current_user_info(current_user: CompanyRep = Depends(get_current_user)):
    """Get current logged-in user info"""
    try:
        return {
            "id": str(current_user.id),
            "rep_id": current_user.rep_id,
            "name": current_user.name,
            "email": current_user.email if hasattr(current_user, 'email') else None,
            "role": current_user.role if hasattr(current_user, 'role') else 'user',
            "rep_type": current_user.rep_type if hasattr(current_user, 'rep_type') else None,
            "mobile": current_user.mobile if hasattr(current_user, 'mobile') else None
        }
    except Exception as e:
        import traceback
        print(f"ERROR in /api/auth/me: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error getting user info: {str(e)}")

@app.post("/api/auth/change-password")
def change_password(
    old_password: str = Form(...),
    new_password: str = Form(...),
    current_user: CompanyRep = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Change user password"""
    if not current_user.password_hash or not verify_password(old_password, current_user.password_hash):
        raise HTTPException(status_code=400, detail="Invalid old password")
    
    current_user.password_hash = get_password_hash(new_password)
    db.commit()
    return {"message": "Password changed successfully"}

@app.get("/api/health")
def health_check(db: Session = Depends(get_db)):
    """Health check endpoint that verifies database tables"""
    try:
        tables = {}
        tables["customers"] = db.execute(text("SELECT COUNT(*) FROM customers")).scalar()
        tables["brokers"] = db.execute(text("SELECT COUNT(*) FROM brokers")).scalar()
        tables["projects"] = db.execute(text("SELECT COUNT(*) FROM projects")).scalar()
        tables["inventory"] = db.execute(text("SELECT COUNT(*) FROM inventory")).scalar()
        tables["transactions"] = db.execute(text("SELECT COUNT(*) FROM transactions")).scalar()
        tables["installments"] = db.execute(text("SELECT COUNT(*) FROM installments")).scalar()
        tables["company_reps"] = db.execute(text("SELECT COUNT(*) FROM company_reps")).scalar()
        
        # Check new tables (might not exist)
        try:
            tables["interactions"] = db.execute(text("SELECT COUNT(*) FROM interactions")).scalar()
        except: tables["interactions"] = "table not found"
        try:
            tables["campaigns"] = db.execute(text("SELECT COUNT(*) FROM campaigns")).scalar()
        except: tables["campaigns"] = "table not found"
        try:
            tables["leads"] = db.execute(text("SELECT COUNT(*) FROM leads")).scalar()
        except: tables["leads"] = "table not found"
        try:
            tables["receipts"] = db.execute(text("SELECT COUNT(*) FROM receipts")).scalar()
        except: tables["receipts"] = "table not found"
        try:
            tables["payments"] = db.execute(text("SELECT COUNT(*) FROM payments")).scalar()
        except: tables["payments"] = "table not found"
        try:
            tables["creditors"] = db.execute(text("SELECT COUNT(*) FROM creditors")).scalar()
        except: tables["creditors"] = "table not found"
        try:
            tables["payments"] = db.execute(text("SELECT COUNT(*) FROM payments")).scalar()
        except: tables["payments"] = "table not found"
        try:
            tables["media_files"] = db.execute(text("SELECT COUNT(*) FROM media_files")).scalar()
        except: tables["media_files"] = "table not found"
        
        return {"status": "healthy", "tables": tables}
    except Exception as e:
        return {"status": "error", "error": str(e)}

# ============================================
# CUSTOMERS API
# ============================================
@app.get("/api/customers")
def list_customers(db: Session = Depends(get_db),
                   current_user: CompanyRep = Depends(get_current_user)):
    iso = get_rep_isolation_filter(current_user, db)
    q = db.query(Customer).order_by(Customer.created_at.desc())
    if iso["isolated"]:
        rt = iso["rep_type"]
        if rt == "indirect":
            return []  # indirect reps don't see customers
        q = q.filter(Customer.assigned_rep_id.in_(iso["team_rep_uuids"]))
    customers = q.all()
    return [{
        "id": str(c.id), "customer_id": c.customer_id, "name": c.name,
        "mobile": c.mobile, "address": c.address, "cnic": c.cnic,
        "email": c.email,
        "additional_mobiles": c.additional_mobiles or [],
        "source": c.source, "source_other": c.source_other,
        "occupation": c.occupation, "occupation_other": c.occupation_other,
        "interested_project_id": str(c.interested_project_id) if c.interested_project_id else None,
        "interested_project_other": c.interested_project_other,
        "area": c.area, "city": c.city,
        "country_code": c.country_code or "+92",
        "notes": c.notes,
        "assigned_rep_id": str(c.assigned_rep_id) if c.assigned_rep_id else None,
        "temperature": c.temperature,
        "created_at": str(c.created_at)
    } for c in customers]

@app.get("/api/customers/{cid}")
def get_customer(cid: str, db: Session = Depends(get_db),
                 current_user: CompanyRep = Depends(get_current_user)):
    c = db.query(Customer).filter(
        (Customer.id == cid) | (Customer.customer_id == cid) | (Customer.mobile == cid)
    ).first()
    if not c:
        raise HTTPException(404, "Customer not found")
    broker = db.query(Broker).filter(Broker.mobile == c.mobile).first()
    return {
        "id": str(c.id), "customer_id": c.customer_id, "name": c.name,
        "mobile": c.mobile, "address": c.address, "cnic": c.cnic,
        "email": c.email,
        "additional_mobiles": c.additional_mobiles or [],
        "source": c.source, "source_other": c.source_other,
        "occupation": c.occupation, "occupation_other": c.occupation_other,
        "interested_project_id": str(c.interested_project_id) if c.interested_project_id else None,
        "interested_project_other": c.interested_project_other,
        "area": c.area, "city": c.city,
        "country_code": c.country_code or "+92",
        "notes": c.notes,
        "assigned_rep_id": str(c.assigned_rep_id) if c.assigned_rep_id else None,
        "temperature": c.temperature,
        "created_at": str(c.created_at),
        "is_also_broker": broker is not None,
        "broker_id": broker.broker_id if broker else None
    }

@app.post("/api/customers")
def create_customer(data: dict, db: Session = Depends(get_db),
                    current_user: CompanyRep = Depends(get_current_user)):
    if current_user.role == "user":
        raise HTTPException(403, "Sales reps can create leads only. Customer creation is restricted.")
    # Normalize primary mobile
    if data.get("mobile"):
        data["mobile"] = normalize_mobile(data["mobile"]) or data["mobile"]
    if db.query(Customer).filter(Customer.mobile == data["mobile"]).first():
        raise HTTPException(400, f"Customer with mobile {data['mobile']} already exists")
    # Check for duplicate mobile in leads + additional mobiles
    warnings = []
    dupes = check_duplicate_mobile(db, data.get("mobile"))
    for d in dupes:
        if d["type"] == "lead":
            warnings.append(f"Lead {d['entity_id']} ({d['name']}) has same mobile — assigned to {d.get('assigned_rep', 'unassigned')}")
    # Check additional mobiles for duplicates
    additional = [m for m in (data.get("additional_mobiles") or []) if m and m.strip()]
    for am in additional:
        am_dupes = check_duplicate_mobile(db, am)
        for d in am_dupes:
            warnings.append(f"Additional mobile {am} matches {d['type']} {d['entity_id']} ({d['name']})")
    # Resolve interested_project_id
    interested_project_id = None
    if data.get("interested_project_id") and data["interested_project_id"] != "other":
        p = db.query(Project).filter(
            (Project.id == data["interested_project_id"]) | (Project.project_id == data["interested_project_id"])
        ).first()
        interested_project_id = p.id if p else None
    c = Customer(
        name=data["name"], mobile=data["mobile"], address=data.get("address"),
        cnic=data.get("cnic"), email=data.get("email"),
        additional_mobiles=additional if additional else [],
        source=data.get("source"), source_other=data.get("source_other"),
        occupation=data.get("occupation"), occupation_other=data.get("occupation_other"),
        interested_project_id=interested_project_id,
        interested_project_other=data.get("interested_project_other"),
        area=data.get("area"), city=data.get("city"),
        country_code=data.get("country_code", "+92"),
        notes=data.get("notes"),
        assigned_rep_id=current_user.id,
        temperature=data.get("temperature")
    )
    db.add(c); db.commit(); db.refresh(c)
    result = {"message": "Customer created", "id": str(c.id), "customer_id": c.customer_id}
    if warnings:
        result["warnings"] = warnings
    return result

@app.put("/api/customers/{cid}")
def update_customer(cid: str, data: dict, db: Session = Depends(get_db),
                    current_user: CompanyRep = Depends(get_current_user)):
    c = db.query(Customer).filter((Customer.id == cid) | (Customer.customer_id == cid)).first()
    if not c: raise HTTPException(404, "Customer not found")
    for k in ["name", "mobile", "address", "cnic", "email",
              "additional_mobiles", "source", "source_other", "occupation", "occupation_other",
              "interested_project_other", "area", "city", "country_code", "notes", "temperature"]:
        if k in data: setattr(c, k, data[k])
    # Handle interested_project_id separately (FK resolution)
    if "interested_project_id" in data:
        if data["interested_project_id"] and data["interested_project_id"] != "other":
            p = db.query(Project).filter(
                (Project.id == data["interested_project_id"]) | (Project.project_id == data["interested_project_id"])
            ).first()
            c.interested_project_id = p.id if p else None
        else:
            c.interested_project_id = None
    # Filter empty strings from additional_mobiles
    if c.additional_mobiles:
        c.additional_mobiles = [m for m in c.additional_mobiles if m and m.strip()]
    db.commit()
    return {"message": "Customer updated"}

@app.delete("/api/customers/{cid}")
def delete_customer(cid: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    c = find_entity(db, Customer, "customer_id", cid)
    if not c: raise HTTPException(404, "Customer not found")

    # Creator role cannot delete
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    # Admin can delete directly
    if current_user.role == "admin":
        db.delete(c); db.commit()
        return {"message": "Customer deleted"}

    # Non-admin: create deletion request
    req = DeletionRequest(entity_type="customer", entity_id=c.id, entity_name=c.name, requested_by=current_user.id)
    db.add(req); db.commit(); db.refresh(req)
    return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

@app.get("/api/customers/template/download")
def download_customer_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['name*', 'mobile*', 'address', 'cnic', 'email', 'source', 'occupation', 'area', 'city', 'notes'])
    writer.writerow(['Ahmed Khan', '0300-1234567', 'DHA Phase 5', '35201-1234567-1', 'ahmed@email.com', 'Walk-in', 'Businessman', 'DHA Phase 5', 'Karachi', 'Interested in plots'])
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=customers_template.csv"})

@app.post("/api/customers/bulk-import")
async def bulk_import_customers(file: UploadFile = File(...), db: Session = Depends(get_db),
                                current_user: CompanyRep = Depends(get_current_user)):
    if current_user.role == "user":
        raise HTTPException(403, "Sales reps can create leads only. Customer import is restricted.")
    content = await file.read()
    reader = csv.DictReader(io.StringIO(content.decode('utf-8-sig')))
    results = {"success": 0, "errors": [], "created": []}
    for i, row in enumerate(reader, 2):
        try:
            name, mobile = row.get('name*', '').strip(), row.get('mobile*', '').strip()
            if not name or not mobile: results["errors"].append(f"Row {i}: Name and mobile required"); continue
            dupes = check_duplicate_mobile(db, mobile)
            if dupes:
                dup = dupes[0]
                results["errors"].append(f"Row {i}: Mobile exists in {dup['type']} {dup['entity_id']} ({dup['name']})")
                continue
            c = Customer(name=name, mobile=mobile, address=row.get('address'), cnic=row.get('cnic'), email=row.get('email'),
                         source=row.get('source'), occupation=row.get('occupation'),
                         area=row.get('area'), city=row.get('city'), notes=row.get('notes'),
                         assigned_rep_id=current_user.id)
            db.add(c); db.flush()
            results["success"] += 1
            results["created"].append({"customer_id": c.customer_id, "name": name})
        except Exception as e: results["errors"].append(f"Row {i}: {e}")
    db.commit()
    return results

# ============================================
# BROKERS API
# ============================================
@app.get("/api/brokers")
def list_brokers(db: Session = Depends(get_db),
                 current_user: CompanyRep = Depends(get_current_user)):
    iso = get_rep_isolation_filter(current_user, db)
    q = db.query(Broker).order_by(Broker.created_at.desc())
    if iso["isolated"]:
        rt = iso["rep_type"]
        if rt == "direct":
            return []  # direct reps don't see brokers
        q = q.filter(Broker.assigned_rep_id.in_(iso["team_rep_uuids"]))
    brokers = q.all()
    result = []
    for b in brokers:
        customer = db.query(Customer).filter(Customer.mobile == b.mobile).first()
        # Get broker stats from transactions
        deals = db.query(Transaction).filter(Transaction.broker_id == b.id).all()
        total_sales = sum(float(t.total_value or 0) for t in deals)
        commission_earned = sum(float(t.total_value or 0) * float(t.broker_commission_rate or 0) / 100 for t in deals)
        result.append({
            "id": str(b.id), "broker_id": b.broker_id, "name": b.name, "mobile": b.mobile,
            "cnic": b.cnic, "email": b.email, "company": b.company, "address": b.address,
            "commission_rate": float(b.commission_rate or 2.0), "bank_name": b.bank_name,
            "bank_account": b.bank_account, "bank_iban": b.bank_iban, "notes": b.notes,
            "status": b.status, "created_at": str(b.created_at),
            "is_also_customer": customer is not None,
            "customer_id": customer.customer_id if customer else None,
            "stats": {
                "total_deals": len(deals), "total_sales_value": total_sales,
                "commission_earned": commission_earned, "commission_paid": 0, "commission_pending": commission_earned
            }
        })
    return result

@app.get("/api/brokers/summary")
def get_brokers_summary(db: Session = Depends(get_db)):
    total = db.query(Broker).count()
    active = db.query(Broker).filter(Broker.status == "active").count()
    deals = db.query(Transaction).filter(Transaction.broker_id.isnot(None)).all()
    total_commission = sum(float(t.total_value or 0) * float(t.broker_commission_rate or 0) / 100 for t in deals)
    total_deals_value = sum(float(t.total_value or 0) for t in deals)
    
    # Deals this month
    from datetime import datetime
    current_month_start = date.today().replace(day=1)
    deals_this_month = [d for d in deals if d.booking_date and d.booking_date >= current_month_start]
    deals_this_month_value = sum(float(t.total_value or 0) for t in deals_this_month)
    
    # Top performers (by deal value)
    broker_stats = {}
    for t in deals:
        if t.broker_id:
            bid = str(t.broker_id)
            if bid not in broker_stats:
                broker_stats[bid] = {"total_value": 0, "deal_count": 0}
            broker_stats[bid]["total_value"] += float(t.total_value or 0)
            broker_stats[bid]["deal_count"] += 1
    
    top_performers = []
    for bid, stats in sorted(broker_stats.items(), key=lambda x: x[1]["total_value"], reverse=True)[:5]:
        broker = db.query(Broker).filter(Broker.id == bid).first()
        if broker:
            top_performers.append({
                "broker_id": broker.broker_id, "name": broker.name,
                "total_value": stats["total_value"], "deal_count": stats["deal_count"]
            })
    
    return {
        "total_brokers": total, "active_brokers": active, "total_deals": len(deals),
        "total_deals_value": total_deals_value,
        "total_commission_owed": total_commission, "total_commission_paid": 0,
        "deals_this_month": len(deals_this_month), "deals_this_month_value": deals_this_month_value,
        "top_performers": top_performers
    }

@app.get("/api/brokers/{bid}")
def get_broker(bid: str, db: Session = Depends(get_db)):
    b = db.query(Broker).filter((Broker.id == bid) | (Broker.broker_id == bid) | (Broker.mobile == bid)).first()
    if not b: raise HTTPException(404, "Broker not found")
    customer = db.query(Customer).filter(Customer.mobile == b.mobile).first()
    deals = db.query(Transaction).filter(Transaction.broker_id == b.id).all()
    return {
        "id": str(b.id), "broker_id": b.broker_id, "name": b.name, "mobile": b.mobile,
        "cnic": b.cnic, "email": b.email, "company": b.company, "address": b.address,
        "commission_rate": float(b.commission_rate or 2.0), "bank_name": b.bank_name,
        "bank_account": b.bank_account, "bank_iban": b.bank_iban, "notes": b.notes,
        "status": b.status, "is_also_customer": customer is not None,
        "transactions": [{"transaction_id": t.transaction_id, "total_value": float(t.total_value)} for t in deals]
    }

@app.post("/api/brokers")
def create_broker(data: dict, db: Session = Depends(get_db)):
    if db.query(Broker).filter(Broker.mobile == data["mobile"]).first():
        raise HTTPException(400, f"Broker with mobile {data['mobile']} already exists")
    customer = db.query(Customer).filter(Customer.mobile == data["mobile"]).first()
    b = Broker(name=data["name"], mobile=data["mobile"], cnic=data.get("cnic"), email=data.get("email"),
               company=data.get("company"), address=data.get("address"),
               commission_rate=data.get("commission_rate", 2.0), bank_name=data.get("bank_name"),
               bank_account=data.get("bank_account"), bank_iban=data.get("bank_iban"), notes=data.get("notes"),
               linked_customer_id=customer.id if customer else None)
    db.add(b); db.commit(); db.refresh(b)
    return {"message": "Broker created", "id": str(b.id), "broker_id": b.broker_id,
            "linked_to_customer": customer is not None}

@app.put("/api/brokers/{bid}")
def update_broker(bid: str, data: dict, db: Session = Depends(get_db)):
    b = db.query(Broker).filter((Broker.id == bid) | (Broker.broker_id == bid)).first()
    if not b: raise HTTPException(404, "Broker not found")
    for k in ["name", "mobile", "cnic", "email", "company", "address", "commission_rate",
              "bank_name", "bank_account", "bank_iban", "notes", "status"]:
        if k in data and data[k] is not None: setattr(b, k, data[k])
    db.commit()
    return {"message": "Broker updated"}

@app.delete("/api/brokers/{bid}")
def delete_broker(bid: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    b = find_entity(db, Broker, "broker_id", bid)
    if not b: raise HTTPException(404, "Broker not found")

    # Creator role cannot delete
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    # Admin can delete directly
    if current_user.role == "admin":
        db.delete(b); db.commit()
        return {"message": "Broker deleted"}

    # Non-admin: create deletion request
    req = DeletionRequest(entity_type="broker", entity_id=b.id, entity_name=b.name, requested_by=current_user.id)
    db.add(req); db.commit(); db.refresh(req)
    return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

@app.get("/api/brokers/template/download")
def download_broker_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['name*', 'mobile*', 'company', 'commission_rate', 'cnic', 'email', 'bank_name', 'bank_account', 'notes'])
    writer.writerow(['Rizwan Properties', '0300-1111111', 'Rizwan Real Estate', '2.5', '', 'rizwan@email.com', 'HBL', '1234567890', ''])
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=brokers_template.csv"})

@app.post("/api/brokers/bulk-import")
async def bulk_import_brokers(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    reader = csv.DictReader(io.StringIO(content.decode('utf-8-sig')))
    results = {"success": 0, "errors": [], "created": []}
    for i, row in enumerate(reader, 2):
        try:
            name, mobile = row.get('name*', '').strip(), row.get('mobile*', '').strip()
            if not name or not mobile: results["errors"].append(f"Row {i}: Name and mobile required"); continue
            # Check for duplicate brokers and leads (exclude customers — brokers link to matching customers)
            dupes = check_duplicate_mobile(db, mobile)
            broker_or_lead_dupes = [d for d in dupes if d["type"] in ("broker", "lead")]
            if broker_or_lead_dupes:
                dup = broker_or_lead_dupes[0]
                results["errors"].append(f"Row {i}: Mobile exists in {dup['type']} {dup['entity_id']} ({dup['name']})")
                continue
            customer = db.query(Customer).filter(Customer.mobile == mobile).first()
            rate = row.get('commission_rate', '').strip()
            b = Broker(name=name, mobile=mobile, company=row.get('company'),
                       commission_rate=float(rate) if rate else 2.0, cnic=row.get('cnic'),
                       email=row.get('email'), bank_name=row.get('bank_name'),
                       bank_account=row.get('bank_account'), notes=row.get('notes'),
                       linked_customer_id=customer.id if customer else None)
            db.add(b); db.flush()
            results["success"] += 1
            results["created"].append({"broker_id": b.broker_id, "name": name})
        except Exception as e: results["errors"].append(f"Row {i}: {e}")
    db.commit()
    return results

# ============================================
# PROJECTS API
# ============================================
@app.get("/api/projects")
def list_projects(db: Session = Depends(get_db)):
    try:
        projects = db.query(Project).order_by(Project.created_at.desc()).all()
        result = []
        for p in projects:
            inventory = db.query(Inventory).filter(Inventory.project_id == p.id).all()
            available = [i for i in inventory if i.status == "available"]
            sold = [i for i in inventory if i.status == "sold"]
            total_area = sum(float(i.area_marla) for i in inventory)
            avg_rate = sum(float(i.rate_per_marla) for i in inventory) / len(inventory) if inventory else 0
            result.append({
                "id": str(p.id), "project_id": p.project_id, "name": p.name,
                "location": p.location, "description": p.description, "status": p.status,
                "stats": {
                    "total_units": len(inventory), "available": len(available), "sold": len(sold),
                    "total_area": total_area, "avg_rate": round(avg_rate, 0),
                    "total_value": sum(float(i.area_marla) * float(i.rate_per_marla) for i in inventory),
                    "sold_value": sum(float(i.area_marla) * float(i.rate_per_marla) for i in sold)
                }
            })
        return result
    except Exception as e:
        error_msg = str(e)
        if "does not exist" in error_msg.lower() or "column" in error_msg.lower():
            raise HTTPException(
                status_code=500,
                detail=f"Database schema error: Missing columns. Please run the migration script: python migrate_vector_schema.py. Error: {error_msg[:200]}"
            )
        raise HTTPException(status_code=500, detail=f"Error loading projects: {error_msg[:200]}")

@app.get("/api/projects/{pid}")
def get_project(pid: str, db: Session = Depends(get_db)):
    p = db.query(Project).filter((Project.id == pid) | (Project.project_id == pid)).first()
    if not p: raise HTTPException(404, "Project not found")
    inventory = db.query(Inventory).filter(Inventory.project_id == p.id).all()
    return {
        "id": str(p.id), "project_id": p.project_id, "name": p.name,
        "location": p.location, "description": p.description, "status": p.status,
        "inventory": [{
            "id": str(i.id), "inventory_id": i.inventory_id, "unit_number": i.unit_number,
            "unit_type": i.unit_type, "block": i.block, "area_marla": float(i.area_marla),
            "rate_per_marla": float(i.rate_per_marla), "total_value": float(i.area_marla) * float(i.rate_per_marla),
            "status": i.status
        } for i in inventory]
    }

@app.post("/api/projects")
def create_project(data: dict, db: Session = Depends(get_db)):
    p = Project(name=data["name"], location=data.get("location"), description=data.get("description"))
    db.add(p); db.commit(); db.refresh(p)
    return {"message": "Project created", "id": str(p.id), "project_id": p.project_id}

@app.put("/api/projects/{pid}")
def update_project(pid: str, data: dict, db: Session = Depends(get_db)):
    p = db.query(Project).filter((Project.id == pid) | (Project.project_id == pid)).first()
    if not p: raise HTTPException(404, "Project not found")
    for k in ["name", "location", "description", "status"]:
        if k in data and data[k] is not None: setattr(p, k, data[k])
    db.commit()
    return {"message": "Project updated"}

@app.delete("/api/projects/{pid}")
def delete_project(pid: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    p = find_entity(db, Project, "project_id", pid)
    if not p: raise HTTPException(404, "Project not found")

    # Creator role cannot delete
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    # Admin can delete directly
    if current_user.role == "admin":
        try:
            db.delete(p); db.commit()
        except Exception:
            db.rollback()
            raise HTTPException(400, "Cannot delete project with linked inventory. Remove inventory first.")
        return {"message": "Project deleted"}

    # Non-admin: create deletion request
    req = DeletionRequest(entity_type="project", entity_id=p.id, entity_name=p.name, requested_by=current_user.id)
    db.add(req); db.commit(); db.refresh(req)
    return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

# ============================================
# COMPANY REPS API
# ============================================
@app.get("/api/company-reps")
def list_reps(db: Session = Depends(get_db)):
    reps = db.query(CompanyRep).order_by(CompanyRep.created_at.desc()).all()
    return [{"id": str(r.id), "rep_id": r.rep_id, "name": r.name, "mobile": r.mobile,
             "email": r.email, "status": r.status, "role": r.role,
             "rep_type": r.rep_type, "title": r.title, "reports_to": r.reports_to} for r in reps]

@app.post("/api/company-reps")
def create_rep(data: dict, db: Session = Depends(get_db), current_user: CompanyRep = Depends(require_role(["admin"]))):
    password = data.get("password")
    password_hash = get_password_hash(password) if password else None
    role = data.get("role", "user")  # Default to "user" role
    
    r = CompanyRep(
        name=data["name"],
        mobile=data.get("mobile"),
        email=data.get("email"),
        password_hash=password_hash,
        role=role,
        rep_type=data.get("rep_type"),
        title=data.get("title"),
        reports_to=data.get("reports_to")
    )
    db.add(r); db.commit(); db.refresh(r)
    return {"message": "Rep created", "id": str(r.id), "rep_id": r.rep_id, "role": r.role, "rep_type": r.rep_type}

@app.put("/api/company-reps/{rid}")
def update_rep(rid: str, data: dict, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    r = db.query(CompanyRep).filter((CompanyRep.id == rid) | (CompanyRep.rep_id == rid)).first()
    if not r: raise HTTPException(404, "Rep not found")

    # Only admin can change roles or status
    if ("role" in data or "status" in data) and current_user.role != "admin":
        raise HTTPException(403, "Only admin can change roles or status")
    # Non-admin/cco can only update their own profile (name, mobile, email, password)
    if str(r.id) != str(current_user.id) and current_user.role not in ["admin", "cco"]:
        raise HTTPException(403, "Can only update your own profile")
    
    for k in ["name", "mobile", "email", "status", "role", "rep_type", "title", "reports_to"]:
        if k in data: setattr(r, k, data[k])
    
    if "password" in data and data["password"]:
        r.password_hash = get_password_hash(data["password"])
    
    db.commit()
    return {"message": "Rep updated"}

@app.delete("/api/company-reps/{rid}")
def delete_rep(rid: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    r = find_entity(db, CompanyRep, "rep_id", rid)
    if not r: raise HTTPException(404, "Rep not found")

    # Only admin, cco, or manager can delete (or request deletion)
    if current_user.role not in ["admin", "cco", "manager"]:
        raise HTTPException(403, "Insufficient permissions to delete reps")

    # Admin can delete directly
    if current_user.role == "admin":
        db.delete(r); db.commit()
        return {"message": "Rep deleted"}

    # Non-admin: create deletion request
    req = DeletionRequest(entity_type="company_rep", entity_id=r.id, entity_name=r.name, requested_by=current_user.id)
    db.add(req); db.commit(); db.refresh(req)
    return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

# ============================================
# INVENTORY API
# ============================================
@app.get("/api/inventory")
def list_inventory(project_id: str = None, status: str = None, db: Session = Depends(get_db)):
    q = db.query(Inventory)
    if project_id:
        p = db.query(Project).filter((Project.id == project_id) | (Project.project_id == project_id)).first()
        if p: q = q.filter(Inventory.project_id == p.id)
    if status: q = q.filter(Inventory.status == status)
    inventory = q.order_by(Inventory.created_at.desc()).all()
    result = []
    for i in inventory:
        p = db.query(Project).filter(Project.id == i.project_id).first()
        result.append({
            "id": str(i.id), "inventory_id": i.inventory_id, "project_id": str(i.project_id),
            "project_name": p.name if p else None, "unit_number": i.unit_number,
            "unit_type": i.unit_type, "block": i.block, "area_marla": float(i.area_marla),
            "rate_per_marla": float(i.rate_per_marla),
            "total_value": float(i.area_marla) * float(i.rate_per_marla),
            "status": i.status, "factor_details": i.factor_details, "notes": i.notes
        })
    return result

@app.get("/api/inventory/available")
def list_available_inventory(db: Session = Depends(get_db)):
    inventory = db.query(Inventory).filter(Inventory.status == "available").order_by(Inventory.created_at.desc()).all()
    result = []
    for i in inventory:
        p = db.query(Project).filter(Project.id == i.project_id).first()
        result.append({
            "id": str(i.id), "inventory_id": i.inventory_id, "project_id": str(i.project_id),
            "project_name": p.name if p else None, "unit_number": i.unit_number,
            "unit_type": i.unit_type, "block": i.block, "area_marla": float(i.area_marla),
            "rate_per_marla": float(i.rate_per_marla),
            "total_value": float(i.area_marla) * float(i.rate_per_marla),
            "status": i.status
        })
    return result

@app.get("/api/inventory/template/download")
def download_inventory_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['project_id*', 'unit_number*', 'unit_type', 'block', 'area_marla*', 'rate_per_marla*', 'factor_details', 'notes'])
    writer.writerow(['PRJ-0001', 'A-101', 'plot', 'A', '5', '500000', 'Corner', 'Prime location'])
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=inventory_template.csv"})

@app.get("/api/inventory/summary")
def get_inventory_summary(db: Session = Depends(get_db)):
    total_projects = db.query(Project).count()
    total_available = db.query(Inventory).filter(Inventory.status == "available").count()
    total_sold = db.query(Inventory).filter(Inventory.status == "sold").count()
    available_value = db.query(func.sum(Inventory.area_marla * Inventory.rate_per_marla)).filter(Inventory.status == "available").scalar() or 0
    sold_value = db.query(func.sum(Inventory.area_marla * Inventory.rate_per_marla)).filter(Inventory.status == "sold").scalar() or 0
    return {
        "total_projects": total_projects, "total_available": total_available,
        "total_sold": total_sold, "available_value": float(available_value), "sold_value": float(sold_value)
    }

@app.get("/api/inventory/{iid}")
def get_inventory_item(iid: str, db: Session = Depends(get_db)):
    i = db.query(Inventory).filter((Inventory.id == iid) | (Inventory.inventory_id == iid)).first()
    if not i: raise HTTPException(404, "Inventory not found")
    p = db.query(Project).filter(Project.id == i.project_id).first()
    return {
        "id": str(i.id), "inventory_id": i.inventory_id, "project_id": str(i.project_id),
        "project_name": p.name if p else None, "unit_number": i.unit_number,
        "unit_type": i.unit_type, "block": i.block, "area_marla": float(i.area_marla),
        "rate_per_marla": float(i.rate_per_marla), "status": i.status,
        "factor_details": i.factor_details, "notes": i.notes
    }

@app.post("/api/inventory")
def create_inventory(data: dict, db: Session = Depends(get_db)):
    p = db.query(Project).filter((Project.id == data["project_id"]) | (Project.project_id == data["project_id"])).first()
    if not p: raise HTTPException(404, "Project not found")
    i = Inventory(project_id=p.id, unit_number=data["unit_number"], unit_type=data.get("unit_type", "plot"),
                  block=data.get("block"), area_marla=data["area_marla"], rate_per_marla=data["rate_per_marla"],
                  factor_details=data.get("factor_details"), notes=data.get("notes"))
    db.add(i); db.commit(); db.refresh(i)

    # Trigger Vector sync if project is linked
    sync_vector_branches_from_orbit(p.id, db)
    db.commit()

    return {"message": "Inventory created", "id": str(i.id), "inventory_id": i.inventory_id}

@app.put("/api/inventory/{iid}")
def update_inventory(iid: str, data: dict, db: Session = Depends(get_db)):
    i = db.query(Inventory).filter((Inventory.id == iid) | (Inventory.inventory_id == iid)).first()
    if not i: raise HTTPException(404, "Inventory not found")
    project_id = i.project_id
    for k in ["unit_number", "unit_type", "block", "area_marla", "rate_per_marla", "status", "factor_details", "notes"]:
        if k in data and data[k] is not None: setattr(i, k, data[k])
    db.commit()

    # Trigger Vector sync if project is linked
    if project_id:
        sync_vector_branches_from_orbit(project_id, db)
        db.commit()

    return {"message": "Inventory updated"}

@app.delete("/api/inventory/{iid}")
def delete_inventory(iid: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    i = find_entity(db, Inventory, "inventory_id", iid)
    if not i: raise HTTPException(404, "Inventory not found")

    # Creator role cannot delete
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    # Admin can delete directly
    if current_user.role == "admin":
        project_id = i.project_id
        db.delete(i); db.commit()

        # Trigger Vector sync if project is linked
        if project_id:
            sync_vector_branches_from_orbit(project_id, db)
            db.commit()

        return {"message": "Inventory deleted"}

    # Non-admin: create deletion request
    entity_name = i.unit_number if i.unit_number else str(i.id)
    req = DeletionRequest(entity_type="inventory", entity_id=i.id, entity_name=entity_name, requested_by=current_user.id)
    db.add(req); db.commit(); db.refresh(req)
    return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

@app.post("/api/inventory/bulk-import")
async def bulk_import_inventory(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    reader = csv.DictReader(io.StringIO(content.decode('utf-8-sig')))
    results = {"success": 0, "errors": []}
    for i, row in enumerate(reader, 2):
        try:
            p = db.query(Project).filter(Project.project_id == row.get('project_id*')).first()
            if not p: results["errors"].append(f"Row {i}: Project not found"); continue
            inv = Inventory(project_id=p.id, unit_number=row['unit_number*'],
                            unit_type=row.get('unit_type', 'plot'), block=row.get('block'),
                            area_marla=float(row['area_marla*']), rate_per_marla=float(row['rate_per_marla*']),
                            factor_details=row.get('factor_details'), notes=row.get('notes'))
            db.add(inv); db.flush()
            results["success"] += 1
        except Exception as e: results["errors"].append(f"Row {i}: {e}")
    db.commit()
    return results

# ============================================
# TRANSACTIONS API
# ============================================
@app.get("/api/transactions")
def list_transactions(db: Session = Depends(get_db),
                      current_user: CompanyRep = Depends(get_current_user)):
    try:
        q = db.query(Transaction).order_by(Transaction.created_at.desc())
        iso = get_rep_isolation_filter(current_user, db)
        if iso["isolated"]:
            rt = iso["rep_type"]
            if rt == "direct":
                q = q.filter(Transaction.company_rep_id.in_(iso["team_rep_uuids"]))
            elif rt == "indirect":
                broker_ids = [b.id for b in db.query(Broker.id).filter(Broker.assigned_rep_id.in_(iso["team_rep_uuids"])).all()]
                q = q.filter(Transaction.broker_id.in_(broker_ids)) if broker_ids else q.filter(literal(False))
            else:  # both
                broker_ids = [b.id for b in db.query(Broker.id).filter(Broker.assigned_rep_id.in_(iso["team_rep_uuids"])).all()]
                q = q.filter(or_(
                    Transaction.company_rep_id.in_(iso["team_rep_uuids"]),
                    Transaction.broker_id.in_(broker_ids) if broker_ids else literal(False)
                ))
        txns = q.all()
        result = []
        for t in txns:
            c = db.query(Customer).filter(Customer.id == t.customer_id).first()
            b = db.query(Broker).filter(Broker.id == t.broker_id).first() if t.broker_id else None
            p = db.query(Project).filter(Project.id == t.project_id).first()
            installments = db.query(Installment).filter(Installment.transaction_id == t.id).all()
            paid = sum(float(i.amount_paid or 0) for i in installments)
            result.append({
                "id": str(t.id), "transaction_id": t.transaction_id,
                "customer_name": c.name if c else None, "customer_mobile": c.mobile if c else None,
                "broker_name": b.name if b else None, "project_name": p.name if p else None,
                "unit_number": t.unit_number, "block": t.block,
                "area_marla": float(t.area_marla), "rate_per_marla": float(t.rate_per_marla),
                "total_value": float(t.total_value), "total_paid": paid,
                "balance": float(t.total_value) - paid, "status": t.status,
                "booking_date": str(t.booking_date) if t.booking_date else None
            })
        return result
    except Exception as e:
        print(f"Error listing transactions: {e}")
        import traceback
        traceback.print_exc()
        return []

@app.get("/api/transactions/summary")
def get_transactions_summary(project_id: str = None, db: Session = Depends(get_db)):
    try:
        q = db.query(Transaction)
        if project_id and project_id.strip():
            p = db.query(Project).filter(Project.project_id == project_id).first()
            if not p:
                try:
                    p = db.query(Project).filter(Project.id == project_id).first()
                except: pass
            if p: q = q.filter(Transaction.project_id == p.id)
        txns = q.all()
        total_value = sum(float(t.total_value or 0) for t in txns)
        current_month_start = date.today().replace(day=1)
        this_month = [t for t in txns if t.booking_date and t.booking_date >= current_month_start]
        this_month_value = sum(float(t.total_value or 0) for t in this_month)
        return {
            "total_transactions": len(txns), "total_value": total_value,
            "this_month_count": len(this_month), "this_month_value": this_month_value
        }
    except Exception as e:
        return {"total_transactions": 0, "total_value": 0, "this_month_count": 0, "this_month_value": 0}

@app.get("/api/transactions/template/download")
def download_transactions_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['customer_mobile*', 'project_id*', 'unit_number*', 'broker_mobile', 'area_marla', 'rate_per_marla', 'installment_cycle', 'num_installments', 'first_due_date', 'booking_date', 'notes'])
    writer.writerow(['0300-1234567', 'PRJ-0001', 'A-101', '0301-2345678', '5', '500000', 'bi-annual', '4', '2025-06-01', '2025-01-15', 'Sample booking'])
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=transactions_template.csv"})

@app.get("/api/transactions/{tid}")
def get_transaction(tid: str, db: Session = Depends(get_db)):
    try:
        tid_uuid = uuid.UUID(str(tid))
        t = db.query(Transaction).filter((Transaction.id == tid_uuid) | (Transaction.transaction_id == tid)).first()
    except ValueError:
        t = db.query(Transaction).filter(Transaction.transaction_id == tid).first()
    if not t: raise HTTPException(404, "Transaction not found")
    c = db.query(Customer).filter(Customer.id == t.customer_id).first()
    b = db.query(Broker).filter(Broker.id == t.broker_id).first() if t.broker_id else None
    p = db.query(Project).filter(Project.id == t.project_id).first()
    r = db.query(CompanyRep).filter(CompanyRep.id == t.company_rep_id).first() if t.company_rep_id else None
    installments = db.query(Installment).filter(Installment.transaction_id == t.id).order_by(Installment.installment_number).all()
    return {
        "id": str(t.id), "transaction_id": t.transaction_id,
        "customer": {"id": str(c.id), "name": c.name, "mobile": c.mobile} if c else None,
        "broker": {"id": str(b.id), "name": b.name, "commission_rate": float(t.broker_commission_rate)} if b else None,
        "project": {"id": str(p.id), "name": p.name} if p else None,
        "company_rep": {"id": str(r.id), "name": r.name} if r else None,
        "customer_name": c.name if c else None,
        "broker_name": b.name if b else None,
        "project_name": p.name if p else None,
        "unit_number": t.unit_number, "block": t.block,
        "area_marla": float(t.area_marla), "rate_per_marla": float(t.rate_per_marla),
        "total_value": float(t.total_value),
        "installment_cycle": t.installment_cycle, "num_installments": t.num_installments,
        "first_due_date": str(t.first_due_date), "booking_date": str(t.booking_date) if t.booking_date else None,
        "status": t.status, "notes": t.notes,
        "installments": [{
            "id": str(i.id), "number": i.installment_number, "due_date": str(i.due_date),
            "amount": float(i.amount), "amount_paid": float(i.amount_paid or 0),
            "balance": float(i.amount) - float(i.amount_paid or 0), "status": i.status
        } for i in installments]
    }

@app.post("/api/transactions")
def create_transaction(data: dict, db: Session = Depends(get_db),
                       current_user: CompanyRep = Depends(get_current_user)):
    customer = find_entity(db, Customer, "customer_id", data["customer_id"]) or db.query(Customer).filter(
        Customer.mobile == data["customer_id"]
    ).first()
    if not customer: raise HTTPException(404, "Customer not found")
    if current_user.role == "user" and str(customer.assigned_rep_id or "") != str(current_user.id):
        raise HTTPException(403, "Sales reps can create transactions only for customers assigned to them.")

    inventory = find_entity(db, Inventory, "inventory_id", data["inventory_id"])
    if not inventory: raise HTTPException(404, "Inventory not found")
    if inventory.status != "available": raise HTTPException(400, f"Unit not available (status: {inventory.status})")
    
    project = db.query(Project).filter(Project.id == inventory.project_id).first()
    broker = (find_entity(db, Broker, "broker_id", data["broker_id"]) or db.query(Broker).filter(Broker.mobile == data["broker_id"]).first()) if data.get("broker_id") else None
    rep = find_entity(db, CompanyRep, "rep_id", data["company_rep_id"]) if data.get("company_rep_id") else None
    
    area = float(data.get("area_marla") or inventory.area_marla)
    rate = float(data.get("rate_per_marla") or inventory.rate_per_marla)
    total = area * rate
    first_due = date.fromisoformat(data["first_due_date"]) if isinstance(data["first_due_date"], str) else data["first_due_date"]
    
    t = Transaction(
        customer_id=customer.id, broker_id=broker.id if broker else None,
        project_id=project.id, inventory_id=inventory.id, company_rep_id=rep.id if rep else None,
        broker_commission_rate=data.get("broker_commission_rate", broker.commission_rate if broker else 0),
        unit_number=inventory.unit_number, block=inventory.block,
        area_marla=area, rate_per_marla=rate, total_value=total,
        installment_cycle=data.get("installment_cycle", "bi-annual"),
        num_installments=data.get("num_installments", 4), first_due_date=first_due,
        notes=data.get("notes"),
        booking_date=date.fromisoformat(data["booking_date"]) if data.get("booking_date") else date.today()
    )
    db.add(t); db.commit(); db.refresh(t)
    
    cycle_months = {"monthly": 1, "quarterly": 3, "bi-annual": 6, "annual": 12}.get(t.installment_cycle, 6)
    installment_amount = total / t.num_installments
    for n in range(t.num_installments):
        due = first_due + relativedelta(months=n * cycle_months)
        db.add(Installment(transaction_id=t.id, installment_number=n + 1, due_date=due, amount=installment_amount))
    
    inventory.status = "sold"
    db.commit()

    # Trigger Vector sync if project is linked
    sync_vector_branches_from_orbit(project.id, db)
    db.commit()

    return {"message": "Transaction created", "id": str(t.id), "transaction_id": t.transaction_id}

@app.put("/api/transactions/{tid}")
def update_transaction(tid: str, data: dict, db: Session = Depends(get_db)):
    try:
        tid_uuid = uuid.UUID(str(tid))
        t = db.query(Transaction).filter((Transaction.id == tid_uuid) | (Transaction.transaction_id == tid)).first()
    except ValueError:
        t = db.query(Transaction).filter(Transaction.transaction_id == tid).first()
    if not t: raise HTTPException(404, "Transaction not found")
    for k in ["broker_commission_rate", "status", "notes"]:
        if k in data and data[k] is not None: setattr(t, k, data[k])
    db.commit()

    # Trigger Vector sync if project is linked
    if t.project_id:
        sync_vector_branches_from_orbit(t.project_id, db)
        db.commit()

    return {"message": "Transaction updated"}

@app.post("/api/transactions/bulk-import")
async def bulk_import_transactions(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    reader = csv.DictReader(io.StringIO(content.decode('utf-8-sig')))
    results = {"success": 0, "errors": []}
    for i, row in enumerate(reader, 2):
        try:
            customer = db.query(Customer).filter(Customer.mobile == row.get('customer_mobile*')).first()
            if not customer: results["errors"].append(f"Row {i}: Customer not found"); continue
            inventory = db.query(Inventory).filter(Inventory.inventory_id == row.get('inventory_id*')).first()
            if not inventory: results["errors"].append(f"Row {i}: Inventory not found"); continue
            if inventory.status != "available": results["errors"].append(f"Row {i}: Unit not available"); continue
            
            project = db.query(Project).filter(Project.id == inventory.project_id).first()
            broker = db.query(Broker).filter(Broker.mobile == row.get('broker_mobile')).first() if row.get('broker_mobile') else None
            area = float(row.get('area_marla') or inventory.area_marla)
            rate = float(row.get('rate_per_marla') or inventory.rate_per_marla)
            first_due = date.fromisoformat(row['first_due_date*'])
            cycle = row.get('installment_cycle', 'bi-annual')
            num_inst = int(row.get('num_installments') or 4)
            
            t = Transaction(
                customer_id=customer.id, broker_id=broker.id if broker else None,
                project_id=project.id, inventory_id=inventory.id,
                broker_commission_rate=float(row.get('broker_commission_rate') or (broker.commission_rate if broker else 0)),
                unit_number=inventory.unit_number, block=inventory.block,
                area_marla=area, rate_per_marla=rate, total_value=area * rate,
                installment_cycle=cycle, num_installments=num_inst, first_due_date=first_due,
                booking_date=date.fromisoformat(row.get('booking_date')) if row.get('booking_date') else date.today(),
                notes=row.get('notes')
            )
            db.add(t); db.flush()
            
            cycle_months = {"monthly": 1, "quarterly": 3, "bi-annual": 6, "annual": 12}.get(cycle, 6)
            for n in range(num_inst):
                due = first_due + relativedelta(months=n * cycle_months)
                db.add(Installment(transaction_id=t.id, installment_number=n + 1, due_date=due, amount=(area * rate) / num_inst))
            
            inventory.status = "sold"
            results["success"] += 1
        except Exception as e: results["errors"].append(f"Row {i}: {e}")
    db.commit()
    return results

# ============================================
# INSTALLMENTS API
# ============================================
@app.get("/api/installments/{tid}")
def get_installments(tid: str, db: Session = Depends(get_db)):
    t = db.query(Transaction).filter((Transaction.id == tid) | (Transaction.transaction_id == tid)).first()
    if not t: raise HTTPException(404, "Transaction not found")
    installments = db.query(Installment).filter(Installment.transaction_id == t.id).order_by(Installment.installment_number).all()
    return [{"id": str(i.id), "number": i.installment_number, "due_date": str(i.due_date),
             "amount": float(i.amount), "amount_paid": float(i.amount_paid or 0),
             "balance": float(i.amount) - float(i.amount_paid or 0), "status": i.status} for i in installments]

@app.put("/api/installments/{iid}")
def update_installment(iid: str, data: dict, db: Session = Depends(get_db)):
    i = db.query(Installment).filter(Installment.id == iid).first()
    if not i: raise HTTPException(404, "Installment not found")
    if "due_date" in data: i.due_date = date.fromisoformat(data["due_date"]) if isinstance(data["due_date"], str) else data["due_date"]
    if "amount" in data: i.amount = data["amount"]
    if "amount_paid" in data: i.amount_paid = data["amount_paid"]
    if "notes" in data: i.notes = data["notes"]
    if float(i.amount_paid or 0) >= float(i.amount): i.status = "paid"
    elif float(i.amount_paid or 0) > 0: i.status = "partial"
    db.commit()
    return {"message": "Installment updated"}

# ============================================
# INTERACTIONS API
# ============================================
def ensure_interaction_target_search_indexes(db: Session):
    global _interaction_target_indexes_ready
    if _interaction_target_indexes_ready:
        return
    statements = [
        "CREATE INDEX IF NOT EXISTS idx_customers_mobile ON customers(mobile)",
        "CREATE INDEX IF NOT EXISTS idx_brokers_mobile ON brokers(mobile)",
        "CREATE INDEX IF NOT EXISTS idx_leads_mobile ON leads(mobile)",
        "CREATE INDEX IF NOT EXISTS idx_leads_name ON leads(LOWER(name))",
    ]
    for stmt in statements:
        try:
            db.execute(text(stmt))
            db.commit()
        except Exception:
            db.rollback()
    _interaction_target_indexes_ready = True


@app.get("/api/interactions/targets/search")
def search_interaction_targets(
    entity_type: str = Query(...),
    q: str = Query(...),
    limit: int = 15,
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user),
):
    et = (entity_type or "").strip().lower()
    if et not in {"customer", "broker", "lead"}:
        raise HTTPException(400, "entity_type must be customer, broker, or lead")

    query_text = (q or "").strip()
    if len(query_text) < 2:
        raise HTTPException(400, "q must be at least 2 characters")

    limit = max(1, min(int(limit), 50))
    offset = max(0, int(offset))

    ensure_interaction_target_search_indexes(db)

    iso = get_rep_isolation_filter(current_user, db)
    q_lower = query_text.lower()
    q_digits = "".join(ch for ch in query_text if ch.isdigit())

    if et == "customer":
        base = db.query(
            Customer.id.label("id"),
            Customer.customer_id.label("entity_id"),
            literal("customer").label("entity_type"),
            Customer.name.label("name"),
            Customer.mobile.label("mobile"),
            func.coalesce(Customer.additional_mobiles, literal("[]").cast(JSONB)).label("additional_mobiles"),
            Customer.city.label("city"),
            literal(None).cast(String).label("source"),
            literal(None).cast(String).label("company"),
            literal(None).cast(String).label("assigned_rep"),
        )
        if iso["isolated"]:
            rt = iso["rep_type"]
            if rt == "indirect":
                return []
            base = base.filter(Customer.assigned_rep_id.in_(iso["team_rep_uuids"]))
        id_col = Customer.customer_id
        mobile_col = Customer.mobile
        name_col = Customer.name
    elif et == "broker":
        base = db.query(
            Broker.id.label("id"),
            Broker.broker_id.label("entity_id"),
            literal("broker").label("entity_type"),
            Broker.name.label("name"),
            Broker.mobile.label("mobile"),
            literal("[]").cast(JSONB).label("additional_mobiles"),
            literal(None).cast(String).label("city"),
            literal(None).cast(String).label("source"),
            Broker.company.label("company"),
            literal(None).cast(String).label("assigned_rep"),
        )
        if iso["isolated"]:
            rt = iso["rep_type"]
            if rt == "direct":
                return []
            base = base.filter(Broker.assigned_rep_id.in_(iso["team_rep_uuids"]))
        id_col = Broker.broker_id
        mobile_col = Broker.mobile
        name_col = Broker.name
    else:
        base = db.query(
            Lead.id.label("id"),
            Lead.lead_id.label("entity_id"),
            literal("lead").label("entity_type"),
            Lead.name.label("name"),
            Lead.mobile.label("mobile"),
            func.coalesce(Lead.additional_mobiles, literal("[]").cast(JSONB)).label("additional_mobiles"),
            Lead.city.label("city"),
            Lead.source.label("source"),
            literal(None).cast(String).label("company"),
            CompanyRep.name.label("assigned_rep"),
        ).outerjoin(CompanyRep, Lead.assigned_rep_id == CompanyRep.id)
        # Hide converted leads — once a lead becomes a customer, interactions go on the customer record
        base = base.filter(or_(Lead.converted_customer_id.is_(None), Lead.status != "converted"))
        if iso["isolated"]:
            base = base.filter(Lead.assigned_rep_id.in_(iso["team_rep_uuids"]))
        id_col = Lead.lead_id
        mobile_col = Lead.mobile
        name_col = Lead.name

    mobile_digits_col = func.regexp_replace(func.coalesce(mobile_col, ""), "[^0-9]", "", "g")
    if q_digits:
        exact_mobile_cond = or_(mobile_col == query_text, mobile_digits_col == q_digits)
        prefix_mobile_cond = or_(mobile_col.ilike(f"{query_text}%"), mobile_digits_col.like(f"{q_digits}%"))
    else:
        exact_mobile_cond = mobile_col == query_text
        prefix_mobile_cond = mobile_col.ilike(f"{query_text}%")

    name_partial_cond = func.lower(func.coalesce(name_col, "")).like(f"%{q_lower}%")
    id_partial_cond = func.lower(func.coalesce(id_col, "")).like(f"%{q_lower}%")

    exact_q = base.filter(exact_mobile_cond).add_columns(literal(1).label("priority"))
    prefix_q = base.filter(prefix_mobile_cond, ~exact_mobile_cond).add_columns(literal(2).label("priority"))
    name_q = base.filter(name_partial_cond, ~exact_mobile_cond, ~prefix_mobile_cond).add_columns(literal(3).label("priority"))
    id_q = base.filter(id_partial_cond, ~exact_mobile_cond, ~prefix_mobile_cond, ~name_partial_cond).add_columns(literal(4).label("priority"))

    ranked = union_all(
        exact_q.statement,
        prefix_q.statement,
        name_q.statement,
        id_q.statement,
    ).alias("ranked_targets")

    rows = (
        db.query(ranked)
        .order_by(ranked.c.priority.asc(), ranked.c.name.asc())
        .offset(offset)
        .limit(limit)
        .all()
    )

    return [
        {
            "id": str(r.id),
            "entity_id": r.entity_id,
            "entity_type": r.entity_type,
            "name": r.name,
            "mobile": r.mobile,
            "additional_mobiles": r.additional_mobiles or [],
            "city": r.city,
            "source": r.source,
            "company": r.company,
            "assigned_rep": r.assigned_rep,
        }
        for r in rows
    ]


@app.get("/api/interactions")
def list_interactions(rep_id: str = None, customer_id: str = None, broker_id: str = None,
                      lead_id: str = None, interaction_type: str = None, period: str = None,
                      limit: int = 100, db: Session = Depends(get_db),
                      current_user: CompanyRep = Depends(get_current_user)):
    q = db.query(Interaction)
    iso = get_rep_isolation_filter(current_user, db)
    if iso["isolated"]:
        q = q.filter(Interaction.company_rep_id.in_(iso["team_rep_uuids"]))
    if rep_id:
        r = db.query(CompanyRep).filter((CompanyRep.id == rep_id) | (CompanyRep.rep_id == rep_id)).first()
        if r: q = q.filter(Interaction.company_rep_id == r.id)
    if customer_id:
        c = db.query(Customer).filter((Customer.id == customer_id) | (Customer.customer_id == customer_id)).first()
        if c: q = q.filter(Interaction.customer_id == c.id)
    if broker_id:
        b = db.query(Broker).filter((Broker.id == broker_id) | (Broker.broker_id == broker_id)).first()
        if b: q = q.filter(Interaction.broker_id == b.id)
    if lead_id:
        try:
            lid_uuid = uuid.UUID(str(lead_id))
            ld = db.query(Lead).filter((Lead.id == lid_uuid) | (Lead.lead_id == lead_id)).first()
        except (ValueError, AttributeError):
            ld = db.query(Lead).filter(Lead.lead_id == lead_id).first()
        if ld: q = q.filter(Interaction.lead_id == ld.id)
    # Type filter (call, message, whatsapp, meeting)
    if interaction_type:
        q = q.filter(Interaction.interaction_type == interaction_type)
    # Period filter (today, week, month)
    if period:
        from datetime import timedelta
        today = date.today()
        if period == "today":
            q = q.filter(func.date(Interaction.created_at) == today)
        elif period == "week":
            week_start = today - timedelta(days=today.weekday())
            q = q.filter(func.date(Interaction.created_at) >= week_start)
        elif period == "month":
            month_start = today.replace(day=1)
            q = q.filter(func.date(Interaction.created_at) >= month_start)

    interactions = q.order_by(Interaction.created_at.desc()).limit(limit).all()

    # Pre-fetch related entities to avoid N+1 queries
    rep_ids = set(i.company_rep_id for i in interactions if i.company_rep_id)
    cust_ids = set(i.customer_id for i in interactions if i.customer_id)
    broker_ids = set(i.broker_id for i in interactions if i.broker_id)
    lead_ids = set(i.lead_id for i in interactions if i.lead_id)
    reps_map = {r.id: r for r in db.query(CompanyRep).filter(CompanyRep.id.in_(rep_ids)).all()} if rep_ids else {}
    custs_map = {c.id: c for c in db.query(Customer).filter(Customer.id.in_(cust_ids)).all()} if cust_ids else {}
    brokers_map = {b.id: b for b in db.query(Broker).filter(Broker.id.in_(broker_ids)).all()} if broker_ids else {}
    leads_map = {l.id: l for l in db.query(Lead).filter(Lead.id.in_(lead_ids)).all()} if lead_ids else {}

    result = []
    for i in interactions:
        rep = reps_map.get(i.company_rep_id)
        cust = custs_map.get(i.customer_id)
        broker = brokers_map.get(i.broker_id)
        ld = leads_map.get(i.lead_id)
        result.append({
            "id": str(i.id), "interaction_id": i.interaction_id,
            "rep_name": rep.name if rep else None, "rep_id": rep.rep_id if rep else None,
            "rep_uuid": str(rep.id) if rep else None,
            "customer_name": cust.name if cust else None, "customer_id": cust.customer_id if cust else None,
            "customer_mobile": cust.mobile if cust else None,
            "broker_name": broker.name if broker else None, "broker_id": broker.broker_id if broker else None,
            "broker_mobile": broker.mobile if broker else None,
            "lead_name": ld.name if ld else None, "lead_id": ld.lead_id if ld else None,
            "lead_uuid": str(ld.id) if ld else None,
            "lead_mobile": ld.mobile if ld else None,
            "lead_additional_mobiles": ld.additional_mobiles if ld else [],
            "lead_temperature": ld.temperature if ld else None,
            "lead_assigned_rep_id": str(ld.assigned_rep_id) if ld and ld.assigned_rep_id else None,
            "interaction_type": i.interaction_type, "status": i.status,
            "notes": i.notes, "next_follow_up": str(i.next_follow_up) if i.next_follow_up else None,
            "created_at": str(i.created_at)
        })
    return result

@app.get("/api/interactions/pending-followups")
def get_pending_followups(db: Session = Depends(get_db),
                          current_user: CompanyRep = Depends(get_current_user)):
    """Return interactions with pending follow-ups (due today or overdue), enriched with entity details."""
    today = date.today()
    q = db.query(Interaction).filter(
        Interaction.next_follow_up.isnot(None),
        Interaction.next_follow_up <= today
    )
    # Role isolation
    iso = get_rep_isolation_filter(current_user, db)
    if iso["isolated"]:
        q = q.filter(Interaction.company_rep_id.in_(iso["team_rep_uuids"]))

    interactions = q.order_by(Interaction.next_follow_up.asc()).all()

    # Pre-fetch related entities
    rep_ids = set(i.company_rep_id for i in interactions if i.company_rep_id)
    cust_ids = set(i.customer_id for i in interactions if i.customer_id)
    broker_ids = set(i.broker_id for i in interactions if i.broker_id)
    lead_ids = set(i.lead_id for i in interactions if i.lead_id)
    reps_map = {r.id: r for r in db.query(CompanyRep).filter(CompanyRep.id.in_(rep_ids)).all()} if rep_ids else {}
    custs_map = {c.id: c for c in db.query(Customer).filter(Customer.id.in_(cust_ids)).all()} if cust_ids else {}
    brokers_map = {b.id: b for b in db.query(Broker).filter(Broker.id.in_(broker_ids)).all()} if broker_ids else {}
    leads_map = {l.id: l for l in db.query(Lead).filter(Lead.id.in_(lead_ids)).all()} if lead_ids else {}

    result = []
    for i in interactions:
        rep = reps_map.get(i.company_rep_id)
        cust = custs_map.get(i.customer_id)
        broker = brokers_map.get(i.broker_id)
        ld = leads_map.get(i.lead_id)
        # Determine entity type and details
        if ld:
            entity_type = "lead"
            entity_id = ld.lead_id
            entity_uuid = str(ld.id)
            entity_name = ld.name
            entity_mobile = ld.mobile
            entity_additional_mobiles = ld.additional_mobiles or []
            entity_temperature = ld.temperature
        elif cust:
            entity_type = "customer"
            entity_id = cust.customer_id
            entity_uuid = str(cust.id)
            entity_name = cust.name
            entity_mobile = cust.mobile
            entity_additional_mobiles = cust.additional_mobiles if hasattr(cust, 'additional_mobiles') and cust.additional_mobiles else []
            entity_temperature = None
        elif broker:
            entity_type = "broker"
            entity_id = broker.broker_id
            entity_uuid = str(broker.id)
            entity_name = broker.name
            entity_mobile = broker.mobile
            entity_additional_mobiles = []
            entity_temperature = None
        else:
            entity_type = "unknown"
            entity_id = None
            entity_uuid = None
            entity_name = None
            entity_mobile = None
            entity_additional_mobiles = []
            entity_temperature = None

        result.append({
            "id": str(i.id), "interaction_id": i.interaction_id,
            "entity_type": entity_type, "entity_id": entity_id,
            "entity_uuid": entity_uuid, "entity_name": entity_name,
            "entity_mobile": entity_mobile,
            "entity_additional_mobiles": entity_additional_mobiles,
            "entity_temperature": entity_temperature,
            "rep_name": rep.name if rep else None, "rep_id": rep.rep_id if rep else None,
            "rep_uuid": str(rep.id) if rep else None,
            "lead_assigned_rep_id": str(ld.assigned_rep_id) if ld and ld.assigned_rep_id else (str(rep.id) if rep else None),
            "interaction_type": i.interaction_type, "status": i.status,
            "notes": i.notes,
            "next_follow_up": str(i.next_follow_up) if i.next_follow_up else None,
            "created_at": str(i.created_at)
        })
    return result

@app.get("/api/interactions/summary")
def get_interactions_summary(db: Session = Depends(get_db),
                             current_user: CompanyRep = Depends(get_current_user)):
    from datetime import datetime, timedelta
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    # Apply role isolation — sales reps see only their own, managers see team
    base_q = db.query(Interaction)
    iso = get_rep_isolation_filter(current_user, db)
    if iso["isolated"]:
        base_q = base_q.filter(Interaction.company_rep_id.in_(iso["team_rep_uuids"]))

    total = base_q.count()
    total_calls = base_q.filter(Interaction.interaction_type == "call").count()
    total_messages = base_q.filter(Interaction.interaction_type == "message").count()
    total_whatsapp = base_q.filter(Interaction.interaction_type == "whatsapp").count()

    today_count = base_q.filter(func.date(Interaction.created_at) == today).count()
    week_count = base_q.filter(func.date(Interaction.created_at) >= week_start).count()
    month_count = base_q.filter(func.date(Interaction.created_at) >= month_start).count()

    # Pending follow-ups
    pending_followups = base_q.filter(
        Interaction.next_follow_up.isnot(None),
        Interaction.next_follow_up <= today
    ).count()

    return {
        "total_interactions": total, "total_calls": total_calls,
        "total_messages": total_messages, "total_whatsapp": total_whatsapp,
        "today": today_count, "this_week": week_count, "this_month": month_count,
        "pending_followups": pending_followups
    }

@app.post("/api/interactions")
def create_interaction(data: dict, db: Session = Depends(get_db),
                       current_user: CompanyRep = Depends(get_current_user)):
    try:
        rep_uuid = uuid.UUID(str(data["company_rep_id"]))
        rep = db.query(CompanyRep).filter((CompanyRep.id == rep_uuid) | (CompanyRep.rep_id == data["company_rep_id"])).first()
    except (ValueError, AttributeError):
        rep = db.query(CompanyRep).filter(CompanyRep.rep_id == data["company_rep_id"]).first()
    if not rep: raise HTTPException(404, "Company rep not found")

    customer = None
    broker = None
    lead = None
    if data.get("customer_id"):
        try:
            cid = uuid.UUID(str(data["customer_id"]))
            customer = db.query(Customer).filter((Customer.id == cid) | (Customer.customer_id == data["customer_id"])).first()
        except (ValueError, AttributeError):
            customer = db.query(Customer).filter((Customer.customer_id == data["customer_id"]) | (Customer.mobile == data["customer_id"])).first()
    if data.get("broker_id"):
        try:
            bid = uuid.UUID(str(data["broker_id"]))
            broker = db.query(Broker).filter((Broker.id == bid) | (Broker.broker_id == data["broker_id"])).first()
        except (ValueError, AttributeError):
            broker = db.query(Broker).filter((Broker.broker_id == data["broker_id"]) | (Broker.mobile == data["broker_id"])).first()
    if data.get("lead_id"):
        try:
            lid_uuid = uuid.UUID(str(data["lead_id"]))
            lead = db.query(Lead).filter((Lead.id == lid_uuid) | (Lead.lead_id == data["lead_id"])).first()
        except (ValueError, AttributeError):
            lead = db.query(Lead).filter(Lead.lead_id == data["lead_id"]).first()

    # Auto-link customer/broker if lead has been synced
    if lead and not customer and lead.converted_customer_id:
        customer = db.query(Customer).filter(Customer.id == lead.converted_customer_id).first()
    if lead and not broker and lead.converted_broker_id:
        broker = db.query(Broker).filter(Broker.id == lead.converted_broker_id).first()

    if not customer and not broker and not lead:
        raise HTTPException(400, "Must specify customer, broker, or lead")

    selected_contact_number = (data.get("contact_number") or "").strip()
    notes = (data.get("notes") or "").strip()
    if selected_contact_number:
        prefix = f"[Spoke On] {selected_contact_number}"
        notes = f"{prefix}\n{notes}" if notes else prefix

    i = Interaction(
        company_rep_id=rep.id,
        customer_id=customer.id if customer else None,
        broker_id=broker.id if broker else None,
        lead_id=lead.id if lead else None,
        interaction_type=data["interaction_type"],
        status=data.get("status"),
        notes=notes or None,
        next_follow_up=date.fromisoformat(data["next_follow_up"]) if data.get("next_follow_up") else None
    )
    db.add(i); db.commit(); db.refresh(i)

    # Update lead's last_contacted_at if interaction is for a lead
    if lead:
        lead.last_contacted_at = datetime.utcnow()
        lead.is_stale = False
        db.commit()

    return {"message": "Interaction created", "id": str(i.id), "interaction_id": i.interaction_id}

@app.put("/api/interactions/{iid}")
def update_interaction(iid: str, data: dict, db: Session = Depends(get_db)):
    i = db.query(Interaction).filter((Interaction.id == iid) | (Interaction.interaction_id == iid)).first()
    if not i: raise HTTPException(404, "Interaction not found")
    for k in ["status", "notes"]:
        if k in data and data[k] is not None: setattr(i, k, data[k])
    if "next_follow_up" in data:
        i.next_follow_up = date.fromisoformat(data["next_follow_up"]) if data["next_follow_up"] else None
    db.commit()
    return {"message": "Interaction updated"}

@app.delete("/api/interactions/{iid}")
def delete_interaction(iid: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    i = find_entity(db, Interaction, "interaction_id", iid)
    if not i: raise HTTPException(404, "Interaction not found")

    # Creator role cannot delete
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    # Admin can delete directly
    if current_user.role == "admin":
        db.delete(i); db.commit()
        return {"message": "Interaction deleted"}

    # Non-admin: create deletion request
    req = DeletionRequest(entity_type="interaction", entity_id=i.id, entity_name=str(i.id), requested_by=current_user.id)
    db.add(req); db.commit(); db.refresh(req)
    return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

# ============================================
# CAMPAIGNS API
# ============================================
@app.get("/api/campaigns")
def list_campaigns(db: Session = Depends(get_db)):
    try:
        campaigns = db.query(Campaign).order_by(Campaign.created_at.desc()).all()
        result = []
        for c in campaigns:
            leads = db.query(Lead).filter(Lead.campaign_id == c.id).all()
            lead_stats = {"total": len(leads), "new": 0, "contacted": 0, "qualified": 0, "converted": 0, "lost": 0}
            for l in leads:
                if l.status in lead_stats: lead_stats[l.status] += 1
            result.append({
                "id": str(c.id), "campaign_id": c.campaign_id, "name": c.name,
                "source": c.source, "start_date": str(c.start_date) if c.start_date else None,
                "end_date": str(c.end_date) if c.end_date else None,
                "budget": float(c.budget) if c.budget else None, "status": c.status,
                "notes": c.notes, "lead_stats": lead_stats
            })
        return result
    except Exception as e:
        print(f"Error listing campaigns: {e}")
        return []

@app.get("/api/campaigns/summary")
def get_campaigns_summary(db: Session = Depends(get_db)):
    try:
        total = db.query(Campaign).count()
        active = db.query(Campaign).filter(Campaign.status == "active").count()
        total_leads = db.query(Lead).count()
        converted = db.query(Lead).filter(Lead.status == "converted").count()
        total_budget = db.query(func.sum(Campaign.budget)).scalar() or 0
        return {
            "total_campaigns": total, "active_campaigns": active,
            "total_leads": total_leads, "converted_leads": converted,
            "total_budget": float(total_budget),
            "conversion_rate": round(converted / total_leads * 100, 1) if total_leads > 0 else 0
        }
    except Exception as e:
        print(f"Error getting campaigns summary: {e}")
        return {"total_campaigns": 0, "active_campaigns": 0, "total_leads": 0, "converted_leads": 0, "total_budget": 0, "conversion_rate": 0}

@app.post("/api/campaigns")
def create_campaign(data: dict, db: Session = Depends(get_db)):
    try:
        c = Campaign(
            name=data["name"], source=data.get("source"),
            start_date=date.fromisoformat(data["start_date"]) if data.get("start_date") else None,
            end_date=date.fromisoformat(data["end_date"]) if data.get("end_date") else None,
            budget=data.get("budget") if data.get("budget") else None, 
            notes=data.get("notes")
        )
        db.add(c); db.commit(); db.refresh(c)
        return {"message": "Campaign created", "id": str(c.id), "campaign_id": c.campaign_id}
    except Exception as e:
        db.rollback()
        print(f"Error creating campaign: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Error creating campaign: {str(e)}")

@app.put("/api/campaigns/{cid}")
def update_campaign(cid: str, data: dict, db: Session = Depends(get_db)):
    c = db.query(Campaign).filter((Campaign.id == cid) | (Campaign.campaign_id == cid)).first()
    if not c: raise HTTPException(404, "Campaign not found")
    for k in ["name", "source", "budget", "status", "notes"]:
        if k in data and data[k] is not None: setattr(c, k, data[k])
    if "start_date" in data: c.start_date = date.fromisoformat(data["start_date"]) if data["start_date"] else None
    if "end_date" in data: c.end_date = date.fromisoformat(data["end_date"]) if data["end_date"] else None
    db.commit()
    return {"message": "Campaign updated"}

@app.delete("/api/campaigns/{cid}")
def delete_campaign(cid: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    c = find_entity(db, Campaign, "campaign_id", cid)
    if not c: raise HTTPException(404, "Campaign not found")

    # Creator role cannot delete
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    # Admin can delete directly
    if current_user.role == "admin":
        db.delete(c); db.commit()
        return {"message": "Campaign deleted"}

    # Non-admin: create deletion request
    req = DeletionRequest(entity_type="campaign", entity_id=c.id, entity_name=c.name, requested_by=current_user.id)
    db.add(req); db.commit(); db.refresh(req)
    return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

# ============================================
# LEADS API
# ============================================
@app.get("/api/leads")
def list_leads(campaign_id: str = None, rep_id: str = None, status: str = None,
               limit: int = 100, db: Session = Depends(get_db),
               current_user: CompanyRep = Depends(get_current_user)):
    try:
        q = db.query(Lead)
        # Data isolation
        iso = get_rep_isolation_filter(current_user, db)
        if iso["isolated"]:
            q = q.filter(Lead.assigned_rep_id.in_(iso["team_rep_uuids"]))
        if campaign_id and campaign_id.strip():
            c = db.query(Campaign).filter((Campaign.id == campaign_id) | (Campaign.campaign_id == campaign_id)).first()
            if c: q = q.filter(Lead.campaign_id == c.id)
        if rep_id and rep_id.strip():
            r = db.query(CompanyRep).filter((CompanyRep.id == rep_id) | (CompanyRep.rep_id == rep_id)).first()
            if r: q = q.filter(Lead.assigned_rep_id == r.id)
        if status and status.strip(): q = q.filter(Lead.status == status)

        leads = q.order_by(Lead.created_at.desc()).limit(limit).all()
        # Pre-fetch related entities to avoid N+1
        camp_ids = set(l.campaign_id for l in leads if l.campaign_id)
        rep_ids = set(l.assigned_rep_id for l in leads if l.assigned_rep_id)
        camps_map = {c.id: c for c in db.query(Campaign).filter(Campaign.id.in_(camp_ids)).all()} if camp_ids else {}
        reps_map = {r.id: r for r in db.query(CompanyRep).filter(CompanyRep.id.in_(rep_ids)).all()} if rep_ids else {}
        result = []
        for l in leads:
            campaign = camps_map.get(l.campaign_id)
            rep = reps_map.get(l.assigned_rep_id)
            result.append({
                "id": str(l.id), "lead_id": l.lead_id, "name": l.name,
                "mobile": l.mobile, "email": l.email,
                "campaign_name": campaign.name if campaign else None,
                "campaign_id": campaign.campaign_id if campaign else None,
                "assigned_rep": rep.name if rep else None, "rep_id": rep.rep_id if rep else None,
                "status": l.status, "lead_type": l.lead_type or "prospect", "notes": l.notes,
                "pipeline_stage": l.pipeline_stage or "New",
                "converted_customer_id": str(l.converted_customer_id) if l.converted_customer_id else None,
                "converted_broker_id": str(l.converted_broker_id) if l.converted_broker_id else None,
                "additional_mobiles": l.additional_mobiles or [],
                "source": l.source, "source_other": l.source_other,
                "occupation": l.occupation, "occupation_other": l.occupation_other,
                "interested_project_id": str(l.interested_project_id) if l.interested_project_id else None,
                "interested_project_other": l.interested_project_other,
                "area": l.area, "city": l.city,
                "country_code": l.country_code or "+92",
                "source_details": l.source_details,
                "temperature": l.temperature,
                "created_at": str(l.created_at)
            })
        return result
    except Exception as e:
        print(f"Error listing leads: {e}")
        return []

@app.post("/api/leads")
def create_lead(data: dict, db: Session = Depends(get_db),
                current_user: CompanyRep = Depends(get_current_user)):
    try:
        campaign = None
        rep = None
        campaign_name = None
        if data.get("campaign_id"):
            campaign = db.query(Campaign).filter((Campaign.id == data["campaign_id"]) | (Campaign.campaign_id == data["campaign_id"])).first()
            campaign_name = campaign.name if campaign else None
        if data.get("assigned_rep_id"):
            rep = db.query(CompanyRep).filter((CompanyRep.id == data["assigned_rep_id"]) | (CompanyRep.rep_id == data["assigned_rep_id"])).first()

        # Check for duplicates — block creation if mobile matches existing record
        raw_mobile = data.get("mobile", "").strip() if data.get("mobile") else ""
        mobile = normalize_mobile(raw_mobile) or raw_mobile
        if mobile:
            dupes = check_duplicate_mobile(db, mobile)
            if dupes:
                # Send notifications before blocking
                notify_duplicate_attempt(db, dupes, current_user, "single", campaign_name)
                db.commit()
                dup = dupes[0]
                raise HTTPException(
                    status_code=409,
                    detail={
                        "message": "Duplicate mobile number detected",
                        "duplicate": {
                            "type": dup["type"],
                            "entity_id": dup["entity_id"],
                            "name": dup["name"],
                            "mobile": dup["mobile"],
                            "assigned_rep": dup.get("assigned_rep"),
                            "assigned_rep_id": dup.get("assigned_rep_id")
                        }
                    }
                )

        # Resolve interested_project_id
        interested_project_id = None
        if data.get("interested_project_id") and data["interested_project_id"] != "other":
            p = db.query(Project).filter(
                (Project.id == data["interested_project_id"]) | (Project.project_id == data["interested_project_id"])
            ).first()
            interested_project_id = p.id if p else None
        # Filter additional mobiles
        additional = [m for m in (data.get("additional_mobiles") or []) if m and m.strip()]
        l = Lead(
            campaign_id=campaign.id if campaign else None,
            assigned_rep_id=rep.id if rep else current_user.id,
            name=data["name"], mobile=mobile or None, email=data.get("email"),
            source_details=data.get("source_details"), status=data.get("status", "new"),
            lead_type=data.get("lead_type", "prospect"), notes=data.get("notes"),
            pipeline_stage=data.get("pipeline_stage", "New"),
            additional_mobiles=additional if additional else [],
            source=data.get("source"), source_other=data.get("source_other"),
            occupation=data.get("occupation"), occupation_other=data.get("occupation_other"),
            interested_project_id=interested_project_id,
            interested_project_other=data.get("interested_project_other"),
            area=data.get("area"), city=data.get("city"),
            country_code=data.get("country_code", "+92"),
            temperature=data.get("temperature")
        )
        db.add(l); db.commit(); db.refresh(l)
        return {"message": "Lead created", "id": str(l.id), "lead_id": l.lead_id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        # Auto-recover from stale lead_id_seq (UniqueViolation on lead_id)
        if "leads_lead_id_key" in str(e):
            print(f"[auto-recover] lead_id_seq stale, syncing and retrying...")
            sync_lead_id_sequence(db)
            try:
                l2 = Lead(
                    campaign_id=campaign.id if campaign else None,
                    assigned_rep_id=rep.id if rep else current_user.id,
                    name=data["name"], mobile=mobile or None, email=data.get("email"),
                    source_details=data.get("source_details"), status=data.get("status", "new"),
                    lead_type=data.get("lead_type", "prospect"), notes=data.get("notes"),
                    pipeline_stage=data.get("pipeline_stage", "New"),
                    additional_mobiles=additional if additional else [],
                    source=data.get("source"), source_other=data.get("source_other"),
                    occupation=data.get("occupation"), occupation_other=data.get("occupation_other"),
                    interested_project_id=interested_project_id,
                    interested_project_other=data.get("interested_project_other"),
                    area=data.get("area"), city=data.get("city"),
                    country_code=data.get("country_code", "+92"),
                    temperature=data.get("temperature")
                )
                db.add(l2); db.commit(); db.refresh(l2)
                print(f"[auto-recover] Retry succeeded — created {l2.lead_id}")
                return {"message": "Lead created", "id": str(l2.id), "lead_id": l2.lead_id}
            except Exception as e2:
                db.rollback()
                print(f"[auto-recover] Retry also failed: {e2}")
                raise HTTPException(500, f"Error creating lead after sequence fix: {str(e2)}")
        print(f"Error creating lead: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Error creating lead: {str(e)}")

@app.get("/api/leads/template/download")
def download_leads_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['name*', 'mobile', 'email', 'assigned_rep', 'assigned_broker', 'pre_assigned_customer', 'lead_type', 'source_details', 'source', 'occupation', 'area', 'city', 'notes'])
    writer.writerow(['John Doe', '0300-1234567', 'john@email.com', 'REP-0001', 'BRK-0001', 'CUST-0001', 'prospect', 'Facebook form', 'Social Media', 'Businessman', 'DHA Phase 5', 'Karachi', 'Interested in Phase 2'])
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=leads_template.csv"})

@app.post("/api/leads/bulk-import")
async def bulk_import_leads(file: UploadFile = File(...), campaign_id: str = None, db: Session = Depends(get_db),
                            current_user: CompanyRep = Depends(get_current_user)):
    campaign = None
    campaign_name = None
    if campaign_id:
        campaign = db.query(Campaign).filter((Campaign.id == campaign_id) | (Campaign.campaign_id == campaign_id)).first()
        campaign_name = campaign.name if campaign else None

    content = await file.read()
    reader = csv.DictReader(io.StringIO(content.decode('utf-8-sig')))
    results = {"success": 0, "errors": [], "duplicates": [], "warnings": []}
    all_dup_records = []  # For notification at end
    for i, row in enumerate(reader, 2):
        try:
            name = row.get('name', row.get('name*', '')).strip()
            if not name: results["errors"].append(f"Row {i}: Name required"); continue

            mobile = normalize_mobile((row.get('mobile') or '').strip()) or (row.get('mobile') or '').strip()
            # Check for duplicates if mobile provided (include converted leads to find best owner)
            if mobile:
                dupes = check_duplicate_mobile(db, mobile, include_converted=True)
                if dupes:
                    for dup in dupes:
                        results["duplicates"].append({
                            "row": i, "uploaded_name": name, "uploaded_mobile": mobile,
                            "system_type": dup["type"], "system_entity_id": dup["entity_id"],
                            "system_name": dup["name"], "system_mobile": dup["mobile"],
                            "system_assigned_rep": dup.get("assigned_rep") or "N/A"
                        })
                        all_dup_records.append(dup)
                    # Create ONE task per duplicate row — find best assignee across all matches
                    best_rep = None
                    for dup in dupes:
                        rid = dup.get("assigned_rep_id")
                        if rid:
                            best_rep = db.query(CompanyRep).filter(CompanyRep.rep_id == rid).first()
                            if best_rep:
                                break
                    assignee = best_rep or current_user
                    primary_dup = dupes[0]
                    dup_summary = ", ".join(f"{d['type']} {d['entity_id']}" for d in dupes)
                    task = Task(
                        title=f"Follow up with {primary_dup['name']} — showed interest in {campaign_name or 'campaign'}",
                        description=f"Duplicate detected during bulk import. {name} (mobile: {mobile}) matches: {dup_summary}. Campaign: {campaign_name or 'N/A'}.",
                        task_type="follow_up",
                        priority="high",
                        status="pending",
                        assignee_id=assignee.id,
                        created_by=current_user.id,
                        due_date=date.today() + timedelta(days=2)
                    )
                    db.add(task)
                    notify_rep_id = best_rep.rep_id if best_rep else current_user.rep_id
                    create_notification(
                        db, notify_rep_id, "follow_up",
                        f"New interest: {primary_dup['name']}",
                        f"{name} showed interest in {campaign_name or 'campaign'}. A follow-up task has been auto-created.",
                        category="lead", entity_type="lead", entity_id=primary_dup["entity_id"]
                    )
                    continue  # Skip this row

            rep = None
            if row.get('assigned_rep'):
                rep = db.query(CompanyRep).filter(
                    (CompanyRep.rep_id == row['assigned_rep']) | (CompanyRep.name == row['assigned_rep'])
                ).first()
                if not rep:
                    results["warnings"].append(f"Row {i}: assigned_rep '{row['assigned_rep']}' not found, skipping assignment")

            # Lookup assigned_broker (optional)
            broker = None
            if row.get('assigned_broker'):
                broker = db.query(Broker).filter(
                    (Broker.broker_id == row['assigned_broker']) | (Broker.name == row['assigned_broker'])
                ).first()
                if not broker:
                    results["warnings"].append(f"Row {i}: assigned_broker '{row['assigned_broker']}' not found, skipping assignment")

            # Lookup pre_assigned_customer (optional)
            pre_customer = None
            if row.get('pre_assigned_customer'):
                pre_customer = db.query(Customer).filter(
                    (Customer.customer_id == row['pre_assigned_customer']) | (Customer.mobile == row['pre_assigned_customer'])
                ).first()
                if not pre_customer:
                    results["warnings"].append(f"Row {i}: pre_assigned_customer '{row['pre_assigned_customer']}' not found, skipping")

            l = Lead(
                campaign_id=campaign.id if campaign else None,
                assigned_rep_id=rep.id if rep else None,
                name=name, mobile=mobile or None, email=row.get('email'),
                source_details=row.get('source_details'), notes=row.get('notes'),
                source=row.get('source'), occupation=row.get('occupation'),
                area=row.get('area'), city=row.get('city'),
                metadata={
                    "assigned_broker_id": str(broker.id) if broker else None,
                    "assigned_broker_name": broker.name if broker else None,
                    "pre_assigned_customer_id": str(pre_customer.id) if pre_customer else None,
                    "pre_assigned_customer_name": pre_customer.name if pre_customer else None,
                }
            )
            db.add(l); db.flush()
            results["success"] += 1
        except Exception as e: results["errors"].append(f"Row {i}: {e}")
    db.commit()
    # Auto-sync lead_id_seq after bulk import to prevent stale sequence errors
    if results["success"] > 0:
        sync_lead_id_sequence(db)
    # Notify admin/CCO about duplicates (not individual reps — individual already notified above)
    if all_dup_records:
        notify_duplicate_attempt(db, all_dup_records, current_user, "bulk", campaign_name)
        db.commit()
    return results

@app.put("/api/leads/{lid}")
def update_lead(lid: str, data: dict, db: Session = Depends(get_db),
                current_user: CompanyRep = Depends(get_current_user)):
    l = db.query(Lead).filter((Lead.id == lid) | (Lead.lead_id == lid)).first()
    if not l: raise HTTPException(404, "Lead not found")
    for k in ["name", "mobile", "email", "source_details", "status", "lead_type", "notes", "pipeline_stage",
              "additional_mobiles", "source", "source_other", "occupation", "occupation_other",
              "interested_project_other", "area", "city", "country_code", "temperature"]:
        if k in data: setattr(l, k, data[k])
    if "interested_project_id" in data:
        if data["interested_project_id"] and data["interested_project_id"] != "other":
            p = db.query(Project).filter(
                (Project.id == data["interested_project_id"]) | (Project.project_id == data["interested_project_id"])
            ).first()
            l.interested_project_id = p.id if p else None
        else:
            l.interested_project_id = None
    if "assigned_rep_id" in data:
        rep = None
        if data["assigned_rep_id"]:
            try:
                rid = uuid.UUID(str(data["assigned_rep_id"]))
                rep = db.query(CompanyRep).filter((CompanyRep.id == rid) | (CompanyRep.rep_id == data["assigned_rep_id"])).first()
            except (ValueError, AttributeError):
                rep = db.query(CompanyRep).filter(CompanyRep.rep_id == data["assigned_rep_id"]).first()
        l.assigned_rep_id = rep.id if rep else None
    db.commit()
    return {"message": "Lead updated"}

@app.post("/api/leads/{lid}/convert")
def convert_lead(lid: str, data: dict, db: Session = Depends(get_db),
                 current_user: CompanyRep = Depends(get_current_user)):
    """Sync a lead to the customer or broker database. Does NOT mark as converted — that only happens when pipeline_stage is set to Won."""
    if current_user.role not in ["admin", "director", "cco", "coo", "manager"]:
        raise HTTPException(403, "Only manager/director/cco/coo/admin can sync leads to customer/broker DB.")
    l = db.query(Lead).filter((Lead.id == lid) | (Lead.lead_id == lid)).first()
    if not l: raise HTTPException(404, "Lead not found")

    convert_to = data.get("convert_to")  # 'customer' or 'broker'
    if convert_to == "customer":
        linked_existing, entity_id, _ = _sync_lead_to_customer_record(l, db)
        db.commit()
        return {"message": "Lead synced to customer DB", "customer_id": str(l.converted_customer_id),
                "linked_existing": linked_existing, "entity_id": entity_id}

    elif convert_to == "broker":
        existing = None
        if l.mobile:
            normalized = normalize_mobile(l.mobile)
            if normalized:
                patterns = [normalized]
                if normalized.startswith("0"):
                    patterns.append("+92" + normalized[1:])
                    patterns.append("92" + normalized[1:])
                existing = db.query(Broker).filter(or_(*[Broker.mobile.ilike(p) for p in patterns])).first()
        linked_existing = False
        if existing:
            l.converted_broker_id = existing.id
            linked_existing = True
            entity_id = existing.broker_id
        else:
            b = Broker(name=l.name, mobile=l.mobile or f"lead-{l.lead_id}", email=l.email)
            db.add(b); db.flush()
            l.converted_broker_id = b.id
            entity_id = b.broker_id
        l.lead_type = "broker"
        db.commit()
        return {"message": "Lead synced to broker DB", "broker_id": str(l.converted_broker_id),
                "linked_existing": linked_existing, "entity_id": entity_id}

    raise HTTPException(400, "Invalid conversion type")

@app.delete("/api/leads/{lid}")
def delete_lead(lid: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    l = find_entity(db, Lead, "lead_id", lid)
    if not l: raise HTTPException(404, "Lead not found")

    # Creator role cannot delete
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    # Admin can delete directly
    if current_user.role == "admin":
        db.delete(l); db.commit()
        return {"message": "Lead deleted"}

    # Non-admin: create deletion request
    entity_name = l.name if l.name else str(l.id)
    req = DeletionRequest(entity_type="lead", entity_id=l.id, entity_name=entity_name, requested_by=current_user.id)
    db.add(req); db.commit(); db.refresh(req)
    return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

# ============================================
# RECEIPTS API
# ============================================
@app.get("/api/receipts")
def list_receipts(customer_id: str = None, transaction_id: str = None, limit: int = 100,
                  db: Session = Depends(get_db),
                  current_user: CompanyRep = Depends(get_current_user)):
    try:
        q = db.query(Receipt)
        # Data isolation: direct/both reps see receipts for their customers only
        iso = get_rep_isolation_filter(current_user, db)
        if iso["isolated"]:
            rt = iso["rep_type"]
            if rt == "indirect":
                return []  # indirect reps don't see receipts
            own_cust_ids = [c.id for c in db.query(Customer.id).filter(Customer.assigned_rep_id.in_(iso["team_rep_uuids"])).all()]
            q = q.filter(Receipt.customer_id.in_(own_cust_ids)) if own_cust_ids else q.filter(literal(False))
        if customer_id and customer_id.strip():
            c = db.query(Customer).filter((Customer.id == customer_id) | (Customer.customer_id == customer_id)).first()
            if c: q = q.filter(Receipt.customer_id == c.id)
        if transaction_id and transaction_id.strip():
            t = find_entity(db, Transaction, "transaction_id", transaction_id)
            if t: q = q.filter(Receipt.transaction_id == t.id)
        
        receipts = q.order_by(Receipt.created_at.desc()).limit(limit).all()
        result = []
        for r in receipts:
            cust = db.query(Customer).filter(Customer.id == r.customer_id).first()
            txn = db.query(Transaction).filter(Transaction.id == r.transaction_id).first() if r.transaction_id else None
            proj = db.query(Project).filter(Project.id == txn.project_id).first() if txn else None
            rep = db.query(CompanyRep).filter(CompanyRep.id == r.created_by_rep_id).first() if r.created_by_rep_id else None
            
            # Get allocations
            allocations = db.query(ReceiptAllocation).filter(ReceiptAllocation.receipt_id == r.id).all()
            alloc_details = []
            for a in allocations:
                inst = db.query(Installment).filter(Installment.id == a.installment_id).first()
                alloc_details.append({
                    "id": str(a.id), "installment_number": inst.installment_number if inst else None,
                    "amount": float(a.amount)
                })
            
            entry = {
                "id": str(r.id), "receipt_id": r.receipt_id,
                "customer_name": cust.name if cust else None, "customer_id": cust.customer_id if cust else None,
                "transaction_id": txn.transaction_id if txn else None,
                "project_name": proj.name if proj else None, "unit_number": txn.unit_number if txn else None,
                "amount": float(r.amount), "payment_method": r.payment_method,
                "reference_number": r.reference_number,
                "payment_date": str(r.payment_date) if r.payment_date else None,
                "notes": r.notes, "created_by": rep.name if rep else None,
                "created_at": str(r.created_at), "allocations": alloc_details
            }
            # Add classification if receipt is linked to a transaction
            if txn:
                try:
                    try:
                        from app.services.receipt_classification_service import classify_receipt as _classify
                    except ImportError:
                        from services.receipt_classification_service import classify_receipt as _classify
                    entry["classification"] = _classify(txn, db)
                except Exception as exc:
                    logging.getLogger("orbit.receipt_classification").warning(
                        "Classification failed for receipt %s: %s", r.receipt_id, exc
                    )
                    entry["classification_error"] = str(exc)
            result.append(entry)
        return result
    except Exception as e:
        print(f"Error listing receipts: {e}")
        return []

@app.get("/api/receipts/summary")
def get_receipts_summary(db: Session = Depends(get_db)):
    try:
        today = date.today()
        month_start = today.replace(day=1)
        
        total_receipts = db.query(Receipt).count()
        total_amount = db.query(func.sum(Receipt.amount)).scalar() or 0
        today_count = db.query(Receipt).filter(Receipt.payment_date == today).count()
        today_amount = db.query(func.sum(Receipt.amount)).filter(Receipt.payment_date == today).scalar() or 0
        month_count = db.query(Receipt).filter(Receipt.payment_date >= month_start).count()
        month_amount = db.query(func.sum(Receipt.amount)).filter(Receipt.payment_date >= month_start).scalar() or 0
        
        # By payment method
        by_method = {}
        for method in ['cash', 'cheque', 'bank_transfer', 'online']:
            amt = db.query(func.sum(Receipt.amount)).filter(Receipt.payment_method == method).scalar() or 0
            by_method[method] = float(amt)
        
        return {
            "total_receipts": total_receipts, "total_amount": float(total_amount),
            "today_count": today_count, "today_amount": float(today_amount),
            "month_count": month_count, "month_amount": float(month_amount),
            "by_method": by_method
        }
    except Exception as e:
        print(f"Error getting receipts summary: {e}")
        return {"total_receipts": 0, "total_amount": 0, "today_count": 0, "today_amount": 0, "month_count": 0, "month_amount": 0, "by_method": {}}

@app.get("/api/receipts/customer/{cid}/transactions")
def get_customer_transactions_for_receipt(cid: str, db: Session = Depends(get_db)):
    """Get customer's transactions with installment details for receipt allocation"""
    c = db.query(Customer).filter((Customer.id == cid) | (Customer.customer_id == cid) | (Customer.mobile == cid)).first()
    if not c: raise HTTPException(404, "Customer not found")
    
    txns = db.query(Transaction).filter(Transaction.customer_id == c.id).all()
    result = []
    for t in txns:
        proj = db.query(Project).filter(Project.id == t.project_id).first()
        installments = db.query(Installment).filter(Installment.transaction_id == t.id).order_by(Installment.installment_number).all()
        
        inst_list = []
        for i in installments:
            balance = float(i.amount) - float(i.amount_paid or 0)
            if balance > 0:  # Only show unpaid installments
                inst_list.append({
                    "id": str(i.id), "number": i.installment_number,
                    "due_date": str(i.due_date), "amount": float(i.amount),
                    "paid": float(i.amount_paid or 0), "balance": balance, "status": i.status
                })
        
        if inst_list:  # Only include transactions with pending installments
            result.append({
                "id": str(t.id), "transaction_id": t.transaction_id,
                "project_name": proj.name if proj else None, "unit_number": t.unit_number,
                "total_value": float(t.total_value),
                "total_paid": sum(float(i.amount_paid or 0) for i in installments),
                "balance": float(t.total_value) - sum(float(i.amount_paid or 0) for i in installments),
                "installments": inst_list
            })
    return result

@app.post("/api/receipts")
def create_receipt(data: dict, db: Session = Depends(get_db)):
    customer = find_entity(db, Customer, "customer_id", data["customer_id"]) or db.query(Customer).filter(
        Customer.mobile == data["customer_id"]
    ).first()
    if not customer: raise HTTPException(404, "Customer not found")
    transaction = None
    if data.get("transaction_id"):
        transaction = find_entity(db, Transaction, "transaction_id", data["transaction_id"])
    
    rep = None
    if data.get("created_by_rep_id"):
        rep = find_entity(db, CompanyRep, "rep_id", data["created_by_rep_id"])
    
    r = Receipt(
        customer_id=customer.id,
        transaction_id=transaction.id if transaction else None,
        amount=data["amount"],
        payment_method=data.get("payment_method"),
        reference_number=data.get("reference_number"),
        payment_date=date.fromisoformat(data["payment_date"]) if data.get("payment_date") else date.today(),
        notes=data.get("notes"),
        created_by_rep_id=rep.id if rep else None
    )
    db.add(r); db.flush()
    
    # Auto-allocate to installments if transaction specified
    allocations = data.get("allocations", [])
    if allocations:
        for alloc in allocations:
            inst = db.query(Installment).filter(Installment.id == alloc["installment_id"]).first()
            if inst:
                db.add(ReceiptAllocation(receipt_id=r.id, installment_id=inst.id, amount=alloc["amount"]))
                inst.amount_paid = float(inst.amount_paid or 0) + float(alloc["amount"])
                if float(inst.amount_paid) >= float(inst.amount):
                    inst.status = "paid"
                elif float(inst.amount_paid) > 0:
                    inst.status = "partial"
    elif transaction:
        # Auto-allocate to oldest unpaid installments
        remaining = float(data["amount"])
        installments = db.query(Installment).filter(
            Installment.transaction_id == transaction.id,
            Installment.status != "paid"
        ).order_by(Installment.installment_number).all()
        
        for inst in installments:
            if remaining <= 0: break
            balance = float(inst.amount) - float(inst.amount_paid or 0)
            alloc_amount = min(remaining, balance)
            
            db.add(ReceiptAllocation(receipt_id=r.id, installment_id=inst.id, amount=alloc_amount))
            inst.amount_paid = float(inst.amount_paid or 0) + alloc_amount
            if float(inst.amount_paid) >= float(inst.amount):
                inst.status = "paid"
            elif float(inst.amount_paid) > 0:
                inst.status = "partial"
            remaining -= alloc_amount
    
    db.commit(); db.refresh(r)

    # ── Receipt classification (read-only against payment rules) ───────
    result = {"message": "Receipt created", "id": str(r.id), "receipt_id": r.receipt_id}
    if transaction:
        try:
            try:
                from app.services.receipt_classification_service import classify_receipt as _classify
            except ImportError:
                from services.receipt_classification_service import classify_receipt as _classify
            result["classification"] = _classify(transaction, db)
        except Exception as exc:
            logging.getLogger("orbit.receipt_classification").warning(
                "Classification failed for new receipt %s: %s", r.receipt_id, exc
            )
            result["classification_error"] = str(exc)
    return result

@app.get("/api/receipts/{rid}")
def get_receipt(rid: str, db: Session = Depends(get_db),
                current_user: CompanyRep = Depends(get_current_user)):
    """Get single receipt with classification details."""
    r = db.query(Receipt).filter(
        (Receipt.id == rid) | (Receipt.receipt_id == rid)
    ).first()
    if not r:
        raise HTTPException(404, "Receipt not found")

    cust = db.query(Customer).filter(Customer.id == r.customer_id).first()
    txn = db.query(Transaction).filter(Transaction.id == r.transaction_id).first() if r.transaction_id else None
    proj = db.query(Project).filter(Project.id == txn.project_id).first() if txn else None
    rep = db.query(CompanyRep).filter(CompanyRep.id == r.created_by_rep_id).first() if r.created_by_rep_id else None

    allocations = db.query(ReceiptAllocation).filter(ReceiptAllocation.receipt_id == r.id).all()
    alloc_details = []
    for a in allocations:
        inst = db.query(Installment).filter(Installment.id == a.installment_id).first()
        alloc_details.append({
            "id": str(a.id), "installment_number": inst.installment_number if inst else None,
            "amount": float(a.amount)
        })

    result = {
        "id": str(r.id), "receipt_id": r.receipt_id,
        "customer_name": cust.name if cust else None, "customer_id": cust.customer_id if cust else None,
        "transaction_id": txn.transaction_id if txn else None,
        "project_name": proj.name if proj else None, "unit_number": txn.unit_number if txn else None,
        "amount": float(r.amount), "payment_method": r.payment_method,
        "reference_number": r.reference_number,
        "payment_date": str(r.payment_date) if r.payment_date else None,
        "notes": r.notes, "created_by": rep.name if rep else None,
        "created_at": str(r.created_at), "allocations": alloc_details
    }

    # Classification
    if txn:
        try:
            try:
                from app.services.receipt_classification_service import classify_receipt as _classify
            except ImportError:
                from services.receipt_classification_service import classify_receipt as _classify
            result["classification"] = _classify(txn, db)
        except Exception as exc:
            logging.getLogger("orbit.receipt_classification").warning(
                "Classification failed for receipt %s: %s", rid, exc
            )
            result["classification_error"] = str(exc)

    return result

@app.delete("/api/receipts/{rid}")
def delete_receipt(rid: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    r = find_entity(db, Receipt, "receipt_id", rid)
    if not r: raise HTTPException(404, "Receipt not found")

    # Creator role cannot delete
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    # Admin can delete directly
    if current_user.role == "admin":
        # Reverse allocations
        allocations = db.query(ReceiptAllocation).filter(ReceiptAllocation.receipt_id == r.id).all()
        for a in allocations:
            inst = db.query(Installment).filter(Installment.id == a.installment_id).first()
            if inst:
                inst.amount_paid = max(0, float(inst.amount_paid or 0) - float(a.amount))
                if float(inst.amount_paid) >= float(inst.amount):
                    inst.status = "paid"
                elif float(inst.amount_paid) > 0:
                    inst.status = "partial"
                else:
                    inst.status = "pending"
            db.delete(a)

        db.delete(r); db.commit()
        return {"message": "Receipt deleted"}

    # Non-admin: create deletion request
    req = DeletionRequest(entity_type="receipt", entity_id=r.id, entity_name=r.receipt_id, requested_by=current_user.id)
    db.add(req); db.commit(); db.refresh(req)
    return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

# ============================================
# EOI COLLECTION API
# ============================================
VALID_EOI_STATUSES = {"active", "converted", "cancelled", "refunded"}

def _eoi_slab_label(marlas_val: Optional[float]) -> str:
    if marlas_val is None:
        return "Unknown"
    if marlas_val >= 3 and marlas_val <= 5:
        return "3-5"
    if marlas_val > 5 and marlas_val <= 8:
        return "5-8"
    if marlas_val > 8:
        return "8+"
    return "Other"

@app.get("/api/eoi")
def list_eoi_collections(
    project_id: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    rep_id: Optional[str] = None,
    broker_id: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 200,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    try:
        q = db.query(EOICollection)
        isolation = get_rep_isolation_filter(current_user, db)
        scoped_rep_ids = None
        if isolation["isolated"]:
            scoped_rep_ids = isolation["team_rep_uuids"]
            q = q.filter(EOICollection.created_by.in_(scoped_rep_ids))

        if project_id:
            proj = _resolve_eoi_project(db, project_id)
            if proj:
                q = q.filter(EOICollection.project_id == proj.id)
            else:
                q = q.filter(literal(False))

        if status:
            status_list = [s.strip().lower() for s in status.split(",") if s.strip()]
            valid = [s for s in status_list if s in VALID_EOI_STATUSES]
            if valid:
                q = q.filter(EOICollection.status.in_(valid))

        if date_from:
            q = q.filter(EOICollection.eoi_date >= date.fromisoformat(date_from))
        if date_to:
            q = q.filter(EOICollection.eoi_date <= date.fromisoformat(date_to))

        if rep_id:
            rep_obj = db.query(CompanyRep).filter(
                (CompanyRep.rep_id == rep_id) | (func.cast(CompanyRep.id, String) == rep_id)
            ).first()
            if rep_obj:
                if not isolation["isolated"] or rep_obj.id in (scoped_rep_ids or []):
                    q = q.filter(EOICollection.created_by == rep_obj.id)
                else:
                    q = q.filter(literal(False))

        if broker_id:
            broker = _resolve_eoi_broker(db, broker_id)
            if broker:
                q = q.filter(EOICollection.broker_id == broker.id)
            elif broker_id.lower() == "direct":
                q = q.filter(EOICollection.broker_id.is_(None))
            else:
                q = q.filter(literal(False))

        if search:
            term = f"%{search.strip()}%"
            q = q.filter(
                or_(
                    EOICollection.eoi_id.ilike(term),
                    EOICollection.party_name.ilike(term),
                    EOICollection.party_mobile.ilike(term),
                    EOICollection.reference_number.ilike(term),
                    EOICollection.broker_name.ilike(term),
                    EOICollection.unit_number.ilike(term),
                )
            )

        rows = q.order_by(EOICollection.eoi_date.desc(), EOICollection.created_at.desc()).limit(limit).all()
        payload = [_eoi_row_dict(row, db) for row in rows]
        return {
            "items": payload,
            "total": len(payload),
            "filters": {
                "project_id": project_id,
                "status": status,
                "date_from": date_from,
                "date_to": date_to,
                "rep_id": rep_id,
                "broker_id": broker_id,
                "search": search,
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"EOI list error: {str(e)}")

@app.post("/api/eoi")
def create_eoi_collection(
    data: dict,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    try:
        proj_raw = data.get("project_id")
        if not proj_raw:
            raise HTTPException(400, "project_id is required")
        proj = _resolve_eoi_project(db, proj_raw)
        if not proj:
            raise HTTPException(404, "Project not found")

        amount = float(data.get("amount") or 0)
        if amount <= 0:
            raise HTTPException(400, "amount must be greater than 0")

        party_name = (data.get("party_name") or "").strip()
        if not party_name:
            raise HTTPException(400, "party_name is required")

        status = (data.get("status") or "active").lower()
        if status not in VALID_EOI_STATUSES:
            raise HTTPException(400, "Invalid status")

        customer = _resolve_eoi_customer(db, data.get("customer_id"))
        broker = _resolve_eoi_broker(db, data.get("broker_id"))
        inventory = _resolve_eoi_inventory(db, data.get("inventory_id"))

        eoi = EOICollection(
            eoi_id=_generate_eoi_code(db),
            project_id=proj.id,
            customer_id=customer.id if customer else None,
            broker_id=broker.id if broker else None,
            party_name=party_name,
            party_mobile=normalize_mobile(data.get("party_mobile")) if data.get("party_mobile") else None,
            party_cnic=data.get("party_cnic"),
            broker_name=(broker.name if broker else data.get("broker_name")),
            amount=amount,
            marlas=float(data["marlas"]) if data.get("marlas") not in (None, "") else None,
            unit_number=data.get("unit_number"),
            inventory_id=inventory.id if inventory else None,
            payment_method=data.get("payment_method"),
            reference_number=data.get("reference_number"),
            payment_received=bool(data.get("payment_received", False)),
            payment_received_at=datetime.utcnow() if data.get("payment_received") else None,
            eoi_date=date.fromisoformat(data["eoi_date"]) if data.get("eoi_date") else date.today(),
            notes=data.get("notes"),
            status=status,
            created_by=current_user.id,
        )
        db.add(eoi)
        db.commit()
        db.refresh(eoi)
        return {"message": "EOI created", "item": _eoi_row_dict(eoi, db)}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"EOI create error: {str(e)}")

@app.put("/api/eoi/{eoi_id}")
def update_eoi_collection(
    eoi_id: str,
    data: dict,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    try:
        eoi = _resolve_eoi_record(db, eoi_id)
        if not eoi:
            raise HTTPException(404, "EOI not found")

        isolation = get_rep_isolation_filter(current_user, db)
        if isolation["isolated"] and eoi.created_by not in isolation["team_rep_uuids"]:
            raise HTTPException(403, "Access denied")

        if eoi.status == "converted":
            raise HTTPException(400, "Converted EOI cannot be edited")

        if "project_id" in data and data["project_id"]:
            proj = _resolve_eoi_project(db, data["project_id"])
            if not proj:
                raise HTTPException(404, "Project not found")
            eoi.project_id = proj.id

        if "customer_id" in data:
            customer = _resolve_eoi_customer(db, data.get("customer_id"))
            eoi.customer_id = customer.id if customer else None
        if "broker_id" in data:
            broker = _resolve_eoi_broker(db, data.get("broker_id"))
            eoi.broker_id = broker.id if broker else None
            if broker:
                eoi.broker_name = broker.name
        if "inventory_id" in data:
            inv = _resolve_eoi_inventory(db, data.get("inventory_id"))
            eoi.inventory_id = inv.id if inv else None
        if "party_name" in data and data["party_name"]:
            eoi.party_name = data["party_name"].strip()
        for key in ["party_cnic", "broker_name", "unit_number", "payment_method", "reference_number", "notes"]:
            if key in data:
                setattr(eoi, key, data[key])
        if "party_mobile" in data:
            eoi.party_mobile = normalize_mobile(data["party_mobile"]) if data["party_mobile"] else None
        if "amount" in data:
            amt = float(data["amount"] or 0)
            if amt <= 0:
                raise HTTPException(400, "amount must be greater than 0")
            eoi.amount = amt
        if "marlas" in data:
            eoi.marlas = float(data["marlas"]) if data["marlas"] not in (None, "") else None
        if "payment_received" in data:
            eoi.payment_received = bool(data["payment_received"])
            if data["payment_received"] and not eoi.payment_received_at:
                eoi.payment_received_at = datetime.utcnow()
            elif not data["payment_received"]:
                eoi.payment_received_at = None
        if "eoi_date" in data and data["eoi_date"]:
            eoi.eoi_date = date.fromisoformat(data["eoi_date"])
        if "status" in data and data["status"]:
            status = data["status"].lower()
            if status not in VALID_EOI_STATUSES:
                raise HTTPException(400, "Invalid status")
            eoi.status = status

        eoi.updated_at = func.now()
        db.commit()
        db.refresh(eoi)
        return {"message": "EOI updated", "item": _eoi_row_dict(eoi, db)}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"EOI update error: {str(e)}")

@app.post("/api/eoi/{eoi_id}/cancel")
def cancel_eoi_collection(
    eoi_id: str,
    data: dict = None,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    eoi = _resolve_eoi_record(db, eoi_id)
    if not eoi:
        raise HTTPException(404, "EOI not found")
    if eoi.status == "converted":
        raise HTTPException(400, "Converted EOI cannot be cancelled")
    eoi.status = "cancelled"
    reason = (data or {}).get("reason") if data else None
    if reason:
        eoi.notes = ((eoi.notes or "") + f"\n[Cancelled] {reason}").strip()
    eoi.updated_at = func.now()
    db.commit()
    db.refresh(eoi)
    return {"message": "EOI cancelled", "item": _eoi_row_dict(eoi, db)}

@app.post("/api/eoi/{eoi_id}/refund")
def refund_eoi_collection(
    eoi_id: str,
    data: dict = None,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    eoi = _resolve_eoi_record(db, eoi_id)
    if not eoi:
        raise HTTPException(404, "EOI not found")
    if eoi.status == "converted":
        raise HTTPException(400, "Converted EOI cannot be refunded")
    eoi.status = "refunded"
    reason = (data or {}).get("reason") if data else None
    if reason:
        eoi.notes = ((eoi.notes or "") + f"\n[Refunded] {reason}").strip()
    eoi.updated_at = func.now()
    db.commit()
    db.refresh(eoi)
    return {"message": "EOI refunded", "item": _eoi_row_dict(eoi, db)}

@app.post("/api/eoi/{eoi_id}/convert")
def convert_eoi_to_transaction(
    eoi_id: str,
    data: dict,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    try:
        eoi = _resolve_eoi_record(db, eoi_id)
        if not eoi:
            raise HTTPException(404, "EOI not found")
        if eoi.status != "active":
            raise HTTPException(400, "Only active EOI can be converted")

        customer = _resolve_eoi_customer(db, data.get("customer_id")) or (
            db.query(Customer).filter(Customer.id == eoi.customer_id).first() if eoi.customer_id else None
        )
        if not customer:
            raise HTTPException(400, "customer_id is required for conversion")

        inv = _resolve_eoi_inventory(db, data.get("inventory_id")) or (
            db.query(Inventory).filter(Inventory.id == eoi.inventory_id).first() if eoi.inventory_id else None
        )
        if not inv:
            raise HTTPException(400, "inventory_id is required for conversion")

        proj = db.query(Project).filter(Project.id == eoi.project_id).first()
        if not proj:
            raise HTTPException(404, "EOI project not found")

        broker = _resolve_eoi_broker(db, data.get("broker_id")) or (
            db.query(Broker).filter(Broker.id == eoi.broker_id).first() if eoi.broker_id else None
        )

        area = float(data.get("area_marla") or inv.area_marla or eoi.marlas or 0)
        rate = float(data.get("rate_per_marla") or inv.rate_per_marla or 0)
        if area <= 0 or rate <= 0:
            raise HTTPException(400, "area_marla and rate_per_marla must be provided")

        num_inst = int(data.get("num_installments") or 4)
        if num_inst <= 0:
            raise HTTPException(400, "num_installments must be > 0")

        first_due = date.fromisoformat(data["first_due_date"]) if data.get("first_due_date") else date.today()
        booking_date = date.fromisoformat(data["booking_date"]) if data.get("booking_date") else (eoi.eoi_date or date.today())

        txn = Transaction(
            customer_id=customer.id,
            broker_id=broker.id if broker else None,
            project_id=proj.id,
            inventory_id=inv.id,
            company_rep_id=current_user.id,
            broker_commission_rate=float(data.get("broker_commission_rate") or 2.0),
            unit_number=data.get("unit_number") or inv.unit_number,
            block=data.get("block") or inv.block,
            area_marla=area,
            rate_per_marla=rate,
            total_value=area * rate,
            installment_cycle=data.get("installment_cycle") or "bi-annual",
            num_installments=num_inst,
            first_due_date=first_due,
            status=data.get("transaction_status") or "active",
            notes=((data.get("notes") or "") + f" Converted from {eoi.eoi_id}").strip(),
            booking_date=booking_date,
        )
        db.add(txn)
        db.flush()

        cycle = data.get("installment_cycle") or "bi-annual"
        cycle_months = {"monthly": 1, "quarterly": 3, "bi-annual": 6, "annual": 12}.get(cycle, 6)
        installment_amount = round((area * rate) / num_inst, 2)
        for n in range(num_inst):
            due = first_due + relativedelta(months=cycle_months * n)
            amt = installment_amount if n < num_inst - 1 else round((area * rate) - installment_amount * (num_inst - 1), 2)
            db.add(Installment(
                transaction_id=txn.id,
                installment_number=n + 1,
                due_date=due,
                amount=amt,
            ))

        inv.status = "sold"

        # Carry token as first receipt + allocation to earliest unpaid installments
        if float(eoi.amount or 0) > 0:
            r = Receipt(
                customer_id=customer.id,
                transaction_id=txn.id,
                amount=float(eoi.amount),
                payment_method=eoi.payment_method,
                reference_number=eoi.reference_number,
                payment_date=eoi.eoi_date or date.today(),
                notes=f"Auto-created from EOI {eoi.eoi_id}",
                created_by_rep_id=current_user.id
            )
            db.add(r)
            db.flush()

            remaining = float(eoi.amount)
            installments = db.query(Installment).filter(
                Installment.transaction_id == txn.id
            ).order_by(Installment.installment_number).all()
            for inst in installments:
                if remaining <= 0:
                    break
                balance = float(inst.amount) - float(inst.amount_paid or 0)
                alloc_amount = min(remaining, balance)
                db.add(ReceiptAllocation(receipt_id=r.id, installment_id=inst.id, amount=alloc_amount))
                inst.amount_paid = float(inst.amount_paid or 0) + alloc_amount
                if float(inst.amount_paid) >= float(inst.amount):
                    inst.status = "paid"
                elif float(inst.amount_paid) > 0:
                    inst.status = "partial"
                remaining -= alloc_amount

        eoi.status = "converted"
        eoi.payment_received = True  # Auto-sync: conversion implies payment received
        if not eoi.payment_received_at:
            eoi.payment_received_at = datetime.utcnow()
        eoi.converted_transaction_id = txn.id
        eoi.updated_at = func.now()
        db.commit()
        db.refresh(eoi)
        return {
            "message": "EOI converted to transaction",
            "eoi": _eoi_row_dict(eoi, db),
            "transaction": {"id": str(txn.id), "transaction_id": txn.transaction_id}
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"EOI conversion error: {str(e)}")

@app.get("/api/eoi/dashboard")
def get_eoi_dashboard(
    project_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    rep_id: Optional[str] = None,
    broker_id: Optional[str] = None,
    slab: Optional[str] = None,
    export_format: Optional[str] = "json",
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    try:
        q = db.query(EOICollection)
        isolation = get_rep_isolation_filter(current_user, db)
        scoped_rep_ids = None
        if isolation["isolated"]:
            scoped_rep_ids = isolation["team_rep_uuids"]
            q = q.filter(EOICollection.created_by.in_(scoped_rep_ids))

        if project_id:
            proj = _resolve_eoi_project(db, project_id)
            if proj:
                q = q.filter(EOICollection.project_id == proj.id)
            else:
                q = q.filter(literal(False))
        if date_from:
            q = q.filter(EOICollection.eoi_date >= date.fromisoformat(date_from))
        if date_to:
            q = q.filter(EOICollection.eoi_date <= date.fromisoformat(date_to))
        if status:
            statuses = [s.strip().lower() for s in status.split(",") if s.strip()]
            valid = [s for s in statuses if s in VALID_EOI_STATUSES]
            if valid:
                q = q.filter(EOICollection.status.in_(valid))
        if rep_id:
            rep_obj = db.query(CompanyRep).filter(
                (CompanyRep.rep_id == rep_id) | (func.cast(CompanyRep.id, String) == rep_id)
            ).first()
            if rep_obj:
                if not isolation["isolated"] or rep_obj.id in (scoped_rep_ids or []):
                    q = q.filter(EOICollection.created_by == rep_obj.id)
                else:
                    q = q.filter(literal(False))
        if broker_id:
            broker = _resolve_eoi_broker(db, broker_id)
            if broker:
                q = q.filter(EOICollection.broker_id == broker.id)
            elif broker_id.lower() == "direct":
                q = q.filter(EOICollection.broker_id.is_(None))
            else:
                q = q.filter(literal(False))

        rows = q.order_by(EOICollection.eoi_date.desc(), EOICollection.created_at.desc()).all()
        items = [_eoi_row_dict(row, db) for row in rows]
        if slab:
            items = [i for i in items if _eoi_slab_label(i.get("marlas")) == slab]

        total_amount = sum(i["amount"] for i in items)
        marlas_total = sum((i["marlas"] or 0) for i in items)

        status_breakdown = {k: {"count": 0, "amount": 0.0} for k in ["active", "converted", "cancelled", "refunded"]}
        broker_map = {}
        rep_map = {}
        slab_map = {"3-5": {"count": 0, "amount": 0.0}, "5-8": {"count": 0, "amount": 0.0}, "8+": {"count": 0, "amount": 0.0}, "Unknown": {"count": 0, "amount": 0.0}, "Other": {"count": 0, "amount": 0.0}}
        for item in items:
            st = item["status"]
            if st in status_breakdown:
                status_breakdown[st]["count"] += 1
                status_breakdown[st]["amount"] += item["amount"]

            bname = item["broker_name"] or "Direct"
            if bname not in broker_map:
                broker_map[bname] = {"broker_name": bname, "count": 0, "amount": 0.0}
            broker_map[bname]["count"] += 1
            broker_map[bname]["amount"] += item["amount"]

            rep_name = item["created_by_name"] or "Unknown"
            rep_key = item["created_by_rep_id"] or "unknown"
            if rep_key not in rep_map:
                rep_map[rep_key] = {"rep_id": rep_key, "rep_name": rep_name, "count": 0, "amount": 0.0}
            rep_map[rep_key]["count"] += 1
            rep_map[rep_key]["amount"] += item["amount"]

            slab_name = _eoi_slab_label(item.get("marlas"))
            slab_map[slab_name]["count"] += 1
            slab_map[slab_name]["amount"] += item["amount"]

        broker_leaderboard = sorted(broker_map.values(), key=lambda x: (x["amount"], x["count"]), reverse=True)
        rep_summary = sorted(rep_map.values(), key=lambda x: (x["amount"], x["count"]), reverse=True)
        slab_breakdown = [{"slab": k, "count": v["count"], "amount": round(v["amount"], 2)} for k, v in slab_map.items()]

        if export_format == "csv":
            import csv as csv_mod
            from io import StringIO
            output = StringIO()
            if items:
                writer = csv_mod.DictWriter(output, fieldnames=list(items[0].keys()))
                writer.writeheader()
                writer.writerows(items)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=eoi_dashboard.csv"}
            )

        if export_format == "excel":
            import openpyxl
            from io import BytesIO
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "EOI Dashboard"
            if items:
                headers = list(items[0].keys())
                ws.append(headers)
                for row in items:
                    ws.append([row.get(h, "") for h in headers])
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=eoi_dashboard.xlsx"}
            )

        return {
            "summary": {
                "total_eois_count": len(items),
                "total_eois_amount": round(total_amount, 2),
                "average_eoi_amount": round((total_amount / len(items)), 2) if items else 0.0,
                "total_marlas": round(marlas_total, 2),
                "status_breakdown": {
                    k: {"count": v["count"], "amount": round(v["amount"], 2)} for k, v in status_breakdown.items()
                },
            },
            "leaderboard": {
                "brokers": [{"broker_name": b["broker_name"], "count": b["count"], "amount": round(b["amount"], 2)} for b in broker_leaderboard],
                "reps": [{"rep_id": r["rep_id"], "rep_name": r["rep_name"], "count": r["count"], "amount": round(r["amount"], 2)} for r in rep_summary],
            },
            "marlas_slabs": slab_breakdown,
            "drilldown": {
                "items": items,
                "total": len(items),
            },
            "filters": {
                "project_id": project_id,
                "date_from": date_from,
                "date_to": date_to,
                "status": status,
                "rep_id": rep_id,
                "broker_id": broker_id,
                "slab": slab,
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"EOI dashboard error: {str(e)}")

# ============================================
# ZAKAT DISBURSEMENT MODULE
# ============================================

ZAKAT_ALLOWED_REPS = ["REP-0002", "REP-0010", "REP-0011", "REP-0012", "REP-0013"]
ZAKAT_APPROVER_REPS = ["REP-0009", "REP-0010"]  # CEO, CFO
ZAKAT_DISBURSER_REPS = ["REP-0002", "REP-0011", "REP-0012", "REP-0013"]  # Malik, Luqman, Mujtaba, Afaaq
ZAKAT_CANCELLER_REPS = ["REP-0002", "REP-0009", "REP-0010"]  # Malik, CEO, CFO
ZAKAT_FUNDS_APPROVER_REPS = ["REP-0010", "REP-0002"]  # CFO + Malik override
_zakat_schema_ready = False

def _ensure_zakat_workflow_schema(db: Session):
    global _zakat_schema_ready
    if _zakat_schema_ready:
        return
    db.execute(text("""
        ALTER TABLE zakat_records
        ADD COLUMN IF NOT EXISTS approval_status VARCHAR(20) DEFAULT 'pending',
        ADD COLUMN IF NOT EXISTS approved_amount NUMERIC(15,2),
        ADD COLUMN IF NOT EXISTS approved_by_rep_id UUID REFERENCES company_reps(id),
        ADD COLUMN IF NOT EXISTS approved_at TIMESTAMP,
        ADD COLUMN IF NOT EXISTS approval_decision VARCHAR(20),
        ADD COLUMN IF NOT EXISTS approval_notes TEXT,
        ADD COLUMN IF NOT EXISTS case_status VARCHAR(20) DEFAULT 'pending',
        ADD COLUMN IF NOT EXISTS disbursement_approval_status VARCHAR(20) DEFAULT 'pending',
        ADD COLUMN IF NOT EXISTS disbursement_approved_by_rep_id UUID REFERENCES company_reps(id),
        ADD COLUMN IF NOT EXISTS disbursement_approved_at TIMESTAMP,
        ADD COLUMN IF NOT EXISTS disbursement_approval_notes TEXT;
    """))
    db.execute(text("""
        ALTER TABLE zakat_records DROP CONSTRAINT IF EXISTS zakat_records_status_check;
        ALTER TABLE zakat_records
        ADD CONSTRAINT zakat_records_status_check
        CHECK (status IN ('active','pending','approved','rejected','disbursed','cancelled'));
    """))
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS zakat_disbursements (
            id UUID PRIMARY KEY DEFAULT uuid_generate_v4(),
            zakat_record_id UUID NOT NULL REFERENCES zakat_records(id) ON DELETE CASCADE,
            amount NUMERIC(15,2) NOT NULL,
            disbursed_by_rep_id UUID REFERENCES company_reps(id) ON DELETE SET NULL,
            payment_id UUID REFERENCES payments(id) ON DELETE SET NULL,
            payment_method VARCHAR(50),
            reference_number VARCHAR(100),
            receipt_number VARCHAR(100),
            notes TEXT,
            disbursement_date DATE,
            disbursed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_zakat_disbursements_zakat_record_id ON zakat_disbursements(zakat_record_id);
    """))
    db.commit()
    _zakat_schema_ready = True

def _check_zakat_access(current_user, db: Session):
    _ensure_zakat_workflow_schema(db)
    if current_user and current_user.rep_id in ZAKAT_ALLOWED_REPS:
        return True
    raise HTTPException(status_code=403, detail="Zakat module access denied")

def _check_zakat_approver(current_user):
    if current_user and current_user.rep_id in ZAKAT_APPROVER_REPS:
        return True
    raise HTTPException(status_code=403, detail="Only CEO/CFO can approve Zakat requests")

def _check_zakat_disburser(current_user):
    if current_user and current_user.rep_id in ZAKAT_DISBURSER_REPS:
        return True
    raise HTTPException(status_code=403, detail="Only authorized disbursers can disburse Zakat")

def _check_zakat_canceller(current_user):
    if current_user and current_user.rep_id in ZAKAT_CANCELLER_REPS:
        return True
    raise HTTPException(status_code=403, detail="Only authorized users can cancel approved Zakat records")

def _check_zakat_funds_approver(current_user):
    if current_user and current_user.rep_id in ZAKAT_FUNDS_APPROVER_REPS:
        return True
    raise HTTPException(status_code=403, detail="Only CFO can approve disbursement readiness")

def _generate_zakat_code(db: Session) -> str:
    count = db.query(func.count(ZakatRecord.id)).scalar() or 0
    candidate = f"ZKT-{str(count + 1).zfill(5)}"
    while db.query(ZakatRecord).filter(ZakatRecord.zakat_id == candidate).first():
        count += 1
        candidate = f"ZKT-{str(count + 1).zfill(5)}"
    return candidate

def _generate_beneficiary_code(db: Session) -> str:
    count = db.query(func.count(ZakatBeneficiary.id)).scalar() or 0
    candidate = f"ZBN-{str(count + 1).zfill(5)}"
    while db.query(ZakatBeneficiary).filter(ZakatBeneficiary.beneficiary_id == candidate).first():
        count += 1
        candidate = f"ZBN-{str(count + 1).zfill(5)}"
    return candidate

def _generate_payment_code(db: Session) -> str:
    count = db.query(func.count(Payment.id)).scalar() or 0
    candidate = f"PAY-{str(count + 1).zfill(5)}"
    while db.query(Payment).filter(Payment.payment_id == candidate).first():
        count += 1
        candidate = f"PAY-{str(count + 1).zfill(5)}"
    return candidate

def _zakat_disbursement_dict(d: ZakatDisbursement, db: Session):
    disburser = db.query(CompanyRep).filter(CompanyRep.id == d.disbursed_by_rep_id).first() if d.disbursed_by_rep_id else None
    return {
        "id": str(d.id),
        "amount": float(d.amount) if d.amount else 0,
        "disbursed_by_rep_id": disburser.rep_id if disburser else None,
        "disbursed_by_name": disburser.name if disburser else None,
        "payment_method": d.payment_method,
        "reference_number": d.reference_number,
        "receipt_number": d.receipt_number,
        "notes": d.notes,
        "disbursement_date": d.disbursement_date.isoformat() if d.disbursement_date else None,
        "disbursed_at": d.disbursed_at.isoformat() if d.disbursed_at else None,
        "created_at": d.created_at.isoformat() if d.created_at else None,
    }

def _zakat_row_dict(z: ZakatRecord, db: Session):
    ben = db.query(ZakatBeneficiary).filter(ZakatBeneficiary.id == z.beneficiary_id).first() if z.beneficiary_id else None
    creator = db.query(CompanyRep).filter(CompanyRep.id == z.created_by).first() if z.created_by else None
    disburser = db.query(CompanyRep).filter(CompanyRep.id == z.disbursed_by).first() if z.disbursed_by else None
    approver = db.query(CompanyRep).filter(CompanyRep.id == z.approved_by_rep_id).first() if z.approved_by_rep_id else None
    disbursement_approver = db.query(CompanyRep).filter(CompanyRep.id == z.disbursement_approved_by_rep_id).first() if z.disbursement_approved_by_rep_id else None
    disbursements = db.query(ZakatDisbursement).filter(ZakatDisbursement.zakat_record_id == z.id).order_by(ZakatDisbursement.created_at.asc()).all()
    disbursed_total = sum(float(d.amount or 0) for d in disbursements)
    approved_amount = float(z.approved_amount) if z.approved_amount is not None else None
    requested_amount = float(z.amount) if z.amount else 0
    remaining_amount = (approved_amount - disbursed_total) if approved_amount is not None else 0
    approval_type = None
    if z.approval_status == "approved" and approved_amount is not None:
        approval_type = "full" if approved_amount >= requested_amount else "partial"

    history_rows = []
    if z.beneficiary_cnic or z.beneficiary_mobile:
        history_q = db.query(ZakatRecord).filter(ZakatRecord.id != z.id)
        if z.beneficiary_cnic and z.beneficiary_mobile:
            history_q = history_q.filter(
                or_(
                    ZakatRecord.beneficiary_cnic == z.beneficiary_cnic,
                    ZakatRecord.beneficiary_mobile == z.beneficiary_mobile
                )
            )
        elif z.beneficiary_cnic:
            history_q = history_q.filter(ZakatRecord.beneficiary_cnic == z.beneficiary_cnic)
        else:
            history_q = history_q.filter(ZakatRecord.beneficiary_mobile == z.beneficiary_mobile)
        history_rows = history_q.order_by(ZakatRecord.created_at.desc()).limit(10).all()

    return {
        "id": str(z.id),
        "zakat_id": z.zakat_id,
        "beneficiary_id": str(z.beneficiary_id) if z.beneficiary_id else None,
        "beneficiary_display_id": ben.beneficiary_id if ben else None,
        "beneficiary_name": z.beneficiary_name,
        "beneficiary_cnic": z.beneficiary_cnic,
        "beneficiary_mobile": z.beneficiary_mobile,
        "beneficiary_address": z.beneficiary_address,
        "amount": requested_amount,
        "requested_amount": requested_amount,
        "category": z.category,
        "purpose": z.purpose,
        "approval_reference": z.approval_reference,
        "approved_by": z.approved_by,
        "approval_status": z.approval_status or "pending",
        "approved_amount": approved_amount,
        "approved_by_rep_id": approver.rep_id if approver else None,
        "approved_by_name": approver.name if approver else (z.approved_by or None),
        "approved_at": z.approved_at.isoformat() if z.approved_at else None,
        "approval_decision": z.approval_decision or approval_type,
        "approval_notes": z.approval_notes,
        "case_status": z.case_status or "pending",
        "disbursement_approval_status": z.disbursement_approval_status or "pending",
        "disbursement_approved_by_rep_id": disbursement_approver.rep_id if disbursement_approver else None,
        "disbursement_approved_by_name": disbursement_approver.name if disbursement_approver else None,
        "disbursement_approved_at": z.disbursement_approved_at.isoformat() if z.disbursement_approved_at else None,
        "disbursement_approval_notes": z.disbursement_approval_notes,
        "disbursed_total": disbursed_total,
        "remaining_amount": max(remaining_amount, 0),
        "disbursement_count": len(disbursements),
        "can_disburse": (
            z.approval_status == "approved" and
            (z.disbursement_approval_status or "pending") == "approved" and
            remaining_amount > 0
        ),
        "payment_method": z.payment_method,
        "reference_number": z.reference_number,
        "receipt_number": z.receipt_number,
        "disbursed_by": str(z.disbursed_by) if z.disbursed_by else None,
        "disbursed_by_rep_id": disburser.rep_id if disburser else None,
        "disbursed_by_name": disburser.name if disburser else None,
        "disbursement_date": z.disbursement_date.isoformat() if z.disbursement_date else None,
        "status": z.status,
        "notes": z.notes,
        "payment_id": str(z.payment_id) if z.payment_id else None,
        "created_by": str(z.created_by) if z.created_by else None,
        "created_by_rep_id": creator.rep_id if creator else None,
        "created_by_name": creator.name if creator else None,
        "disbursements": [_zakat_disbursement_dict(d, db) for d in disbursements],
        "beneficiary_history": [
            {
                "id": str(h.id),
                "zakat_id": h.zakat_id,
                "requested_amount": float(h.amount) if h.amount else 0,
                "approved_amount": float(h.approved_amount) if h.approved_amount is not None else None,
                "approval_status": h.approval_status or "pending",
                "case_status": h.case_status or "pending",
                "created_at": h.created_at.isoformat() if h.created_at else None,
            }
            for h in history_rows
        ],
        "created_at": z.created_at.isoformat() if z.created_at else None,
        "updated_at": z.updated_at.isoformat() if z.updated_at else None,
    }

# --- Beneficiary endpoints ---

@app.get("/api/zakat/beneficiaries")
def list_zakat_beneficiaries(
    search: str = None, status: str = None, limit: int = 200,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    _check_zakat_access(current_user, db)
    q = db.query(ZakatBeneficiary)
    if status:
        q = q.filter(ZakatBeneficiary.status == status)
    if search:
        s = f"%{search}%"
        q = q.filter(
            (ZakatBeneficiary.name.ilike(s)) |
            (ZakatBeneficiary.cnic.ilike(s)) |
            (ZakatBeneficiary.mobile.ilike(s)) |
            (ZakatBeneficiary.beneficiary_id.ilike(s))
        )
    bens = q.order_by(ZakatBeneficiary.created_at.desc()).limit(limit).all()
    return [
        {
            "id": str(b.id), "beneficiary_id": b.beneficiary_id, "name": b.name,
            "cnic": b.cnic, "mobile": b.mobile, "address": b.address,
            "notes": b.notes, "status": b.status,
            "created_at": b.created_at.isoformat() if b.created_at else None,
        }
        for b in bens
    ]

@app.post("/api/zakat/beneficiaries")
def create_zakat_beneficiary(data: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    _check_zakat_access(current_user, db)
    if not data.get("name"):
        raise HTTPException(400, "Beneficiary name is required")
    b = ZakatBeneficiary(
        beneficiary_id=_generate_beneficiary_code(db),
        name=data["name"],
        cnic=data.get("cnic"),
        mobile=data.get("mobile"),
        address=data.get("address"),
        notes=data.get("notes"),
    )
    db.add(b); db.commit(); db.refresh(b)
    return {"message": "Beneficiary created", "item": {
        "id": str(b.id), "beneficiary_id": b.beneficiary_id, "name": b.name,
        "cnic": b.cnic, "mobile": b.mobile, "address": b.address,
    }}

@app.put("/api/zakat/beneficiaries/{bid}")
def update_zakat_beneficiary(bid: str, data: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    _check_zakat_access(current_user, db)
    try:
        b = db.query(ZakatBeneficiary).filter(
            (ZakatBeneficiary.id == uuid.UUID(bid)) | (ZakatBeneficiary.beneficiary_id == bid)
        ).first()
    except ValueError:
        b = db.query(ZakatBeneficiary).filter(ZakatBeneficiary.beneficiary_id == bid).first()
    if not b:
        raise HTTPException(404, "Beneficiary not found")
    for field in ["name", "cnic", "mobile", "address", "notes", "status"]:
        if field in data:
            setattr(b, field, data[field])
    b.updated_at = func.now()
    db.commit(); db.refresh(b)
    return {"message": "Beneficiary updated", "item": {
        "id": str(b.id), "beneficiary_id": b.beneficiary_id, "name": b.name,
        "cnic": b.cnic, "mobile": b.mobile, "address": b.address,
    }}

# --- Zakat record endpoints ---

@app.get("/api/zakat")
def list_zakat_records(
    status: str = None, category: str = None, search: str = None,
    date_from: str = None, date_to: str = None,
    disbursed_by: str = None, beneficiary_id: str = None,
    limit: int = 200,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    _check_zakat_access(current_user, db)
    q = db.query(ZakatRecord)
    if status:
        statuses = [s.strip() for s in status.split(",")]
        q = q.filter(ZakatRecord.status.in_(statuses))
    if category:
        categories = [c.strip() for c in category.split(",")]
        q = q.filter(ZakatRecord.category.in_(categories))
    if search:
        s = f"%{search}%"
        q = q.filter(
            (ZakatRecord.zakat_id.ilike(s)) |
            (ZakatRecord.beneficiary_name.ilike(s)) |
            (ZakatRecord.beneficiary_cnic.ilike(s)) |
            (ZakatRecord.beneficiary_mobile.ilike(s)) |
            (ZakatRecord.receipt_number.ilike(s)) |
            (ZakatRecord.reference_number.ilike(s))
        )
    if date_from:
        q = q.filter(ZakatRecord.created_at >= date_from)
    if date_to:
        q = q.filter(ZakatRecord.created_at <= date_to + " 23:59:59")
    if disbursed_by:
        rep = db.query(CompanyRep).filter(
            (CompanyRep.rep_id == disbursed_by) | (CompanyRep.id == disbursed_by)
        ).first()
        if rep:
            q = q.filter(ZakatRecord.disbursed_by == rep.id)
    if beneficiary_id:
        try:
            q = q.filter(ZakatRecord.beneficiary_id == uuid.UUID(beneficiary_id))
        except ValueError:
            ben = db.query(ZakatBeneficiary).filter(ZakatBeneficiary.beneficiary_id == beneficiary_id).first()
            if ben:
                q = q.filter(ZakatRecord.beneficiary_id == ben.id)
    records = q.order_by(ZakatRecord.created_at.desc()).limit(limit).all()
    total = q.count()
    return {
        "items": [_zakat_row_dict(r, db) for r in records],
        "total": total,
    }

@app.post("/api/zakat")
def create_zakat_record(data: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    _check_zakat_access(current_user, db)
    if not data.get("beneficiary_name"):
        raise HTTPException(400, "Beneficiary name is required")
    if not data.get("amount") or float(data["amount"]) <= 0:
        raise HTTPException(400, "Amount must be greater than 0")
    if not data.get("category"):
        raise HTTPException(400, "Category is required")

    # Resolve or create beneficiary
    ben_id = None
    if data.get("beneficiary_id"):
        try:
            ben = db.query(ZakatBeneficiary).filter(
                (ZakatBeneficiary.id == uuid.UUID(data["beneficiary_id"])) |
                (ZakatBeneficiary.beneficiary_id == data["beneficiary_id"])
            ).first()
        except ValueError:
            ben = db.query(ZakatBeneficiary).filter(ZakatBeneficiary.beneficiary_id == data["beneficiary_id"]).first()
        if ben:
            ben_id = ben.id
    elif data.get("create_beneficiary"):
        new_ben = ZakatBeneficiary(
            beneficiary_id=_generate_beneficiary_code(db),
            name=data["beneficiary_name"],
            cnic=data.get("beneficiary_cnic"),
            mobile=data.get("beneficiary_mobile"),
            address=data.get("beneficiary_address"),
        )
        db.add(new_ben); db.flush()
        ben_id = new_ben.id

    rep = db.query(CompanyRep).filter(CompanyRep.rep_id == current_user.rep_id).first()

    z = ZakatRecord(
        zakat_id=_generate_zakat_code(db),
        beneficiary_id=ben_id,
        beneficiary_name=data["beneficiary_name"],
        beneficiary_cnic=data.get("beneficiary_cnic"),
        beneficiary_mobile=data.get("beneficiary_mobile"),
        beneficiary_address=data.get("beneficiary_address"),
        amount=float(data["amount"]),
        category=data["category"],
        purpose=data.get("purpose"),
        approval_reference=data.get("approval_reference"),
        approved_by=data.get("approved_by"),
        payment_method=data.get("payment_method"),
        reference_number=data.get("reference_number"),
        receipt_number=data.get("receipt_number"),
        notes=data.get("notes"),
        status="pending",
        approval_status="pending",
        case_status="pending",
        disbursement_approval_status="pending",
        created_by=rep.id if rep else None,
    )
    db.add(z); db.commit(); db.refresh(z)

    # Auto-create disbursement task for tracking
    try:
        from app.services.task_service import task_service as _zts
        amount_str = f"PKR {z.amount:,.0f}" if z.amount else "PKR -"
        _zts.create_task(
            db, creator_id=rep.id if rep else z.created_by,
            title=f"Disburse {amount_str} to {z.beneficiary_name}",
            description=f"Zakat ID: {z.zakat_id}\nBeneficiary: {z.beneficiary_name}\nCNIC: {z.beneficiary_cnic or '-'}\nMobile: {z.beneficiary_mobile or '-'}\nCategory: {z.category}\nPurpose: {z.purpose or '-'}\nApproval: {z.approval_reference or '-'}",
            task_type="collection",
            priority="high",
            assignee_id=rep.id if rep else None,
            due_date=date.today() + timedelta(days=7),
        )
    except Exception:
        pass  # Non-critical — task creation failure shouldn't block zakat record

    for approver_rep_id in ZAKAT_APPROVER_REPS:
        create_notification(
            db,
            approver_rep_id,
            "zakat_approval_required",
            f"Zakat approval required: {z.zakat_id}",
            f"{z.beneficiary_name} requested PKR {float(z.amount):,.0f}. Please review and approve/reject.",
            category="zakat",
            entity_type="zakat",
            entity_id=z.id,
            data={"zakat_id": z.zakat_id},
        )
    db.commit()
    return {"message": "Zakat record created", "item": _zakat_row_dict(z, db)}

@app.put("/api/zakat/{zid}")
def update_zakat_record(zid: str, data: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    _check_zakat_access(current_user, db)
    try:
        z = db.query(ZakatRecord).filter(
            (ZakatRecord.id == uuid.UUID(zid)) | (ZakatRecord.zakat_id == zid)
        ).first()
    except ValueError:
        z = db.query(ZakatRecord).filter(ZakatRecord.zakat_id == zid).first()
    if not z:
        raise HTTPException(404, "Zakat record not found")
    if (z.approval_status or "pending") != "pending":
        raise HTTPException(400, "Cannot edit post-approval/rejection record")

    # Resolve beneficiary if changed
    if data.get("beneficiary_id"):
        try:
            ben = db.query(ZakatBeneficiary).filter(
                (ZakatBeneficiary.id == uuid.UUID(data["beneficiary_id"])) |
                (ZakatBeneficiary.beneficiary_id == data["beneficiary_id"])
            ).first()
        except ValueError:
            ben = db.query(ZakatBeneficiary).filter(ZakatBeneficiary.beneficiary_id == data["beneficiary_id"]).first()
        if ben:
            z.beneficiary_id = ben.id

    for field in [
        "beneficiary_name", "beneficiary_cnic", "beneficiary_mobile", "beneficiary_address",
        "amount", "category", "purpose", "approval_reference", "approved_by",
        "payment_method", "reference_number", "receipt_number", "notes",
    ]:
        if field in data:
            val = data[field]
            if field == "amount":
                val = float(val) if val else 0
            setattr(z, field, val)
    z.updated_at = func.now()
    db.commit(); db.refresh(z)
    return {"message": "Zakat record updated", "item": _zakat_row_dict(z, db)}

@app.post("/api/zakat/{zid}/approve")
def approve_zakat(zid: str, data: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    _check_zakat_access(current_user, db)
    _check_zakat_approver(current_user)
    try:
        z = db.query(ZakatRecord).filter(
            (ZakatRecord.id == uuid.UUID(zid)) | (ZakatRecord.zakat_id == zid)
        ).first()
    except ValueError:
        z = db.query(ZakatRecord).filter(ZakatRecord.zakat_id == zid).first()
    if not z:
        raise HTTPException(404, "Zakat record not found")
    if z.status == "cancelled":
        raise HTTPException(400, "Cancelled record cannot be approved/rejected")

    action = (data.get("action") or "approve").strip().lower()
    approver = db.query(CompanyRep).filter(CompanyRep.rep_id == current_user.rep_id).first()
    creator = db.query(CompanyRep).filter(CompanyRep.id == z.created_by).first() if z.created_by else None

    if action == "reject":
        z.approval_status = "rejected"
        z.approval_decision = "rejected"
        z.case_status = "closed"
        z.status = "rejected"
        z.approved_amount = 0
        z.approved_by_rep_id = approver.id if approver else None
        z.approved_by = approver.name if approver else current_user.name
        z.approved_at = datetime.utcnow()
        z.approval_notes = data.get("notes")
        z.disbursement_approval_status = "pending"
        z.disbursement_approved_by_rep_id = None
        z.disbursement_approved_at = None
        z.disbursement_approval_notes = None
        z.updated_at = func.now()
        db.commit(); db.refresh(z)
        for rep_id in ZAKAT_ALLOWED_REPS:
            create_notification(
                db,
                rep_id,
                "zakat_rejected",
                f"Zakat rejected: {z.zakat_id}",
                f"{z.zakat_id} was rejected by {z.approved_by}.",
                category="zakat",
                entity_type="zakat",
                entity_id=z.id,
                data={"zakat_id": z.zakat_id},
            )
        db.commit()
        return {"message": "Zakat request rejected", "item": _zakat_row_dict(z, db)}

    approved_amount = data.get("approved_amount")
    if approved_amount is None:
        raise HTTPException(400, "approved_amount is required")
    approved_amount = float(approved_amount)
    if approved_amount <= 0:
        raise HTTPException(400, "approved_amount must be greater than 0")
    if approved_amount > float(z.amount):
        raise HTTPException(400, "approved_amount cannot exceed requested amount")

    existing_disbursed_total = db.query(func.coalesce(func.sum(ZakatDisbursement.amount), 0)).filter(
        ZakatDisbursement.zakat_record_id == z.id
    ).scalar() or 0
    if float(existing_disbursed_total) > 0 and (z.approval_status or "pending") == "approved":
        raise HTTPException(400, "Approval can only be revised before first disbursement")
    if float(existing_disbursed_total) > approved_amount:
        raise HTTPException(400, "approved_amount cannot be less than already disbursed amount")

    close_case = bool(data.get("close_case", False))
    z.approval_status = "approved"
    z.approved_amount = approved_amount
    z.approved_by_rep_id = approver.id if approver else None
    z.approved_by = approver.name if approver else current_user.name
    z.approved_at = datetime.utcnow()
    z.approval_notes = data.get("notes")
    z.approval_decision = "full" if approved_amount >= float(z.amount) else "partial"
    z.case_status = "closed" if close_case else "open"
    z.disbursement_approval_status = "pending"
    z.disbursement_approved_by_rep_id = None
    z.disbursement_approved_at = None
    z.disbursement_approval_notes = None
    z.status = "approved"
    z.updated_at = func.now()
    db.commit(); db.refresh(z)

    create_notification(
        db,
        "REP-0010",
        "zakat_disbursement_approval_required",
        f"Pending disbursement approval: {z.zakat_id}",
        f"{z.zakat_id} approved for PKR {approved_amount:,.0f}. Please approve or postpone disbursement.",
        category="zakat",
        entity_type="zakat",
        entity_id=z.id,
        data={"zakat_id": z.zakat_id, "approved_amount": approved_amount},
    )
    db.commit()
    return {"message": "Zakat request approved", "item": _zakat_row_dict(z, db)}

@app.post("/api/zakat/{zid}/approve-disbursement")
def approve_zakat_disbursement(zid: str, data: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    _check_zakat_access(current_user, db)
    _check_zakat_funds_approver(current_user)
    try:
        z = db.query(ZakatRecord).filter(
            (ZakatRecord.id == uuid.UUID(zid)) | (ZakatRecord.zakat_id == zid)
        ).first()
    except ValueError:
        z = db.query(ZakatRecord).filter(ZakatRecord.zakat_id == zid).first()
    if not z:
        raise HTTPException(404, "Zakat record not found")
    if z.approval_status != "approved":
        raise HTTPException(400, "Case must be approved by CEO/CFO first")
    if z.status in ["cancelled", "rejected"]:
        raise HTTPException(400, f"Cannot approve disbursement for status '{z.status}'")

    action = (data.get("action") or "approve").strip().lower()
    if action not in ["approve", "postpone"]:
        raise HTTPException(400, "Action must be 'approve' or 'postpone'")

    approver = db.query(CompanyRep).filter(CompanyRep.rep_id == current_user.rep_id).first()
    z.disbursement_approval_status = "approved" if action == "approve" else "postponed"
    z.disbursement_approved_by_rep_id = approver.id if approver else None
    z.disbursement_approved_at = datetime.utcnow()
    z.disbursement_approval_notes = data.get("notes")
    z.updated_at = func.now()
    db.commit(); db.refresh(z)

    if action == "approve":
        for rep_id in ZAKAT_DISBURSER_REPS:
            create_notification(
                db,
                rep_id,
                "zakat_ready_for_disbursement",
                f"Approved disbursement by CFO: {z.zakat_id}",
                f"{z.zakat_id} is approved by CFO and ready for disbursement.",
                category="zakat",
                entity_type="zakat",
                entity_id=z.id,
                data={"zakat_id": z.zakat_id, "approved_amount": float(z.approved_amount or 0)},
            )
    else:
        creator = db.query(CompanyRep).filter(CompanyRep.id == z.created_by).first() if z.created_by else None
        if creator and creator.rep_id:
            create_notification(
                db,
                creator.rep_id,
                "zakat_disbursement_postponed",
                f"Disbursement postponed: {z.zakat_id}",
                f"CFO postponed disbursement readiness for {z.zakat_id}.",
                category="zakat",
                entity_type="zakat",
                entity_id=z.id,
                data={"zakat_id": z.zakat_id},
            )
    db.commit()
    msg = "Zakat marked ready for disbursement" if action == "approve" else "Zakat disbursement postponed"
    return {"message": msg, "item": _zakat_row_dict(z, db)}

@app.post("/api/zakat/{zid}/disburse")
def disburse_zakat(zid: str, data: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    _check_zakat_access(current_user, db)
    _check_zakat_disburser(current_user)
    try:
        z = db.query(ZakatRecord).filter(
            (ZakatRecord.id == uuid.UUID(zid)) | (ZakatRecord.zakat_id == zid)
        ).first()
    except ValueError:
        z = db.query(ZakatRecord).filter(ZakatRecord.zakat_id == zid).first()
    if not z:
        raise HTTPException(404, "Zakat record not found")
    if z.status in ["cancelled", "rejected"]:
        raise HTTPException(400, f"Cannot disburse a record with status '{z.status}'")
    if z.approval_status != "approved":
        raise HTTPException(400, "Approval is required before disbursement")
    if (z.disbursement_approval_status or "pending") != "approved":
        raise HTTPException(400, "CFO disbursement approval is required before disbursement")

    # Resolve disbursed_by rep
    disburser = None
    if data.get("disbursed_by"):
        try:
            did = uuid.UUID(str(data["disbursed_by"]))
            disburser = db.query(CompanyRep).filter(
                (CompanyRep.rep_id == data["disbursed_by"]) | (CompanyRep.id == did)
            ).first()
        except ValueError:
            disburser = db.query(CompanyRep).filter(CompanyRep.rep_id == data["disbursed_by"]).first()
    if not disburser:
        # Default to current user
        disburser = db.query(CompanyRep).filter(CompanyRep.rep_id == current_user.rep_id).first()

    amount = data.get("amount")
    if amount is None:
        raise HTTPException(400, "Disbursed amount is required")
    amount = float(amount)
    if amount <= 0:
        raise HTTPException(400, "Disbursed amount must be greater than 0")

    approved_amount = float(z.approved_amount or 0)
    disbursed_total = db.query(func.coalesce(func.sum(ZakatDisbursement.amount), 0)).filter(
        ZakatDisbursement.zakat_record_id == z.id
    ).scalar() or 0
    remaining = approved_amount - float(disbursed_total)
    if remaining <= 0:
        raise HTTPException(400, "Approved amount is already fully disbursed")
    if amount > remaining:
        raise HTTPException(400, f"Disbursement exceeds remaining approved amount ({remaining:,.2f})")

    payment_method = data.get("payment_method") or z.payment_method
    reference_number = data.get("reference_number") or z.reference_number
    receipt_number = data.get("receipt_number") or z.receipt_number
    disbursement_date = date.fromisoformat(data["disbursement_date"]) if data.get("disbursement_date") else date.today()

    # Auto-create Payment record
    payment = Payment(
        payment_id=_generate_payment_code(db),
        payment_type="zakat",
        payee_type="beneficiary",
        amount=amount,
        payment_method=payment_method,
        reference_number=reference_number,
        payment_date=disbursement_date,
        notes=f"Zakat disbursement {z.zakat_id} to {z.beneficiary_name} (PKR {amount:,.0f})",
        status="completed",
    )
    db.add(payment); db.flush()

    d = ZakatDisbursement(
        zakat_record_id=z.id,
        amount=amount,
        disbursed_by_rep_id=disburser.id if disburser else None,
        payment_id=payment.id,
        payment_method=payment_method,
        reference_number=reference_number,
        receipt_number=receipt_number,
        notes=data.get("notes"),
        disbursement_date=disbursement_date,
        disbursed_at=datetime.utcnow(),
    )
    db.add(d)

    # Update zakat record
    z.status = "approved"
    z.disbursed_by = disburser.id if disburser else None
    z.payment_method = payment_method
    z.reference_number = reference_number
    z.receipt_number = receipt_number
    z.disbursement_date = disbursement_date
    z.payment_id = payment.id
    if data.get("notes"):
        z.notes = (z.notes + "\n" if z.notes else "") + data["notes"]
    z.updated_at = func.now()

    db.commit(); db.refresh(z)
    return {"message": "Zakat disbursement recorded", "item": _zakat_row_dict(z, db)}

@app.post("/api/zakat/{zid}/cancel")
def cancel_zakat(zid: str, data: dict = {}, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    _check_zakat_access(current_user, db)
    try:
        z = db.query(ZakatRecord).filter(
            (ZakatRecord.id == uuid.UUID(zid)) | (ZakatRecord.zakat_id == zid)
        ).first()
    except ValueError:
        z = db.query(ZakatRecord).filter(ZakatRecord.zakat_id == zid).first()
    if not z:
        raise HTTPException(404, "Zakat record not found")
    if z.approval_status == "approved":
        _check_zakat_canceller(current_user)
        disbursed_total = db.query(func.coalesce(func.sum(ZakatDisbursement.amount), 0)).filter(
            ZakatDisbursement.zakat_record_id == z.id
        ).scalar() or 0
        if float(disbursed_total) > 0:
            raise HTTPException(400, "Cannot cancel an already disbursed approved record")

    z.status = "cancelled"
    z.case_status = "closed"
    reason = data.get("reason", "")
    if reason:
        z.notes = (z.notes + "\n" if z.notes else "") + f"Cancelled: {reason}"
    z.updated_at = func.now()
    db.commit(); db.refresh(z)
    return {"message": "Zakat record cancelled", "item": _zakat_row_dict(z, db)}

# --- Zakat Dashboard ---

@app.get("/api/zakat/dashboard")
def get_zakat_dashboard(
    status: str = None, category: str = None,
    date_from: str = None, date_to: str = None,
    quarter: str = None,
    export_format: str = None,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)
):
    _check_zakat_access(current_user, db)
    try:
        q = db.query(ZakatRecord)
        if status:
            q = q.filter(ZakatRecord.status.in_([s.strip() for s in status.split(",")]))
        if category:
            q = q.filter(ZakatRecord.category.in_([c.strip() for c in category.split(",")]))
        if date_from:
            q = q.filter(ZakatRecord.created_at >= date_from)
        if date_to:
            q = q.filter(ZakatRecord.created_at <= date_to + " 23:59:59")

        all_records = q.all()

        disbursement_totals = {
            str(x.zakat_record_id): float(x.total_amount or 0)
            for x in db.query(
                ZakatDisbursement.zakat_record_id,
                func.sum(ZakatDisbursement.amount).label("total_amount")
            ).group_by(ZakatDisbursement.zakat_record_id).all()
        }

        requested_total_value = sum(float(r.amount or 0) for r in all_records)
        approved_records = [r for r in all_records if (r.approval_status or "pending") == "approved"]
        approved_total_value = sum(float(r.approved_amount or 0) for r in approved_records)
        disbursed_records = [r for r in all_records if disbursement_totals.get(str(r.id), 0) > 0]
        disbursed_total_value = sum(disbursement_totals.get(str(r.id), 0) for r in all_records)
        pending_records = [r for r in all_records if (r.approval_status or "pending") == "pending"]
        open_records = [r for r in approved_records if (r.case_status or "pending") == "open"]
        closed_records = [r for r in all_records if (r.case_status or "pending") == "closed"]
        rejected_records = [r for r in all_records if (r.approval_status or "pending") == "rejected"]
        funds_pending_records = [r for r in approved_records if (r.disbursement_approval_status or "pending") in ["pending", "postponed"]]
        funds_approved_records = [r for r in approved_records if (r.disbursement_approval_status or "pending") == "approved"]
        ready_for_disbursement_records = [
            r for r in funds_approved_records
            if (float(r.approved_amount or 0) - disbursement_totals.get(str(r.id), 0)) > 0
        ]

        summary = {
            "requested_total_count": len(all_records),
            "requested_total_value": requested_total_value,
            "approved_total_count": len(approved_records),
            "approved_total_value": approved_total_value,
            "disbursed_total_count": len(disbursed_records),
            "disbursed_total_value": disbursed_total_value,
            "pending_count": len(pending_records),
            "pending_value": sum(float(r.amount or 0) for r in pending_records),
            "rejected_count": len(rejected_records),
            "rejected_value": sum(float(r.amount or 0) for r in rejected_records),
            "open_cases_count": len(open_records),
            "open_cases_value": sum(float(r.approved_amount or 0) for r in open_records),
            "closed_cases_count": len(closed_records),
            "closed_cases_value": sum(float((r.approved_amount if r.approved_amount is not None else r.amount) or 0) for r in closed_records),
            "funds_pending_count": len(funds_pending_records),
            "funds_pending_value": sum(float(r.approved_amount or 0) for r in funds_pending_records),
            "funds_approved_count": len(funds_approved_records),
            "funds_approved_value": sum(float(r.approved_amount or 0) for r in funds_approved_records),
            "ready_for_disbursement_count": len(ready_for_disbursement_records),
            "ready_for_disbursement_value": sum(
                max(float(r.approved_amount or 0) - disbursement_totals.get(str(r.id), 0), 0)
                for r in ready_for_disbursement_records
            ),
            # Legacy keys for backward compatibility in UI
            "total_records": len(all_records),
            "total_amount": requested_total_value,
            "active_count": len(open_records),
            "active_amount": sum(float(r.approved_amount or 0) for r in open_records),
            "disbursed_count": len(disbursed_records),
            "disbursed_amount": disbursed_total_value,
            "cancelled_count": len([r for r in all_records if r.status == "cancelled"]),
            "cancelled_amount": sum(float(r.amount or 0) for r in all_records if r.status == "cancelled"),
        }

        # By category
        cat_map = {}
        for r in all_records:
            cat = r.category or "other"
            if cat not in cat_map:
                cat_map[cat] = {"category": cat, "count": 0, "amount": 0}
            cat_map[cat]["count"] += 1
            cat_map[cat]["amount"] += float(r.amount)
        by_category = sorted(cat_map.values(), key=lambda x: x["amount"], reverse=True)

        # By quarter (calendar quarters)
        quarter_map = {}
        for r in all_records:
            dt = r.disbursement_date or (r.created_at.date() if r.created_at else None)
            if dt:
                q_num = (dt.month - 1) // 3 + 1
                q_label = f"Q{q_num} {dt.year}"
                if q_label not in quarter_map:
                    quarter_map[q_label] = {"quarter": q_label, "count": 0, "amount": 0}
                quarter_map[q_label]["count"] += 1
                quarter_map[q_label]["amount"] += float(r.amount)
        by_quarter = sorted(quarter_map.values(), key=lambda x: x["quarter"])

        # By rep (zakat_disbursements.disbursed_by_rep_id)
        rep_map = {}
        disbursement_rows = db.query(ZakatDisbursement).all()
        for d in disbursement_rows:
            if d.disbursed_by_rep_id:
                rep = db.query(CompanyRep).filter(CompanyRep.id == d.disbursed_by_rep_id).first()
                key = str(d.disbursed_by_rep_id)
                if key not in rep_map:
                    rep_map[key] = {
                        "rep_id": rep.rep_id if rep else "Unknown",
                        "rep_name": rep.name if rep else "Unknown",
                        "count": 0, "amount": 0,
                    }
                rep_map[key]["count"] += 1
                rep_map[key]["amount"] += float(d.amount or 0)
        by_rep = sorted(rep_map.values(), key=lambda x: x["amount"], reverse=True)

        # Export
        if export_format in ("csv", "excel"):
            import io, csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                "Zakat ID", "Beneficiary", "CNIC", "Mobile", "Address",
                "Requested Amount (PKR)", "Approved Amount (PKR)", "Disbursed Amount (PKR)",
                "Category", "Purpose", "Approval Reference", "Approved By",
                "Approval Status", "Disbursement Approval Status", "Case Status", "Status", "Notes", "Created At",
            ])
            for r in all_records:
                writer.writerow([
                    r.zakat_id, r.beneficiary_name, r.beneficiary_cnic or "",
                    r.beneficiary_mobile or "", r.beneficiary_address or "",
                    float(r.amount or 0), float(r.approved_amount or 0), disbursement_totals.get(str(r.id), 0),
                    r.category, r.purpose or "",
                    r.approval_reference or "", r.approved_by or "",
                    r.approval_status or "pending", r.disbursement_approval_status or "pending", r.case_status or "pending",
                    r.status, r.notes or "",
                    r.created_at.isoformat() if r.created_at else "",
                ])
            from starlette.responses import StreamingResponse
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=zakat_register.csv"},
            )

        return {
            "summary": summary,
            "by_category": by_category,
            "by_quarter": by_quarter,
            "by_rep": by_rep,
            "drilldown": {
                "items": [_zakat_row_dict(r, db) for r in all_records[:200]],
                "total": len(all_records),
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Zakat dashboard error: {str(e)}")

# ============================================
# CREDITORS API
# ============================================
@app.get("/api/creditors")
def list_creditors(db: Session = Depends(get_db)):
    creditors = db.query(Creditor).order_by(Creditor.created_at.desc()).all()
    return [{
        "id": str(c.id), "creditor_id": c.creditor_id, "name": c.name,
        "mobile": c.mobile, "company": c.company, "bank_name": c.bank_name,
        "bank_account": c.bank_account, "bank_iban": c.bank_iban,
        "notes": c.notes, "status": c.status, "created_at": str(c.created_at)
    } for c in creditors]

@app.get("/api/creditors/{cid}")
def get_creditor(cid: str, db: Session = Depends(get_db)):
    c = db.query(Creditor).filter((Creditor.id == cid) | (Creditor.creditor_id == cid)).first()
    if not c: raise HTTPException(404, "Creditor not found")
    return {
        "id": str(c.id), "creditor_id": c.creditor_id, "name": c.name,
        "mobile": c.mobile, "company": c.company, "bank_name": c.bank_name,
        "bank_account": c.bank_account, "bank_iban": c.bank_iban,
        "notes": c.notes, "status": c.status, "created_at": str(c.created_at)
    }

@app.post("/api/creditors")
def create_creditor(data: dict, db: Session = Depends(get_db)):
    c = Creditor(
        name=data["name"], mobile=data.get("mobile"), company=data.get("company"),
        bank_name=data.get("bank_name"), bank_account=data.get("bank_account"),
        bank_iban=data.get("bank_iban"), notes=data.get("notes")
    )
    db.add(c); db.commit(); db.refresh(c)
    return {"message": "Creditor created", "id": str(c.id), "creditor_id": c.creditor_id}

@app.put("/api/creditors/{cid}")
def update_creditor(cid: str, data: dict, db: Session = Depends(get_db)):
    c = db.query(Creditor).filter((Creditor.id == cid) | (Creditor.creditor_id == cid)).first()
    if not c: raise HTTPException(404, "Creditor not found")
    for k in ["name", "mobile", "company", "bank_name", "bank_account", "bank_iban", "notes", "status"]:
        if k in data and data[k] is not None: setattr(c, k, data[k])
    db.commit()
    return {"message": "Creditor updated"}

@app.delete("/api/creditors/{cid}")
def delete_creditor(cid: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    c = find_entity(db, Creditor, "creditor_id", cid)
    if not c: raise HTTPException(404, "Creditor not found")

    # Creator role cannot delete
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    # Admin can delete directly
    if current_user.role == "admin":
        db.delete(c); db.commit()
        return {"message": "Creditor deleted"}

    # Non-admin: create deletion request
    req = DeletionRequest(entity_type="creditor", entity_id=c.id, entity_name=c.name, requested_by=current_user.id)
    db.add(req); db.commit(); db.refresh(req)
    return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

# ============================================
# PAYMENTS API
# ============================================
@app.get("/api/payments")
def list_payments(
    broker_id: str = None, rep_id: str = None, creditor_id: str = None,
    payment_type: str = None, status: str = None, start_date: str = None,
    end_date: str = None, limit: int = 100, db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    q = db.query(Payment)
    # Data isolation: indirect/both reps see payments for their brokers only
    iso = get_rep_isolation_filter(current_user, db)
    if iso["isolated"]:
        rt = iso["rep_type"]
        if rt == "direct":
            return []  # direct reps don't see payments
        own_broker_ids = [b.id for b in db.query(Broker.id).filter(Broker.assigned_rep_id.in_(iso["team_rep_uuids"])).all()]
        q = q.filter(Payment.broker_id.in_(own_broker_ids)) if own_broker_ids else q.filter(literal(False))
    if broker_id:
        b = db.query(Broker).filter((Broker.id == broker_id) | (Broker.broker_id == broker_id)).first()
        if b: q = q.filter(Payment.broker_id == b.id)
    if rep_id:
        r = db.query(CompanyRep).filter((CompanyRep.id == rep_id) | (CompanyRep.rep_id == rep_id)).first()
        if r: q = q.filter(Payment.company_rep_id == r.id)
    if creditor_id:
        c = db.query(Creditor).filter((Creditor.id == creditor_id) | (Creditor.creditor_id == creditor_id)).first()
        if c: q = q.filter(Payment.creditor_id == c.id)
    if payment_type: q = q.filter(Payment.payment_type == payment_type)
    if status: q = q.filter(Payment.status == status)
    if start_date: q = q.filter(Payment.payment_date >= date.fromisoformat(start_date))
    if end_date: q = q.filter(Payment.payment_date <= date.fromisoformat(end_date))
    
    payments = q.order_by(Payment.created_at.desc()).limit(limit).all()
    result = []
    for p in payments:
        broker = db.query(Broker).filter(Broker.id == p.broker_id).first() if p.broker_id else None
        rep = db.query(CompanyRep).filter(CompanyRep.id == p.company_rep_id).first() if p.company_rep_id else None
        creditor = db.query(Creditor).filter(Creditor.id == p.creditor_id).first() if p.creditor_id else None
        txn = db.query(Transaction).filter(Transaction.id == p.transaction_id).first() if p.transaction_id else None
        approver = db.query(CompanyRep).filter(CompanyRep.id == p.approved_by_rep_id).first() if p.approved_by_rep_id else None
        
        result.append({
            "id": str(p.id), "payment_id": p.payment_id, "payment_type": p.payment_type,
            "payee_type": p.payee_type, "broker_name": broker.name if broker else None,
            "broker_id": broker.broker_id if broker else None,
            "rep_name": rep.name if rep else None, "rep_id": rep.rep_id if rep else None,
            "creditor_name": creditor.name if creditor else None,
            "creditor_id": creditor.creditor_id if creditor else None,
            "transaction_id": txn.transaction_id if txn else None,
            "amount": float(p.amount), "payment_method": p.payment_method,
            "reference_number": p.reference_number,
            "payment_date": str(p.payment_date) if p.payment_date else None,
            "notes": p.notes, "approved_by": approver.name if approver else None,
            "status": p.status, "created_at": str(p.created_at)
        })
    return result

@app.get("/api/payments/summary")
def get_payments_summary(db: Session = Depends(get_db)):
    today = date.today()
    month_start = today.replace(day=1)
    
    total_payments = db.query(Payment).count()
    total_amount = db.query(func.sum(Payment.amount)).filter(Payment.status == "completed").scalar() or 0
    today_count = db.query(Payment).filter(Payment.payment_date == today, Payment.status == "completed").count()
    today_amount = db.query(func.sum(Payment.amount)).filter(Payment.payment_date == today, Payment.status == "completed").scalar() or 0
    month_count = db.query(Payment).filter(Payment.payment_date >= month_start, Payment.status == "completed").count()
    month_amount = db.query(func.sum(Payment.amount)).filter(Payment.payment_date >= month_start, Payment.status == "completed").scalar() or 0
    
    # By payment type
    by_type = {}
    for ptype in ['broker_commission', 'rep_incentive', 'creditor', 'other']:
        amt = db.query(func.sum(Payment.amount)).filter(Payment.payment_type == ptype, Payment.status == "completed").scalar() or 0
        by_type[ptype] = float(amt)
    
    # Pending payments
    pending_count = db.query(Payment).filter(Payment.status == "pending").count()
    pending_amount = db.query(func.sum(Payment.amount)).filter(Payment.status == "pending").scalar() or 0
    
    return {
        "total_payments": total_payments, "total_amount": float(total_amount),
        "today_count": today_count, "today_amount": float(today_amount),
        "month_count": month_count, "month_amount": float(month_amount),
        "by_type": by_type, "pending_count": pending_count, "pending_amount": float(pending_amount)
    }

@app.get("/api/payments/available-commissions")
def get_available_commissions(broker_id: str = None, rep_id: str = None, db: Session = Depends(get_db)):
    """Calculate available commissions for brokers/reps"""
    result = {}
    
    if broker_id:
        broker = db.query(Broker).filter((Broker.id == broker_id) | (Broker.broker_id == broker_id)).first()
        if not broker: raise HTTPException(404, "Broker not found")
        
        # Get all transactions for this broker
        txns = db.query(Transaction).filter(Transaction.broker_id == broker.id).all()
        total_commission_earned = sum(float(t.total_value or 0) * float(t.broker_commission_rate or 0) / 100 for t in txns)
        
        # Get total paid
        total_paid = db.query(func.sum(Payment.amount)).filter(
            Payment.broker_id == broker.id,
            Payment.payment_type == "broker_commission",
            Payment.status == "completed"
        ).scalar() or 0
        
        result["broker"] = {
            "broker_id": broker.broker_id, "name": broker.name,
            "total_commission_earned": total_commission_earned,
            "total_commission_paid": float(total_paid),
            "available_commission": total_commission_earned - float(total_paid),
            "transactions": [{"transaction_id": t.transaction_id, "commission": float(t.total_value or 0) * float(t.broker_commission_rate or 0) / 100} for t in txns]
        }
    
    if rep_id:
        rep = db.query(CompanyRep).filter((CompanyRep.id == rep_id) | (CompanyRep.rep_id == rep_id)).first()
        if not rep: raise HTTPException(404, "Rep not found")
        
        # For reps, we'd need to define incentive structure - for now, return basic info
        total_paid = db.query(func.sum(Payment.amount)).filter(
            Payment.company_rep_id == rep.id,
            Payment.payment_type == "rep_incentive",
            Payment.status == "completed"
        ).scalar() or 0
        
        result["rep"] = {
            "rep_id": rep.rep_id, "name": rep.name,
            "total_incentive_paid": float(total_paid)
        }
    
    return result

@app.get("/api/payments/broker/{broker_id}")
def get_broker_payments(broker_id: str, db: Session = Depends(get_db)):
    broker = db.query(Broker).filter((Broker.id == broker_id) | (Broker.broker_id == broker_id)).first()
    if not broker: raise HTTPException(404, "Broker not found")
    
    payments = db.query(Payment).filter(
        Payment.broker_id == broker.id
    ).order_by(Payment.created_at.desc()).all()
    
    return [{
        "id": str(p.id), "payment_id": p.payment_id, "amount": float(p.amount),
        "payment_type": p.payment_type, "payment_method": p.payment_method,
        "payment_date": str(p.payment_date), "status": p.status,
        "transaction_id": db.query(Transaction).filter(Transaction.id == p.transaction_id).first().transaction_id if p.transaction_id else None,
        "created_at": str(p.created_at)
    } for p in payments]

@app.get("/api/payments/rep/{rep_id}")
def get_rep_payments(rep_id: str, db: Session = Depends(get_db)):
    rep = db.query(CompanyRep).filter((CompanyRep.id == rep_id) | (CompanyRep.rep_id == rep_id)).first()
    if not rep: raise HTTPException(404, "Rep not found")
    
    payments = db.query(Payment).filter(
        Payment.company_rep_id == rep.id
    ).order_by(Payment.created_at.desc()).all()
    
    return [{
        "id": str(p.id), "payment_id": p.payment_id, "amount": float(p.amount),
        "payment_type": p.payment_type, "payment_method": p.payment_method,
        "payment_date": str(p.payment_date), "status": p.status,
        "created_at": str(p.created_at)
    } for p in payments]

@app.get("/api/payments/{pid}")
def get_payment(pid: str, db: Session = Depends(get_db)):
    p = db.query(Payment).filter((Payment.id == pid) | (Payment.payment_id == pid)).first()
    if not p: raise HTTPException(404, "Payment not found")
    
    broker = db.query(Broker).filter(Broker.id == p.broker_id).first() if p.broker_id else None
    rep = db.query(CompanyRep).filter(CompanyRep.id == p.company_rep_id).first() if p.company_rep_id else None
    creditor = db.query(Creditor).filter(Creditor.id == p.creditor_id).first() if p.creditor_id else None
    txn = db.query(Transaction).filter(Transaction.id == p.transaction_id).first() if p.transaction_id else None
    approver = db.query(CompanyRep).filter(CompanyRep.id == p.approved_by_rep_id).first() if p.approved_by_rep_id else None
    
    # Get allocations
    allocations = db.query(PaymentAllocation).filter(PaymentAllocation.payment_id == p.id).all()
    alloc_details = []
    for a in allocations:
        t = db.query(Transaction).filter(Transaction.id == a.transaction_id).first()
        alloc_details.append({
            "id": str(a.id), "transaction_id": t.transaction_id if t else None,
            "amount": float(a.amount)
        })
    
    return {
        "id": str(p.id), "payment_id": p.payment_id, "payment_type": p.payment_type,
        "payee_type": p.payee_type,
        "broker": {"id": str(broker.id), "name": broker.name, "broker_id": broker.broker_id} if broker else None,
        "rep": {"id": str(rep.id), "name": rep.name, "rep_id": rep.rep_id} if rep else None,
        "creditor": {"id": str(creditor.id), "name": creditor.name, "creditor_id": creditor.creditor_id} if creditor else None,
        "transaction": {"id": str(txn.id), "transaction_id": txn.transaction_id} if txn else None,
        "amount": float(p.amount), "payment_method": p.payment_method,
        "reference_number": p.reference_number,
        "payment_date": str(p.payment_date) if p.payment_date else None,
        "notes": p.notes, "approved_by": {"id": str(approver.id), "name": approver.name} if approver else None,
        "status": p.status, "created_at": str(p.created_at),
        "allocations": alloc_details
    }

@app.post("/api/payments")
def create_payment(data: dict, db: Session = Depends(get_db)):
    broker = None
    rep = None
    creditor = None
    transaction = None
    approver = None
    
    if data.get("broker_id"):
        broker = db.query(Broker).filter(
            (Broker.id == data["broker_id"]) | (Broker.broker_id == data["broker_id"]) | (Broker.mobile == data["broker_id"])
        ).first()
        if not broker: raise HTTPException(404, "Broker not found")
    
    if data.get("company_rep_id"):
        rep = db.query(CompanyRep).filter(
            (CompanyRep.id == data["company_rep_id"]) | (CompanyRep.rep_id == data["company_rep_id"])
        ).first()
        if not rep: raise HTTPException(404, "Rep not found")
    
    if data.get("creditor_id"):
        creditor = db.query(Creditor).filter(
            (Creditor.id == data["creditor_id"]) | (Creditor.creditor_id == data["creditor_id"])
        ).first()
        if not creditor: raise HTTPException(404, "Creditor not found")
    
    if data.get("transaction_id"):
        transaction = db.query(Transaction).filter(
            (Transaction.id == data["transaction_id"]) | (Transaction.transaction_id == data["transaction_id"])
        ).first()
    
    if data.get("approved_by_rep_id"):
        approver = db.query(CompanyRep).filter(
            (CompanyRep.id == data["approved_by_rep_id"]) | (CompanyRep.rep_id == data["approved_by_rep_id"])
        ).first()
    
    # Determine payee_type
    payee_type = None
    if broker: payee_type = "broker"
    elif rep: payee_type = "company_rep"
    elif creditor: payee_type = "creditor"
    
    p = Payment(
        payment_id=_generate_payment_code(db),
        payment_type=data["payment_type"],
        payee_type=payee_type,
        broker_id=broker.id if broker else None,
        company_rep_id=rep.id if rep else None,
        creditor_id=creditor.id if creditor else None,
        transaction_id=transaction.id if transaction else None,
        amount=data["amount"],
        payment_method=data.get("payment_method"),
        reference_number=data.get("reference_number"),
        payment_date=date.fromisoformat(data["payment_date"]) if data.get("payment_date") else date.today(),
        notes=data.get("notes"),
        approved_by_rep_id=approver.id if approver else None,
        status=data.get("status", "completed")
    )
    db.add(p); db.flush()
    
    # Handle allocations
    allocations = data.get("allocations", [])
    if allocations:
        for alloc in allocations:
            txn = db.query(Transaction).filter(
                (Transaction.id == alloc["transaction_id"]) | (Transaction.transaction_id == alloc["transaction_id"])
            ).first()
            if txn:
                db.add(PaymentAllocation(payment_id=p.id, transaction_id=txn.id, amount=alloc["amount"]))
    elif transaction:
        # Auto-allocate to transaction if only one transaction specified
        db.add(PaymentAllocation(payment_id=p.id, transaction_id=transaction.id, amount=data["amount"]))
    
    db.commit(); db.refresh(p)
    return {"message": "Payment created", "id": str(p.id), "payment_id": p.payment_id}

@app.put("/api/payments/{pid}")
def update_payment(pid: str, data: dict, db: Session = Depends(get_db)):
    p = db.query(Payment).filter((Payment.id == pid) | (Payment.payment_id == pid)).first()
    if not p: raise HTTPException(404, "Payment not found")
    
    for k in ["payment_type", "amount", "payment_method", "reference_number", "notes", "status"]:
        if k in data and data[k] is not None: setattr(p, k, data[k])
    if "payment_date" in data: p.payment_date = date.fromisoformat(data["payment_date"]) if data["payment_date"] else None
    
    db.commit()
    return {"message": "Payment updated"}

@app.delete("/api/payments/{pid}")
def delete_payment(pid: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    p = find_entity(db, Payment, "payment_id", pid)
    if not p: raise HTTPException(404, "Payment not found")

    # Creator role cannot delete
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    # Admin can delete directly
    if current_user.role == "admin":
        # Delete allocations
        allocations = db.query(PaymentAllocation).filter(PaymentAllocation.payment_id == p.id).all()
        for a in allocations:
            db.delete(a)

        db.delete(p); db.commit()
        return {"message": "Payment deleted"}

    # Non-admin: create deletion request
    req = DeletionRequest(entity_type="payment", entity_id=p.id, entity_name=p.payment_id, requested_by=current_user.id)
    db.add(req); db.commit(); db.refresh(req)
    return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

# ============================================
# DASHBOARD API
# ============================================
@app.get("/api/dashboard/summary")
def get_dashboard_summary(db: Session = Depends(get_db)):
    """Overall system statistics"""
    try:
        today = date.today()
        month_start = today.replace(day=1)
        
        # Customers
        total_customers = db.query(Customer).count()
        active_customers = db.query(Customer).join(Transaction).distinct().count()
        
        # Transactions
        total_transactions = db.query(Transaction).count()
        total_sale_value = db.query(func.sum(Transaction.total_value)).scalar() or 0
        
        # Receipts
        total_received = db.query(func.sum(Receipt.amount)).scalar() or 0
        month_received = db.query(func.sum(Receipt.amount)).filter(Receipt.payment_date >= month_start).scalar() or 0
        
        # Payments
        total_paid = db.query(func.sum(Payment.amount)).filter(Payment.status == "completed").scalar() or 0
        month_paid = db.query(func.sum(Payment.amount)).filter(Payment.payment_date >= month_start, Payment.status == "completed").scalar() or 0
        
        # Outstanding breakdown: Overdue vs Future Receivable
        today = date.today()
        total_overdue = 0
        total_future_receivable = 0
        
        all_transactions = db.query(Transaction).all()
        for txn in all_transactions:
            installments = db.query(Installment).filter(Installment.transaction_id == txn.id).all()
            for inst in installments:
                balance = float(inst.amount) - float(inst.amount_paid or 0)
                if balance > 0:
                    if inst.due_date <= today:
                        total_overdue += balance
                    else:
                        total_future_receivable += balance
        
        total_outstanding = float(total_sale_value) - float(total_received)
        
        # Projects
        total_projects = db.query(Project).count()
        active_projects = db.query(Project).filter(Project.status == "active").count()
        
        # Brokers
        total_brokers = db.query(Broker).count()
        active_brokers = db.query(Broker).filter(Broker.status == "active").count()
        
        # Inventory
        total_units = db.query(Inventory).count()
        available_units = db.query(Inventory).filter(Inventory.status == "available").count()
        sold_units = db.query(Inventory).filter(Inventory.status == "sold").count()
        
        # This month transactions
        this_month_txns = db.query(Transaction).filter(Transaction.booking_date >= month_start).count()
        this_month_sale = db.query(func.sum(Transaction.total_value)).filter(Transaction.booking_date >= month_start).scalar() or 0
        
        return {
            "customers": {"total": total_customers, "active": active_customers},
            "transactions": {"total": total_transactions, "total_value": float(total_sale_value), "this_month": this_month_txns, "this_month_value": float(this_month_sale)},
            "financials": {
                "total_sale": float(total_sale_value),
                "total_received": float(total_received),
                "total_outstanding": float(total_outstanding),
                "total_overdue": float(total_overdue),
                "future_receivable": float(total_future_receivable),
                "total_paid": float(total_paid),
                "month_received": float(month_received),
                "month_paid": float(month_paid)
            },
            "projects": {"total": total_projects, "active": active_projects},
            "brokers": {"total": total_brokers, "active": active_brokers},
            "inventory": {"total": total_units, "available": available_units, "sold": sold_units}
        }
    except Exception as e:
        error_msg = str(e)
        if "does not exist" in error_msg.lower() or "column" in error_msg.lower():
            raise HTTPException(
                status_code=500,
                detail=f"Database schema error: Missing columns. Please run the migration script: python migrate_vector_schema.py. Error: {error_msg[:200]}"
            )
        raise HTTPException(status_code=500, detail=f"Error loading dashboard summary: {error_msg[:200]}")

@app.get("/api/dashboard/customer-stats")
def get_customer_stats(db: Session = Depends(get_db)):
    """Customer financial summaries"""
    customers = db.query(Customer).all()
    result = []
    
    for c in customers:
        txns = db.query(Transaction).filter(Transaction.customer_id == c.id).all()
        total_sale = sum(float(t.total_value or 0) for t in txns)
        
        # Calculate received
        total_received = 0
        for t in txns:
            installments = db.query(Installment).filter(Installment.transaction_id == t.id).all()
            total_received += sum(float(i.amount_paid or 0) for i in installments)
        
        # Calculate overdue (installments past due date)
        today = date.today()
        overdue = 0
        future_receivable = 0
        
        for t in txns:
            installments = db.query(Installment).filter(Installment.transaction_id == t.id).all()
            for i in installments:
                balance = float(i.amount) - float(i.amount_paid or 0)
                if balance > 0:
                    if i.due_date < today:
                        overdue += balance
                    else:
                        future_receivable += balance
        
        # Interaction count
        interactions = db.query(Interaction).filter(Interaction.customer_id == c.id).count()
        
        result.append({
            "customer_id": c.customer_id,
            "name": c.name,
            "mobile": c.mobile,
            "total_sale": total_sale,
            "total_received": total_received,
            "overdue": overdue,
            "future_receivable": future_receivable,
            "outstanding": total_sale - total_received,
            "interaction_count": interactions,
            "transaction_count": len(txns)
        })
    
    return result

@app.get("/api/dashboard/project-stats")
def get_project_stats(db: Session = Depends(get_db)):
    """Project performance statistics"""
    try:
        projects = db.query(Project).all()
        result = []
        
        for p in projects:
            inventory = db.query(Inventory).filter(Inventory.project_id == p.id).all()
            txns = db.query(Transaction).filter(Transaction.project_id == p.id).all()
            
            # Inventory stats
            available = [i for i in inventory if i.status == "available"]
            sold = [i for i in inventory if i.status == "sold"]
            total_marlas = sum(float(i.area_marla) for i in inventory)
            available_marlas = sum(float(i.area_marla) for i in available)
            sold_marlas = sum(float(i.area_marla) for i in sold)
            
            # Financial stats
            total_sale = sum(float(t.total_value or 0) for t in txns)
            total_received = 0
            overdue = 0
            future_receivable = 0
            today = date.today()
            
            for t in txns:
                installments = db.query(Installment).filter(Installment.transaction_id == t.id).all()
                for i in installments:
                    paid = float(i.amount_paid or 0)
                    total_received += paid
                    balance = float(i.amount) - paid
                    if balance > 0:
                        if i.due_date < today:
                            overdue += balance
                        else:
                            future_receivable += balance
            
            result.append({
                "project_id": p.project_id,
                "name": p.name,
                "location": p.location,
                "inventory": {
                    "total_units": len(inventory),
                    "available_units": len(available),
                    "sold_units": len(sold),
                    "total_marlas": total_marlas,
                    "available_marlas": available_marlas,
                    "sold_marlas": sold_marlas
                },
                "financials": {
                    "total_sale": total_sale,
                    "total_received": total_received,
                    "overdue": overdue,
                    "future_receivable": future_receivable,
                    "outstanding": total_sale - total_received
                },
                "transaction_count": len(txns)
            })
        
        return result
    except Exception as e:
        error_msg = str(e)
        if "does not exist" in error_msg.lower() or "column" in error_msg.lower():
            raise HTTPException(
                status_code=500,
                detail=f"Database schema error: Missing columns. Please run the migration script: python migrate_vector_schema.py. Error: {error_msg[:200]}"
            )
        raise HTTPException(status_code=500, detail=f"Error loading project stats: {error_msg[:200]}")

@app.get("/api/dashboard/broker-stats")
def get_broker_stats(db: Session = Depends(get_db)):
    """Broker performance statistics"""
    brokers = db.query(Broker).all()
    result = []
    
    for b in brokers:
        txns = db.query(Transaction).filter(Transaction.broker_id == b.id).all()
        total_sale = sum(float(t.total_value or 0) for t in txns)
        total_commission_earned = sum(float(t.total_value or 0) * float(t.broker_commission_rate or 0) / 100 for t in txns)
        
        # Commission paid
        total_commission_paid = db.query(func.sum(Payment.amount)).filter(
            Payment.broker_id == b.id,
            Payment.payment_type == "broker_commission",
            Payment.status == "completed"
        ).scalar() or 0
        
        # Interactions
        interactions = db.query(Interaction).filter(Interaction.broker_id == b.id).count()
        
        result.append({
            "broker_id": b.broker_id,
            "name": b.name,
            "mobile": b.mobile,
            "total_transactions": len(txns),
            "total_sale_value": total_sale,
            "commission": {
                "rate": float(b.commission_rate or 0),
                "total_earned": total_commission_earned,
                "total_paid": float(total_commission_paid),
                "pending": total_commission_earned - float(total_commission_paid)
            },
            "interaction_count": interactions
        })
    
    return result

@app.get("/api/dashboard/revenue-trends")
def get_revenue_trends(start_date: str = None, end_date: str = None, db: Session = Depends(get_db)):
    """Time-series revenue data"""
    if not start_date:
        start_date = (date.today() - timedelta(days=90)).isoformat()
    if not end_date:
        end_date = date.today().isoformat()
    
    start = date.fromisoformat(start_date)
    end = date.fromisoformat(end_date)
    
    # Get receipts by date
    receipts = db.query(Receipt).filter(
        Receipt.payment_date >= start,
        Receipt.payment_date <= end
    ).order_by(Receipt.payment_date).all()
    
    # Group by date
    daily_receipts = {}
    for r in receipts:
        date_str = str(r.payment_date)
        if date_str not in daily_receipts:
            daily_receipts[date_str] = 0
        daily_receipts[date_str] += float(r.amount)
    
    # Get payments by date
    payments = db.query(Payment).filter(
        Payment.payment_date >= start,
        Payment.payment_date <= end,
        Payment.status == "completed"
    ).order_by(Payment.payment_date).all()
    
    daily_payments = {}
    for p in payments:
        date_str = str(p.payment_date)
        if date_str not in daily_payments:
            daily_payments[date_str] = 0
        daily_payments[date_str] += float(p.amount)
    
    # Generate all dates in range
    trends = []
    current = start
    while current <= end:
        date_str = current.isoformat()
        trends.append({
            "date": date_str,
            "receipts": daily_receipts.get(date_str, 0),
            "payments": daily_payments.get(date_str, 0),
            "net": daily_receipts.get(date_str, 0) - daily_payments.get(date_str, 0)
        })
        current += timedelta(days=1)
    
    return trends

@app.get("/api/dashboard/top-receivables")
def get_top_receivables(limit: int = 10, db: Session = Depends(get_db)):
    """Top receivables by customer with overdue breakdown"""
    try:
        today = date.today()
        customers = db.query(Customer).all()
        result = []
        
        for c in customers:
            txns = db.query(Transaction).filter(Transaction.customer_id == c.id).all()
            total_sale = sum(float(t.total_value or 0) for t in txns)
            
            # Calculate received
            total_received = 0
            overdue_amount = 0
            future_amount = 0
            overdue_installments = []
            future_installments = []
            
            for t in txns:
                installments = db.query(Installment).filter(Installment.transaction_id == t.id).all()
                for i in installments:
                    paid = float(i.amount_paid or 0)
                    total_received += paid
                    balance = float(i.amount) - paid
                    if balance > 0:
                        inst_data = {
                            "installment_number": i.installment_number,
                            "due_date": str(i.due_date),
                            "amount": float(i.amount),
                            "amount_paid": paid,
                            "balance": balance,
                            "status": i.status,
                            "transaction_id": str(t.transaction_id),
                            "project_name": db.query(Project).filter(Project.id == t.project_id).first().name if t.project_id else None
                        }
                        if i.due_date <= today:
                            overdue_amount += balance
                            overdue_installments.append(inst_data)
                        else:
                            future_amount += balance
                            future_installments.append(inst_data)
            
            outstanding = total_sale - total_received
            if outstanding > 0:
                result.append({
                    "customer_id": c.customer_id,
                    "customer_name": c.name,
                    "mobile": c.mobile,
                    "total_sale": total_sale,
                    "total_received": total_received,
                    "total_outstanding": outstanding,
                    "overdue": overdue_amount,
                    "future_receivable": future_amount,
                    "overdue_installments": overdue_installments,
                    "future_installments": future_installments,
                    "transaction_count": len(txns)
                })
        
        # Sort by total outstanding descending
        result.sort(key=lambda x: x["total_outstanding"], reverse=True)
        return result[:limit]
    except Exception as e:
        error_msg = str(e)
        if "does not exist" in error_msg.lower() or "column" in error_msg.lower():
            raise HTTPException(
                status_code=500,
                detail=f"Database schema error: Missing columns. Please run the migration script: python migrate_vector_schema.py. Error: {error_msg[:200]}"
            )
        raise HTTPException(status_code=500, detail=f"Error loading top receivables: {error_msg[:200]}")

@app.get("/api/dashboard/project-inventory")
def get_project_inventory(db: Session = Depends(get_db)):
    """Project-wise inventory breakdown with detailed metrics"""
    try:
        projects = db.query(Project).all()
        result = []
        
        for p in projects:
            inventory = db.query(Inventory).filter(Inventory.project_id == p.id).all()
            available = [i for i in inventory if i.status == "available"]
            sold = [i for i in inventory if i.status == "sold"]
            
            # Calculate totals
            total_units = len(inventory)
            available_units = len(available)
            sold_units = len(sold)
            
            total_marlas = sum(float(i.area_marla or 0) for i in inventory)
            available_marlas = sum(float(i.area_marla or 0) for i in available)
            sold_marlas = sum(float(i.area_marla or 0) for i in sold)
            
            total_value = sum(float(i.area_marla or 0) * float(i.rate_per_marla or 0) for i in inventory)
            available_value = sum(float(i.area_marla or 0) * float(i.rate_per_marla or 0) for i in available)
            sold_value = sum(float(i.area_marla or 0) * float(i.rate_per_marla or 0) for i in sold)
            
            # Average rates
            avg_rate = total_value / total_marlas if total_marlas > 0 else 0
            available_avg_rate = available_value / available_marlas if available_marlas > 0 else 0
            
            # Unit type breakdown
            unit_types = {}
            for i in inventory:
                ut = i.unit_type or "plot"
                if ut not in unit_types:
                    unit_types[ut] = {"total": 0, "available": 0, "sold": 0}
                unit_types[ut]["total"] += 1
                if i.status == "available":
                    unit_types[ut]["available"] += 1
                elif i.status == "sold":
                    unit_types[ut]["sold"] += 1
            
            result.append({
                "project_id": p.project_id,
                "name": p.name,
                "location": p.location,
                "status": p.status,
                "summary": {
                    "total_units": total_units,
                    "available_units": available_units,
                    "sold_units": sold_units,
                    "utilization_rate": (sold_units / total_units * 100) if total_units > 0 else 0
                },
                "area": {
                    "total_marlas": total_marlas,
                    "available_marlas": available_marlas,
                    "sold_marlas": sold_marlas,
                    "available_percentage": (available_marlas / total_marlas * 100) if total_marlas > 0 else 0
                },
                "value": {
                    "total_value": total_value,
                    "available_value": available_value,
                    "sold_value": sold_value,
                    "avg_rate_per_marla": avg_rate,
                    "available_avg_rate": available_avg_rate
                },
                "unit_types": unit_types
            })
        
        return result
    except Exception as e:
        error_msg = str(e)
        if "does not exist" in error_msg.lower() or "column" in error_msg.lower():
            raise HTTPException(
                status_code=500,
                detail=f"Database schema error: Missing columns. Please run the migration script: python migrate_vector_schema.py. Error: {error_msg[:200]}"
            )
        raise HTTPException(status_code=500, detail=f"Error loading project inventory: {error_msg[:200]}")


@app.get("/api/dashboard/sales-kpis")
def get_sales_kpis(
    month: Optional[int] = None,
    year: Optional[int] = None,
    rep_id: Optional[str] = None,
    project_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Sales KPIs: tokens, partial down payments, closed won, achieved revenue, unit/target.
    Role-scoped: user sees self, manager sees team, admin/cco/director/coo sees all."""
    try:
        today = date.today()
        kpi_month = month or today.month
        kpi_year = year or today.year

        isolation = get_rep_isolation_filter(current_user, db)

        # Build rep scope
        scoped_rep_ids = None  # None = all
        if isolation["isolated"]:
            scoped_rep_ids = isolation["team_rep_uuids"]
        if rep_id:
            rep_obj = db.query(CompanyRep).filter(
                (CompanyRep.rep_id == rep_id) | (func.cast(CompanyRep.id, String) == rep_id)
            ).first()
            if rep_obj:
                if not isolation["isolated"] or rep_obj.id in (scoped_rep_ids or []):
                    scoped_rep_ids = [rep_obj.id]

        # --- Transactions for token/partial/down-payment classification ---
        txn_q = db.query(Transaction).filter(Transaction.status != "cancelled")
        if scoped_rep_ids is not None:
            txn_q = txn_q.filter(Transaction.company_rep_id.in_(scoped_rep_ids))
        if project_id:
            proj = db.query(Project).filter(
                (Project.project_id == project_id) | (func.cast(Project.id, String) == project_id)
            ).first()
            if proj:
                txn_q = txn_q.filter(Transaction.project_id == proj.id)
        all_txns = txn_q.all()

        # Pre-fetch cumulative receipts per transaction (join through installments)
        txn_ids = [t.id for t in all_txns]
        receipt_sums = {}
        if txn_ids:
            alloc_q = db.query(
                Installment.transaction_id,
                func.sum(ReceiptAllocation.amount).label("total_paid")
            ).join(
                Installment, ReceiptAllocation.installment_id == Installment.id
            ).filter(
                Installment.transaction_id.in_(txn_ids)
            )
            alloc_rows = alloc_q.group_by(Installment.transaction_id).all()
            receipt_sums = {row.transaction_id: float(row.total_paid or 0) for row in alloc_rows}

        # Classify each transaction by cumulative paid percentage
        tokens_amount = 0.0
        partial_dp_amount = 0.0
        down_payment_complete_count = 0
        total_received = 0.0

        for t in all_txns:
            tv = float(t.total_value or 0)
            paid = receipt_sums.get(t.id, 0.0)
            total_received += paid
            if tv <= 0:
                continue
            pct = (paid / tv) * 100.0
            # Default threshold: 10% booking amount
            threshold = 10.0
            if pct < threshold:
                tokens_amount += paid
            elif pct < 100.0:
                partial_dp_amount += paid
            if pct >= threshold:
                down_payment_complete_count += 1

        # --- Closed Won Cases (converted leads) ---
        lead_q = db.query(func.count(Lead.id)).filter(Lead.status == "converted")
        if scoped_rep_ids is not None:
            lead_q = lead_q.filter(Lead.assigned_rep_id.in_(scoped_rep_ids))
        closed_won = lead_q.scalar() or 0

        # --- Project units sold vs target ---
        inv_q = db.query(func.count(Inventory.id)).filter(Inventory.status == "sold")
        if project_id and 'proj' in dir() and proj:
            inv_q = inv_q.filter(Inventory.project_id == proj.id)
        units_sold = inv_q.scalar() or 0

        # Monthly target from MonthlyRepTarget
        target_q = db.query(
            func.sum(MonthlyRepTarget.transaction_target).label("txn_target"),
            func.sum(MonthlyRepTarget.revenue_target).label("rev_target")
        ).filter(
            MonthlyRepTarget.target_month == kpi_month,
            MonthlyRepTarget.target_year == kpi_year
        )
        if scoped_rep_ids is not None:
            target_q = target_q.filter(MonthlyRepTarget.rep_id.in_(scoped_rep_ids))
        target_row = target_q.first()
        units_target = int(target_row.txn_target or 0) if target_row else 0
        achievement_pct = round(units_sold / units_target * 100, 1) if units_target > 0 else None

        return {
            "tokens": round(tokens_amount, 2),
            "partial_down_payments": round(partial_dp_amount, 2),
            "closed_won_cases": closed_won,
            "achieved_revenue": round(total_received, 2),
            "project_units_sold": units_sold,
            "project_units_target": units_target if units_target > 0 else None,
            "project_units_target_achievement_pct": achievement_pct,
            "period": {"month": kpi_month, "year": kpi_year},
            "scope": "team" if isolation.get("isolated") else "global"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sales KPI error: {str(e)}")


@app.get("/api/customers/{customer_id}/details")
def get_customer_details(customer_id: str, db: Session = Depends(get_db)):
    """Comprehensive customer details for modal popup"""
    # Find customer by UUID-safe helper first, then mobile fallback.
    customer = find_entity(db, Customer, "customer_id", customer_id) or db.query(Customer).filter(
        Customer.mobile == customer_id
    ).first()
    
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    today = date.today()
    
    # Transactions
    txns = db.query(Transaction).filter(Transaction.customer_id == customer.id).all()
    total_sale = sum(float(t.total_value or 0) for t in txns)
    
    # Receipts
    receipts = db.query(Receipt).filter(Receipt.customer_id == customer.id).order_by(Receipt.payment_date.desc()).all()
    total_received = sum(float(r.amount) for r in receipts)
    
    # Installments breakdown
    overdue_installments = []
    future_installments = []
    paid_installments = []
    total_overdue = 0
    total_future = 0
    transaction_progress = []
    
    for t in txns:
        project = db.query(Project).filter(Project.id == t.project_id).first()
        installments = db.query(Installment).filter(Installment.transaction_id == t.id).order_by(Installment.due_date).all()
        progress_item = {
            "transaction_id": str(t.transaction_id),
            "transaction_uuid": str(t.id),
            "project_name": project.name if project else None,
            "unit_number": t.unit_number,
            "total_value": float(t.total_value or 0),
            "booking_date": str(t.booking_date) if t.booking_date else None,
            "status": t.status,
            "installments_total": len(installments),
            "installments_paid": 0,
            "installments_overdue": 0,
            "installments_pending": 0,
            "cumulative_paid": 0.0,
            "paid_percent": 0.0,
            "threshold_percent": None,
            "classification": None,
            "payment_rule_source": None,
            "classification_error": None
        }
        
        for i in installments:
            paid = float(i.amount_paid or 0)
            balance = float(i.amount) - paid
            
            inst_data = {
                "id": str(i.id),
                "installment_number": i.installment_number,
                "due_date": str(i.due_date),
                "amount": float(i.amount),
                "amount_paid": paid,
                "balance": balance,
                "status": i.status,
                "transaction_id": str(t.transaction_id),
                "project_name": project.name if project else None,
                "unit_number": t.unit_number
            }
            
            if balance <= 0:
                paid_installments.append(inst_data)
                progress_item["installments_paid"] += 1
            elif i.due_date <= today:
                overdue_installments.append(inst_data)
                total_overdue += balance
                progress_item["installments_overdue"] += 1
            else:
                future_installments.append(inst_data)
                total_future += balance
                progress_item["installments_pending"] += 1

            progress_item["cumulative_paid"] += paid

        # Add per-transaction payment progress classification.
        try:
            try:
                from app.services.receipt_classification_service import classify_receipt as _classify_receipt
            except ImportError:
                from services.receipt_classification_service import classify_receipt as _classify_receipt
            c = _classify_receipt(t, db)
            progress_item["classification"] = c.get("classification")
            progress_item["threshold_percent"] = c.get("threshold_percent")
            progress_item["payment_rule_source"] = c.get("payment_rule_source")
            progress_item["paid_percent"] = c.get("paid_percent", 0.0)
        except Exception as exc:
            # Keep endpoint resilient while surfacing what failed for UI diagnostics.
            progress_item["classification_error"] = str(exc)
            tv = float(t.total_value or 0)
            progress_item["paid_percent"] = round((progress_item["cumulative_paid"] / tv) * 100, 2) if tv > 0 else 0.0

        transaction_progress.append(progress_item)
    
    # Interactions
    interactions = db.query(Interaction).filter(Interaction.customer_id == customer.id).order_by(Interaction.created_at.desc()).limit(20).all()
    
    return {
        "customer": {
            "id": str(customer.id),
            "customer_id": customer.customer_id,
            "name": customer.name,
            "mobile": customer.mobile,
            "email": customer.email,
            "address": customer.address,
            "cnic": customer.cnic,
            "notes": customer.notes,
            "additional_mobiles": customer.additional_mobiles or [],
            "source": customer.source, "source_other": customer.source_other,
            "occupation": customer.occupation, "occupation_other": customer.occupation_other,
            "interested_project_id": str(customer.interested_project_id) if customer.interested_project_id else None,
            "interested_project_other": customer.interested_project_other,
            "area": customer.area, "city": customer.city,
            "country_code": customer.country_code or "+92",
            "assigned_rep_id": str(customer.assigned_rep_id) if customer.assigned_rep_id else None,
            "temperature": customer.temperature,
            "created_at": str(customer.created_at) if customer.created_at else None
        },
        "financials": {
            "total_sale": total_sale,
            "total_received": total_received,
            "total_outstanding": total_sale - total_received,
            "total_overdue": total_overdue,
            "future_receivable": total_future,
            "transaction_count": len(txns)
        },
        "installments": {
            "overdue": overdue_installments,
            "future": future_installments,
            "paid": paid_installments
        },
        "transaction_progress": transaction_progress,
        "receipts": [{
            "id": str(r.id),
            "receipt_id": r.receipt_id,
            "amount": float(r.amount),
            "payment_date": str(r.payment_date),
            "payment_method": r.payment_method,
            "reference_number": r.reference_number,
            "notes": r.notes
        } for r in receipts],
        "interactions": [{
            "id": str(i.id),
            "interaction_id": i.interaction_id,
            "interaction_type": i.interaction_type,
            "status": i.status,
            "notes": i.notes,
            "created_at": str(i.created_at),
            "next_follow_up": str(i.next_follow_up) if i.next_follow_up else None,
            "rep_name": db.query(CompanyRep).filter(CompanyRep.id == i.company_rep_id).first().name if i.company_rep_id else None
        } for i in interactions]
    }

@app.get("/api/brokers/{broker_id}/details")
def get_broker_details(broker_id: str, db: Session = Depends(get_db)):
    """Comprehensive broker details for modal popup"""
    # Find broker by UUID-safe helper first, then mobile fallback.
    broker = find_entity(db, Broker, "broker_id", broker_id) or db.query(Broker).filter(
        Broker.mobile == broker_id
    ).first()
    
    if not broker:
        raise HTTPException(status_code=404, detail="Broker not found")
    
    # Transactions
    txns = db.query(Transaction).filter(Transaction.broker_id == broker.id).all()
    total_sale = sum(float(t.total_value or 0) for t in txns)
    
    # Commission calculations
    total_commission_earned = sum(float(t.total_value or 0) * float(t.broker_commission_rate or 0) / 100 for t in txns)
    
    # Commission paid
    payments = db.query(Payment).filter(
        Payment.broker_id == broker.id,
        Payment.payment_type == "broker_commission"
    ).order_by(Payment.payment_date.desc()).all()
    
    total_commission_paid = sum(float(p.amount) for p in payments if p.status == "completed")
    
    # Brokered transactions details
    brokered_transactions = []
    for t in txns:
        project = db.query(Project).filter(Project.id == t.project_id).first()
        customer = db.query(Customer).filter(Customer.id == t.customer_id).first()
        commission = float(t.total_value or 0) * float(t.broker_commission_rate or 0) / 100
        
        brokered_transactions.append({
            "transaction_id": str(t.transaction_id),
            "booking_date": str(t.booking_date),
            "project_name": project.name if project else None,
            "customer_name": customer.name if customer else None,
            "total_value": float(t.total_value),
            "commission_rate": float(t.broker_commission_rate or 0),
            "commission_amount": commission,
            "status": t.status
        })
    
    # Interactions
    interactions = db.query(Interaction).filter(Interaction.broker_id == broker.id).order_by(Interaction.created_at.desc()).limit(20).all()
    
    return {
        "broker": {
            "id": str(broker.id),
            "broker_id": broker.broker_id,
            "name": broker.name,
            "mobile": broker.mobile,
            "email": broker.email,
            "company": broker.company,
            "address": broker.address,
            "cnic": broker.cnic,
            "commission_rate": float(broker.commission_rate or 0),
            "bank_name": broker.bank_name,
            "bank_account": broker.bank_account,
            "bank_iban": broker.bank_iban,
            "status": broker.status,
            "notes": broker.notes,
            "created_at": str(broker.created_at) if broker.created_at else None
        },
        "performance": {
            "total_transactions": len(txns),
            "total_sale_value": total_sale,
            "commission": {
                "rate": float(broker.commission_rate or 0),
                "total_earned": total_commission_earned,
                "total_paid": total_commission_paid,
                "pending": total_commission_earned - total_commission_paid
            }
        },
        "brokered_transactions": brokered_transactions,
        "payments": [{
            "id": str(p.id),
            "payment_id": p.payment_id,
            "amount": float(p.amount),
            "payment_date": str(p.payment_date),
            "payment_method": p.payment_method,
            "status": p.status,
            "reference_number": p.reference_number
        } for p in payments],
        "interactions": [{
            "id": str(i.id),
            "interaction_id": i.interaction_id,
            "interaction_type": i.interaction_type,
            "status": i.status,
            "notes": i.notes,
            "created_at": str(i.created_at),
            "next_follow_up": str(i.next_follow_up) if i.next_follow_up else None,
            "rep_name": db.query(CompanyRep).filter(CompanyRep.id == i.company_rep_id).first().name if i.company_rep_id else None
        } for i in interactions]
    }

# ============================================
# REPORTS API
# ============================================
@app.get("/api/reports/customers/detailed/{customer_id}")
def get_customer_detailed_report(customer_id: str, db: Session = Depends(get_db)):
    """Get detailed customer financial report"""
    try:
        try:
            from app.reports import get_customer_detailed_report as get_report
        except ImportError:
            from reports import get_customer_detailed_report as get_report
        report = get_report(customer_id, db)
        if not report:
            raise HTTPException(status_code=404, detail="Customer not found")
        return report
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating report: {str(e)}")

@app.get("/api/reports/customers/list")
def get_customers_list_report(db: Session = Depends(get_db)):
    """Get customer list with financials"""
    try:
        try:
            from app.reports import get_customer_detailed_report as get_report
        except ImportError:
            from reports import get_customer_detailed_report as get_report
        customers = db.query(Customer).all()
        result = []
        for c in customers:
            try:
                report = get_report(str(c.id), db)
                if report:
                    result.append({
                        "customer_id": report["customer"]["customer_id"],
                        "name": report["customer"]["name"],
                        "mobile": report["customer"]["mobile"],
                        "total_sale": report["financials"]["total_sale"],
                        "total_received": report["financials"]["total_received"],
                        "overdue": report["financials"]["overdue"],
                        "future_receivable": report["financials"]["future_receivable"],
                        "outstanding": report["financials"]["outstanding"],
                        "interaction_count": report["interactions"]["total_count"]
                    })
            except Exception:
                continue  # Skip customers with errors
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating customer list: {str(e)}")

@app.get("/api/reports/projects/{project_id}")
def get_project_report(project_id: str, db: Session = Depends(get_db)):
    """Get project financial report"""
    try:
        try:
            from app.reports import get_project_detailed_report as get_report
        except ImportError:
            from reports import get_project_detailed_report as get_report
        report = get_report(project_id, db)
        if not report:
            raise HTTPException(status_code=404, detail="Project not found")
        return report
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating report: {str(e)}")

@app.get("/api/reports/brokers/{broker_id}")
def get_broker_report(broker_id: str, db: Session = Depends(get_db)):
    """Get detailed broker report"""
    try:
        try:
            from app.reports import get_broker_detailed_report as get_report
        except ImportError:
            from reports import get_broker_detailed_report as get_report
        report = get_report(broker_id, db)
        if not report:
            raise HTTPException(status_code=404, detail="Broker not found")
        return report
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating report: {str(e)}")

@app.get("/api/reports/customers/pdf/{customer_id}")
def get_customer_pdf(customer_id: str, db: Session = Depends(get_db)):
    """Generate PDF report for customer"""
    try:
        try:
            from app.reports import get_customer_detailed_report as get_report
            from app.report_generator import generate_customer_pdf
        except ImportError:
            from reports import get_customer_detailed_report as get_report
            from report_generator import generate_customer_pdf
        report = get_report(customer_id, db)
        if not report:
            raise HTTPException(status_code=404, detail="Customer not found")
        pdf_buffer = generate_customer_pdf(report)
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=customer_{report['customer']['customer_id']}_report.pdf"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

@app.get("/api/reports/customers/excel")
def get_customers_excel(db: Session = Depends(get_db)):
    """Generate Excel report for all customers"""
    try:
        try:
            from app.reports import get_customer_detailed_report as get_report
            from app.report_generator import generate_customer_excel
        except ImportError:
            from reports import get_customer_detailed_report as get_report
            from report_generator import generate_customer_excel
        from io import BytesIO
        import zipfile
        
        customers = db.query(Customer).all()
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for c in customers:
                try:
                    report = get_report(str(c.id), db)
                    if report:
                        excel_buffer = generate_customer_excel(report)
                        zip_file.writestr(f"customer_{report['customer']['customer_id']}.xlsx", excel_buffer.getvalue())
                except Exception:
                    continue  # Skip customers with errors
        
        zip_buffer.seek(0)
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=customers_report.zip"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")

@app.get("/api/reports/projects/pdf/{project_id}")
def get_project_pdf(project_id: str, db: Session = Depends(get_db)):
    """Generate PDF report for project"""
    try:
        try:
            from app.reports import get_project_detailed_report as get_report
            from app.report_generator import generate_project_pdf
        except ImportError:
            from reports import get_project_detailed_report as get_report
            from report_generator import generate_project_pdf
        report = get_report(project_id, db)
        if not report:
            raise HTTPException(status_code=404, detail="Project not found")
        pdf_buffer = generate_project_pdf(report)
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=project_{report['project']['project_id']}_report.pdf"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

@app.get("/api/reports/projects/excel")
def get_projects_excel(db: Session = Depends(get_db)):
    """Generate Excel report for all projects"""
    try:
        try:
            from app.reports import get_project_detailed_report as get_report
            from app.report_generator import generate_customer_excel  # Reuse customer format
        except ImportError:
            from reports import get_project_detailed_report as get_report
            from report_generator import generate_customer_excel  # Reuse customer format
        from io import BytesIO
        import zipfile
        
        projects = db.query(Project).all()
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for p in projects:
                try:
                    report = get_report(str(p.id), db)
                    if report:
                        # Convert project report to similar format
                        excel_buffer = generate_customer_excel({
                            "customer": {"customer_id": report["project"]["project_id"], "name": report["project"]["name"]},
                            "financials": report["financials"],
                            "transactions": report["transactions"],
                            "interactions": {"total_count": 0, "history": []}
                        })
                        zip_file.writestr(f"project_{report['project']['project_id']}.xlsx", excel_buffer.getvalue())
                except Exception:
                    continue  # Skip projects with errors
        
        zip_buffer.seek(0)
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=projects_report.zip"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")

@app.get("/api/reports/brokers/pdf/{broker_id}")
def get_broker_pdf(broker_id: str, db: Session = Depends(get_db)):
    """Generate PDF report for broker"""
    try:
        try:
            from app.reports import get_broker_detailed_report as get_report
            from app.report_generator import generate_broker_pdf
        except ImportError:
            from reports import get_broker_detailed_report as get_report
            from report_generator import generate_broker_pdf
        report = get_report(broker_id, db)
        if not report:
            raise HTTPException(status_code=404, detail="Broker not found")
        pdf_buffer = generate_broker_pdf(report)
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=broker_{report['broker']['broker_id']}_report.pdf"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

@app.get("/api/reports/brokers/excel")
def get_brokers_excel(db: Session = Depends(get_db)):
    """Generate Excel report for all brokers"""
    try:
        try:
            from app.reports import get_broker_detailed_report as get_report
            from app.report_generator import generate_customer_excel  # Reuse format
        except ImportError:
            from reports import get_broker_detailed_report as get_report
            from report_generator import generate_customer_excel  # Reuse format
        from io import BytesIO
        import zipfile
        
        brokers = db.query(Broker).all()
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for b in brokers:
                try:
                    report = get_report(str(b.id), db)
                    if report:
                        excel_buffer = generate_customer_excel({
                            "customer": {"customer_id": report["broker"]["broker_id"], "name": report["broker"]["name"]},
                            "financials": {
                                "total_sale": report["financials"]["total_sale_value"],
                                "total_received": report["financials"].get("total_received", 0),
                                "overdue": report["financials"].get("due", 0),
                                "future_receivable": report["financials"].get("future_receivable", 0),
                                "outstanding": report["financials"].get("outstanding", 0)
                            },
                            "transactions": report["transactions"],
                            "interactions": report["interactions"]
                        })
                        zip_file.writestr(f"broker_{report['broker']['broker_id']}.xlsx", excel_buffer.getvalue())
                except Exception:
                    continue  # Skip brokers with errors
        
        zip_buffer.seek(0)
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=brokers_report.zip"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")

# ============================================
# MEDIA API
# ============================================
import shutil
from pathlib import Path

MEDIA_ROOT = Path(os.getenv("MEDIA_ROOT", "media"))
MEDIA_ROOT.mkdir(exist_ok=True)
for entity_type in ["transactions", "interactions", "projects", "receipts", "payments", "eois", "zakats"]:
    (MEDIA_ROOT / entity_type).mkdir(exist_ok=True)

def get_file_type(filename: str) -> str:
    """Determine file type from extension"""
    ext = filename.lower().split('.')[-1] if '.' in filename else ''
    if ext in ['pdf']: return 'pdf'
    elif ext in ['xlsx', 'xls']: return 'excel'
    elif ext in ['jpg', 'jpeg', 'png', 'gif', 'webp']: return 'image'
    elif ext in ['mp3', 'wav', 'm4a']: return 'audio'
    elif ext in ['mp4', 'avi', 'mov', 'mkv']: return 'video'
    else: return 'other'

@app.get("/api/media/library")
def get_media_library(
    entity_type: str = Query(None),
    file_type: str = Query(None),
    search: str = Query(None),
    limit: int = Query(100),
    offset: int = Query(0),
    db: Session = Depends(get_db)
):
    """Get all media files with search and filters"""
    try:
        query = db.query(MediaFile)
        
        # Apply filters
        if entity_type:
            query = query.filter(MediaFile.entity_type == entity_type)
        if file_type:
            query = query.filter(MediaFile.file_type == file_type)
        if search:
            search_term = f"%{search.lower()}%"
            query = query.filter(
                (func.lower(MediaFile.file_name).like(search_term)) |
                (func.lower(MediaFile.description).like(search_term))
            )
        
        # Get total count
        total = query.count()
        
        # Apply pagination and ordering
        files = query.order_by(MediaFile.created_at.desc()).offset(offset).limit(limit).all()
        
        # Build response with entity information
        result = []
        for f in files:
            # Get entity details
            entity_name = None
            entity_ref = None
            
            if f.entity_type == "transaction":
                entity = db.query(Transaction).filter(Transaction.id == f.entity_id).first()
                if entity:
                    customer = db.query(Customer).filter(Customer.id == entity.customer_id).first() if entity.customer_id else None
                    project = db.query(Project).filter(Project.id == entity.project_id).first() if entity.project_id else None
                    entity_name = f"{(customer.name if customer else 'N/A')} - {(project.name if project else 'N/A')}"
                    entity_ref = entity.transaction_id
            elif f.entity_type == "receipt":
                entity = db.query(Receipt).filter(Receipt.id == f.entity_id).first()
                if entity:
                    customer = db.query(Customer).filter(Customer.id == entity.customer_id).first() if entity.customer_id else None
                    entity_name = f"Receipt for {customer.name if customer else 'N/A'}"
                    entity_ref = entity.receipt_id
            elif f.entity_type == "payment":
                entity = db.query(Payment).filter(Payment.id == f.entity_id).first()
                if entity:
                    if entity.broker_id:
                        broker = db.query(Broker).filter(Broker.id == entity.broker_id).first()
                        entity_name = f"Payment to {broker.name if broker else 'N/A'}"
                    elif entity.company_rep_id:
                        rep = db.query(CompanyRep).filter(CompanyRep.id == entity.company_rep_id).first()
                        entity_name = f"Payment to {rep.name if rep else 'N/A'}"
                    elif entity.creditor_id:
                        creditor = db.query(Creditor).filter(Creditor.id == entity.creditor_id).first()
                        entity_name = f"Payment to {creditor.name if creditor else 'N/A'}"
                    else:
                        entity_name = "Payment"
                    entity_ref = entity.payment_id
            elif f.entity_type == "project":
                try:
                    entity = db.query(Project).filter(Project.id == f.entity_id).first()
                    if entity:
                        entity_name = entity.name
                        entity_ref = entity.project_id
                except Exception as e:
                    # If lookup fails, still return the file but without entity info
                    entity_name = None
                    entity_ref = None
            elif f.entity_type == "interaction":
                entity = db.query(Interaction).filter(Interaction.id == f.entity_id).first()
                if entity:
                    if entity.customer_id:
                        customer = db.query(Customer).filter(Customer.id == entity.customer_id).first()
                        entity_name = f"Interaction with {customer.name if customer else 'N/A'}"
                    elif entity.broker_id:
                        broker = db.query(Broker).filter(Broker.id == entity.broker_id).first()
                        entity_name = f"Interaction with {broker.name if broker else 'N/A'}"
                    else:
                        entity_name = "Interaction"
                    entity_ref = entity.interaction_id
            elif f.entity_type == "eoi":
                entity = db.query(EOICollection).filter(EOICollection.id == f.entity_id).first()
                if entity:
                    entity_name = entity.party_name
                    entity_ref = entity.eoi_id
            
            # Get uploader info
            rep = db.query(CompanyRep).filter(CompanyRep.id == f.uploaded_by_rep_id).first() if f.uploaded_by_rep_id else None
            
            result.append({
                "id": str(f.id),
                "file_id": f.file_id,
                "file_name": f.file_name,
                "file_type": f.file_type,
                "file_size": f.file_size,
                "description": f.description,
                "entity_type": f.entity_type,
                "entity_id": str(f.entity_id),
                "entity_name": entity_name,
                "entity_ref": entity_ref,
                "uploaded_by": rep.name if rep else None,
                "created_at": str(f.created_at)
            })
        
        return {
            "total": total,
            "files": result,
            "limit": limit,
            "offset": offset
        }
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Error in get_media_library: {error_detail}")
        raise HTTPException(status_code=500, detail=f"Error loading media library: {str(e)}")

@app.post("/api/media/upload")
async def upload_media(
    entity_type: str = Form(...),
    entity_id: str = Form(...),
    file: UploadFile = File(...),
    description: str = Form(None),
    uploaded_by_rep_id: str = Form(None),
    db: Session = Depends(get_db)
):
    """Upload media file"""
    if entity_type not in ["transaction", "interaction", "project", "receipt", "payment", "customer", "broker", "task", "report", "eoi", "zakat"]:
        raise HTTPException(400, "Invalid entity_type")

    # Verify entity exists
    entity = None
    if entity_type == "transaction":
        entity = db.query(Transaction).filter((Transaction.id == entity_id) | (Transaction.transaction_id == entity_id)).first()
    elif entity_type == "interaction":
        entity = db.query(Interaction).filter((Interaction.id == entity_id) | (Interaction.interaction_id == entity_id)).first()
    elif entity_type == "project":
        entity = db.query(Project).filter((Project.id == entity_id) | (Project.project_id == entity_id)).first()
    elif entity_type == "receipt":
        entity = db.query(Receipt).filter((Receipt.id == entity_id) | (Receipt.receipt_id == entity_id)).first()
    elif entity_type == "payment":
        entity = db.query(Payment).filter((Payment.id == entity_id) | (Payment.payment_id == entity_id)).first()
    elif entity_type == "customer":
        entity = db.query(Customer).filter((Customer.id == entity_id) | (Customer.customer_id == entity_id)).first()
    elif entity_type == "broker":
        entity = db.query(Broker).filter((Broker.id == entity_id) | (Broker.broker_id == entity_id)).first()
    elif entity_type == "task":
        entity = db.query(Task).filter((Task.id == entity_id) | (Task.task_id == entity_id)).first()
    elif entity_type == "report":
        # Reports are stored against rep.id; skip entity lookup, use entity_id directly
        entity = db.query(CompanyRep).filter((CompanyRep.id == entity_id) | (CompanyRep.rep_id == entity_id)).first()
    elif entity_type == "eoi":
        entity = db.query(EOICollection).filter((EOICollection.id == entity_id) | (EOICollection.eoi_id == entity_id)).first()
    elif entity_type == "zakat":
        entity = db.query(ZakatRecord).filter((ZakatRecord.id == entity_id) | (ZakatRecord.zakat_id == entity_id)).first()

    if not entity:
        raise HTTPException(404, f"{entity_type} not found")

    # Get rep if provided
    rep = None
    if uploaded_by_rep_id:
        rep = find_entity(db, CompanyRep, "rep_id", uploaded_by_rep_id)
    
    # Save file
    file_ext = file.filename.split('.')[-1] if '.' in file.filename else ''
    unique_filename = f"{uuid.uuid4()}.{file_ext}"
    file_path = MEDIA_ROOT / f"{entity_type}s" / unique_filename
    
    # Ensure directory exists
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    file_size = file_path.stat().st_size
    file_type = get_file_type(file.filename)
    
    # Store absolute path to avoid resolution issues
    file_path_absolute = file_path.resolve()
    
    # Generate file_id from sequence
    try:
        file_id_result = db.execute(text("SELECT nextval('media_file_id_seq')")).scalar()
        file_id = f"MED-{str(file_id_result).zfill(5)}"
    except Exception:
        # Fallback if sequence doesn't exist yet
        file_id_result = db.execute(text("SELECT COUNT(*) + 1 FROM media_files")).scalar()
        file_id = f"MED-{str(file_id_result).zfill(5)}"
    
    # Create database record
    # Ensure entity_id is a UUID object
    if isinstance(entity.id, uuid.UUID):
        entity_uuid = entity.id
    elif isinstance(entity.id, str):
        entity_uuid = uuid.UUID(entity.id)
    else:
        entity_uuid = uuid.UUID(str(entity.id))
    
    media = MediaFile(
        file_id=file_id,
        entity_type=entity_type,
        entity_id=entity_uuid,
        file_name=file.filename,
        file_path=str(file_path_absolute),
        file_type=file_type,
        file_size=file_size,
        uploaded_by_rep_id=rep.id if rep else None,
        description=description
    )
    db.add(media); db.commit(); db.refresh(media)
    
    return {
        "message": "File uploaded",
        "id": str(media.id),
        "file_id": media.file_id,
        "file_name": media.file_name,
        "file_size": media.file_size
    }

@app.get("/api/media/{file_id}/download")
def download_media_file(file_id: str, db: Session = Depends(get_db)):
    """Download media file"""
    # Try to match by file_id (string) first, or by id (UUID) if file_id is a valid UUID
    try:
        # Check if file_id is a valid UUID
        file_id_uuid = uuid.UUID(file_id)
        # If it's a UUID, try matching both id and file_id
        media = db.query(MediaFile).filter(
            (MediaFile.id == file_id_uuid) | (MediaFile.file_id == file_id)
        ).first()
    except (ValueError, TypeError):
        # If not a valid UUID, only match by file_id (string)
        media = db.query(MediaFile).filter(MediaFile.file_id == file_id).first()
    
    if not media:
        raise HTTPException(404, "File not found")
    
    # Resolve file path - handle both absolute and relative paths
    file_path_str = media.file_path
    
    # Normalize path separators
    file_path_str = file_path_str.replace('\\', '/')
    
    # Try multiple resolution strategies
    file_path = None
    
    # Strategy 1: If it's already an absolute path and exists
    test_path = Path(file_path_str)
    if test_path.is_absolute() and test_path.exists():
        file_path = test_path
    
    # Strategy 2: Try relative to MEDIA_ROOT (most common for old files)
    if not file_path or not file_path.exists():
        # Remove 'media/' prefix if present
        if file_path_str.startswith('media/'):
            relative_part = file_path_str[6:]  # Remove 'media/' prefix
            file_path = MEDIA_ROOT / relative_part
        else:
            # Use as-is relative to MEDIA_ROOT
            file_path = MEDIA_ROOT / file_path_str
    
    # Strategy 3: Try relative to current working directory
    if not file_path.exists():
        file_path = Path.cwd() / file_path_str
    
    # Strategy 4: Try as absolute path from root
    if not file_path.exists() and not test_path.is_absolute():
        # Try making it absolute from MEDIA_ROOT
        if not file_path_str.startswith(str(MEDIA_ROOT)):
            file_path = MEDIA_ROOT / file_path_str.lstrip('/')
    
    if not file_path.exists():
        # Last attempt: try resolving MEDIA_ROOT and building path
        media_root_abs = MEDIA_ROOT.resolve()
        if file_path_str.startswith('media/'):
            relative_part = file_path_str[6:]
            file_path = media_root_abs / relative_part
        else:
            file_path = media_root_abs / file_path_str.lstrip('/')
    
    if not file_path.exists():
        raise HTTPException(404, f"File not found on disk. Tried: {file_path}. Stored path: {file_path_str}. MEDIA_ROOT: {MEDIA_ROOT}")
    
    return FileResponse(
        path=str(file_path),
        filename=media.file_name,
        media_type="application/octet-stream"
    )

@app.get("/api/media/{entity_type}/{entity_id}")
def list_media_files(entity_type: str, entity_id: str, db: Session = Depends(get_db)):
    """List media files for an entity"""
    if entity_type not in ["transaction", "interaction", "project", "receipt", "payment", "customer", "broker", "task", "report", "eoi", "zakat"]:
        raise HTTPException(400, "Invalid entity_type")

    # Get entity
    entity = None
    if entity_type == "transaction":
        entity = db.query(Transaction).filter((Transaction.id == entity_id) | (Transaction.transaction_id == entity_id)).first()
    elif entity_type == "interaction":
        entity = db.query(Interaction).filter((Interaction.id == entity_id) | (Interaction.interaction_id == entity_id)).first()
    elif entity_type == "project":
        entity = db.query(Project).filter((Project.id == entity_id) | (Project.project_id == entity_id)).first()
    elif entity_type == "receipt":
        entity = db.query(Receipt).filter((Receipt.id == entity_id) | (Receipt.receipt_id == entity_id)).first()
    elif entity_type == "payment":
        entity = db.query(Payment).filter((Payment.id == entity_id) | (Payment.payment_id == entity_id)).first()
    elif entity_type == "customer":
        entity = db.query(Customer).filter((Customer.id == entity_id) | (Customer.customer_id == entity_id)).first()
    elif entity_type == "broker":
        entity = db.query(Broker).filter((Broker.id == entity_id) | (Broker.broker_id == entity_id)).first()
    elif entity_type == "task":
        entity = db.query(Task).filter((Task.id == entity_id) | (Task.task_id == entity_id)).first()
    elif entity_type == "report":
        entity = db.query(CompanyRep).filter((CompanyRep.id == entity_id) | (CompanyRep.rep_id == entity_id)).first()
    elif entity_type == "eoi":
        entity = db.query(EOICollection).filter((EOICollection.id == entity_id) | (EOICollection.eoi_id == entity_id)).first()
    elif entity_type == "zakat":
        entity = db.query(ZakatRecord).filter((ZakatRecord.id == entity_id) | (ZakatRecord.zakat_id == entity_id)).first()

    if not entity:
        raise HTTPException(404, f"{entity_type} not found")

    files = db.query(MediaFile).filter(
        MediaFile.entity_type == entity_type,
        MediaFile.entity_id == entity.id
    ).order_by(MediaFile.created_at.desc()).all()
    
    result = []
    for f in files:
        rep = db.query(CompanyRep).filter(CompanyRep.id == f.uploaded_by_rep_id).first() if f.uploaded_by_rep_id else None
        result.append({
            "id": str(f.id),
            "file_id": f.file_id,
            "file_name": f.file_name,
            "file_type": f.file_type,
            "file_size": f.file_size,
            "description": f.description,
            "uploaded_by": rep.name if rep else None,
            "created_at": str(f.created_at)
        })
    
    return result

@app.get("/api/media/{file_id}")
def get_media_file_info(file_id: str, db: Session = Depends(get_db)):
    """Get media file metadata"""
    # Try to match by file_id (string) first, or by id (UUID) if file_id is a valid UUID
    try:
        # Check if file_id is a valid UUID
        file_id_uuid = uuid.UUID(file_id)
        # If it's a UUID, try matching both id and file_id
        media = db.query(MediaFile).filter(
            (MediaFile.id == file_id_uuid) | (MediaFile.file_id == file_id)
        ).first()
    except (ValueError, TypeError):
        # If not a valid UUID, only match by file_id (string)
        media = db.query(MediaFile).filter(MediaFile.file_id == file_id).first()
    
    if not media:
        raise HTTPException(404, "File not found")
    
    rep = db.query(CompanyRep).filter(CompanyRep.id == media.uploaded_by_rep_id).first() if media.uploaded_by_rep_id else None
    
    return {
        "id": str(media.id),
        "file_id": media.file_id,
        "entity_type": media.entity_type,
        "entity_id": str(media.entity_id),
        "file_name": media.file_name,
        "file_type": media.file_type,
        "file_size": media.file_size,
        "description": media.description,
        "uploaded_by": rep.name if rep else None,
        "created_at": str(media.created_at)
    }

@app.delete("/api/media/{file_id}")
def delete_media_file(file_id: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    """Delete media file"""
    # Try to match by file_id (string) first, or by id (UUID) if file_id is a valid UUID
    try:
        # Check if file_id is a valid UUID
        file_id_uuid = uuid.UUID(file_id)
        # If it's a UUID, try matching both id and file_id
        media = db.query(MediaFile).filter(
            (MediaFile.id == file_id_uuid) | (MediaFile.file_id == file_id)
        ).first()
    except (ValueError, TypeError):
        # If not a valid UUID, only match by file_id (string)
        media = db.query(MediaFile).filter(MediaFile.file_id == file_id).first()

    if not media:
        raise HTTPException(404, "File not found")

    # Creator role cannot delete
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    # Admin can delete directly
    if current_user.role == "admin":
        # Delete file from disk
        file_path = Path(media.file_path)
        if file_path.exists():
            file_path.unlink()

        db.delete(media); db.commit()
        return {"message": "File deleted"}

    # Non-admin: create deletion request
    entity_name = media.file_name if media.file_name else str(media.id)
    req = DeletionRequest(entity_type="media", entity_id=media.id, entity_name=entity_name, requested_by=current_user.id)
    db.add(req); db.commit(); db.refresh(req)
    return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

# ============================================
# DELETION REQUESTS API
# ============================================

@app.get("/api/deletion-requests")
def list_deletion_requests(status: str = None, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    q = db.query(DeletionRequest).order_by(DeletionRequest.requested_at.desc())
    if status:
        q = q.filter(DeletionRequest.status == status)
    if current_user.role != "admin":
        q = q.filter(DeletionRequest.requested_by == current_user.id)
    requests = q.limit(200).all()
    result = []
    for r in requests:
        requester = db.query(CompanyRep).filter(CompanyRep.id == r.requested_by).first()
        reviewer = db.query(CompanyRep).filter(CompanyRep.id == r.reviewed_by).first() if r.reviewed_by else None
        result.append({
            "id": str(r.id), "request_id": r.request_id, "entity_type": r.entity_type,
            "entity_id": str(r.entity_id), "entity_name": r.entity_name,
            "requested_by": requester.name if requester else "Unknown",
            "requested_by_id": str(r.requested_by) if r.requested_by else None,
            "requested_at": str(r.requested_at), "reason": r.reason, "status": r.status,
            "reviewed_by": reviewer.name if reviewer else None,
            "reviewed_at": str(r.reviewed_at) if r.reviewed_at else None,
            "rejection_reason": r.rejection_reason
        })
    return result

@app.get("/api/deletion-requests/pending-count")
def get_pending_deletion_count(db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    count = db.query(DeletionRequest).filter(DeletionRequest.status == "pending").count()
    return {"count": count}

@app.post("/api/deletion-requests/{request_id}/approve")
def approve_deletion_request(request_id: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(require_role(["admin"]))):
    req = find_entity(db, DeletionRequest, "request_id", request_id)
    if not req: raise HTTPException(404, "Deletion request not found")
    if req.status != "pending": raise HTTPException(400, "Request is not pending")

    # Map entity types to models
    entity_map = {
        "customer": Customer, "broker": Broker, "project": Project,
        "company_rep": CompanyRep, "inventory": Inventory, "interaction": Interaction,
        "campaign": Campaign, "lead": Lead, "receipt": Receipt,
        "creditor": Creditor, "payment": Payment, "media": MediaFile
    }
    model_class = entity_map.get(req.entity_type)
    if not model_class:
        raise HTTPException(400, f"Unknown entity type: {req.entity_type}")

    entity = db.query(model_class).filter(model_class.id == req.entity_id).first()
    if not entity:
        req.status = "approved"
        req.reviewed_by = current_user.id
        req.reviewed_at = datetime.utcnow()
        db.commit()
        return {"message": "Approved but entity already deleted"}

    # Handle special cascading cases
    if req.entity_type == "receipt":
        allocations = db.query(ReceiptAllocation).filter(ReceiptAllocation.receipt_id == entity.id).all()
        for a in allocations:
            inst = db.query(Installment).filter(Installment.id == a.installment_id).first()
            if inst:
                inst.amount_paid = max(0, float(inst.amount_paid or 0) - float(a.amount))
                if float(inst.amount_paid) >= float(inst.amount): inst.status = "paid"
                elif float(inst.amount_paid) > 0: inst.status = "partial"
                else: inst.status = "pending"
            db.delete(a)
    elif req.entity_type == "payment":
        allocations = db.query(PaymentAllocation).filter(PaymentAllocation.payment_id == entity.id).all()
        for a in allocations:
            db.delete(a)
    elif req.entity_type == "inventory":
        project_id = entity.project_id
        db.delete(entity)
        db.flush()
        if project_id:
            try: sync_vector_branches_from_orbit(project_id, db)
            except: pass
        req.status = "approved"
        req.reviewed_by = current_user.id
        req.reviewed_at = datetime.utcnow()
        db.commit()
        return {"message": f"{req.entity_type} deletion approved and executed"}

    db.delete(entity)
    req.status = "approved"
    req.reviewed_by = current_user.id
    req.reviewed_at = datetime.utcnow()
    db.commit()
    return {"message": f"{req.entity_type} deletion approved and executed"}

@app.post("/api/deletion-requests/{request_id}/reject")
def reject_deletion_request(request_id: str, data: dict, db: Session = Depends(get_db), current_user: CompanyRep = Depends(require_role(["admin"]))):
    req = find_entity(db, DeletionRequest, "request_id", request_id)
    if not req: raise HTTPException(404, "Deletion request not found")
    if req.status != "pending": raise HTTPException(400, "Request is not pending")
    req.status = "rejected"
    req.reviewed_by = current_user.id
    req.reviewed_at = datetime.utcnow()
    req.rejection_reason = data.get("reason", "")
    db.commit()
    return {"message": "Deletion request rejected"}

# ============================================
# VECTOR API ENDPOINTS
# ============================================

# Vector Projects Endpoints
@app.get("/api/vector/projects")
def get_vector_projects(
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """List all Vector projects (standalone + linked) — lightweight, no heavy metadata"""
    # Join to get linked project names in one query instead of N+1
    projects = db.query(VectorProject).all()
    # Batch-fetch linked project names
    linked_ids = [p.linked_project_id for p in projects if p.linked_project_id]
    linked_map = {}
    if linked_ids:
        linked_projects = db.query(Project).filter(Project.id.in_(linked_ids)).all()
        linked_map = {p.id: p.name for p in linked_projects}

    result = []
    for p in projects:
        # Extract only the lightweight flags the frontend needs for filtering/display
        meta = p.vector_metadata or {}
        has_vector_data = bool(meta.get('plots') or meta.get('annos') or meta.get('shapes')
                               or meta.get('labels') or meta.get('branches') or meta.get('projectMetadata'))

        result.append({
            "id": str(p.id),
            "name": p.name,
            "map_name": p.map_name,
            "map_size": p.map_size,
            "linked_project_id": str(p.linked_project_id) if p.linked_project_id else None,
            "linked_project_name": linked_map.get(p.linked_project_id),
            "has_map": bool(p.map_pdf_base64),
            "has_vector_data": has_vector_data,
            # Lightweight summary instead of full vector_metadata
            "vector_metadata": {
                "isAutoGenerated": meta.get('isAutoGenerated', False),
                "sourceProjectId": meta.get('sourceProjectId'),
                "mapType": meta.get('mapType'),
            } if has_vector_data else None,
            "created_at": str(p.created_at),
            "updated_at": str(p.updated_at)
        })
    return result

@app.get("/api/vector/projects/{project_id}")
def get_vector_project(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Get Vector project details - returns full project data in JSON format for frontend"""
    try:
        try:
            project_uuid = uuid.UUID(project_id)
            project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
        except (ValueError, TypeError):
            raise HTTPException(400, "Invalid project ID")

        if not project:
            raise HTTPException(404, "Vector project not found")

        # Get all related data
        annotations = db.query(VectorAnnotation).filter(VectorAnnotation.project_id == project_uuid).all()
        shapes = db.query(VectorShape).filter(VectorShape.project_id == project_uuid).all()
        labels = db.query(VectorLabel).filter(VectorLabel.project_id == project_uuid).all()
        branches = db.query(VectorBranch).filter(VectorBranch.project_id == project_uuid).all()
        creator_notes = db.query(VectorCreatorNote).filter(VectorCreatorNote.project_id == project_uuid).all()
        change_logs = db.query(VectorChangeLog).filter(VectorChangeLog.project_id == project_uuid).order_by(VectorChangeLog.timestamp.desc()).all()
        legend = db.query(VectorLegend).filter(VectorLegend.project_id == project_uuid).first()

        # Extract plots and metadata from vector_metadata
        plots = project.vector_metadata.get('plots', []) if project.vector_metadata else []
        plot_offsets = project.vector_metadata.get('plotOffsets', {}) if project.vector_metadata else {}
        plot_rotations = project.vector_metadata.get('plotRotations', {}) if project.vector_metadata else {}

        # Quick debug summary
        print(f"LOAD: {project.name} — {len(plots)} plots")

        # Convert annotations to frontend format
        # PRIORITY: Use vector_metadata if it has annotations, otherwise use separate table
        annos = []

        if project.vector_metadata and 'annos' in project.vector_metadata:
            metadata_annos = project.vector_metadata.get('annos')
            if isinstance(metadata_annos, list):
                annos = metadata_annos
        elif len(annotations) > 0:
            for a in annotations:
                annos.append({
                    "id": a.annotation_id,  # Use annotation_id as id (timestamp-based)
                    "note": a.note,
                    "cat": a.category or "",
                    "color": a.color or "#6366f1",
                    "fontSize": a.font_size or 12,
                    "rotation": a.rotation or 0,
                    "plotIds": a.plot_ids or [],
                    "plotNums": a.plot_nums or []
                })
        else:
            pass

        # Convert shapes to frontend format
        # PRIORITY: Use vector_metadata if it has shapes, otherwise use separate table
        shapes_list = []
        if project.vector_metadata and project.vector_metadata.get('shapes') and len(project.vector_metadata.get('shapes', [])) > 0:
            shapes_list = project.vector_metadata.get('shapes', [])
        elif len(shapes) > 0:
            for s in shapes:
                shape_data = s.data if isinstance(s.data, dict) else {}
                shape_obj = {
                    "id": s.shape_id,
                    "type": s.type or "rectangle",
                    "x": s.x or 0,
                    "y": s.y or 0,
                    "width": s.width or 50,
                    "height": s.height or 50,
                    "color": s.color or "#6366f1"
                }
                # Merge additional shape data
                if shape_data:
                    shape_obj.update(shape_data)
                shapes_list.append(shape_obj)

        # Convert labels to frontend format
        # PRIORITY: Use vector_metadata if it has labels, otherwise use separate table
        labels_list = []
        if project.vector_metadata and project.vector_metadata.get('labels') and len(project.vector_metadata.get('labels', [])) > 0:
            labels_list = project.vector_metadata.get('labels', [])
        elif len(labels) > 0:
            for l in labels:
                labels_list.append({
                    "id": l.label_id,
                    "text": l.text or "",
                    "x": l.x or 0,
                    "y": l.y or 0,
                    "size": l.size or 12,
                    "color": l.color or "#000000"
                })

        # Convert branches to frontend format
        # PRIORITY: Use vector_metadata if it has branches, otherwise use separate table
        branches_list = []
        if project.vector_metadata and project.vector_metadata.get('branches') and len(project.vector_metadata.get('branches', [])) > 0:
            branches_list = project.vector_metadata.get('branches', [])
        elif len(branches) > 0:
            for b in branches:
                branches_list.append({
                    "id": str(b.id),
                    "name": b.name,
                    "timestamp": str(b.timestamp) if b.timestamp else None,
                    "annoCount": b.anno_count or 0,
                    "data": b.data or {}
                })

        # Convert creator notes
        # PRIORITY: Use vector_metadata if it has creator notes, otherwise use separate table
        creator_notes_list = []
        if project.vector_metadata and project.vector_metadata.get('creatorNotes') and len(project.vector_metadata.get('creatorNotes', [])) > 0:
            creator_notes_list = project.vector_metadata.get('creatorNotes', [])
        elif len(creator_notes) > 0:
            for n in creator_notes:
                creator_notes_list.append({
                    "id": str(n.id),
                    "text": n.text,
                    "timestamp": str(n.timestamp) if n.timestamp else None
                })

        # Convert change log
        # PRIORITY: Use vector_metadata if it has change log, otherwise use separate table
        change_log_list = []
        if project.vector_metadata and project.vector_metadata.get('changeLog') and len(project.vector_metadata.get('changeLog', [])) > 0:
            change_log_list = project.vector_metadata.get('changeLog', [])
        elif len(change_logs) > 0:
            for log in change_logs:
                change_log_list.append({
                    "timestamp": str(log.timestamp) if log.timestamp else None,
                    "action": log.action or "",
                    "details": log.details or ""
                })

        # Get legend
        # PRIORITY: Use vector_metadata if it has legend data, otherwise use separate table
        if project.vector_metadata and project.vector_metadata.get('legend'):
            legend_data = project.vector_metadata.get('legend', {})
            # Ensure it has all required fields
            if not isinstance(legend_data, dict):
                legend_data = {}
            legend_data = {
                "visible": legend_data.get("visible", True) if isinstance(legend_data.get("visible"), bool) else (legend_data.get("visible", "true") == "true" or legend_data.get("visible", True) == True),
                "minimized": legend_data.get("minimized", False) if isinstance(legend_data.get("minimized"), bool) else (legend_data.get("minimized", "false") == "true" or legend_data.get("minimized", False) == True),
                "position": legend_data.get("position", "bottom-right"),
                "manualEntries": legend_data.get("manualEntries", [])
            }
        elif legend:
            legend_data = {
                "visible": legend.visible if legend else True,
                "minimized": legend.minimized if legend else False,
                "position": legend.position or "bottom-right",
                "manualEntries": legend.manual_entries or [] if legend else []
            }
        else:
            legend_data = {
                "visible": True,
                "minimized": False,
                "position": "bottom-right",
                "manualEntries": []
            }

        # Get inventory - merge from both sources if available
        # Priority: ORBIT data (linked project) + Vector metadata inventory
        inventory = {}

        # First, load from vector_metadata if exists (this preserves user-entered data)
        if project.vector_metadata and project.vector_metadata.get('inventory'):
            inventory = dict(project.vector_metadata.get('inventory', {}))

        # Then, overlay/merge with ORBIT data if linked (ORBIT data takes precedence)
        if project.linked_project_id:
            inventory_items = db.query(Inventory).filter(Inventory.project_id == project.linked_project_id).all()
            for inv in inventory_items:
                # Calculate total value from area and rate
                total_value = float(inv.area_marla or 0) * float(inv.rate_per_marla or 0)

                # Build composite key to match plot naming convention (e.g., "1-IB1")
                unit = inv.unit_number.strip() if inv.unit_number else ''
                blk = inv.block.strip() if inv.block else ''
                composite_key = f"{unit}-{blk}" if blk else unit

                # Merge with existing data (ORBIT data overwrites but preserves extra fields)
                existing = inventory.get(composite_key, {})
                inv_data = {
                    **existing,  # Keep any extra fields from vector_metadata
                    "marla": float(inv.area_marla) if inv.area_marla else existing.get('marla'),
                    "totalValue": total_value if total_value > 0 else existing.get('totalValue', 0),
                    "ratePerMarla": float(inv.rate_per_marla) if inv.rate_per_marla else existing.get('ratePerMarla'),
                    "dimensions": inv.factor_details or existing.get('dimensions', ''),
                    "block": blk,
                    "unitNumber": unit,
                    "unitType": inv.unit_type or existing.get('unitType', 'plot'),
                    "status": inv.status or existing.get('status', ''),
                    "notes": inv.notes or existing.get('notes', ''),
                    "plotCoordinates": inv.plot_coordinates or existing.get('plotCoordinates'),
                    "isManualPlot": inv.is_manual_plot == "true"
                }

                # Store by composite key (primary)
                inventory[composite_key] = inv_data
                # Also store by unit-only for fallback lookups
                if blk and unit not in inventory:
                    inventory[unit] = inv_data

        print(f"LOAD: {len(inventory)} inventory items")

        # Build project metadata
        project_metadata = project.vector_metadata.get('projectMetadata', {}) if project.vector_metadata else {
            "created": str(project.created_at),
            "lastModified": str(project.updated_at),
            "version": "8.2",
            "createdBy": "",
            "description": "",
            "notes": []
        }

        # Get linked project info
        linked_project_name = None
        if project.linked_project_id:
            linked_project = db.query(Project).filter(Project.id == project.linked_project_id).first()
            if linked_project:
                linked_project_name = linked_project.name

        # Return full project data in frontend JSON format
        return {
            "projectName": project.name,
            "mapName": project.map_name or "Map",
            "pdfBase64": project.map_pdf_base64,
            "plots": plots,
            "annos": annos,  # Preserve order from database
            "inventory": inventory,
            "shapes": shapes_list,
            "labels": labels_list,
            "branches": branches_list,
            "plotOffsets": plot_offsets,
            "plotRotations": plot_rotations,
            "creatorNotes": creator_notes_list,
            "changeLog": change_log_list,
            "legend": legend_data,
            "projectMetadata": project_metadata,
            "systemBranches": project.system_branches or {"sales": {}, "inventory": {}},
            "linkedProjectId": str(project.linked_project_id) if project.linked_project_id else None,
            "linkedProjectName": linked_project_name
        }
    except HTTPException:
        raise  # Re-raise HTTP exceptions as-is
    except Exception as e:
        import traceback
        print(f"ERROR in get_vector_project: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(500, f"Failed to load project: {str(e)}")

@app.post("/api/vector/projects")
def create_vector_project(
    name: str = Form(...),
    map_name: str = Form(None),
    map_pdf_base64: str = Form(None),
    map_size: str = Form(None),
    vector_metadata: str = Form(None),
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Create Vector project (standalone)"""
    import json
    
    parsed_metadata = json.loads(vector_metadata) if vector_metadata else None
    if parsed_metadata:
        # Debug: Log what's being saved
        plots_in_save = parsed_metadata.get('plots', [])
        annos_in_save = parsed_metadata.get('annos', [])
        print(f"DEBUG CREATE: Creating project with {len(plots_in_save)} plots and {len(annos_in_save) if isinstance(annos_in_save, list) else 'non-list'} annotations")
        if isinstance(plots_in_save, list) and len(plots_in_save) > 0:
            print(f"DEBUG CREATE: First 3 plot IDs: {[p.get('id') for p in plots_in_save[:3]]}")
        if isinstance(annos_in_save, list) and len(annos_in_save) > 0:
            print(f"DEBUG CREATE: First annotation plotIds (first 5): {annos_in_save[0].get('plotIds', [])[:5] if annos_in_save[0] else 'none'}")
    
    project = VectorProject(
        name=name,
        map_name=map_name,
        map_pdf_base64=map_pdf_base64,
        map_size=json.loads(map_size) if map_size else None,
        vector_metadata=parsed_metadata,
        system_branches={"sales": {}, "inventory": {}}
    )
    db.add(project)
    db.commit()
    db.refresh(project)
    
    return {
        "id": str(project.id),
        "name": project.name,
        "map_name": project.map_name,
        "map_size": project.map_size,
        "system_branches": project.system_branches,
        "created_at": str(project.created_at)
    }

@app.put("/api/vector/projects/{project_id}")
def update_vector_project(
    project_id: str,
    name: str = Form(None),
    map_name: str = Form(None),
    map_pdf_base64: str = Form(None),
    map_size: str = Form(None),
    vector_metadata: str = Form(None),
    system_branches: str = Form(None),
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Update Vector project"""
    import json
    
    try:
        project_uuid = uuid.UUID(project_id)
        project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")
    
    if not project:
        raise HTTPException(404, "Vector project not found")
    
    if name is not None:
        project.name = name
    if map_name is not None:
        project.map_name = map_name
    if map_pdf_base64 is not None:
        project.map_pdf_base64 = map_pdf_base64
    if map_size is not None:
        project.map_size = json.loads(map_size) if map_size else None
    if vector_metadata is not None:
        parsed_metadata = json.loads(vector_metadata) if vector_metadata else None
        if parsed_metadata:
            # Debug: Log what's being saved
            plots_in_save = parsed_metadata.get('plots', [])
            annos_in_save = parsed_metadata.get('annos', [])
            print(f"DEBUG UPDATE: Updating project with {len(plots_in_save)} plots and {len(annos_in_save) if isinstance(annos_in_save, list) else 'non-list'} annotations")
            if isinstance(plots_in_save, list) and len(plots_in_save) > 0:
                print(f"DEBUG UPDATE: First 3 plot IDs: {[p.get('id') for p in plots_in_save[:3]]}")
            if isinstance(annos_in_save, list) and len(annos_in_save) > 0:
                print(f"DEBUG UPDATE: First annotation plotIds (first 5): {annos_in_save[0].get('plotIds', [])[:5] if annos_in_save[0] else 'none'}")
        project.vector_metadata = parsed_metadata
    if system_branches is not None:
        project.system_branches = json.loads(system_branches) if system_branches else None
    
    project.updated_at = func.now()
    db.commit()
    db.refresh(project)
    
    return {
        "id": str(project.id),
        "name": project.name,
        "map_name": project.map_name,
        "map_size": project.map_size,
        "system_branches": project.system_branches,
        "updated_at": str(project.updated_at)
    }

@app.delete("/api/vector/projects/{project_id}")
def delete_vector_project(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Delete Vector project"""
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    try:
        project_uuid = uuid.UUID(project_id)
        project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

    if not project:
        raise HTTPException(404, "Vector project not found")

    # Non-admin: create deletion request
    if current_user.role != "admin":
        req = DeletionRequest(entity_type="vector_project", entity_id=project.id, entity_name=project.name, requested_by=current_user.id)
        db.add(req); db.commit(); db.refresh(req)
        return {"message": "Deletion request submitted", "request_id": req.request_id, "pending": True}

    db.delete(project)
    db.commit()
    return {"message": "Vector project deleted"}

@app.post("/api/vector/projects/{project_id}/link")
def link_vector_project(
    project_id: str,
    linked_project_id: str = Form(...),
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Link Vector project to Radius project"""
    import json
    
    try:
        vector_project_uuid = uuid.UUID(project_id)
        radius_project_uuid = uuid.UUID(linked_project_id)
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")
    
    vector_project = db.query(VectorProject).filter(VectorProject.id == vector_project_uuid).first()
    if not vector_project:
        raise HTTPException(404, "Vector project not found")
    
    radius_project = db.query(Project).filter(Project.id == radius_project_uuid).first()
    if not radius_project:
        raise HTTPException(404, "Radius project not found")
    
    vector_project.linked_project_id = radius_project_uuid
    # Initialize system branches if not exists
    if not vector_project.system_branches:
        vector_project.system_branches = {"sales": {}, "inventory": {}}

    # AUTO-GENERATE ANNOTATIONS from ORBIT data
    # Get Vector plots for matching
    plots = vector_project.vector_metadata.get('plots', []) if vector_project.vector_metadata else []
    print(f"DEBUG LINK: Vector project has {len(plots)} plots")

    # Build a map of plot number to plot ID with multiple matching strategies
    plot_map = {}
    for p in plots:
        plot_num = p.get('n', '')
        if plot_num:
            # Store multiple variations for flexible matching
            plot_map[plot_num] = p.get('id')
            plot_map[plot_num.upper()] = p.get('id')
            plot_map[plot_num.lower()] = p.get('id')
            plot_map[plot_num.strip()] = p.get('id')
            plot_map[plot_num.strip().upper()] = p.get('id')
            plot_map[plot_num.strip().lower()] = p.get('id')
            # Also try removing common prefixes/suffixes
            clean_num = plot_num.strip().upper().replace('PLOT', '').replace('PLOT#', '').replace('#', '').replace('-', '').strip()
            if clean_num:
                plot_map[clean_num] = p.get('id')

    print(f"DEBUG LINK: Plot map has {len(plot_map)} entries, sample keys: {list(plot_map.keys())[:10]}")

    # Get inventory and transactions from linked ORBIT project
    inventory_items = db.query(Inventory).filter(Inventory.project_id == radius_project_uuid).all()
    transactions = db.query(Transaction).filter(Transaction.project_id == radius_project_uuid).all()
    print(f"DEBUG LINK: Found {len(inventory_items)} inventory items and {len(transactions)} transactions")

    # Build set of sold unit numbers
    sold_units = set()
    for txn in transactions:
        if txn.unit_number:
            sold_units.add(txn.unit_number.upper())
            sold_units.add(txn.unit_number.strip().upper())
    for item in inventory_items:
        if item.status and item.status.lower() == 'sold' and item.unit_number:
            sold_units.add(item.unit_number.upper())
            sold_units.add(item.unit_number.strip().upper())
    print(f"DEBUG LINK: Found {len(sold_units)} sold units")

    # Build annotation arrays
    sold_plot_ids = []
    sold_plot_nums = []
    available_plot_ids = []
    available_plot_nums = []
    unmatched_units = []

    # Helper function to find matching plot ID
    def find_plot_id(unit_num):
        if not unit_num:
            return None
        # Try various matching strategies
        variations = [
            unit_num,
            unit_num.upper(),
            unit_num.lower(),
            unit_num.strip(),
            unit_num.strip().upper(),
            unit_num.strip().lower(),
            unit_num.strip().upper().replace('PLOT', '').replace('#', '').replace('-', '').strip(),
        ]
        for v in variations:
            if v in plot_map:
                return plot_map[v]
        return None

    for item in inventory_items:
        unit_num = item.unit_number
        if not unit_num:
            continue
        plot_id = find_plot_id(unit_num)
        if plot_id:
            if unit_num.upper() in sold_units or unit_num.strip().upper() in sold_units:
                sold_plot_ids.append(plot_id)
                sold_plot_nums.append(unit_num)
            else:
                available_plot_ids.append(plot_id)
                available_plot_nums.append(unit_num)
        else:
            unmatched_units.append(unit_num)

    if unmatched_units:
        print(f"DEBUG LINK: {len(unmatched_units)} units didn't match any plots: {unmatched_units[:10]}")
    print(f"DEBUG LINK: Matched {len(sold_plot_ids)} sold plots, {len(available_plot_ids)} available plots")

    # Create auto-generated annotations
    import time
    current_time = int(time.time() * 1000)

    sales_annotation = None
    if sold_plot_ids:
        sales_annotation = {
            "id": f"auto_sold_{current_time}",
            "note": "SOLD",
            "cat": "SOLD",
            "color": "#ef4444",
            "plotIds": sold_plot_ids,
            "plotNums": sold_plot_nums,
            "fontSize": 12,
            "rotation": 0,
            "isAutoGenerated": True
        }

    inventory_annotation = None
    if available_plot_ids:
        inventory_annotation = {
            "id": f"auto_inventory_{current_time}",
            "note": "INVENTORY",
            "cat": "INVENTORY",
            "color": "#22c55e",
            "plotIds": available_plot_ids,
            "plotNums": available_plot_nums,
            "fontSize": 12,
            "rotation": 0,
            "isAutoGenerated": True
        }

    # Check if this is a "raw" link (no inventory/transactions in ORBIT yet)
    is_raw_link = len(inventory_items) == 0 and len(transactions) == 0

    # Create raw branch with all manual plots if no ORBIT data exists yet
    raw_annotation = None
    if is_raw_link and plots:
        # Create a raw branch containing all plots from the Vector map
        # This serves as a template for when ORBIT data is populated
        all_plot_ids = [p.get('id') for p in plots if p.get('id')]
        all_plot_nums = [p.get('n') for p in plots if p.get('n')]
        raw_annotation = {
            "id": f"raw_linked_{current_time}",
            "note": "LINKED (Awaiting ORBIT Data)",
            "cat": "RAW",
            "color": "#94a3b8",  # Gray color for pending
            "plotIds": all_plot_ids,
            "plotNums": all_plot_nums,
            "fontSize": 12,
            "rotation": 0,
            "isAutoGenerated": True,
            "isRawLink": True
        }

    # Store in system_branches
    # Keep existing manual annotations (WIP) intact - they are in vector_metadata.annos
    vector_project.system_branches = {
        "sales": {
            "annotations": [sales_annotation] if sales_annotation else [],
            "lastSyncAt": datetime.utcnow().isoformat(),
            "transactionCount": len(transactions),
            "plotCount": len(sold_plot_ids)
        },
        "inventory": {
            "annotations": [inventory_annotation] if inventory_annotation else [],
            "lastSyncAt": datetime.utcnow().isoformat(),
            "unitCount": len(inventory_items),
            "plotCount": len(available_plot_ids)
        },
        "raw": {
            "annotations": [raw_annotation] if raw_annotation else [],
            "lastSyncAt": datetime.utcnow().isoformat(),
            "isAwaitingData": is_raw_link,
            "plotCount": len(plots) if is_raw_link else 0
        }
    }

    vector_project.updated_at = func.now()
    db.commit()

    # Create or update reconciliation record
    recon = db.query(VectorReconciliation).filter(VectorReconciliation.project_id == vector_project_uuid).first()
    sync_status = "awaiting_data" if is_raw_link else "synced"
    if not recon:
        recon = VectorReconciliation(
            project_id=vector_project_uuid,
            linked_project_id=radius_project_uuid,
            sync_status=sync_status,
            last_sync_at=datetime.utcnow()
        )
        db.add(recon)
    else:
        recon.linked_project_id = radius_project_uuid
        recon.sync_status = sync_status
        recon.last_sync_at = datetime.utcnow()

    db.commit()

    # Build response message based on link type
    if is_raw_link:
        message = "Vector project linked to ORBIT project (awaiting inventory/transaction data)"
    else:
        message = "Vector project linked to ORBIT project with auto-generated annotations"

    return {
        "message": message,
        "vector_project_id": str(vector_project.id),
        "linked_project_id": str(radius_project.id),
        "linked_project_name": radius_project.name,
        "sold_plots": len(sold_plot_ids),
        "inventory_plots": len(available_plot_ids),
        "is_raw_link": is_raw_link,
        "total_vector_plots": len(plots)
    }

@app.post("/api/vector/projects/{project_id}/unlink")
def unlink_vector_project(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Unlink Vector project from Radius project"""
    try:
        project_uuid = uuid.UUID(project_id)
        project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")
    
    if not project:
        raise HTTPException(404, "Vector project not found")
    
    project.linked_project_id = None
    project.updated_at = func.now()
    db.commit()
    
    # Update reconciliation record
    recon = db.query(VectorReconciliation).filter(VectorReconciliation.project_id == project_uuid).first()
    if recon:
        recon.linked_project_id = None
        recon.sync_status = "standalone"
        db.commit()
    
    return {"message": "Vector project unlinked"}

@app.get("/api/vector/projects/{project_id}/sync-status")
def get_sync_status(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Get sync status for Vector project"""
    try:
        project_uuid = uuid.UUID(project_id)
        project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")
    
    if not project:
        raise HTTPException(404, "Vector project not found")
    
    recon = db.query(VectorReconciliation).filter(VectorReconciliation.project_id == project_uuid).first()
    
    # Check if incomplete (no source map)
    is_incomplete = not bool(project.map_pdf_base64)
    
    status = "standalone"
    if project.linked_project_id:
        status = recon.sync_status if recon else "linked"
    
    return {
        "status": status,
        "is_incomplete": is_incomplete,
        "linked_project_id": str(project.linked_project_id) if project.linked_project_id else None,
        "last_sync_at": str(recon.last_sync_at) if recon and recon.last_sync_at else None,
        "has_map": bool(project.map_pdf_base64)
    }

# Vector Annotations Endpoints
@app.get("/api/vector/projects/{project_id}/annotations")
def get_vector_annotations(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Get annotations for Vector project"""
    try:
        project_uuid = uuid.UUID(project_id)
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")
    
    annotations = db.query(VectorAnnotation).filter(VectorAnnotation.project_id == project_uuid).all()
    return [{
        "id": str(a.id),
        "annotation_id": a.annotation_id,
        "note": a.note,
        "category": a.category,
        "color": a.color,
        "font_size": a.font_size,
        "rotation": a.rotation,
        "plot_ids": a.plot_ids,
        "plot_nums": a.plot_nums
    } for a in annotations]

@app.post("/api/vector/projects/{project_id}/annotations")
def create_vector_annotation(
    project_id: str,
    annotation_id: str = Form(...),
    note: str = Form(None),
    category: str = Form(None),
    color: str = Form(None),
    font_size: int = Form(None),
    rotation: int = Form(0),
    plot_ids: str = Form(None),
    plot_nums: str = Form(None),
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Create annotation for Vector project"""
    import json
    
    try:
        project_uuid = uuid.UUID(project_id)
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")
    
    project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    if not project:
        raise HTTPException(404, "Vector project not found")
    
    annotation = VectorAnnotation(
        project_id=project_uuid,
        annotation_id=annotation_id,
        note=note,
        category=category,
        color=color,
        font_size=font_size,
        rotation=rotation,
        plot_ids=json.loads(plot_ids) if plot_ids else [],
        plot_nums=json.loads(plot_nums) if plot_nums else []
    )
    db.add(annotation)
    db.commit()
    db.refresh(annotation)
    
    return {
        "id": str(annotation.id),
        "annotation_id": annotation.annotation_id,
        "note": annotation.note,
        "plot_ids": annotation.plot_ids,
        "plot_nums": annotation.plot_nums
    }

@app.put("/api/vector/projects/{project_id}/annotations/{anno_id}")
def update_vector_annotation(
    project_id: str,
    anno_id: str,
    note: str = Form(None),
    category: str = Form(None),
    color: str = Form(None),
    font_size: int = Form(None),
    rotation: int = Form(None),
    plot_ids: str = Form(None),
    plot_nums: str = Form(None),
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Update annotation"""
    import json
    
    try:
        project_uuid = uuid.UUID(project_id)
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")
    
    annotation = db.query(VectorAnnotation).filter(
        VectorAnnotation.project_id == project_uuid,
        VectorAnnotation.annotation_id == anno_id
    ).first()
    
    if not annotation:
        raise HTTPException(404, "Annotation not found")
    
    if note is not None:
        annotation.note = note
    if category is not None:
        annotation.category = category
    if color is not None:
        annotation.color = color
    if font_size is not None:
        annotation.font_size = font_size
    if rotation is not None:
        annotation.rotation = rotation
    if plot_ids is not None:
        annotation.plot_ids = json.loads(plot_ids) if plot_ids else []
    if plot_nums is not None:
        annotation.plot_nums = json.loads(plot_nums) if plot_nums else []
    
    annotation.updated_at = func.now()
    db.commit()
    db.refresh(annotation)
    
    return {
        "id": str(annotation.id),
        "annotation_id": annotation.annotation_id,
        "note": annotation.note,
        "plot_ids": annotation.plot_ids,
        "plot_nums": annotation.plot_nums
    }

@app.delete("/api/vector/projects/{project_id}/annotations/{anno_id}")
def delete_vector_annotation(
    project_id: str,
    anno_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Delete annotation"""
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")

    try:
        project_uuid = uuid.UUID(project_id)
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

    annotation = db.query(VectorAnnotation).filter(
        VectorAnnotation.project_id == project_uuid,
        VectorAnnotation.annotation_id == anno_id
    ).first()

    if not annotation:
        raise HTTPException(404, "Annotation not found")

    db.delete(annotation)
    db.commit()
    return {"message": "Annotation deleted"}

# Vector Shapes, Labels, Legend, Branches endpoints (similar pattern)
@app.get("/api/vector/projects/{project_id}/shapes")
def get_vector_shapes(project_id: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    try:
        project_uuid = uuid.UUID(project_id)
        shapes = db.query(VectorShape).filter(VectorShape.project_id == project_uuid).all()
        return [{"id": str(s.id), "shape_id": s.shape_id, "type": s.type, "x": s.x, "y": s.y, "width": s.width, "height": s.height, "color": s.color, "data": s.data} for s in shapes]
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

@app.post("/api/vector/projects/{project_id}/shapes")
def create_vector_shape(project_id: str, shape_id: str = Form(...), type: str = Form(None), x: int = Form(None), y: int = Form(None), width: int = Form(None), height: int = Form(None), color: str = Form(None), data: str = Form(None), db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    import json
    try:
        project_uuid = uuid.UUID(project_id)
        project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
        if not project:
            raise HTTPException(404, "Vector project not found")
        shape = VectorShape(project_id=project_uuid, shape_id=shape_id, type=type, x=x, y=y, width=width, height=height, color=color, data=json.loads(data) if data else None)
        db.add(shape)
        db.commit()
        db.refresh(shape)
        return {"id": str(shape.id), "shape_id": shape.shape_id}
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

@app.get("/api/vector/projects/{project_id}/labels")
def get_vector_labels(project_id: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    try:
        project_uuid = uuid.UUID(project_id)
        labels = db.query(VectorLabel).filter(VectorLabel.project_id == project_uuid).all()
        return [{"id": str(l.id), "label_id": l.label_id, "text": l.text, "x": l.x, "y": l.y, "size": l.size, "color": l.color} for l in labels]
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

@app.post("/api/vector/projects/{project_id}/labels")
def create_vector_label(project_id: str, label_id: str = Form(...), text: str = Form(None), x: int = Form(None), y: int = Form(None), size: int = Form(None), color: str = Form(None), db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    try:
        project_uuid = uuid.UUID(project_id)
        project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
        if not project:
            raise HTTPException(404, "Vector project not found")
        label = VectorLabel(project_id=project_uuid, label_id=label_id, text=text, x=x, y=y, size=size, color=color)
        db.add(label)
        db.commit()
        db.refresh(label)
        return {"id": str(label.id), "label_id": label.label_id}
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

@app.get("/api/vector/projects/{project_id}/legend")
def get_vector_legend(project_id: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    try:
        project_uuid = uuid.UUID(project_id)
        legend = db.query(VectorLegend).filter(VectorLegend.project_id == project_uuid).first()
        if not legend:
            return {"visible": True, "minimized": False, "position": None, "manual_entries": []}
        return {"id": str(legend.id), "visible": legend.visible == "true", "minimized": legend.minimized == "true", "position": legend.position, "manual_entries": legend.manual_entries}
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

@app.put("/api/vector/projects/{project_id}/legend")
def update_vector_legend(project_id: str, visible: str = Form(None), minimized: str = Form(None), position: str = Form(None), manual_entries: str = Form(None), db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    import json
    try:
        project_uuid = uuid.UUID(project_id)
        project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
        if not project:
            raise HTTPException(404, "Vector project not found")
        legend = db.query(VectorLegend).filter(VectorLegend.project_id == project_uuid).first()
        if not legend:
            legend = VectorLegend(project_id=project_uuid, visible="true", minimized="false")
            db.add(legend)
        if visible is not None:
            legend.visible = visible
        if minimized is not None:
            legend.minimized = minimized
        if position is not None:
            legend.position = position
        if manual_entries is not None:
            legend.manual_entries = json.loads(manual_entries) if manual_entries else None
        legend.updated_at = func.now()
        db.commit()
        db.refresh(legend)
        return {"id": str(legend.id), "visible": legend.visible == "true", "minimized": legend.minimized == "true", "position": legend.position, "manual_entries": legend.manual_entries}
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

@app.get("/api/vector/projects/{project_id}/branches")
def get_vector_branches(project_id: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    try:
        project_uuid = uuid.UUID(project_id)
        branches = db.query(VectorBranch).filter(VectorBranch.project_id == project_uuid).all()
        return [{"id": str(b.id), "name": b.name, "timestamp": str(b.timestamp) if b.timestamp else None, "anno_count": b.anno_count, "data": b.data} for b in branches]
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

@app.post("/api/vector/projects/{project_id}/branches")
def create_vector_branch(project_id: str, name: str = Form(...), timestamp: str = Form(None), anno_count: int = Form(None), data: str = Form(None), db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    import json
    from dateutil.parser import parse as parse_date
    try:
        project_uuid = uuid.UUID(project_id)
        project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
        if not project:
            raise HTTPException(404, "Vector project not found")
        branch = VectorBranch(project_id=project_uuid, name=name, timestamp=parse_date(timestamp) if timestamp else datetime.utcnow(), anno_count=anno_count, data=json.loads(data) if data else None)
        db.add(branch)
        db.commit()
        db.refresh(branch)
        return {"id": str(branch.id), "name": branch.name}
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

@app.post("/api/vector/projects/{project_id}/backup")
def create_vector_backup(project_id: str, backup_data: str = Form(...), db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    import json
    try:
        project_uuid = uuid.UUID(project_id)
        project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
        if not project:
            raise HTTPException(404, "Vector project not found")
        backup = VectorProjectBackup(project_id=project_uuid, backup_data=json.loads(backup_data), created_by=current_user.id)
        db.add(backup)
        db.commit()
        db.refresh(backup)
        return {"id": str(backup.id), "created_at": str(backup.created_at)}
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

@app.get("/api/vector/projects/{project_id}/backups")
def get_vector_backups(project_id: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    try:
        project_uuid = uuid.UUID(project_id)
        backups = db.query(VectorProjectBackup).filter(VectorProjectBackup.project_id == project_uuid).order_by(VectorProjectBackup.created_at.desc()).all()
        return [{"id": str(b.id), "created_at": str(b.created_at), "created_by": str(b.created_by)} for b in backups]
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

@app.post("/api/vector/projects/{project_id}/backups/{backup_id}/restore")
def restore_vector_backup(project_id: str, backup_id: str, db: Session = Depends(get_db), current_user: CompanyRep = Depends(get_current_user)):
    try:
        project_uuid = uuid.UUID(project_id)
        backup_uuid = uuid.UUID(backup_id)
        backup = db.query(VectorProjectBackup).filter(VectorProjectBackup.id == backup_uuid, VectorProjectBackup.project_id == project_uuid).first()
        if not backup:
            raise HTTPException(404, "Backup not found")
        return {"backup_data": backup.backup_data, "message": "Backup data retrieved"}
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid ID")

# Reconciliation Endpoints
@app.get("/api/vector/projects/{project_id}/reconcile")
def get_reconciliation_report(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Get reconciliation report (discrepancies, missing data)"""
    try:
        project_uuid = uuid.UUID(project_id)
        vector_project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")
    
    if not vector_project:
        raise HTTPException(404, "Vector project not found")
    
    if not vector_project.linked_project_id:
        return {
            "status": "standalone",
            "message": "Vector project is not linked to a Radius project",
            "missing_inventory": [],
            "missing_plots": [],
            "smart_matches": []
        }
    
    # Get Vector annotations and extract plot numbers
    annotations = db.query(VectorAnnotation).filter(VectorAnnotation.project_id == project_uuid).all()
    vector_plot_nums = set()
    for anno in annotations:
        if anno.plot_nums:
            vector_plot_nums.update(anno.plot_nums)
    
    # Get Radius inventory for linked project
    inventory_items = db.query(Inventory).filter(Inventory.project_id == vector_project.linked_project_id).all()
    radius_unit_nums = [item.unit_number for item in inventory_items]
    
    # Find missing data
    missing_inventory = list(vector_plot_nums - set(radius_unit_nums))
    missing_plots = [unit for unit in radius_unit_nums if unit not in vector_plot_nums]
    
    # Smart matching
    smart_matches = smart_match_plot_numbers(list(vector_plot_nums), radius_unit_nums)
    
    # Update reconciliation record
    recon = db.query(VectorReconciliation).filter(VectorReconciliation.project_id == project_uuid).first()
    if not recon:
        recon = VectorReconciliation(
            project_id=project_uuid,
            linked_project_id=vector_project.linked_project_id,
            sync_status="needs-reconciliation" if (missing_inventory or missing_plots) else "synced",
            discrepancies={"missing_inventory": missing_inventory, "missing_plots": missing_plots}
        )
        db.add(recon)
    else:
        recon.discrepancies = {"missing_inventory": missing_inventory, "missing_plots": missing_plots}
        recon.sync_status = "needs-reconciliation" if (missing_inventory or missing_plots) else "synced"
        recon.updated_at = func.now()
    
    db.commit()
    
    return {
        "status": recon.sync_status,
        "missing_inventory": missing_inventory,
        "missing_plots": missing_plots,
        "smart_matches": smart_matches,
        "total_vector_plots": len(vector_plot_nums),
        "total_radius_units": len(radius_unit_nums)
    }

@app.post("/api/vector/projects/{project_id}/reconcile/match")
def smart_match_plots(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Smart match plot/unit numbers (fuzzy matching)"""
    try:
        project_uuid = uuid.UUID(project_id)
        vector_project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")
    
    if not vector_project or not vector_project.linked_project_id:
        raise HTTPException(404, "Vector project not found or not linked")
    
    # Get Vector plot numbers
    annotations = db.query(VectorAnnotation).filter(VectorAnnotation.project_id == project_uuid).all()
    vector_plot_nums = set()
    for anno in annotations:
        if anno.plot_nums:
            vector_plot_nums.update(anno.plot_nums)
    
    # Get Radius unit numbers
    inventory_items = db.query(Inventory).filter(Inventory.project_id == vector_project.linked_project_id).all()
    radius_unit_nums = [item.unit_number for item in inventory_items]
    
    # Perform smart matching
    matches = smart_match_plot_numbers(list(vector_plot_nums), radius_unit_nums)
    
    return {
        "matches": matches,
        "total_matches": len([m for m in matches if m["radius_unit"] is not None]),
        "auto_matched": len([m for m in matches if m["auto_match"]])
    }

@app.post("/api/vector/projects/{project_id}/reconcile/sync-from-projects")
def sync_from_projects(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Sync data from Radius Projects → Vector (one-way)"""
    try:
        project_uuid = uuid.UUID(project_id)
        vector_project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")
    
    if not vector_project or not vector_project.linked_project_id:
        raise HTTPException(404, "Vector project not found or not linked")
    
    # Get Radius inventory
    inventory_items = db.query(Inventory).filter(Inventory.project_id == vector_project.linked_project_id).all()
    
    # Get Radius transactions
    transactions = db.query(Transaction).filter(Transaction.project_id == vector_project.linked_project_id).all()
    
    # Update system-generated branches
    inventory_branch = {}
    for item in inventory_items:
        inventory_branch[item.unit_number] = {
            "unit_number": item.unit_number,
            "area_marla": float(item.area_marla) if item.area_marla else 0,
            "rate_per_marla": float(item.rate_per_marla) if item.rate_per_marla else 0,
            "status": item.status,
            "plot_coordinates": item.plot_coordinates,
            "plot_offset": item.plot_offset
        }
    
    sales_branch = {}
    for txn in transactions:
        sales_branch[txn.transaction_id] = {
            "transaction_id": txn.transaction_id,
            "unit_number": txn.unit_number,
            "total_value": float(txn.total_value) if txn.total_value else 0,
            "status": txn.status,
            "booking_date": str(txn.booking_date) if txn.booking_date else None
        }
    
    # Update system branches
    if not vector_project.system_branches:
        vector_project.system_branches = {}
    vector_project.system_branches["inventory"] = inventory_branch
    vector_project.system_branches["sales"] = sales_branch
    vector_project.updated_at = func.now()
    
    # Update reconciliation record
    recon = db.query(VectorReconciliation).filter(VectorReconciliation.project_id == project_uuid).first()
    if not recon:
        recon = VectorReconciliation(
            project_id=project_uuid,
            linked_project_id=vector_project.linked_project_id,
            sync_status="synced",
            last_sync_at=datetime.utcnow()
        )
        db.add(recon)
    else:
        recon.sync_status = "synced"
        recon.last_sync_at = datetime.utcnow()
        recon.updated_at = func.now()
    
    db.commit()
    
    return {
        "message": "Data synced from Radius Projects to Vector",
        "inventory_items_synced": len(inventory_items),
        "transactions_synced": len(transactions),
        "last_sync_at": str(recon.last_sync_at)
    }

@app.post("/api/vector/projects/{project_id}/sync-branches")
def sync_branches(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Generate TWO separate auto-generated Vector projects from linked ORBIT data.

    Creates/Updates:
    1. Status Map: SOLD vs AVAILABLE (2 annotations)
    2. Customer Map: Customer-wise breakdown with unique colors

    The original Vector project remains UNTOUCHED.
    """
    import time
    import colorsys

    try:
        project_uuid = uuid.UUID(project_id)
        vector_project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

    if not vector_project:
        raise HTTPException(404, "Vector project not found")

    if not vector_project.linked_project_id:
        raise HTTPException(400, "Vector project not linked to an ORBIT project")

    # Check if this is an auto-generated project - if so, use the SOURCE project's data
    is_auto_generated = vector_project.vector_metadata.get('isAutoGenerated', False) if vector_project.vector_metadata else False
    source_project = None
    master_project = vector_project  # The project that has the reconciliation mappings

    if is_auto_generated:
        source_project_id = vector_project.vector_metadata.get('sourceProjectId')
        if source_project_id:
            try:
                source_project = db.query(VectorProject).filter(VectorProject.id == uuid.UUID(source_project_id)).first()
                if source_project:
                    master_project = source_project
                    print(f"DEBUG SYNC: Auto-generated project detected, using source project: {source_project_id}")
            except:
                pass

    # Get linked ORBIT project name
    linked_project = db.query(Project).filter(Project.id == vector_project.linked_project_id).first()
    orbit_project_name = linked_project.name if linked_project else "ORBIT"

    # Helper function to generate unique colors for customers
    def generate_unique_colors(n):
        """Generate n visually distinct colors."""
        colors = []
        for i in range(n):
            hue = i / n
            saturation = 0.7 + (i % 3) * 0.1  # Vary saturation slightly
            lightness = 0.5
            r, g, b = colorsys.hls_to_rgb(hue, lightness, saturation)
            colors.append(f"#{int(r*255):02x}{int(g*255):02x}{int(b*255):02x}")
        return colors

    # Get Vector plots for matching - always use MASTER project's plots
    plots = master_project.vector_metadata.get('plots', []) if master_project.vector_metadata else []
    print(f"DEBUG SYNC: Master project has {len(plots)} plots")
    print(f"DEBUG SYNC: Is auto-generated: {is_auto_generated}, Master project ID: {master_project.id}")

    # Build multiple maps for flexible matching
    # Map 1: Exact plot name -> plot data (for "1-IB1" style names)
    # Map 2: Parsed (unit, block) -> plot data
    plot_by_name = {}  # "1-IB1" -> {id, n}
    plot_by_unit_block = {}  # ("1", "IB1") -> {id, n}
    plot_by_unit_only = {}  # "1" -> {id, n} (for simple cases without blocks)

    for p in plots:
        plot_num = p.get('n', '').strip()
        plot_id = p.get('id')
        if not plot_num or not plot_id:
            continue

        plot_data = {'id': plot_id, 'n': plot_num}

        # Store by exact name (case variations)
        plot_by_name[plot_num] = plot_data
        plot_by_name[plot_num.upper()] = plot_data
        plot_by_name[plot_num.lower()] = plot_data

        # Try to parse "unit-block" format (e.g., "149-IB1", "1-A", "25-IB2")
        if '-' in plot_num:
            parts = plot_num.split('-', 1)
            if len(parts) == 2:
                unit_part = parts[0].strip()
                block_part = parts[1].strip()
                # Store by (unit, block) tuple
                plot_by_unit_block[(unit_part, block_part)] = plot_data
                plot_by_unit_block[(unit_part.upper(), block_part.upper())] = plot_data
                plot_by_unit_block[(unit_part.lower(), block_part.lower())] = plot_data
        else:
            # Simple unit number without block
            plot_by_unit_only[plot_num] = plot_data
            plot_by_unit_only[plot_num.upper()] = plot_data
            plot_by_unit_only[plot_num.lower()] = plot_data

    print(f"DEBUG SYNC: Built maps - by_name: {len(plot_by_name)}, by_unit_block: {len(plot_by_unit_block)}, by_unit_only: {len(plot_by_unit_only)}")

    # Get inventory items from linked ORBIT project
    inventory_items = db.query(Inventory).filter(
        Inventory.project_id == vector_project.linked_project_id
    ).all()

    # Get transactions to identify SOLD items
    transactions = db.query(Transaction).filter(
        Transaction.project_id == vector_project.linked_project_id
    ).all()

    print(f"DEBUG SYNC: Found {len(inventory_items)} inventory items, {len(transactions)} transactions")

    # Build set of sold unit identifiers (unit_number-block or just unit_number)
    # Also build inventory_id -> sold mapping for manual mapping lookup
    sold_identifiers = set()
    sold_inventory_ids = set()  # Track which inventory_ids are sold

    for txn in transactions:
        print(f"DEBUG SYNC: Transaction found - unit: {txn.unit_number}, inventory_id (UUID): {txn.inventory_id}")
        if txn.unit_number:
            # IMPORTANT: Only add unit-only identifier if there's NO block
            # Otherwise only add the composite to prevent cross-block matching
            txn_unit = txn.unit_number.strip().upper()
            txn_block = txn.block.strip().upper() if hasattr(txn, 'block') and txn.block else ''
            if txn_block:
                # Has block - ONLY add composite identifier
                sold_identifiers.add(f"{txn_unit}-{txn_block}")
            else:
                # No block - add unit only
                sold_identifiers.add(txn_unit)

        # Get the inventory item to find its string inventory_id
        if txn.inventory_id:
            txn_inv = db.query(Inventory).filter(Inventory.id == txn.inventory_id).first()
            if txn_inv:
                sold_inventory_ids.add(txn_inv.inventory_id)  # The string like "INV-00001"
                print(f"DEBUG SYNC: Transaction inventory_id (string): {txn_inv.inventory_id}")

    # Also check inventory status for sold items
    for item in inventory_items:
        if item.status and item.status.lower() == 'sold' and item.unit_number:
            sold_inventory_ids.add(item.inventory_id)
            # IMPORTANT: Only add unit-only identifier if there's NO block
            # Otherwise only add the composite to prevent cross-block matching
            item_unit = item.unit_number.strip().upper()
            item_block = item.block.strip().upper() if item.block else ''
            if item_block:
                # Has block - ONLY add composite identifier
                sold_identifiers.add(f"{item_unit}-{item_block}")
            else:
                # No block - add unit only
                sold_identifiers.add(item_unit)

    print(f"DEBUG SYNC: Found {len(sold_identifiers)} sold identifiers, {len(sold_inventory_ids)} sold inventory_ids")
    if sold_inventory_ids:
        print(f"DEBUG SYNC: Sold inventory_ids: {list(sold_inventory_ids)[:5]}")

    # Load manual mappings (inventory_id -> plot_id)
    manual_mappings = {}
    inv_id_to_plot = {}  # Reverse lookup: inventory_id -> plot_data

    # Debug: Check what's in MASTER project's vector_metadata (where mappings are stored)
    print(f"DEBUG SYNC: master_project.vector_metadata type: {type(master_project.vector_metadata)}")
    if master_project.vector_metadata:
        print(f"DEBUG SYNC: master_project.vector_metadata keys: {list(master_project.vector_metadata.keys())}")
        raw_mappings = master_project.vector_metadata.get('plot_inventory_mapping')
        print(f"DEBUG SYNC: plot_inventory_mapping type: {type(raw_mappings)}, value exists: {raw_mappings is not None}")
        if raw_mappings:
            print(f"DEBUG SYNC: plot_inventory_mapping count: {len(raw_mappings)}")

    if master_project.vector_metadata and master_project.vector_metadata.get('plot_inventory_mapping'):
        mappings = master_project.vector_metadata['plot_inventory_mapping']
        for plot_id, mapping_data in mappings.items():
            inv_id = mapping_data.get('inventory_id')
            plot_name = mapping_data.get('plot_name', '')
            if inv_id:
                inv_id_to_plot[inv_id] = {'id': plot_id, 'n': plot_name}
                inv_id_to_plot[inv_id.upper()] = {'id': plot_id, 'n': plot_name}
                inv_id_to_plot[inv_id.lower()] = {'id': plot_id, 'n': plot_name}
        print(f"DEBUG SYNC: Loaded {len(inv_id_to_plot)} manual mappings (from {len(mappings)} plot entries)")
        if inv_id_to_plot:
            sample = list(inv_id_to_plot.items())[:3]
            print(f"DEBUG SYNC: Sample mappings: {sample}")
    else:
        print("DEBUG SYNC: No manual mappings found")

    # Smart matching function
    def find_matching_plot(inventory_item):
        """Find matching Vector plot for an ORBIT inventory item.

        IMPORTANT: When inventory has a block, we MUST match unit+block exactly.
        We only fall back to unit-only matching when inventory has NO block specified.
        This prevents matching "1-IB2" transaction to "1-IB1" plot.
        """
        # Strategy 0: Check manual mapping first (by inventory_id)
        inv_id = inventory_item.inventory_id
        if inv_id:
            if inv_id in inv_id_to_plot:
                print(f"DEBUG SYNC: Manual mapping found for {inv_id}")
                return inv_id_to_plot[inv_id]
            if inv_id.upper() in inv_id_to_plot:
                return inv_id_to_plot[inv_id.upper()]

        unit = inventory_item.unit_number.strip() if inventory_item.unit_number else ''
        blk = inventory_item.block.strip() if inventory_item.block else ''

        if not unit:
            return None

        # Strategy 1: Try exact "unit-block" match (e.g., "1-IB1")
        # This is the ONLY strategy when inventory HAS a block
        if blk:
            composite_key = f"{unit}-{blk}"
            if composite_key in plot_by_name:
                print(f"DEBUG SYNC: Exact unit-block match: {composite_key}")
                return plot_by_name[composite_key]
            if composite_key.upper() in plot_by_name:
                print(f"DEBUG SYNC: Exact unit-block match (upper): {composite_key.upper()}")
                return plot_by_name[composite_key.upper()]

            # Try tuple lookup
            if (unit, blk) in plot_by_unit_block:
                print(f"DEBUG SYNC: Tuple unit-block match: ({unit}, {blk})")
                return plot_by_unit_block[(unit, blk)]
            if (unit.upper(), blk.upper()) in plot_by_unit_block:
                print(f"DEBUG SYNC: Tuple unit-block match (upper): ({unit.upper()}, {blk.upper()})")
                return plot_by_unit_block[(unit.upper(), blk.upper())]

            # IMPORTANT: Do NOT fall through to unit-only matching when block is specified
            # If block is specified but no match found, return None - don't match wrong block
            print(f"DEBUG SYNC: No exact match for unit={unit}, block={blk} - NOT falling back to unit-only")
            return None

        # Strategy 2: Try just unit number ONLY when inventory has NO block
        # This is safe because we know the inventory doesn't have a block requirement
        if unit in plot_by_name:
            print(f"DEBUG SYNC: Unit-only match (no block in inventory): {unit}")
            return plot_by_name[unit]
        if unit.upper() in plot_by_name:
            print(f"DEBUG SYNC: Unit-only match (upper, no block): {unit.upper()}")
            return plot_by_name[unit.upper()]
        if unit in plot_by_unit_only:
            print(f"DEBUG SYNC: Unit-only from simple plots: {unit}")
            return plot_by_unit_only[unit]
        if unit.upper() in plot_by_unit_only:
            print(f"DEBUG SYNC: Unit-only from simple plots (upper): {unit.upper()}")
            return plot_by_unit_only[unit.upper()]

        return None

    # Build annotation arrays
    sold_plot_ids = []
    sold_plot_nums = []
    available_plot_ids = []
    available_plot_nums = []
    matched_items = []
    unmatched_items = []

    for item in inventory_items:
        unit_num = item.unit_number
        block = item.block or ''

        if not unit_num:
            continue

        plot_data = find_matching_plot(item)  # Pass full item for inventory_id lookup

        # Check if this item is SOLD - STRICT matching to prevent cross-block contamination
        # When item has a block, we ONLY check the composite identifier
        item_unit = unit_num.strip().upper()
        item_block = block.strip().upper() if block else ''
        item_identifier = f"{item_unit}-{item_block}" if item_block else item_unit

        is_sold = (
            item.inventory_id in sold_inventory_ids or  # By inventory_id (from transactions) - most reliable
            (item.status and item.status.lower() == 'sold') or  # By status field
            item_identifier in sold_identifiers  # By exact identifier (composite if has block, unit-only otherwise)
            # NOTE: We do NOT fallback to unit-only when block is present - prevents cross-block matching
        )

        if plot_data and is_sold:
            print(f"DEBUG SYNC: Item {item.inventory_id} ({unit_num}) is SOLD, matched to plot {plot_data['n']}")

        if plot_data:
            matched_items.append({
                'unit': unit_num,
                'block': block,
                'plot_id': plot_data['id'],
                'plot_name': plot_data['n'],
                'status': 'sold' if is_sold else 'available'
            })
            if is_sold:
                sold_plot_ids.append(plot_data['id'])
                sold_plot_nums.append(plot_data['n'])
            else:
                available_plot_ids.append(plot_data['id'])
                available_plot_nums.append(plot_data['n'])
        else:
            unmatched_items.append({
                'unit': unit_num,
                'block': block,
                'status': 'sold' if is_sold else 'available',
                'reason': 'No matching plot found in Vector map'
            })

    print(f"DEBUG SYNC: Matched {len(matched_items)} items, {len(unmatched_items)} unmatched")
    if unmatched_items[:5]:
        print(f"DEBUG SYNC: Sample unmatched: {unmatched_items[:5]}")

    # Check if this is still a "raw" link (no inventory/transactions in ORBIT yet)
    is_raw_link = len(inventory_items) == 0 and len(transactions) == 0

    if is_raw_link:
        return {
            "message": f"Linked but awaiting ORBIT data. {len(plots)} Vector plots available.",
            "isAwaitingData": True,
            "needsReconciliation": False,
            "statusMapId": None,
            "customerMapId": None
        }

    # ========================================
    # CREATE/UPDATE AUTO-GENERATED PROJECTS
    # ========================================
    current_time = int(time.time() * 1000)

    # Get existing auto-generated project IDs from MASTER project metadata
    # Re-query master project to get fresh data from DB
    db.refresh(master_project)

    auto_gen_metadata = master_project.vector_metadata.get('auto_generated_maps', {}) if master_project.vector_metadata else {}
    status_map_id = auto_gen_metadata.get('status_map_id')
    customer_map_id = auto_gen_metadata.get('customer_map_id')

    print(f"DEBUG SYNC: Looking for existing auto-generated maps in master project {master_project.id}")
    print(f"DEBUG SYNC: auto_gen_metadata = {auto_gen_metadata}")
    print(f"DEBUG SYNC: Existing status_map_id = {status_map_id}, customer_map_id = {customer_map_id}")

    # Build inventory data for legend calculations
    # KEY FIX: Build inventory using PLOT NAMES (plot.n) as keys, not ORBIT unit-block
    # This ensures LegendPanel can look up vectorState.inventory[plot.n]
    inventory_for_legend = {}

    # First, build a lookup from inventory_id to inventory data
    inv_id_to_data = {}
    for item in inventory_items:
        unit = item.unit_number.strip() if item.unit_number else ''
        blk = item.block.strip() if item.block else ''

        area_marla = float(item.area_marla) if item.area_marla else 0
        rate_per_marla = float(item.rate_per_marla) if item.rate_per_marla else 0
        total_value = area_marla * rate_per_marla

        # Get customer name from transaction if sold
        customer_name = None
        if item.status and item.status.lower() == 'sold':
            txn = next((t for t in transactions if t.inventory_id == item.id), None)
            if txn and txn.customer_id:
                customer = db.query(Customer).filter(Customer.id == txn.customer_id).first()
                if customer:
                    customer_name = customer.name

        inv_data = {
            'marla': area_marla,
            'totalValue': total_value,
            'rate': rate_per_marla,
            'status': item.status or 'available',
            'block': blk,
            'unitNumber': unit,
            'inventoryId': item.inventory_id,
            'customer': customer_name
        }

        # Store by inventory_id for reconciliation lookup
        if item.inventory_id:
            inv_id_to_data[item.inventory_id] = inv_data
            inv_id_to_data[item.inventory_id.upper()] = inv_data
            inv_id_to_data[item.inventory_id.lower()] = inv_data

        # Also store by unit-block combo for fallback
        orbit_key = f"{unit}-{blk}" if blk else unit
        inv_id_to_data[orbit_key] = inv_data
        inv_id_to_data[orbit_key.upper()] = inv_data

    print(f"DEBUG SYNC: Built inv_id_to_data with {len(inv_id_to_data)} entries")

    # Now build inventory_for_legend using PLOT NAMES from the reconciliation mapping
    # This maps Vector plot names to ORBIT inventory data
    plot_inv_mapping = master_project.vector_metadata.get('plot_inventory_mapping', {}) if master_project.vector_metadata else {}
    print(f"DEBUG SYNC: Found {len(plot_inv_mapping)} reconciliation mappings")

    # For each plot, try to find its inventory data
    for plot in plots:
        plot_name = plot.get('n', '')
        plot_id = plot.get('id', '')
        if not plot_name:
            continue

        inv_data = None

        # Strategy 1: Use reconciliation mapping (plot_id -> inventory_id)
        if plot_id in plot_inv_mapping:
            mapping = plot_inv_mapping[plot_id]
            inv_id = mapping.get('inventory_id', '')
            if inv_id and inv_id in inv_id_to_data:
                inv_data = inv_id_to_data[inv_id]
                print(f"DEBUG SYNC: Plot {plot_name} matched via reconciliation to {inv_id}")

        # Strategy 2: Try direct match by plot name
        if not inv_data:
            if plot_name in inv_id_to_data:
                inv_data = inv_id_to_data[plot_name]
            elif plot_name.upper() in inv_id_to_data:
                inv_data = inv_id_to_data[plot_name.upper()]

        # Strategy 3: Parse plot name and try unit-block variations
        if not inv_data and '-' in plot_name:
            parts = plot_name.split('-', 1)
            if len(parts) == 2:
                # Try "unit-block" format
                for key in [f"{parts[0]}-{parts[1]}", f"{parts[1]}-{parts[0]}"]:
                    if key in inv_id_to_data:
                        inv_data = inv_id_to_data[key]
                        break
                    if key.upper() in inv_id_to_data:
                        inv_data = inv_id_to_data[key.upper()]
                        break

        if inv_data:
            inventory_for_legend[plot_name] = inv_data

    print(f"DEBUG SYNC: Built inventory_for_legend with {len(inventory_for_legend)} entries (by plot name)")
    if inventory_for_legend:
        sample_keys = list(inventory_for_legend.keys())[:5]
        print(f"DEBUG SYNC: Sample inventory keys (plot names): {sample_keys}")

    # Base metadata for auto-generated projects (copy from MASTER project)
    base_plots = plots.copy()  # Same plots as master
    base_metadata = {
        'plots': base_plots,
        'plotOffsets': master_project.vector_metadata.get('plotOffsets', {}) if master_project.vector_metadata else {},
        'plotRotations': master_project.vector_metadata.get('plotRotations', {}) if master_project.vector_metadata else {},
        'inventory': inventory_for_legend,  # Include inventory for legend calculations
        'isAutoGenerated': True,
        'sourceProjectId': str(master_project.id),
        'sourceProjectName': master_project.name,
        'lastSyncAt': datetime.utcnow().isoformat()
    }

    # ========================================
    # 1. STATUS MAP (SOLD vs AVAILABLE)
    # ========================================
    status_annotations = []

    # SOLD annotation (red)
    if sold_plot_ids:
        status_annotations.append({
            "id": f"status_sold_{current_time}",
            "note": "SOLD",
            "cat": "SOLD",
            "color": "#ef4444",  # Red
            "plotIds": sold_plot_ids,
            "plotNums": sold_plot_nums,
            "fontSize": 12,
            "rotation": 0,
            "isAutoGenerated": True,
            "source": "ORBIT"
        })

    # AVAILABLE annotation (green)
    if available_plot_ids:
        status_annotations.append({
            "id": f"status_available_{current_time}",
            "note": "AVAILABLE",
            "cat": "AVAILABLE",
            "color": "#22c55e",  # Green
            "plotIds": available_plot_ids,
            "plotNums": available_plot_nums,
            "fontSize": 12,
            "rotation": 0,
            "isAutoGenerated": True,
            "source": "ORBIT"
        })

    status_metadata = {
        **base_metadata,
        'annos': status_annotations,
        'mapType': 'status'
    }

    # Create or update Status Map project
    if status_map_id:
        # Update existing
        status_project = db.query(VectorProject).filter(VectorProject.id == uuid.UUID(status_map_id)).first()
        if status_project:
            status_project.vector_metadata = status_metadata
            status_project.updated_at = func.now()
            print(f"DEBUG SYNC: Updated Status Map project {status_map_id}")
        else:
            status_map_id = None  # Project was deleted, create new

    if not status_map_id:
        # Create new Status Map project (using MASTER project's data)
        status_project = VectorProject(
            name=f"{master_project.name} - Status [Auto]",
            map_name=master_project.map_name,
            map_pdf_base64=master_project.map_pdf_base64,
            map_size=master_project.map_size,
            linked_project_id=master_project.linked_project_id,
            vector_metadata=status_metadata
        )
        db.add(status_project)
        db.flush()  # Get the ID
        status_map_id = str(status_project.id)
        print(f"DEBUG SYNC: Created Status Map project {status_map_id}")

    # ========================================
    # 2. CUSTOMER MAP (Customer-wise breakdown)
    # ========================================

    # Build customer -> plots mapping from transactions
    customer_plots = {}  # customer_name -> {plot_ids: [], plot_nums: []}

    # Get transaction details with customer names
    for txn in transactions:
        customer = db.query(Customer).filter(Customer.id == txn.customer_id).first()
        customer_name = customer.name if customer else "Unknown Customer"

        # Find the plot for this transaction
        inv_item = db.query(Inventory).filter(Inventory.id == txn.inventory_id).first()
        if inv_item:
            plot_data = find_matching_plot(inv_item)
            if plot_data:
                if customer_name not in customer_plots:
                    customer_plots[customer_name] = {'plot_ids': [], 'plot_nums': []}
                customer_plots[customer_name]['plot_ids'].append(plot_data['id'])
                customer_plots[customer_name]['plot_nums'].append(plot_data['n'])

    # Generate unique colors for each customer
    customer_names = list(customer_plots.keys())
    customer_colors = generate_unique_colors(len(customer_names) + 1)  # +1 for AVAILABLE

    customer_annotations = []

    # Add annotation for each customer
    for idx, customer_name in enumerate(customer_names):
        data = customer_plots[customer_name]
        if data['plot_ids']:
            customer_annotations.append({
                "id": f"customer_{idx}_{current_time}",
                "note": customer_name,
                "cat": customer_name,
                "color": customer_colors[idx],
                "plotIds": data['plot_ids'],
                "plotNums": data['plot_nums'],
                "fontSize": 10,
                "rotation": 0,
                "isAutoGenerated": True,
                "source": "ORBIT",
                "customerName": customer_name
            })

    # Add AVAILABLE annotation (for unsold plots)
    if available_plot_ids:
        customer_annotations.append({
            "id": f"customer_available_{current_time}",
            "note": "AVAILABLE",
            "cat": "AVAILABLE",
            "color": "#22c55e",  # Green for available
            "plotIds": available_plot_ids,
            "plotNums": available_plot_nums,
            "fontSize": 10,
            "rotation": 0,
            "isAutoGenerated": True,
            "source": "ORBIT"
        })

    customer_metadata = {
        **base_metadata,
        'annos': customer_annotations,
        'mapType': 'customer',
        'customerCount': len(customer_names)
    }

    # Create or update Customer Map project
    if customer_map_id:
        # Update existing
        customer_project = db.query(VectorProject).filter(VectorProject.id == uuid.UUID(customer_map_id)).first()
        if customer_project:
            customer_project.vector_metadata = customer_metadata
            customer_project.updated_at = func.now()
            print(f"DEBUG SYNC: Updated Customer Map project {customer_map_id}")
        else:
            customer_map_id = None  # Project was deleted, create new

    if not customer_map_id:
        # Create new Customer Map project (using MASTER project's data)
        customer_project = VectorProject(
            name=f"{master_project.name} - Customers [Auto]",
            map_name=master_project.map_name,
            map_pdf_base64=master_project.map_pdf_base64,
            map_size=master_project.map_size,
            linked_project_id=master_project.linked_project_id,
            vector_metadata=customer_metadata
        )
        db.add(customer_project)
        db.flush()
        customer_map_id = str(customer_project.id)
        print(f"DEBUG SYNC: Created Customer Map project {customer_map_id}")

    # ========================================
    # 3. AUTOMATED MASTER (Copy of master with ORBIT updates)
    # ========================================
    # This preserves master's annotations but enriches them with ORBIT status data
    # Any changes to master annotations are synced to this copy

    auto_master_map_id = auto_gen_metadata.get('auto_master_map_id')

    # Get master's original annotations (preserve exact structure)
    master_annos = master_project.vector_metadata.get('annos', []) if master_project.vector_metadata else []

    # Create enriched annotations - copy master annotations and add ORBIT status info
    enriched_annotations = []

    # Build plot_id -> sold status lookup
    sold_plot_id_set = set(sold_plot_ids)

    for anno in master_annos:
        # Deep copy the annotation
        enriched_anno = dict(anno)

        # Enrich plotIds with sold status counts
        anno_sold_count = 0
        anno_available_count = 0
        for pid in anno.get('plotIds', []):
            if pid in sold_plot_id_set:
                anno_sold_count += 1
            else:
                anno_available_count += 1

        # Add ORBIT status info to annotation note (optional enhancement)
        enriched_anno['orbitStats'] = {
            'soldCount': anno_sold_count,
            'availableCount': anno_available_count,
            'totalCount': len(anno.get('plotIds', []))
        }

        enriched_annotations.append(enriched_anno)

    # Also include master's labels, shapes, etc.
    master_labels = master_project.vector_metadata.get('labels', []) if master_project.vector_metadata else []
    master_shapes = master_project.vector_metadata.get('shapes', []) if master_project.vector_metadata else []

    auto_master_metadata = {
        **base_metadata,
        'annos': enriched_annotations,
        'labels': master_labels,
        'shapes': master_shapes,
        'mapType': 'auto_master',
        'preservesMasterAnnotations': True,
        'masterSyncedAt': datetime.utcnow().isoformat()
    }

    # Create or update Automated Master project
    if auto_master_map_id:
        auto_master_project = db.query(VectorProject).filter(VectorProject.id == uuid.UUID(auto_master_map_id)).first()
        if auto_master_project:
            auto_master_project.vector_metadata = auto_master_metadata
            auto_master_project.updated_at = func.now()
            print(f"DEBUG SYNC: Updated Automated Master project {auto_master_map_id}")
        else:
            auto_master_map_id = None

    if not auto_master_map_id:
        auto_master_project = VectorProject(
            name=f"{master_project.name} - Master [Auto]",
            map_name=master_project.map_name,
            map_pdf_base64=master_project.map_pdf_base64,
            map_size=master_project.map_size,
            linked_project_id=master_project.linked_project_id,
            vector_metadata=auto_master_metadata
        )
        db.add(auto_master_project)
        db.flush()
        auto_master_map_id = str(auto_master_project.id)
        print(f"DEBUG SYNC: Created Automated Master project {auto_master_map_id}")

    # ========================================
    # UPDATE MASTER PROJECT METADATA
    # ========================================
    # Store references to auto-generated maps in MASTER project (not auto-generated ones)
    from sqlalchemy.orm.attributes import flag_modified

    if not master_project.vector_metadata:
        master_project.vector_metadata = {}

    new_metadata = dict(master_project.vector_metadata)
    new_metadata['auto_generated_maps'] = {
        'status_map_id': status_map_id,
        'customer_map_id': customer_map_id,
        'auto_master_map_id': auto_master_map_id,
        'last_sync_at': datetime.utcnow().isoformat()
    }
    master_project.vector_metadata = new_metadata
    flag_modified(master_project, 'vector_metadata')
    master_project.updated_at = func.now()

    print(f"DEBUG SYNC: Saving auto_generated_maps to master project: status={status_map_id}, customer={customer_map_id}")

    # Update reconciliation record (always on MASTER project)
    master_project_uuid = master_project.id
    recon = db.query(VectorReconciliation).filter(VectorReconciliation.project_id == master_project_uuid).first()
    sync_status = "needs_reconciliation" if unmatched_items else "synced"
    if recon:
        recon.last_sync_at = datetime.utcnow()
        recon.sync_status = sync_status
        recon.discrepancies = {"unmatched": unmatched_items} if unmatched_items else None
        recon.updated_at = func.now()
    else:
        recon = VectorReconciliation(
            project_id=master_project_uuid,
            linked_project_id=master_project.linked_project_id,
            sync_status=sync_status,
            discrepancies={"unmatched": unmatched_items} if unmatched_items else None,
            last_sync_at=datetime.utcnow()
        )
        db.add(recon)

    db.commit()

    # Verify the auto_generated_maps was saved correctly
    db.refresh(master_project)
    saved_auto_gen = master_project.vector_metadata.get('auto_generated_maps', {}) if master_project.vector_metadata else {}
    print(f"DEBUG SYNC: After commit, auto_generated_maps in DB: {saved_auto_gen}")

    # Build response message
    if unmatched_items:
        message = f"Created/Updated 3 auto-maps: {len(sold_plot_ids)} SOLD + {len(available_plot_ids)} AVAILABLE + Master copy. {len(unmatched_items)} items need reconciliation."
    else:
        message = f"Created/Updated 3 auto-maps: {len(sold_plot_ids)} SOLD ({len(customer_names)} customers) + {len(available_plot_ids)} AVAILABLE + Master copy with {len(master_annos)} annotations."

    return {
        "message": message,
        "statusMapId": status_map_id,
        "customerMapId": customer_map_id,
        "autoMasterMapId": auto_master_map_id,
        "statusMapName": f"{master_project.name} - Status [Auto]",
        "customerMapName": f"{master_project.name} - Customers [Auto]",
        "autoMasterMapName": f"{master_project.name} - Master [Auto]",
        "soldCount": len(sold_plot_ids),
        "availableCount": len(available_plot_ids),
        "customerCount": len(customer_names),
        "masterAnnosCount": len(master_annos),
        "unmatchedCount": len(unmatched_items),
        "isAwaitingData": False,
        "needsReconciliation": len(unmatched_items) > 0
    }


@app.delete("/api/vector/projects/{project_id}/cleanup-auto-maps")
def cleanup_auto_generated_maps(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Delete all auto-generated maps for a master project, keeping only the latest one of each type."""
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")
    if current_user.role != "admin":
        raise HTTPException(403, "Only admin can cleanup auto-generated maps")
    try:
        project_uuid = uuid.UUID(project_id)
        master_project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

    if not master_project:
        raise HTTPException(404, "Vector project not found")

    # Find all auto-generated projects that reference this master
    # Use multiple query strategies to catch all duplicates
    auto_projects = []
    seen_ids = set()

    # Strategy 1: Query by sourceProjectId (text extraction)
    try:
        projects_by_source = db.query(VectorProject).filter(
            VectorProject.vector_metadata.op('->>')('sourceProjectId') == str(project_uuid)
        ).all()
        for p in projects_by_source:
            if p.id not in seen_ids:
                auto_projects.append(p)
                seen_ids.add(p.id)
    except Exception as e:
        print(f"DEBUG CLEANUP: Strategy 1 error: {e}")

    # Strategy 2: Query all projects and filter by isAutoGenerated + matching linked_project
    # This catches projects that might have been created with different metadata structure
    try:
        all_vector_projects = db.query(VectorProject).filter(
            VectorProject.linked_project_id == master_project.linked_project_id,
            VectorProject.id != master_project.id
        ).all()
        for p in all_vector_projects:
            if p.id not in seen_ids and p.vector_metadata:
                if p.vector_metadata.get('isAutoGenerated') == True:
                    # Verify it's related to this master (by name pattern or sourceProjectId)
                    source_id = p.vector_metadata.get('sourceProjectId', '')
                    if source_id == str(project_uuid) or source_id == str(master_project.id):
                        auto_projects.append(p)
                        seen_ids.add(p.id)
                    # Also check by name pattern (e.g., "ProjectName - Status [Auto]")
                    elif master_project.name and p.name and master_project.name in p.name and '[Auto]' in p.name:
                        auto_projects.append(p)
                        seen_ids.add(p.id)
    except Exception as e:
        print(f"DEBUG CLEANUP: Strategy 2 error: {e}")

    print(f"DEBUG CLEANUP: Found {len(auto_projects)} auto-generated projects for cleanup")

    # Group by type
    status_maps = []
    customer_maps = []
    auto_master_maps = []

    for proj in auto_projects:
        if proj.vector_metadata:
            map_type = proj.vector_metadata.get('mapType')
            print(f"DEBUG CLEANUP: Project {proj.id} ({proj.name}) - mapType: {map_type}")
            if map_type == 'status':
                status_maps.append(proj)
            elif map_type == 'customer':
                customer_maps.append(proj)
            elif map_type == 'auto_master':
                auto_master_maps.append(proj)

    # Keep only the most recent of each type (by updated_at)
    deleted_count = 0
    kept_status_id = None
    kept_customer_id = None
    kept_auto_master_id = None

    if status_maps:
        status_maps.sort(key=lambda p: p.updated_at or p.created_at, reverse=True)
        kept_status_id = str(status_maps[0].id)
        for proj in status_maps[1:]:
            db.delete(proj)
            deleted_count += 1

    if customer_maps:
        customer_maps.sort(key=lambda p: p.updated_at or p.created_at, reverse=True)
        kept_customer_id = str(customer_maps[0].id)
        for proj in customer_maps[1:]:
            db.delete(proj)
            deleted_count += 1

    if auto_master_maps:
        auto_master_maps.sort(key=lambda p: p.updated_at or p.created_at, reverse=True)
        kept_auto_master_id = str(auto_master_maps[0].id)
        for proj in auto_master_maps[1:]:
            db.delete(proj)
            deleted_count += 1

    # Update master project's auto_generated_maps reference
    if kept_status_id or kept_customer_id or kept_auto_master_id:
        from sqlalchemy.orm.attributes import flag_modified

        if not master_project.vector_metadata:
            master_project.vector_metadata = {}

        new_metadata = dict(master_project.vector_metadata)
        new_metadata['auto_generated_maps'] = {
            'status_map_id': kept_status_id,
            'customer_map_id': kept_customer_id,
            'auto_master_map_id': kept_auto_master_id,
            'last_sync_at': datetime.utcnow().isoformat()
        }
        master_project.vector_metadata = new_metadata
        flag_modified(master_project, 'vector_metadata')

    db.commit()

    return {
        "message": f"Cleaned up {deleted_count} duplicate auto-generated maps",
        "deleted_count": deleted_count,
        "kept_status_map_id": kept_status_id,
        "kept_customer_map_id": kept_customer_id,
        "kept_auto_master_map_id": kept_auto_master_id
    }


# ============================================
# RECONCILIATION ENDPOINTS
# ============================================

@app.get("/api/vector/projects/{project_id}/reconciliation/template")
def download_reconciliation_template(
    project_id: str,
    format: str = Query("xlsx", description="Output format: xlsx or csv"),
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Download template for mapping Vector plots to ORBIT inventory.

    The template includes:
    - Sheet 1: Vector plots with empty inventory_id column to fill
    - Sheet 2: ORBIT inventory reference for lookup

    Supports: xlsx (default) or csv
    """
    import io

    try:
        project_uuid = uuid.UUID(project_id)
        vector_project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

    if not vector_project:
        raise HTTPException(404, "Vector project not found")

    # Get Vector plots
    plots = vector_project.vector_metadata.get('plots', []) if vector_project.vector_metadata else []

    # Get ORBIT inventory if linked
    orbit_inventory = []
    if vector_project.linked_project_id:
        inventory_items = db.query(Inventory).filter(
            Inventory.project_id == vector_project.linked_project_id
        ).all()
        orbit_inventory = [{
            'inventory_id': item.inventory_id,
            'unit_number': item.unit_number,
            'block': item.block or '',
            'area_marla': float(item.area_marla) if item.area_marla else 0,
            'status': item.status or ''
        } for item in inventory_items]

    from fastapi.responses import StreamingResponse

    if format.lower() == 'xlsx':
        # Generate Excel file with two sheets
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()

            # Sheet 1: Mapping (to be filled)
            ws1 = wb.active
            ws1.title = "Mapping"

            # Headers with styling
            headers = ['plot_name', 'plot_id', 'inventory_id']
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Add Vector plots
            for row, p in enumerate(plots, 2):
                ws1.cell(row=row, column=1, value=p.get('n', ''))
                ws1.cell(row=row, column=2, value=p.get('id', ''))
                ws1.cell(row=row, column=3, value='')  # Empty for user to fill

            # Adjust column widths
            ws1.column_dimensions['A'].width = 20
            ws1.column_dimensions['B'].width = 45
            ws1.column_dimensions['C'].width = 20

            # Sheet 2: ORBIT Inventory Reference
            ws2 = wb.create_sheet(title="ORBIT_Inventory_Reference")

            ref_headers = ['inventory_id', 'unit_number', 'block', 'area_marla', 'status']
            ref_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

            for col, header in enumerate(ref_headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.fill = ref_fill
                cell.font = header_font

            for row, inv in enumerate(orbit_inventory, 2):
                ws2.cell(row=row, column=1, value=inv['inventory_id'])
                ws2.cell(row=row, column=2, value=inv['unit_number'])
                ws2.cell(row=row, column=3, value=inv['block'])
                ws2.cell(row=row, column=4, value=inv['area_marla'])
                ws2.cell(row=row, column=5, value=inv['status'])

            # Adjust column widths
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws2.column_dimensions[col].width = 15

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"reconciliation_template_{vector_project.name.replace(' ', '_')}.xlsx"

            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )

        except ImportError:
            # Fallback to CSV if openpyxl not available
            format = 'csv'

    # CSV format fallback
    import csv
    output = io.StringIO()

    output.write("# RECONCILIATION TEMPLATE\n")
    output.write("# Fill in inventory_id for each plot, then upload this file\n")
    output.write("#\n")

    writer = csv.writer(output)
    writer.writerow(['plot_name', 'plot_id', 'inventory_id'])

    for p in plots:
        writer.writerow([p.get('n', ''), p.get('id', ''), ''])

    output.write("\n# === ORBIT INVENTORY REFERENCE ===\n")
    output.write("# inventory_id,unit_number,block,area_marla,status\n")
    for inv in orbit_inventory:
        output.write(f"# {inv['inventory_id']},{inv['unit_number']},{inv['block']},{inv['area_marla']},{inv['status']}\n")

    output.seek(0)
    filename = f"reconciliation_template_{vector_project.name.replace(' ', '_')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.post("/api/vector/projects/{project_id}/reconciliation/upload")
def upload_reconciliation_mapping(
    project_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Upload completed reconciliation mapping (CSV or Excel).

    Expected columns:
    plot_name, plot_id, inventory_id

    Supports: .csv, .xlsx, .xls
    """
    import csv
    import io

    try:
        project_uuid = uuid.UUID(project_id)
        vector_project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

    if not vector_project:
        raise HTTPException(404, "Vector project not found")

    # Determine file type and parse accordingly
    filename = file.filename.lower() if file.filename else ''
    file_content = file.file.read()

    rows = []

    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        # Parse Excel file
        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
            ws = wb.active

            # Get headers from first row
            headers = []
            header_map = {}  # Map column index to header name

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    # First row is headers - normalize them
                    for col_idx, cell in enumerate(row):
                        if cell:
                            header = str(cell).strip().lower().replace(' ', '_')
                            headers.append(header)
                            header_map[col_idx] = header
                    continue

                # Skip empty rows
                if not any(row):
                    continue

                # Build row dict
                row_dict = {}
                for col_idx, cell in enumerate(row):
                    if col_idx in header_map:
                        row_dict[header_map[col_idx]] = str(cell).strip() if cell else ''

                if row_dict:
                    rows.append(row_dict)

            print(f"DEBUG UPLOAD: Parsed Excel with {len(rows)} rows, headers: {headers}")

        except ImportError:
            raise HTTPException(400, "Excel support requires openpyxl. Please install it or use CSV format.")
        except Exception as e:
            print(f"DEBUG UPLOAD: Excel parse error: {e}")
            raise HTTPException(400, f"Failed to parse Excel file: {str(e)}")
    else:
        # Parse CSV file
        try:
            content = file_content.decode('utf-8')
            reader = csv.DictReader(
                [line for line in content.split('\n') if line.strip() and not line.strip().startswith('#')]
            )
            rows = list(reader)
            print(f"DEBUG UPLOAD: Parsed CSV with {len(rows)} rows")
        except Exception as e:
            print(f"DEBUG UPLOAD: CSV parse error: {e}")
            raise HTTPException(400, f"Failed to parse CSV file: {str(e)}")

    # Process rows into mappings
    mappings = {}
    valid_count = 0
    skipped_count = 0

    for row in rows:
        # Handle various column name formats
        plot_name = (row.get('plot_name') or row.get('plot name') or row.get('plotname') or '').strip()
        plot_id = (row.get('plot_id') or row.get('plot id') or row.get('plotid') or '').strip()
        inventory_id = (row.get('inventory_id') or row.get('inventory id') or row.get('inventoryid') or
                       row.get('inv_id') or row.get('invid') or '').strip()

        if plot_id and inventory_id:
            mappings[plot_id] = {
                'plot_name': plot_name,
                'inventory_id': inventory_id
            }
            valid_count += 1
        elif plot_id:
            skipped_count += 1

    print(f"DEBUG UPLOAD: Created {valid_count} mappings, skipped {skipped_count}")
    if mappings:
        sample_mapping = list(mappings.items())[:3]
        print(f"DEBUG UPLOAD: Sample mappings: {sample_mapping}")

    # Store mapping in vector_metadata - create NEW dict to force SQLAlchemy to detect change
    # (In-place mutations of JSONB may not be detected even with flag_modified)
    from sqlalchemy.orm.attributes import flag_modified

    existing_metadata = dict(vector_project.vector_metadata) if vector_project.vector_metadata else {}
    existing_metadata['plot_inventory_mapping'] = mappings

    # Assign completely new dict to trigger change detection
    vector_project.vector_metadata = existing_metadata
    flag_modified(vector_project, 'vector_metadata')

    vector_project.updated_at = func.now()

    db.commit()

    # Query fresh from DB to verify the save actually worked
    db.expire(vector_project)  # Clear cached state
    db.refresh(vector_project)  # Refresh from DB

    # Verify the save
    verify_mappings = vector_project.vector_metadata.get('plot_inventory_mapping', {}) if vector_project.vector_metadata else {}
    print(f"DEBUG UPLOAD: Verified {len(verify_mappings)} mappings saved to DB")
    if len(verify_mappings) != valid_count:
        print(f"DEBUG UPLOAD: WARNING - Count mismatch! Expected {valid_count}, got {len(verify_mappings)}")

    return {
        "message": f"Mapping uploaded successfully. {valid_count} plots mapped, {skipped_count} skipped (no inventory_id).",
        "mappedCount": valid_count,
        "skippedCount": skipped_count,
        "mappings": mappings
    }


@app.get("/api/vector/projects/{project_id}/reconciliation/mapping")
def get_reconciliation_mapping(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Get current plot-to-inventory mapping."""
    try:
        project_uuid = uuid.UUID(project_id)
        vector_project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

    if not vector_project:
        raise HTTPException(404, "Vector project not found")

    mappings = {}
    if vector_project.vector_metadata:
        mappings = vector_project.vector_metadata.get('plot_inventory_mapping', {})
        print(f"DEBUG GET_MAPPING: Found {len(mappings)} mappings in vector_metadata")
    else:
        print(f"DEBUG GET_MAPPING: No vector_metadata found")

    return {
        "mappings": mappings,
        "count": len(mappings)
    }


@app.delete("/api/vector/projects/{project_id}/reconciliation/mapping")
def clear_reconciliation_mapping(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Clear all plot-to-inventory mappings."""
    if current_user.role == "creator":
        raise HTTPException(403, "Creator role cannot delete records")
    if current_user.role != "admin":
        raise HTTPException(403, "Only admin can clear reconciliation mappings")
    try:
        project_uuid = uuid.UUID(project_id)
        vector_project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")

    if not vector_project:
        raise HTTPException(404, "Vector project not found")

    if vector_project.vector_metadata and 'plot_inventory_mapping' in vector_project.vector_metadata:
        del vector_project.vector_metadata['plot_inventory_mapping']
        # Force SQLAlchemy to detect the change
        from sqlalchemy.orm.attributes import flag_modified
        flag_modified(vector_project, 'vector_metadata')
        vector_project.updated_at = func.now()
        db.commit()

    return {"message": "Mapping cleared successfully"}


@app.post("/api/vector/projects/{project_id}/reconcile/add-to-projects")
def add_to_projects(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Add Vector map data to Radius Projects"""
    try:
        project_uuid = uuid.UUID(project_id)
        vector_project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")
    
    if not vector_project:
        raise HTTPException(404, "Vector project not found")
    
    # Get Vector annotations to extract plot data
    annotations = db.query(VectorAnnotation).filter(VectorAnnotation.project_id == project_uuid).all()
    
    # This endpoint prepares data for adding to Radius Projects
    # Actual creation would be done via existing /api/projects and /api/inventory endpoints
    plot_data = []
    for anno in annotations:
        if anno.plot_nums:
            for plot_num in anno.plot_nums:
                plot_data.append({
                    "plot_number": plot_num,
                    "annotation": anno.note,
                    "category": anno.category,
                    "color": anno.color
                })
    
    return {
        "message": "Vector map data prepared for adding to Radius Projects",
        "plots_count": len(plot_data),
        "plots": plot_data,
        "note": "Use /api/projects and /api/inventory endpoints to create actual records"
    }

@app.get("/api/vector/projects/{project_id}/incomplete-check")
def check_incomplete_project(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Check if Vector project is incomplete (no source map)"""
    try:
        project_uuid = uuid.UUID(project_id)
        project = db.query(VectorProject).filter(VectorProject.id == project_uuid).first()
    except (ValueError, TypeError):
        raise HTTPException(400, "Invalid project ID")
    
    if not project:
        raise HTTPException(404, "Vector project not found")
    
    is_incomplete = not bool(project.map_pdf_base64)
    
    return {
        "is_incomplete": is_incomplete,
        "has_map": bool(project.map_pdf_base64),
        "map_name": project.map_name,
        "message": "Vector project without source map needs attention" if is_incomplete else "Vector project has source map"
    }

@app.post("/api/vector/projects/import")
def import_vector_json(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Import Vector JSON file - 100% compatible with original Vector format"""
    import json
    
    try:
        # Read and parse JSON file
        content = file.file.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8')
        
        vector_data = json.loads(content)
        
        # Extract project data
        project_name = vector_data.get('projectName', file.filename.replace('.json', ''))
        map_name = vector_data.get('mapName', '')
        pdf_base64 = vector_data.get('pdfBase64', '')
        map_size = vector_data.get('mapSize', {})
        project_metadata = vector_data.get('projectMetadata', {})
        
        # Create or update Vector project
        vector_project = VectorProject(
            name=project_name,
            map_name=map_name,
            map_pdf_base64=pdf_base64 if pdf_base64 else None,
            map_size=map_size if map_size else None,
            vector_metadata=project_metadata if project_metadata else None,
            system_branches=vector_data.get('branches', {})
        )
        db.add(vector_project)
        db.flush()  # Get the ID
        
        # Import plots (convert to annotations/shapes as needed)
        plots = vector_data.get('plots', [])
        plot_offsets = vector_data.get('plotOffsets', {})
        plot_rotations = vector_data.get('plotRotations', {})
        
        # Import annotations
        annos = vector_data.get('annos', [])
        for anno in annos:
            vector_anno = VectorAnnotation(
                project_id=vector_project.id,
                annotation_id=str(anno.get('id', uuid.uuid4())),
                note=anno.get('note', ''),
                category=anno.get('cat', ''),
                color=anno.get('color', '#000000'),
                font_size=anno.get('fontSize', 12),
                rotation=anno.get('rotation', 0),
                plot_ids=anno.get('plotIds', []),
                plot_nums=anno.get('plotNums', [])
            )
            db.add(vector_anno)
        
        # Import labels
        labels = vector_data.get('labels', [])
        for label in labels:
            vector_label = VectorLabel(
                project_id=vector_project.id,
                label_id=str(label.get('id', uuid.uuid4())),
                text=label.get('text', ''),
                x=label.get('x', 0),
                y=label.get('y', 0),
                size=label.get('size', 12),
                color=label.get('color', '#000000')
            )
            db.add(vector_label)
        
        # Import shapes
        shapes = vector_data.get('shapes', [])
        for shape in shapes:
            vector_shape = VectorShape(
                project_id=vector_project.id,
                shape_id=str(shape.get('id', uuid.uuid4())),
                type=shape.get('type', ''),
                x=shape.get('x', 0),
                y=shape.get('y', 0),
                width=shape.get('width', 0),
                height=shape.get('height', 0),
                color=shape.get('color', '#000000'),
                data=shape.get('data', {})
            )
            db.add(vector_shape)
        
        # Import legend
        legend = vector_data.get('legend', {})
        if legend:
            vector_legend = VectorLegend(
                project_id=vector_project.id,
                visible=str(legend.get('visible', 'true')).lower(),
                minimized=str(legend.get('minimized', 'false')).lower(),
                position=legend.get('position', ''),
                manual_entries=legend.get('manualEntries', {})
            )
            db.add(vector_legend)
        
        # Import creator notes
        creator_notes = vector_data.get('creatorNotes', [])
        if isinstance(creator_notes, list):
            for note in creator_notes:
                if isinstance(note, dict):
                    vector_note = VectorCreatorNote(
                        project_id=vector_project.id,
                        note=note.get('note', ''),
                        timestamp=datetime.fromisoformat(note.get('timestamp', datetime.now().isoformat())) if isinstance(note.get('timestamp'), str) else datetime.now()
                    )
                    db.add(vector_note)
        
        # Import change log
        change_log = vector_data.get('changeLog', [])
        if isinstance(change_log, list):
            for log_entry in change_log:
                if isinstance(log_entry, dict):
                    vector_log = VectorChangeLog(
                        project_id=vector_project.id,
                        action=log_entry.get('action', ''),
                        details=log_entry.get('details', {}),
                        timestamp=datetime.fromisoformat(log_entry.get('timestamp', datetime.now().isoformat())) if isinstance(log_entry.get('timestamp'), str) else datetime.now()
                    )
                    db.add(vector_log)
        
        # Store plots data in vector_metadata for reference
        if not vector_project.vector_metadata:
            vector_project.vector_metadata = {}
        vector_project.vector_metadata['plots'] = plots
        vector_project.vector_metadata['plotOffsets'] = plot_offsets
        vector_project.vector_metadata['plotRotations'] = plot_rotations
        vector_project.vector_metadata['imported_from'] = file.filename
        vector_project.vector_metadata['import_timestamp'] = datetime.now().isoformat()
        
        db.commit()
        db.refresh(vector_project)
        
        return {
            "message": "Vector project imported successfully",
            "project_id": str(vector_project.id),
            "name": vector_project.name,
            "plots_imported": len(plots),
            "annotations_imported": len(annos),
            "labels_imported": len(labels),
            "shapes_imported": len(shapes)
        }
        
    except json.JSONDecodeError as e:
        raise HTTPException(400, f"Invalid JSON file: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Error importing Vector project: {str(e)}")

@app.get("/api/vector/orphan-tracking")
def get_orphan_tracking(
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Get orphan projects - Vector projects without ORBIT links and vice versa"""
    # Available to all authenticated users (read-only)

    # Get all Vector projects
    vector_projects = db.query(VectorProject).all()

    # Get all ORBIT projects
    orbit_projects = db.query(Project).all()

    # Build a map of ORBIT project ID to linked Vector project for quick lookup
    orbit_to_vector = {}
    for vp in vector_projects:
        if vp.linked_project_id:
            orbit_to_vector[str(vp.linked_project_id)] = {
                "id": str(vp.id),
                "name": vp.name
            }

    # Find Vector projects without ORBIT links (orphan vectors)
    orphan_vectors = []
    linked_vectors = []
    for vp in vector_projects:
        vp_data = {
            "id": str(vp.id),
            "name": vp.name,
            "map_name": vp.map_name,
            "created_at": str(vp.created_at),
            "has_map": bool(vp.map_pdf_base64),
            "type": "vector"
        }
        if not vp.linked_project_id:
            orphan_vectors.append(vp_data)
        else:
            # Find linked ORBIT project name
            linked_orbit = db.query(Project).filter(Project.id == vp.linked_project_id).first()
            vp_data["linked_orbit_id"] = str(vp.linked_project_id)
            vp_data["linked_orbit_name"] = linked_orbit.name if linked_orbit else "Unknown"
            linked_vectors.append(vp_data)

    # Find ORBIT projects without Vector links (orphan orbit)
    orphan_orbit = []
    linked_orbit = []
    for op in orbit_projects:
        inventory_count = db.query(Inventory).filter(Inventory.project_id == op.id).count()
        transaction_count = db.query(Transaction).filter(Transaction.project_id == op.id).count()
        op_data = {
            "id": str(op.id),
            "name": op.name,
            "project_id": op.project_id,
            "created_at": str(op.created_at),
            "inventory_count": inventory_count,
            "transaction_count": transaction_count,
            "type": "orbit"
        }

        # Check if this ORBIT project is linked to a Vector project
        linked_vector = orbit_to_vector.get(str(op.id))
        if linked_vector:
            op_data["linked_vector_id"] = linked_vector["id"]
            op_data["linked_vector_name"] = linked_vector["name"]
            linked_orbit.append(op_data)
        else:
            orphan_orbit.append(op_data)

    return {
        "orphan_vectors": orphan_vectors,
        "orphan_orbit": orphan_orbit,
        "linked_vectors": linked_vectors,
        "linked_orbit": linked_orbit,
        "total_orphan_vectors": len(orphan_vectors),
        "total_orphan_orbit": len(orphan_orbit),
        "total_linked": len(linked_vectors),
        "total_orphans": len(orphan_vectors) + len(orphan_orbit)
    }

@app.get("/api/vector/link-suggestions")
def get_link_suggestions(
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Suggest potential links between orphan Vector and ORBIT projects based on name similarity"""
    # Available to all authenticated users

    # Get orphan Vector projects
    vector_projects = db.query(VectorProject).filter(VectorProject.linked_project_id == None).all()

    # Get orphan ORBIT projects (not linked to any Vector project)
    linked_orbit_ids = [str(vp.linked_project_id) for vp in db.query(VectorProject).filter(VectorProject.linked_project_id != None).all()]
    orbit_projects = db.query(Project).filter(~Project.id.in_(linked_orbit_ids) if linked_orbit_ids else True).all()

    suggestions = []

    for vp in vector_projects:
        vp_name = (vp.name or '').lower()
        for op in orbit_projects:
            op_name = (op.name or '').lower()

            # Check if names are similar (simple substring match)
            if vp_name and op_name and (vp_name in op_name or op_name in vp_name):
                suggestions.append({
                    "vector_project": {
                        "id": str(vp.id),
                        "name": vp.name
                    },
                    "orbit_project": {
                        "id": str(op.id),
                        "name": op.name
                    },
                    "confidence": "high" if vp_name == op_name else "medium"
                })

    return {
        "suggestions": suggestions,
        "total": len(suggestions)
    }

# ============================================
# NOTIFICATION API
# ============================================
@app.get("/api/notifications")
def list_notifications(limit: int = 50, db: Session = Depends(get_db),
                       current_user: CompanyRep = Depends(get_current_user)):
    notifs = db.query(Notification).filter(
        Notification.user_rep_id == current_user.rep_id
    ).order_by(Notification.created_at.desc()).limit(limit).all()
    return [{
        "id": n.id, "notification_id": n.notification_id,
        "title": n.title, "message": n.message, "type": n.type,
        "category": n.category, "entity_type": n.entity_type,
        "entity_id": n.entity_id, "is_read": bool(n.is_read),
        "data": n.data or {}, "created_at": str(n.created_at),
        "read_at": str(n.read_at) if n.read_at else None
    } for n in notifs]

@app.get("/api/notifications/org-feed")
def list_org_notifications(
    limit: int = 200,
    offset: int = 0,
    rep_id: str = "",
    unread_only: bool = False,
    category: str = "",
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    can_view_org_feed = (
        current_user.role in ["admin", "director", "cco", "coo"]
        or current_user.rep_id in CUSTOMER_SYNC_APPROVER_REP_IDS
    )
    if not can_view_org_feed:
        raise HTTPException(403, "Insufficient permissions")

    q = db.query(Notification)
    if rep_id:
        q = q.filter(Notification.user_rep_id == rep_id)
    if unread_only:
        q = q.filter(Notification.is_read == False)
    if category:
        q = q.filter(Notification.category == category)

    total = q.count()
    rows = q.order_by(Notification.created_at.desc()).offset(max(0, offset)).limit(min(max(limit, 1), 1000)).all()
    rep_ids = list({n.user_rep_id for n in rows if n.user_rep_id})
    rep_rows = db.query(CompanyRep).filter(CompanyRep.rep_id.in_(rep_ids)).all() if rep_ids else []
    rep_map = {r.rep_id: r.name for r in rep_rows}

    return {
        "total": total,
        "count": len(rows),
        "items": [{
            "id": n.id,
            "notification_id": n.notification_id,
            "recipient_rep_id": n.user_rep_id,
            "recipient_name": rep_map.get(n.user_rep_id),
            "title": n.title,
            "message": n.message,
            "type": n.type,
            "category": n.category,
            "entity_type": n.entity_type,
            "entity_id": n.entity_id,
            "is_read": bool(n.is_read),
            "data": n.data or {},
            "created_at": str(n.created_at),
            "read_at": str(n.read_at) if n.read_at else None
        } for n in rows]
    }

@app.get("/api/notifications/unread")
def get_unread_notifications(db: Session = Depends(get_db),
                              current_user: CompanyRep = Depends(get_current_user)):
    notifs = db.query(Notification).filter(
        Notification.user_rep_id == current_user.rep_id,
        Notification.is_read == False
    ).order_by(Notification.created_at.desc()).limit(20).all()
    count = db.query(Notification).filter(
        Notification.user_rep_id == current_user.rep_id,
        Notification.is_read == False
    ).count()
    return {
        "count": count,
        "notifications": [{
            "id": n.id, "notification_id": n.notification_id,
            "title": n.title, "message": n.message, "type": n.type,
            "category": n.category,
            "entity_type": n.entity_type,
            "entity_id": n.entity_id,
            "is_read": bool(n.is_read),
            "data": n.data or {},
            "created_at": str(n.created_at)
        } for n in notifs]
    }

@app.post("/api/notifications/{nid}/read")
def mark_notification_read(nid: int, db: Session = Depends(get_db),
                           current_user: CompanyRep = Depends(get_current_user)):
    n = db.query(Notification).filter(
        Notification.id == nid,
        Notification.user_rep_id == current_user.rep_id
    ).first()
    if not n: raise HTTPException(404, "Notification not found")
    n.is_read = True
    n.read_at = datetime.utcnow()
    db.commit()
    return {"message": "Marked as read"}

@app.post("/api/notifications/read-all")
def mark_all_notifications_read(db: Session = Depends(get_db),
                                 current_user: CompanyRep = Depends(get_current_user)):
    db.query(Notification).filter(
        Notification.user_rep_id == current_user.rep_id,
        Notification.is_read == False
    ).update({"is_read": True, "read_at": datetime.utcnow()}, synchronize_session="fetch")
    db.commit()
    return {"message": "All notifications marked as read"}

# ============================================
# DUPLICATE CHECK API
# ============================================
@app.get("/api/duplicate-check")
def duplicate_check(mobile: str = None, db: Session = Depends(get_db),
                    current_user: CompanyRep = Depends(get_current_user)):
    if not mobile:
        return {"found": False, "matches": []}
    matches = check_duplicate_mobile(db, mobile)
    return {"found": len(matches) > 0, "matches": matches}

# ============================================
# UNIFIED SEARCH API
# ============================================
@app.get("/api/search/unified")
def unified_search(q: str = "", search_type: str = "mobile",
                   db: Session = Depends(get_db),
                   current_user: CompanyRep = Depends(get_current_user)):
    if not q or len(q) < 2:
        return {"results": []}

    results = []
    seen_ids = set()  # Deduplicate results from multi-transaction JOINs

    if search_type == "mobile":
        normalized = normalize_mobile(q)
        if normalized:
            suffix = normalized[-7:]  # last 7 digits for fuzzy SQL match
            # Strip non-digits in SQL for reliable matching
            stripped_cust_mobile = func.regexp_replace(Customer.mobile, '[^0-9]', '', 'g')
            stripped_lead_mobile = func.regexp_replace(Lead.mobile, '[^0-9]', '', 'g')

            # Search customers via SQL with JOIN for rep
            cust_rows = db.query(Customer, CompanyRep).outerjoin(
                Transaction, Transaction.customer_id == Customer.id
            ).outerjoin(
                CompanyRep, CompanyRep.id == Transaction.company_rep_id
            ).filter(
                Customer.mobile.isnot(None),
                stripped_cust_mobile.like(f"%{suffix}%")
            ).all()
            for c, rep in cust_rows:
                if str(c.id) in seen_ids:
                    continue
                if normalize_mobile(c.mobile) == normalized:
                    seen_ids.add(str(c.id))
                    results.append({
                        "type": "customer", "id": str(c.id), "entity_id": c.customer_id,
                        "name": c.name, "mobile": c.mobile,
                        "owner_rep": rep.name if rep else None,
                        "owner_rep_id": rep.rep_id if rep else None
                    })

            # ALSO search customer additional_mobiles JSONB array (fix for mobile search bug)
            try:
                add_mob_cust = db.execute(text("""
                    SELECT c.id, c.customer_id, c.name, c.mobile, am.elem,
                           r.name as rep_name, r.rep_id as rep_rep_id
                    FROM customers c
                    LEFT JOIN transactions t ON t.customer_id = c.id
                    LEFT JOIN company_reps r ON r.id = t.company_rep_id
                    CROSS JOIN jsonb_array_elements_text(c.additional_mobiles) AS am(elem)
                    WHERE c.additional_mobiles IS NOT NULL AND c.additional_mobiles != '[]'::jsonb
                      AND regexp_replace(am.elem, '[^0-9]', '', 'g') LIKE :suffix_pattern
                """), {"suffix_pattern": f"%{suffix}%"})
                for row in add_mob_cust:
                    if normalize_mobile(row.elem) == normalized:
                        # Avoid duplicates if primary mobile was already added
                        if str(row.id) not in seen_ids:
                            seen_ids.add(str(row.id))
                            results.append({
                                "type": "customer", "id": str(row.id), "entity_id": row.customer_id,
                                "name": row.name, "mobile": row.elem,
                                "owner_rep": row.rep_name,
                                "owner_rep_id": row.rep_rep_id
                            })
            except Exception as e:
                print(f"[unified-search] additional_mobiles customer search failed: {e}")
                pass

            # Search leads via SQL with JOIN for rep
            lead_rows = db.query(Lead, CompanyRep).outerjoin(
                CompanyRep, CompanyRep.id == Lead.assigned_rep_id
            ).filter(
                Lead.status != "converted",
                Lead.mobile.isnot(None),
                stripped_lead_mobile.like(f"%{suffix}%")
            ).all()
            for l, rep in lead_rows:
                if normalize_mobile(l.mobile) == normalized:
                    results.append({
                        "type": "lead", "id": str(l.id), "entity_id": l.lead_id,
                        "name": l.name, "mobile": l.mobile,
                        "pipeline_stage": l.pipeline_stage,
                        "owner_rep": rep.name if rep else None,
                        "owner_rep_id": rep.rep_id if rep else None
                    })

            # ALSO search lead additional_mobiles JSONB array (fix for mobile search bug)
            try:
                add_mob_lead = db.execute(text("""
                    SELECT l.id, l.lead_id, l.name, l.mobile, am.elem, l.pipeline_stage,
                           r.name as rep_name, r.rep_id as rep_rep_id
                    FROM leads l
                    LEFT JOIN company_reps r ON l.assigned_rep_id = r.id
                    CROSS JOIN jsonb_array_elements_text(l.additional_mobiles) AS am(elem)
                    WHERE l.status != 'converted'
                      AND l.additional_mobiles IS NOT NULL AND l.additional_mobiles != '[]'::jsonb
                      AND regexp_replace(am.elem, '[^0-9]', '', 'g') LIKE :suffix_pattern
                """), {"suffix_pattern": f"%{suffix}%"})
                for row in add_mob_lead:
                    if normalize_mobile(row.elem) == normalized:
                        # Avoid duplicates if primary mobile was already added
                        already = any(r["id"] == str(row.id) and r["type"] == "lead" for r in results)
                        if not already:
                            results.append({
                                "type": "lead", "id": str(row.id), "entity_id": row.lead_id,
                                "name": row.name, "mobile": row.elem,
                                "pipeline_stage": row.pipeline_stage,
                                "owner_rep": row.rep_name,
                                "owner_rep_id": row.rep_rep_id
                            })
            except Exception as e:
                print(f"[unified-search] additional_mobiles lead search failed: {e}")
                pass
    else:
        # Name search via SQL ILIKE
        cust_rows = db.query(Customer, CompanyRep).outerjoin(
            Transaction, Transaction.customer_id == Customer.id
        ).outerjoin(
            CompanyRep, CompanyRep.id == Transaction.company_rep_id
        ).filter(
            Customer.name.ilike(f"%{q}%")
        ).all()
        for c, rep in cust_rows:
            if str(c.id) in seen_ids:
                continue
            seen_ids.add(str(c.id))
            results.append({
                "type": "customer", "id": str(c.id), "entity_id": c.customer_id,
                "name": c.name, "mobile": c.mobile,
                "owner_rep": rep.name if rep else None,
                "owner_rep_id": rep.rep_id if rep else None
            })
        lead_rows = db.query(Lead, CompanyRep).outerjoin(
            CompanyRep, CompanyRep.id == Lead.assigned_rep_id
        ).filter(
            Lead.status != "converted",
            Lead.name.ilike(f"%{q}%")
        ).all()
        for l, rep in lead_rows:
            results.append({
                "type": "lead", "id": str(l.id), "entity_id": l.lead_id,
                "name": l.name, "mobile": l.mobile,
                "pipeline_stage": l.pipeline_stage,
                "owner_rep": rep.name if rep else None,
                "owner_rep_id": rep.rep_id if rep else None
            })

    # Log search if results found belonging to another rep, aggregate notifications per owner
    owner_results = {}  # owner_rep_id -> list of result names
    for r in results:
        if r.get("owner_rep_id") and r["owner_rep_id"] != current_user.rep_id:
            # Log each search hit individually
            log = SearchLog(
                searcher_rep_id=current_user.rep_id,
                search_query=q,
                search_type=search_type,
                matched_entity_type=r["type"],
                matched_entity_id=r["entity_id"],
                matched_entity_name=r["name"],
                owner_rep_id=r["owner_rep_id"]
            )
            db.add(log)
            # Aggregate for notification
            owner_id = r["owner_rep_id"]
            if owner_id not in owner_results:
                owner_results[owner_id] = []
            owner_results[owner_id].append(r["name"])
    # Send one aggregated notification per owner rep
    for owner_id, names in owner_results.items():
        count = len(names)
        preview = ", ".join(names[:3])
        if count > 3:
            preview += f" (+{count - 3} more)"
        create_notification(
            db, owner_id, "search_alert",
            f"{current_user.name} searched your records ({count} found)",
            f"{current_user.name} ({current_user.rep_id}) searched for '{q}' — matched: {preview}",
            category="search_alert",
            data={"searcher_rep_id": current_user.rep_id, "searcher_name": current_user.name, "match_count": count}
        )
    if owner_results:
        try:
            db.commit()
        except Exception as e:
            print(f"[WARN] Search log commit failed: {e}")
            db.rollback()

    return {"results": results}

@app.get("/api/search-log")
def get_search_log(limit: int = 50, db: Session = Depends(get_db),
                   current_user: CompanyRep = Depends(require_role(["admin", "cco"]))):
    logs = db.query(SearchLog).order_by(SearchLog.created_at.desc()).limit(limit).all()
    return [{
        "id": str(l.id), "searcher_rep_id": l.searcher_rep_id,
        "search_query": l.search_query, "search_type": l.search_type,
        "matched_entity_type": l.matched_entity_type,
        "matched_entity_id": l.matched_entity_id,
        "matched_entity_name": l.matched_entity_name,
        "owner_rep_id": l.owner_rep_id,
        "created_at": str(l.created_at)
    } for l in logs]

# ============================================
# PIPELINE STAGES API
# ============================================
@app.get("/api/pipeline-stages")
def list_pipeline_stages(db: Session = Depends(get_db),
                         current_user: CompanyRep = Depends(get_current_user)):
    stages = db.query(PipelineStage).order_by(PipelineStage.display_order).all()
    return [{
        "id": str(s.id), "name": s.name, "display_order": s.display_order,
        "color": s.color, "is_terminal": bool(s.is_terminal),
        "created_at": str(s.created_at)
    } for s in stages]

@app.post("/api/pipeline-stages")
def create_pipeline_stage(data: dict, db: Session = Depends(get_db),
                          current_user: CompanyRep = Depends(require_role(["admin", "cco"]))):
    s = PipelineStage(
        name=data["name"],
        display_order=data.get("display_order", 99),
        color=data.get("color", "#6B7280"),
        is_terminal=bool(data.get("is_terminal", False))
    )
    db.add(s); db.commit(); db.refresh(s)
    return {"message": "Stage created", "id": str(s.id)}

@app.put("/api/pipeline-stages/{sid}")
def update_pipeline_stage(sid: str, data: dict, db: Session = Depends(get_db),
                          current_user: CompanyRep = Depends(require_role(["admin", "cco"]))):
    s = db.query(PipelineStage).filter(PipelineStage.id == sid).first()
    if not s: raise HTTPException(404, "Stage not found")
    for k in ["name", "display_order", "color"]:
        if k in data: setattr(s, k, data[k])
    if "is_terminal" in data:
        s.is_terminal = bool(data["is_terminal"])
    db.commit()
    return {"message": "Stage updated"}

@app.delete("/api/pipeline-stages/{sid}")
def delete_pipeline_stage(sid: str, data: dict = {}, db: Session = Depends(get_db),
                          current_user: CompanyRep = Depends(require_role(["admin", "cco"]))):
    s = db.query(PipelineStage).filter(PipelineStage.id == sid).first()
    if not s: raise HTTPException(404, "Stage not found")
    # Don't delete if leads exist in this stage
    count = db.query(Lead).filter(Lead.pipeline_stage == s.name).count()
    if count > 0:
        raise HTTPException(400, f"Cannot delete stage with {count} leads. Move leads first.")
    db.delete(s); db.commit()
    return {"message": "Stage deleted"}

# ============================================
# LOOKUP VALUES API (admin-managed dropdowns)
# ============================================
@app.get("/api/lookup-values")
def list_lookup_values(category: str = None, db: Session = Depends(get_db),
                       current_user: CompanyRep = Depends(get_current_user)):
    q = db.query(LookupValue).filter(LookupValue.is_active == True)
    if category:
        q = q.filter(LookupValue.category == category)
    values = q.order_by(LookupValue.category, LookupValue.sort_order).all()
    return [{
        "id": str(v.id), "category": v.category, "label": v.label,
        "sort_order": v.sort_order, "is_active": v.is_active,
        "created_at": str(v.created_at)
    } for v in values]

@app.post("/api/lookup-values")
def create_lookup_value(data: dict, db: Session = Depends(get_db),
                        current_user: CompanyRep = Depends(require_role(["admin", "cco"]))):
    existing = db.query(LookupValue).filter(
        LookupValue.category == data["category"],
        LookupValue.label == data["label"]
    ).first()
    if existing:
        raise HTTPException(400, f"Value '{data['label']}' already exists in '{data['category']}'")
    v = LookupValue(
        category=data["category"],
        label=data["label"],
        sort_order=data.get("sort_order", 99)
    )
    db.add(v); db.commit(); db.refresh(v)
    return {"message": "Value created", "id": str(v.id)}

@app.put("/api/lookup-values/{vid}")
def update_lookup_value(vid: str, data: dict, db: Session = Depends(get_db),
                        current_user: CompanyRep = Depends(require_role(["admin", "cco"]))):
    v = db.query(LookupValue).filter(LookupValue.id == vid).first()
    if not v: raise HTTPException(404, "Value not found")
    for k in ["label", "sort_order", "is_active"]:
        if k in data: setattr(v, k, data[k])
    db.commit()
    return {"message": "Value updated"}

@app.delete("/api/lookup-values/{vid}")
def delete_lookup_value(vid: str, db: Session = Depends(get_db),
                        current_user: CompanyRep = Depends(require_role(["admin", "cco"]))):
    v = db.query(LookupValue).filter(LookupValue.id == vid).first()
    if not v: raise HTTPException(404, "Value not found")
    db.delete(v); db.commit()
    return {"message": "Value deleted"}

# ============================================
# LEAD PIPELINE API
# ============================================
@app.get("/api/leads/pipeline")
def get_leads_pipeline(rep_id: str = None, campaign_id: str = None,
                       stale_only: bool = False, search: str = None,
                       stage: str = None, page: int = 1, page_size: int = 50,
                       db: Session = Depends(get_db),
                       current_user: CompanyRep = Depends(get_current_user)):
    # User role only sees their own leads
    effective_rep_id = None
    if current_user.role == "user":
        effective_rep_id = current_user.id
    elif rep_id:
        r = db.query(CompanyRep).filter((CompanyRep.id == rep_id) | (CompanyRep.rep_id == rep_id)).first()
        if r: effective_rep_id = r.id

    # Include active leads + recently Won/Lost (last 90 days) so they appear in terminal stages
    cutoff = datetime.utcnow() - timedelta(days=90)
    base_filter = or_(
        Lead.status.notin_(["converted", "lost"]),
        and_(Lead.status.in_(["converted", "lost"]), Lead.updated_at >= cutoff)
    )
    q = db.query(Lead).filter(base_filter)
    if effective_rep_id:
        q = q.filter(Lead.assigned_rep_id == effective_rep_id)
    if campaign_id:
        camp = db.query(Campaign).filter((Campaign.id == campaign_id) | (Campaign.campaign_id == campaign_id)).first()
        if camp: q = q.filter(Lead.campaign_id == camp.id)
    if stale_only:
        q = q.filter(Lead.is_stale == True)

    # Stage counts (fast SQL) — always computed from base query before search/stage filter
    from sqlalchemy import func as sa_func
    count_q = q.with_entities(
        sa_func.coalesce(Lead.pipeline_stage, 'New').label('stage'),
        sa_func.count(Lead.id)
    ).group_by(sa_func.coalesce(Lead.pipeline_stage, 'New'))
    stage_counts = dict(count_q.all())
    total_count = sum(stage_counts.values())

    # Build pipeline stage metadata with counts
    stages_db = db.query(PipelineStage).order_by(PipelineStage.display_order).all()
    pipeline = []
    for s in stages_db:
        pipeline.append({
            "stage": s.name, "color": s.color, "is_terminal": bool(s.is_terminal),
            "count": stage_counts.get(s.name, 0), "leads": []
        })

    # Apply search filter
    if search and search.strip():
        term = f"%{search.strip()}%"
        q = q.filter(or_(
            Lead.name.ilike(term),
            Lead.mobile.ilike(term),
            Lead.lead_id.ilike(term),
            Lead.email.ilike(term),
            Lead.source.ilike(term)
        ))

    # Apply stage filter
    if stage and stage != 'all':
        q = q.filter(sa_func.coalesce(Lead.pipeline_stage, 'New') == stage)

    # Get filtered total for pagination
    filtered_total = q.count()

    # Paginate
    offset = (max(page, 1) - 1) * page_size
    leads = q.order_by(Lead.created_at.desc()).offset(offset).limit(page_size).all()

    # Pre-fetch reps and campaigns for this page only
    rep_ids = set(l.assigned_rep_id for l in leads if l.assigned_rep_id)
    camp_ids = set(l.campaign_id for l in leads if l.campaign_id)
    reps_map = {r.id: r for r in db.query(CompanyRep).filter(CompanyRep.id.in_(rep_ids)).all()} if rep_ids else {}
    camps_map = {c.id: c for c in db.query(Campaign).filter(Campaign.id.in_(camp_ids)).all()} if camp_ids else {}

    leads_data = []
    for l in leads:
        rep = reps_map.get(l.assigned_rep_id)
        camp = camps_map.get(l.campaign_id)
        days_since_contact = None
        if l.last_contacted_at:
            days_since_contact = (datetime.utcnow() - l.last_contacted_at).days
        leads_data.append({
            "id": str(l.id), "lead_id": l.lead_id, "name": l.name,
            "mobile": l.mobile, "email": l.email,
            "additional_mobiles": l.additional_mobiles or [],
            "pipeline_stage": l.pipeline_stage or "New",
            "status": l.status, "lead_type": l.lead_type,
            "campaign_name": camp.name if camp else None,
            "campaign_source": camp.source if camp else None,
            "assigned_rep": rep.name if rep else None,
            "assigned_rep_id": rep.rep_id if rep else None,
            "days_since_contact": days_since_contact,
            "is_stale": bool(l.is_stale),
            "converted_customer_id": str(l.converted_customer_id) if l.converted_customer_id else None,
            "converted_broker_id": str(l.converted_broker_id) if l.converted_broker_id else None,
            "source": l.source, "source_details": l.source_details,
            "temperature": l.temperature,
            "notes": l.notes, "created_at": str(l.created_at)
        })

    return {
        "pipeline": pipeline, "total": total_count,
        "leads": leads_data, "filtered_total": filtered_total,
        "page": page, "page_size": page_size,
        "total_pages": max(1, -(-filtered_total // page_size))
    }

@app.put("/api/leads/{lid}/stage")
def update_lead_stage(lid: str, data: dict, db: Session = Depends(get_db),
                      current_user: CompanyRep = Depends(get_current_user)):
    l = db.query(Lead).filter((Lead.id == lid) | (Lead.lead_id == lid)).first()
    if not l: raise HTTPException(404, "Lead not found")
    # Viewers cannot modify leads
    if current_user.role == "viewer":
        raise HTTPException(403, "Insufficient permissions")
    # Non-privileged roles can only update their own assigned leads
    if current_user.role in ("user", "creator") and l.assigned_rep_id != current_user.id:
        raise HTTPException(403, "You can only update your own leads")
    new_stage = data.get("stage")
    if not new_stage: raise HTTPException(400, "Stage required")
    # Verify stage exists
    stage = db.query(PipelineStage).filter(PipelineStage.name == new_stage).first()
    if not stage: raise HTTPException(400, f"Invalid stage: {new_stage}")
    l.pipeline_stage = new_stage
    # If terminal stage, update status
    if stage.is_terminal:
        if new_stage == "Won":
            l.status = "converted"  # Sale closed
        elif new_stage == "Lost":
            l.status = "lost"
    db.commit()
    return {"message": f"Lead moved to {new_stage}"}

# ============================================
# LEAD ASSIGNMENT API
# ============================================
@app.post("/api/leads/bulk-assign")
def bulk_assign_leads(data: dict, db: Session = Depends(get_db),
                      current_user: CompanyRep = Depends(require_role(["admin", "manager", "cco"]))):
    lead_ids = data.get("lead_ids", [])
    rep_id = data.get("rep_id")

    # Validate inputs
    if not lead_ids or not rep_id:
        raise HTTPException(400, "lead_ids and rep_id required")
    if not isinstance(lead_ids, list) or len(lead_ids) == 0:
        raise HTTPException(400, "lead_ids must be a non-empty array")

    # Find target rep (support both UUID and rep_id string)
    rep = db.query(CompanyRep).filter(
        (CompanyRep.id == rep_id) | (CompanyRep.rep_id == rep_id)
    ).first()
    if not rep:
        raise HTTPException(404, f"Rep not found with id: {rep_id}")

    # Batch fetch leads to avoid N+1
    leads = db.query(Lead).filter(
        or_(Lead.id.in_(lead_ids), Lead.lead_id.in_(lead_ids))
    ).all()

    if len(leads) == 0:
        raise HTTPException(404, f"No leads found matching the provided IDs")

    # Assign leads
    count = 0
    try:
        for l in leads:
            l.assigned_rep_id = rep.id  # Set to UUID
            count += 1

        # Create notification
        try:
            create_notification(
                db, rep.rep_id, "assignment",
                f"{count} leads assigned to you",
                f"{current_user.name} assigned {count} leads to you",
                category="assignment"
            )
        except Exception as notif_err:
            # Log but don't fail the assignment if notification fails
            print(f"[bulk-assign] Notification creation failed: {notif_err}")

        # Commit all changes
        db.commit()
        return {"message": f"{count} leads assigned to {rep.name}"}

    except Exception as e:
        db.rollback()
        print(f"[bulk-assign] Assignment failed: {str(e)}")
        raise HTTPException(500, f"Assignment failed: {str(e)}")

@app.post("/api/leads/{lid}/request-assignment")
def request_lead_assignment(lid: str, data: dict, db: Session = Depends(get_db),
                            current_user: CompanyRep = Depends(get_current_user)):
    l = db.query(Lead).filter((Lead.id == lid) | (Lead.lead_id == lid)).first()
    if not l: raise HTTPException(404, "Lead not found")

    # Check if there's already a pending request
    existing = db.query(LeadAssignmentRequest).filter(
        LeadAssignmentRequest.lead_id == l.id,
        LeadAssignmentRequest.requested_by == current_user.id,
        LeadAssignmentRequest.status == "pending"
    ).first()
    if existing:
        raise HTTPException(400, "You already have a pending request for this lead")

    req = LeadAssignmentRequest(
        lead_id=l.id,
        requested_by=current_user.id,
        reason=data.get("reason", "")
    )
    db.add(req); db.commit(); db.refresh(req)

    # Notify admin/cco
    admins = db.query(CompanyRep).filter(CompanyRep.role.in_(["admin", "cco"]), CompanyRep.status == "active").all()
    for admin in admins:
        create_notification(
            db, admin.rep_id, "assignment_request",
            f"{current_user.name} requests lead {l.lead_id}",
            f"{current_user.name} wants to be assigned lead {l.name} ({l.lead_id}). Reason: {data.get('reason', 'N/A')}",
            category="assignment_request",
            entity_type="lead",
            entity_id=l.lead_id
        )
    db.commit()
    return {"message": "Assignment request submitted", "request_id": req.request_id}

@app.get("/api/lead-assignment-requests")
def list_assignment_requests(status: str = "pending", db: Session = Depends(get_db),
                             current_user: CompanyRep = Depends(require_role(["admin", "cco"]))):
    q = db.query(LeadAssignmentRequest)
    if status:
        q = q.filter(LeadAssignmentRequest.status == status)
    requests = q.order_by(LeadAssignmentRequest.created_at.desc()).all()
    # Pre-fetch related entities to avoid N+1
    lead_ids = set(r.lead_id for r in requests if r.lead_id)
    rep_ids = set(r.requested_by for r in requests) | set(r.reviewed_by for r in requests if r.reviewed_by)
    leads_map = {l.id: l for l in db.query(Lead).filter(Lead.id.in_(lead_ids)).all()} if lead_ids else {}
    reps_map = {r.id: r for r in db.query(CompanyRep).filter(CompanyRep.id.in_(rep_ids)).all()} if rep_ids else {}
    result = []
    for r in requests:
        lead = leads_map.get(r.lead_id)
        requester = reps_map.get(r.requested_by)
        reviewer = reps_map.get(r.reviewed_by)
        result.append({
            "id": str(r.id), "request_id": r.request_id,
            "lead_id": lead.lead_id if lead else None, "lead_name": lead.name if lead else None,
            "requested_by": requester.name if requester else None,
            "requester_rep_id": requester.rep_id if requester else None,
            "reason": r.reason, "status": r.status,
            "reviewed_by": reviewer.name if reviewer else None,
            "reviewed_at": str(r.reviewed_at) if r.reviewed_at else None,
            "created_at": str(r.created_at)
        })
    return result

@app.post("/api/lead-assignment-requests/{rid}/review")
def review_assignment_request(rid: str, data: dict, db: Session = Depends(get_db),
                              current_user: CompanyRep = Depends(require_role(["admin", "cco"]))):
    req = db.query(LeadAssignmentRequest).filter(
        (LeadAssignmentRequest.id == rid) | (LeadAssignmentRequest.request_id == rid)
    ).first()
    if not req: raise HTTPException(404, "Request not found")
    if req.status != "pending": raise HTTPException(400, "Request is not pending")

    action = data.get("action")  # "approve" or "reject"
    if action == "approve":
        req.status = "approved"
        req.reviewed_by = current_user.id
        req.reviewed_at = datetime.utcnow()

        # Assign the lead
        lead = db.query(Lead).filter(Lead.id == req.lead_id).first()
        if lead:
            lead.assigned_rep_id = req.requested_by

        # Notify requester
        requester = db.query(CompanyRep).filter(CompanyRep.id == req.requested_by).first()
        if requester:
            create_notification(
                db, requester.rep_id, "assignment_approved",
                f"Lead assignment approved",
                f"Your request for lead {lead.lead_id if lead else 'unknown'} ({lead.name if lead else ''}) was approved by {current_user.name}",
                category="assignment"
            )
    elif action == "reject":
        req.status = "rejected"
        req.reviewed_by = current_user.id
        req.reviewed_at = datetime.utcnow()

        requester = db.query(CompanyRep).filter(CompanyRep.id == req.requested_by).first()
        lead = db.query(Lead).filter(Lead.id == req.lead_id).first()
        if requester:
            create_notification(
                db, requester.rep_id, "assignment_rejected",
                f"Lead assignment rejected",
                f"Your request for lead {lead.lead_id if lead else 'unknown'} was rejected by {current_user.name}",
                category="assignment"
            )
    else:
        raise HTTPException(400, "Action must be 'approve' or 'reject'")

    db.commit()
    return {"message": f"Request {action}d"}

# ============================================
# LEAD -> CUSTOMER SYNC REQUEST API
# ============================================
@app.post("/api/leads/{lid}/request-customer-sync")
def request_customer_sync(lid: str, data: dict, db: Session = Depends(get_db),
                          current_user: CompanyRep = Depends(get_current_user)):
    _ensure_lead_sync_requests_table(db)
    lead = db.query(Lead).filter((Lead.id == lid) | (Lead.lead_id == lid)).first()
    if not lead:
        raise HTTPException(404, "Lead not found")
    if lead.converted_customer_id:
        raise HTTPException(400, "Lead is already synced to customer DB")
    # Sales users can request sync only for leads assigned to them.
    if current_user.role == "user" and lead.assigned_rep_id != current_user.id:
        raise HTTPException(403, "You can request sync only for leads assigned to you")

    exists = db.execute(text("""
        SELECT 1 FROM lead_sync_requests
        WHERE lead_id = :lead_id AND requested_by = :requested_by AND status = 'pending'
        LIMIT 1
    """), {"lead_id": str(lead.id), "requested_by": str(current_user.id)}).first()
    if exists:
        raise HTTPException(400, "You already have a pending customer sync request for this lead")

    request_id = f"LSR-{uuid.uuid4().hex[:8].upper()}"
    db.execute(text("""
        INSERT INTO lead_sync_requests (request_id, lead_id, requested_by, reason, status)
        VALUES (:request_id, :lead_id, :requested_by, :reason, 'pending')
    """), {
        "request_id": request_id,
        "lead_id": str(lead.id),
        "requested_by": str(current_user.id),
        "reason": data.get("reason", "")
    })

    approvers = db.query(CompanyRep).filter(
        and_(
            CompanyRep.status == "active",
            or_(
                CompanyRep.rep_id.in_(CUSTOMER_SYNC_APPROVER_REP_IDS),
                CompanyRep.role.in_(["admin", "cco", "director", "coo"])
            )
        )
    ).all()
    for ap in approvers:
        create_notification(
            db, ap.rep_id, "customer_sync_request",
            f"{current_user.name} requested customer sync",
            f"Lead {lead.lead_id} ({lead.name}) requested for customer sync. Reason: {data.get('reason', 'N/A')}",
            category="sync_request",
            entity_type="lead",
            entity_id=lead.lead_id,
            data={"request_id": request_id, "lead_id": lead.lead_id, "requested_by": current_user.rep_id}
        )
    db.commit()
    return {"message": "Customer sync request submitted", "request_id": request_id}

@app.get("/api/lead-sync-requests")
def list_customer_sync_requests(status: str = "pending", db: Session = Depends(get_db),
                                current_user: CompanyRep = Depends(get_current_user)):
    _ensure_lead_sync_requests_table(db)
    if not _is_customer_sync_approver(current_user):
        raise HTTPException(403, "Insufficient permissions")
    rows = db.execute(text("""
        SELECT
            r.id::text AS id, r.request_id, r.status, r.reason,
            r.created_at::text AS created_at, r.reviewed_at::text AS reviewed_at,
            l.lead_id AS lead_code, l.name AS lead_name, l.mobile AS lead_mobile,
            req.rep_id AS requester_rep_id, req.name AS requester_name,
            rev.rep_id AS reviewer_rep_id, rev.name AS reviewer_name
        FROM lead_sync_requests r
        JOIN leads l ON l.id = r.lead_id
        JOIN company_reps req ON req.id = r.requested_by
        LEFT JOIN company_reps rev ON rev.id = r.reviewed_by
        WHERE (:status = '' OR r.status = :status)
        ORDER BY r.created_at DESC
    """), {"status": status or ""}).mappings().all()
    return [dict(row) for row in rows]

@app.post("/api/lead-sync-requests/{rid}/review")
def review_customer_sync_request(rid: str, data: dict, db: Session = Depends(get_db),
                                 current_user: CompanyRep = Depends(get_current_user)):
    _ensure_lead_sync_requests_table(db)
    if not _is_customer_sync_approver(current_user):
        raise HTTPException(403, "Insufficient permissions")
    action = (data.get("action") or "").strip().lower()
    if action not in ["approve", "reject"]:
        raise HTTPException(400, "Action must be 'approve' or 'reject'")

    req = db.execute(text("""
        SELECT id, request_id, lead_id, requested_by, status
        FROM lead_sync_requests
        WHERE id::text = :rid OR request_id = :rid
        LIMIT 1
    """), {"rid": rid}).mappings().first()
    if not req:
        raise HTTPException(404, "Request not found")
    if req["status"] != "pending":
        raise HTTPException(400, "Request is not pending")

    lead = db.query(Lead).filter(Lead.id == req["lead_id"]).first()
    requester = db.query(CompanyRep).filter(CompanyRep.id == req["requested_by"]).first()
    if not lead or not requester:
        raise HTTPException(404, "Linked lead/requester not found")

    if action == "approve":
        linked_existing, entity_id, _ = _sync_lead_to_customer_record(lead, db)
        db.execute(text("""
            UPDATE lead_sync_requests
            SET status='approved', reviewed_by=:reviewed_by, reviewed_at=NOW()
            WHERE id=:id
        """), {"reviewed_by": str(current_user.id), "id": str(req["id"])})
        create_notification(
            db, requester.rep_id, "customer_sync_approved",
            "Customer sync approved",
            f"Your sync request for lead {lead.lead_id} was approved by {current_user.name}. Customer: {entity_id}",
            category="sync_request",
            entity_type="lead",
            entity_id=lead.lead_id
        )
        db.commit()
        return {"message": "Request approved and lead synced to customer DB", "linked_existing": linked_existing, "entity_id": entity_id}

    db.execute(text("""
        UPDATE lead_sync_requests
        SET status='rejected', reviewed_by=:reviewed_by, reviewed_at=NOW()
        WHERE id=:id
    """), {"reviewed_by": str(current_user.id), "id": str(req["id"])})
    create_notification(
        db, requester.rep_id, "customer_sync_rejected",
        "Customer sync request rejected",
        f"Your sync request for lead {lead.lead_id} was rejected by {current_user.name}.",
        category="sync_request",
        entity_type="lead",
        entity_id=lead.lead_id
    )
    db.commit()
    return {"message": "Request rejected"}

@app.post("/api/lead-sync-requests/bulk-review")
def bulk_review_customer_sync_requests(data: dict, db: Session = Depends(get_db),
                                       current_user: CompanyRep = Depends(get_current_user)):
    _ensure_lead_sync_requests_table(db)
    if not _is_customer_sync_approver(current_user):
        raise HTTPException(403, "Insufficient permissions")
    request_ids = data.get("request_ids") or []
    action = (data.get("action") or "").strip().lower()
    if not request_ids or action not in ["approve", "reject"]:
        raise HTTPException(400, "request_ids and valid action required")

    processed = 0
    synced = 0
    for rid in request_ids:
        req = db.execute(text("""
            SELECT id, request_id, lead_id, requested_by, status
            FROM lead_sync_requests
            WHERE id::text = :rid OR request_id = :rid
            LIMIT 1
        """), {"rid": rid}).mappings().first()
        if not req or req["status"] != "pending":
            continue
        lead = db.query(Lead).filter(Lead.id == req["lead_id"]).first()
        requester = db.query(CompanyRep).filter(CompanyRep.id == req["requested_by"]).first()
        if not lead or not requester:
            continue

        if action == "approve":
            _sync_lead_to_customer_record(lead, db)
            db.execute(text("""
                UPDATE lead_sync_requests
                SET status='approved', reviewed_by=:reviewed_by, reviewed_at=NOW()
                WHERE id=:id
            """), {"reviewed_by": str(current_user.id), "id": str(req["id"])})
            create_notification(
                db, requester.rep_id, "customer_sync_approved",
                "Customer sync approved",
                f"Your sync request for lead {lead.lead_id} was approved by {current_user.name}.",
                category="sync_request",
                entity_type="lead",
                entity_id=lead.lead_id
            )
            synced += 1
        else:
            db.execute(text("""
                UPDATE lead_sync_requests
                SET status='rejected', reviewed_by=:reviewed_by, reviewed_at=NOW()
                WHERE id=:id
            """), {"reviewed_by": str(current_user.id), "id": str(req["id"])})
            create_notification(
                db, requester.rep_id, "customer_sync_rejected",
                "Customer sync request rejected",
                f"Your sync request for lead {lead.lead_id} was rejected by {current_user.name}.",
                category="sync_request",
                entity_type="lead",
                entity_id=lead.lead_id
            )
        processed += 1

    db.commit()
    return {"message": f"{processed} requests processed", "processed": processed, "synced": synced, "action": action}

@app.post("/api/leads/bulk-sync-customers")
def bulk_sync_leads_to_customers(data: dict, db: Session = Depends(get_db),
                                 current_user: CompanyRep = Depends(get_current_user)):
    if not _is_customer_sync_approver(current_user):
        raise HTTPException(403, "Insufficient permissions")

    lead_ids = data.get("lead_ids") or []
    if not isinstance(lead_ids, list) or not lead_ids:
        raise HTTPException(400, "lead_ids list is required")
    if len(lead_ids) > 500:
        raise HTTPException(400, "Maximum 500 leads allowed per bulk sync request")

    synced = 0
    linked_existing = 0
    failed = 0
    failures = []
    processed_codes = []

    for raw_id in lead_ids:
        lid = (str(raw_id or "")).strip()
        if not lid:
            continue
        lead = find_entity(db, Lead, "lead_id", lid)
        if not lead:
            failed += 1
            failures.append({"lead_id": lid, "error": "Lead not found"})
            continue

        try:
            linked, entity_id, _ = _sync_lead_to_customer_record(lead, db)
            synced += 1
            if linked:
                linked_existing += 1
            processed_codes.append({"lead_id": lead.lead_id, "customer_id": entity_id, "linked_existing": linked})
        except Exception as e:
            failed += 1
            failures.append({"lead_id": lead.lead_id, "error": str(e)})

    db.commit()
    return {
        "message": "Bulk customer sync completed",
        "synced": synced,
        "linked_existing": linked_existing,
        "failed": failed,
        "processed": processed_codes,
        "failures": failures[:50]
    }

# ============================================
# STALE LEAD API
# ============================================
@app.get("/api/leads/stale")
def list_stale_leads(db: Session = Depends(get_db),
                     current_user: CompanyRep = Depends(require_role(["admin", "manager", "cco"]))):
    stale_rows = db.query(Lead, CompanyRep).outerjoin(
        CompanyRep, CompanyRep.id == Lead.assigned_rep_id
    ).filter(
        Lead.is_stale == True,
        Lead.status.notin_(["converted", "lost"])
    ).order_by(Lead.last_contacted_at.asc().nullsfirst()).all()
    result = []
    for l, rep in stale_rows:
        days = None
        if l.last_contacted_at:
            days = (datetime.utcnow() - l.last_contacted_at).days
        result.append({
            "id": str(l.id), "lead_id": l.lead_id, "name": l.name,
            "mobile": l.mobile, "pipeline_stage": l.pipeline_stage,
            "assigned_rep": rep.name if rep else None,
            "assigned_rep_id": rep.rep_id if rep else None,
            "days_since_contact": days,
            "created_at": str(l.created_at)
        })
    return result

@app.post("/api/leads/check-stale")
def check_stale_leads(db: Session = Depends(get_db),
                      current_user: CompanyRep = Depends(require_role(["admin", "cco"]))):
    """Flag leads with no interaction in 60 days"""
    threshold = datetime.utcnow() - timedelta(days=60)

    # Bulk SQL: flag newly stale (no contact in 60+ days, not already stale)
    newly_stale = db.query(Lead).filter(
        Lead.status.notin_(["converted", "lost"]),
        or_(Lead.last_contacted_at == None, Lead.last_contacted_at < threshold),
        or_(Lead.is_stale == False, Lead.is_stale == None)
    ).update({"is_stale": True}, synchronize_session="fetch")

    # Bulk SQL: unflag recovered leads (contacted recently but still marked stale)
    db.query(Lead).filter(
        Lead.status.notin_(["converted", "lost"]),
        Lead.is_stale == True,
        Lead.last_contacted_at != None,
        Lead.last_contacted_at >= threshold
    ).update({"is_stale": False}, synchronize_session="fetch")

    db.commit()

    # Notify admin/cco about newly stale leads
    if newly_stale > 0:
        admins = db.query(CompanyRep).filter(
            CompanyRep.role.in_(["admin", "cco"]),
            CompanyRep.status == "active"
        ).all()
        for admin in admins:
            create_notification(
                db, admin.rep_id, "stale_lead",
                f"{newly_stale} leads flagged as stale",
                f"{newly_stale} leads have had no interaction in 60+ days and need attention",
                category="stale_lead"
            )
        db.commit()

    return {"message": f"Stale check complete. {newly_stale} newly flagged.", "total_stale": db.query(Lead).filter(Lead.is_stale == True).count()}

@app.post("/api/leads/{lid}/reassign")
def reassign_lead(lid: str, data: dict, db: Session = Depends(get_db),
                  current_user: CompanyRep = Depends(require_role(["admin", "cco"]))):
    l = db.query(Lead).filter((Lead.id == lid) | (Lead.lead_id == lid)).first()
    if not l: raise HTTPException(404, "Lead not found")

    new_rep_id = data.get("rep_id")
    if not new_rep_id: raise HTTPException(400, "rep_id required")

    new_rep = db.query(CompanyRep).filter((CompanyRep.id == new_rep_id) | (CompanyRep.rep_id == new_rep_id)).first()
    if not new_rep: raise HTTPException(404, "Rep not found")

    old_rep = db.query(CompanyRep).filter(CompanyRep.id == l.assigned_rep_id).first() if l.assigned_rep_id else None

    l.assigned_rep_id = new_rep.id

    # Notifications + data in single atomic commit
    create_notification(
        db, new_rep.rep_id, "assignment",
        f"Lead {l.lead_id} assigned to you",
        f"{current_user.name} assigned lead {l.name} ({l.lead_id}) to you",
        category="assignment",
        entity_type="lead",
        entity_id=l.lead_id
    )
    if old_rep:
        create_notification(
            db, old_rep.rep_id, "reassignment",
            f"Lead {l.lead_id} reassigned",
            f"Lead {l.name} ({l.lead_id}) has been reassigned from you to {new_rep.name} by {current_user.name}",
            category="reassignment",
            entity_type="lead",
            entity_id=l.lead_id
        )
    db.commit()
    return {"message": f"Lead reassigned to {new_rep.name}"}


# ============================================
# PAYMENT PLAN MANAGEMENT ENDPOINTS
# ============================================

PAYMENT_PLAN_ROLES = ["admin", "director", "cco", "coo"]


def _plan_to_dict(plan, db=None):
    """Convert payment plan to dict with versions if db provided."""
    d = {
        "id": str(plan.id), "plan_id": plan.plan_id,
        "project_id": str(plan.project_id), "name": plan.name,
        "description": plan.description, "is_default": plan.is_default,
        "is_locked": plan.is_locked, "locked_at": plan.locked_at.isoformat() if plan.locked_at else None,
        "lock_reason": plan.lock_reason, "status": plan.status,
        "created_at": plan.created_at.isoformat() if plan.created_at else None,
    }
    if plan.locked_by:
        d["locked_by"] = str(plan.locked_by)
    if db:
        versions = db.query(ProjectPaymentPlanVersion).filter(
            ProjectPaymentPlanVersion.plan_id == plan.id
        ).order_by(ProjectPaymentPlanVersion.version_number.desc()).all()
        d["versions"] = [{
            "id": str(v.id), "version_number": v.version_number,
            "is_active": v.is_active, "installments": v.installments,
            "num_installments": v.num_installments,
            "installment_cycle": v.installment_cycle, "notes": v.notes,
            "created_at": v.created_at.isoformat() if v.created_at else None,
        } for v in versions]
    return d


@app.get("/api/projects/{project_id}/payment-plans")
def list_payment_plans(
    project_id: str, status: str = "active",
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """List payment plans for a project."""
    proj = db.query(Project).filter(
        (Project.project_id == project_id) | (func.cast(Project.id, String) == project_id)
    ).first()
    if not proj:
        raise HTTPException(404, "Project not found")
    q = db.query(ProjectPaymentPlan).filter(ProjectPaymentPlan.project_id == proj.id)
    if status != "all":
        q = q.filter(ProjectPaymentPlan.status == status)
    plans = q.order_by(ProjectPaymentPlan.created_at.desc()).all()
    return [_plan_to_dict(p, db) for p in plans]


@app.post("/api/projects/{project_id}/payment-plans", status_code=201)
def create_payment_plan(
    project_id: str, data: dict,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(require_role(PAYMENT_PLAN_ROLES))
):
    """Create a new payment plan shell for a project."""
    proj = db.query(Project).filter(
        (Project.project_id == project_id) | (func.cast(Project.id, String) == project_id)
    ).first()
    if not proj:
        raise HTTPException(404, "Project not found")
    # Generate plan_id
    count = db.query(ProjectPaymentPlan).count()
    plan_id = f"PPL-{str(count + 1).zfill(4)}"
    while db.query(ProjectPaymentPlan).filter(ProjectPaymentPlan.plan_id == plan_id).first():
        count += 1
        plan_id = f"PPL-{str(count + 1).zfill(4)}"
    plan = ProjectPaymentPlan(
        plan_id=plan_id, project_id=proj.id,
        name=data.get("name", "Default Plan"),
        description=data.get("description"),
        is_default=data.get("is_default", False),
        created_by=current_user.id
    )
    db.add(plan)
    db.commit()
    db.refresh(plan)
    return _plan_to_dict(plan, db)


@app.get("/api/payment-plans/{plan_id}")
def get_payment_plan(
    plan_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Get payment plan detail with versions."""
    plan = db.query(ProjectPaymentPlan).filter(
        (ProjectPaymentPlan.plan_id == plan_id) | (func.cast(ProjectPaymentPlan.id, String) == plan_id)
    ).first()
    if not plan:
        raise HTTPException(404, "Payment plan not found")
    return _plan_to_dict(plan, db)


@app.put("/api/payment-plans/{plan_id}")
def update_payment_plan(
    plan_id: str, data: dict,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(require_role(PAYMENT_PLAN_ROLES))
):
    """Update mutable metadata. Reject business field changes if locked."""
    plan = db.query(ProjectPaymentPlan).filter(
        (ProjectPaymentPlan.plan_id == plan_id) | (func.cast(ProjectPaymentPlan.id, String) == plan_id)
    ).first()
    if not plan:
        raise HTTPException(404, "Payment plan not found")
    if plan.is_locked and any(k in data for k in ["is_default"]):
        raise HTTPException(409, detail="Payment plan is locked; unlock required before mutation")
    for k in ["name", "description", "is_default", "status"]:
        if k in data:
            setattr(plan, k, data[k])
    plan.updated_at = func.now()
    db.commit()
    db.refresh(plan)
    return _plan_to_dict(plan, db)


@app.delete("/api/payment-plans/{plan_id}")
def delete_payment_plan(
    plan_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(require_role(PAYMENT_PLAN_ROLES))
):
    """Archive or delete a payment plan."""
    plan = db.query(ProjectPaymentPlan).filter(
        (ProjectPaymentPlan.plan_id == plan_id) | (func.cast(ProjectPaymentPlan.id, String) == plan_id)
    ).first()
    if not plan:
        raise HTTPException(404, "Payment plan not found")
    # Check if referenced by transactions
    txn_ref = db.query(Transaction).filter(Transaction.payment_plan_version_id.in_(
        db.query(ProjectPaymentPlanVersion.id).filter(ProjectPaymentPlanVersion.plan_id == plan.id)
    )).first()
    if txn_ref:
        plan.status = "archived"
        db.commit()
        return {"message": "Plan archived (referenced by transactions)"}
    db.delete(plan)
    db.commit()
    return {"message": "Plan deleted"}


@app.post("/api/payment-plans/{plan_id}/versions", status_code=201)
def create_plan_version(
    plan_id: str, data: dict,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(require_role(PAYMENT_PLAN_ROLES))
):
    """Create a new version under a payment plan. Deny if plan is locked."""
    plan = db.query(ProjectPaymentPlan).filter(
        (ProjectPaymentPlan.plan_id == plan_id) | (func.cast(ProjectPaymentPlan.id, String) == plan_id)
    ).first()
    if not plan:
        raise HTTPException(404, "Payment plan not found")
    if plan.is_locked:
        raise HTTPException(409, detail="Payment plan is locked; unlock required before mutation")
    installments = data.get("installments", [])
    if not installments:
        raise HTTPException(400, detail="installments required")
    pct_sum = sum(i.get("percentage", 0) for i in installments)
    if abs(pct_sum - 100.0) > 0.01:
        raise HTTPException(400, detail="Installment percentages must sum to 100")
    max_ver = db.query(func.max(ProjectPaymentPlanVersion.version_number)).filter(
        ProjectPaymentPlanVersion.plan_id == plan.id
    ).scalar() or 0
    ver = ProjectPaymentPlanVersion(
        plan_id=plan.id, version_number=max_ver + 1,
        is_active=True, installments=installments,
        num_installments=data.get("num_installments", len(installments)),
        installment_cycle=data.get("installment_cycle", "bi-annual"),
        notes=data.get("notes"), created_by=current_user.id
    )
    # Deactivate siblings
    db.query(ProjectPaymentPlanVersion).filter(
        ProjectPaymentPlanVersion.plan_id == plan.id
    ).update({"is_active": False})
    db.add(ver)
    db.commit()
    db.refresh(ver)
    return {
        "id": str(ver.id), "version_number": ver.version_number,
        "is_active": ver.is_active, "installments": ver.installments,
        "num_installments": ver.num_installments,
        "percentage_total": round(pct_sum, 2)
    }


@app.get("/api/payment-plans/{plan_id}/versions")
def list_plan_versions(
    plan_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """List versions for a payment plan."""
    plan = db.query(ProjectPaymentPlan).filter(
        (ProjectPaymentPlan.plan_id == plan_id) | (func.cast(ProjectPaymentPlan.id, String) == plan_id)
    ).first()
    if not plan:
        raise HTTPException(404, "Payment plan not found")
    versions = db.query(ProjectPaymentPlanVersion).filter(
        ProjectPaymentPlanVersion.plan_id == plan.id
    ).order_by(ProjectPaymentPlanVersion.version_number.desc()).all()
    return [{
        "id": str(v.id), "version_number": v.version_number,
        "is_active": v.is_active, "installments": v.installments,
        "num_installments": v.num_installments,
        "installment_cycle": v.installment_cycle, "notes": v.notes,
        "created_at": v.created_at.isoformat() if v.created_at else None,
    } for v in versions]


@app.post("/api/payment-plans/{plan_id}/versions/{version_id}/activate")
def activate_plan_version(
    plan_id: str, version_id: str,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(require_role(PAYMENT_PLAN_ROLES))
):
    """Activate a specific version; deactivate siblings. Deny if locked."""
    plan = db.query(ProjectPaymentPlan).filter(
        (ProjectPaymentPlan.plan_id == plan_id) | (func.cast(ProjectPaymentPlan.id, String) == plan_id)
    ).first()
    if not plan:
        raise HTTPException(404, "Payment plan not found")
    if plan.is_locked:
        raise HTTPException(409, detail="Payment plan is locked; unlock required before mutation")
    ver = db.query(ProjectPaymentPlanVersion).filter(
        (func.cast(ProjectPaymentPlanVersion.id, String) == version_id),
        ProjectPaymentPlanVersion.plan_id == plan.id
    ).first()
    if not ver:
        raise HTTPException(404, "Version not found")
    db.query(ProjectPaymentPlanVersion).filter(
        ProjectPaymentPlanVersion.plan_id == plan.id
    ).update({"is_active": False})
    ver.is_active = True
    db.commit()
    return {"message": "Version activated", "active_version_id": str(ver.id)}


@app.post("/api/payment-plans/{plan_id}/lock")
def lock_payment_plan(
    plan_id: str, data: dict = {},
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(require_role(PAYMENT_PLAN_ROLES))
):
    """Lock a payment plan. Prevents version creation and business field mutations."""
    plan = db.query(ProjectPaymentPlan).filter(
        (ProjectPaymentPlan.plan_id == plan_id) | (func.cast(ProjectPaymentPlan.id, String) == plan_id)
    ).first()
    if not plan:
        raise HTTPException(404, "Payment plan not found")
    if plan.is_locked:
        raise HTTPException(409, detail="Payment plan is already locked")
    plan.is_locked = True
    plan.locked_by = current_user.id
    plan.locked_at = func.now()
    plan.lock_reason = data.get("reason", "")
    db.commit()
    db.refresh(plan)
    return _plan_to_dict(plan)


@app.post("/api/payment-plans/{plan_id}/unlock")
def unlock_payment_plan(
    plan_id: str, data: dict = {},
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(require_role(PAYMENT_PLAN_ROLES))
):
    """Unlock a payment plan. Requires reason."""
    plan = db.query(ProjectPaymentPlan).filter(
        (ProjectPaymentPlan.plan_id == plan_id) | (func.cast(ProjectPaymentPlan.id, String) == plan_id)
    ).first()
    if not plan:
        raise HTTPException(404, "Payment plan not found")
    if not plan.is_locked:
        raise HTTPException(409, detail="Payment plan is not locked")
    if not data.get("reason"):
        raise HTTPException(400, detail="Unlock reason is required")
    plan.is_locked = False
    plan.locked_by = None
    plan.locked_at = None
    plan.lock_reason = None
    db.commit()
    db.refresh(plan)
    return _plan_to_dict(plan)


# ============================================
# ANALYTICS ENGINE ENDPOINTS
# ============================================

@app.get("/api/analytics/campaign-metrics")
def analytics_campaign_metrics(
    campaign_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(require_role(["admin", "cco", "manager", "creator"]))
):
    """Comprehensive campaign analytics with funnel, source breakdown, and revenue attribution."""
    try:
        # Build base lead query with optional filters
        lead_q = db.query(Lead)
        if campaign_id:
            camp = db.query(Campaign).filter(
                (Campaign.campaign_id == campaign_id) | (Campaign.id == campaign_id)
            ).first()
            if camp:
                lead_q = lead_q.filter(Lead.campaign_id == camp.id)
        if start_date:
            lead_q = lead_q.filter(Lead.created_at >= start_date)
        if end_date:
            lead_q = lead_q.filter(Lead.created_at <= end_date)

        all_leads = lead_q.all()
        total = len(all_leads)

        # Overall metrics
        converted = sum(1 for l in all_leads if l.status == "converted")
        lost = sum(1 for l in all_leads if l.status == "lost")
        active = total - converted - lost
        conversion_rate = round((converted / total * 100), 1) if total > 0 else 0.0

        # Average days to conversion
        conv_days = []
        for l in all_leads:
            if l.status == "converted" and l.updated_at and l.created_at:
                delta = (l.updated_at - l.created_at).days
                conv_days.append(delta)
        avg_days = round(sum(conv_days) / len(conv_days), 1) if conv_days else 0.0

        # Funnel: count by pipeline_stage, ordered by PipelineStage.display_order
        stages_db = db.query(PipelineStage).order_by(PipelineStage.display_order).all()
        stage_order = {s.name: (s.display_order, s.color) for s in stages_db}
        stage_counts = {}
        for l in all_leads:
            ps = l.pipeline_stage or "New"
            stage_counts[ps] = stage_counts.get(ps, 0) + 1

        funnel = []
        prev_count = total
        for s in stages_db:
            count = stage_counts.pop(s.name, 0)
            pct_total = round(count / total * 100, 1) if total > 0 else 0.0
            pct_prev = round(count / prev_count * 100, 1) if prev_count > 0 else 0.0
            funnel.append({
                "stage": s.name, "count": count, "color": s.color,
                "pct_of_total": pct_total, "pct_of_previous": pct_prev
            })
            if count > 0:
                prev_count = count
        # Any stages not in DB config
        for stage_name, count in stage_counts.items():
            pct_total = round(count / total * 100, 1) if total > 0 else 0.0
            funnel.append({
                "stage": stage_name, "count": count, "color": "#6B7280",
                "pct_of_total": pct_total, "pct_of_previous": 0.0
            })

        # By source: group by campaign.source
        campaigns = db.query(Campaign).all()
        camp_map = {c.id: c for c in campaigns}
        source_data = {}
        for l in all_leads:
            c = camp_map.get(l.campaign_id)
            src = c.source if c else "unknown"
            if src not in source_data:
                source_data[src] = {"leads": 0, "converted": 0, "budget": 0.0}
            source_data[src]["leads"] += 1
            if l.status == "converted":
                source_data[src]["converted"] += 1
            if c and c.budget and src not in source_data.get("_budgeted", set()):
                source_data.setdefault("_camp_ids", set())
                if c.id not in source_data.get("_camp_ids", set()):
                    source_data[src]["budget"] += float(c.budget or 0)
                    source_data.setdefault("_camp_ids", set()).add(c.id)

        source_data.pop("_camp_ids", None)
        by_source = []
        for src, d in source_data.items():
            cr = round(d["converted"] / d["leads"] * 100, 1) if d["leads"] > 0 else 0.0
            cpl = round(d["budget"] / d["leads"], 0) if d["leads"] > 0 and d["budget"] > 0 else 0.0
            by_source.append({
                "source": src, "leads": d["leads"], "converted": d["converted"],
                "conversion_rate": cr, "budget": d["budget"], "cost_per_lead": cpl
            })

        # By campaign: per-campaign breakdown with revenue attribution
        # Pre-fetch revenue: lead -> converted_customer_id -> transactions
        converted_cust_ids = [l.converted_customer_id for l in all_leads if l.converted_customer_id]
        revenue_by_customer = {}
        if converted_cust_ids:
            txn_rows = db.query(
                Transaction.customer_id,
                func.sum(Transaction.total_value).label("revenue"),
                func.count(Transaction.id).label("deals")
            ).filter(Transaction.customer_id.in_(converted_cust_ids)).group_by(Transaction.customer_id).all()
            for row in txn_rows:
                revenue_by_customer[row.customer_id] = {"revenue": float(row.revenue or 0), "deals": int(row.deals)}

        # Build lead-to-campaign mapping for revenue rollup
        camp_revenue = {}
        for l in all_leads:
            if l.converted_customer_id and l.converted_customer_id in revenue_by_customer:
                cid = l.campaign_id
                if cid not in camp_revenue:
                    camp_revenue[cid] = {"revenue": 0.0, "deals": 0}
                camp_revenue[cid]["revenue"] += revenue_by_customer[l.converted_customer_id]["revenue"]
                camp_revenue[cid]["deals"] += revenue_by_customer[l.converted_customer_id]["deals"]

        camp_leads = {}
        for l in all_leads:
            cid = l.campaign_id
            if cid not in camp_leads:
                camp_leads[cid] = {"total": 0, "converted": 0, "lost": 0}
            camp_leads[cid]["total"] += 1
            if l.status == "converted":
                camp_leads[cid]["converted"] += 1
            elif l.status == "lost":
                camp_leads[cid]["lost"] += 1

        by_campaign = []
        for c in campaigns:
            cl = camp_leads.get(c.id, {"total": 0, "converted": 0, "lost": 0})
            if cl["total"] == 0 and not campaign_id:
                continue
            cr_val = round(cl["converted"] / cl["total"] * 100, 1) if cl["total"] > 0 else 0.0
            rev = camp_revenue.get(c.id, {"revenue": 0.0, "deals": 0})
            by_campaign.append({
                "campaign_id": c.campaign_id, "name": c.name, "source": c.source or "unknown",
                "status": c.status, "leads": cl["total"], "converted": cl["converted"],
                "lost": cl["lost"], "conversion_rate": cr_val,
                "budget": float(c.budget or 0),
                "revenue_generated": rev["revenue"], "deals": rev["deals"],
                "roi": round((rev["revenue"] - float(c.budget or 0)) / float(c.budget or 1) * 100, 1) if c.budget and float(c.budget) > 0 else 0.0
            })

        # Total revenue attribution
        total_revenue = sum(c["revenue_generated"] for c in by_campaign)
        total_budget = sum(c["budget"] for c in by_campaign)

        return {
            "overall": {
                "total_leads": total, "converted": converted, "lost": lost,
                "active": active, "conversion_rate": conversion_rate,
                "avg_days_to_conversion": avg_days,
                "total_revenue_attributed": total_revenue,
                "total_budget": total_budget,
                "overall_roi": round((total_revenue - total_budget) / total_budget * 100, 1) if total_budget > 0 else 0.0
            },
            "funnel": funnel,
            "by_source": sorted(by_source, key=lambda x: x["leads"], reverse=True),
            "by_campaign": sorted(by_campaign, key=lambda x: x["leads"], reverse=True)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Analytics error: {str(e)}")


@app.get("/api/analytics/rep-performance")
def analytics_rep_performance(
    campaign_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(require_role(["admin", "cco", "manager"]))
):
    """Rep-wise performance: lead aging, interactions, conversions, follow-ups."""
    try:
        # Build base lead query
        lead_q = db.query(Lead)
        if campaign_id:
            camp = db.query(Campaign).filter(
                (Campaign.campaign_id == campaign_id) | (Campaign.id == campaign_id)
            ).first()
            if camp:
                lead_q = lead_q.filter(Lead.campaign_id == camp.id)

        all_leads = lead_q.all()
        today = date.today()

        # Pre-fetch all reps
        reps = db.query(CompanyRep).filter(CompanyRep.status == "active").all()
        rep_map = {r.id: r for r in reps}

        # Group leads by assigned_rep_id
        rep_leads = {}
        for l in all_leads:
            rid = l.assigned_rep_id
            if rid not in rep_leads:
                rep_leads[rid] = []
            rep_leads[rid].append(l)

        # Pre-fetch interactions grouped by company_rep_id (single query)
        interaction_q = db.query(
            Interaction.company_rep_id,
            Interaction.interaction_type,
            func.count(Interaction.id).label("cnt")
        )
        if campaign_id and 'camp' in dir() and camp:
            lead_ids = [l.id for l in all_leads]
            if lead_ids:
                interaction_q = interaction_q.filter(Interaction.lead_id.in_(lead_ids))
        interaction_q = interaction_q.group_by(Interaction.company_rep_id, Interaction.interaction_type)
        interaction_rows = interaction_q.all()

        int_map = {}
        for row in interaction_rows:
            rid = row.company_rep_id
            if rid not in int_map:
                int_map[rid] = {"total": 0, "call": 0, "message": 0, "whatsapp": 0}
            int_map[rid][row.interaction_type] = int_map[rid].get(row.interaction_type, 0) + int(row.cnt)
            int_map[rid]["total"] += int(row.cnt)

        # Pre-fetch pending followups by rep (single query)
        followup_q = db.query(
            Interaction.company_rep_id,
            func.count(Interaction.id).label("cnt")
        ).filter(
            Interaction.next_follow_up <= today,
            Interaction.next_follow_up.isnot(None)
        )
        if campaign_id and 'camp' in dir() and camp:
            lead_ids = [l.id for l in all_leads]
            if lead_ids:
                followup_q = followup_q.filter(Interaction.lead_id.in_(lead_ids))
        followup_rows = followup_q.group_by(Interaction.company_rep_id).all()
        followup_map = {row.company_rep_id: int(row.cnt) for row in followup_rows}

        # Build per-rep metrics
        result_reps = []
        total_not_attempted = 0
        total_pending = 0
        total_converted_all = 0
        total_assigned_all = 0

        all_rep_ids = set(rep_leads.keys()) | set(r.id for r in reps)
        for rid in all_rep_ids:
            rep = rep_map.get(rid)
            if not rep:
                continue
            leads_list = rep_leads.get(rid, [])
            total_assigned = len(leads_list)
            if total_assigned == 0 and rid not in rep_leads:
                continue

            # Not attempted buckets: leads with last_contacted_at IS NULL and status not terminal
            not_attempted = {"0_7d": 0, "7_14d": 0, "14d_plus": 0, "total": 0}
            converted_count = 0
            lost_count = 0
            response_days = []

            for l in leads_list:
                if l.status == "converted":
                    converted_count += 1
                elif l.status == "lost":
                    lost_count += 1

                if l.last_contacted_at is None and l.status not in ("converted", "lost"):
                    days_since = (today - l.created_at.date()).days if l.created_at else 0
                    not_attempted["total"] += 1
                    if days_since <= 7:
                        not_attempted["0_7d"] += 1
                    elif days_since <= 14:
                        not_attempted["7_14d"] += 1
                    else:
                        not_attempted["14d_plus"] += 1

                # Average response time (days from creation to first contact)
                if l.last_contacted_at and l.created_at:
                    rd = (l.last_contacted_at - l.created_at).total_seconds() / 86400
                    if rd >= 0:
                        response_days.append(rd)

            interactions = int_map.get(rid, {"total": 0, "call": 0, "message": 0, "whatsapp": 0})
            pending = followup_map.get(rid, 0)
            cr = round(converted_count / total_assigned * 100, 1) if total_assigned > 0 else 0.0
            avg_resp = round(sum(response_days) / len(response_days), 1) if response_days else None

            result_reps.append({
                "rep_id": rep.rep_id, "name": rep.name,
                "total_assigned": total_assigned,
                "not_attempted": not_attempted,
                "interactions": interactions,
                "pending_followups": pending,
                "converted": converted_count, "lost": lost_count,
                "active": total_assigned - converted_count - lost_count,
                "conversion_rate": cr,
                "avg_response_days": avg_resp
            })

            total_not_attempted += not_attempted["total"]
            total_pending += pending
            total_converted_all += converted_count
            total_assigned_all += total_assigned

        result_reps.sort(key=lambda x: x["total_assigned"], reverse=True)

        return {
            "reps": result_reps,
            "totals": {
                "total_assigned": total_assigned_all,
                "total_not_attempted": total_not_attempted,
                "total_pending_followups": total_pending,
                "total_converted": total_converted_all,
                "overall_conversion_rate": round(total_converted_all / total_assigned_all * 100, 1) if total_assigned_all > 0 else 0.0
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Rep analytics error: {str(e)}")


@app.get("/api/analytics/leads/drilldown")
def analytics_leads_drilldown(
    stage: Optional[str] = None,
    aging_bucket: Optional[str] = None,
    rep_id: Optional[str] = None,
    campaign_id: Optional[str] = None,
    source: Optional[str] = None,
    pending_followups: Optional[bool] = None,
    show_interactions: Optional[bool] = None,
    export_format: Optional[str] = "json",
    db: Session = Depends(get_db),
    current_user: CompanyRep = Depends(get_current_user)
):
    """Analytics drilldown: filtered lead list with optional CSV/Excel export.
    Supports funnel stage click, aging bucket click, pending follow-ups, interactions view, and role-scoped filtering."""
    try:
        isolation = get_rep_isolation_filter(current_user, db)

        lead_q = db.query(Lead)

        # Role-based isolation
        if isolation["isolated"]:
            lead_q = lead_q.filter(Lead.assigned_rep_id.in_(isolation["team_rep_uuids"]))

        # Rep filter (admin/cco/director/coo can filter any rep; manager within team)
        if rep_id:
            rep_obj = db.query(CompanyRep).filter(
                (CompanyRep.rep_id == rep_id) | (func.cast(CompanyRep.id, String) == rep_id)
            ).first()
            if rep_obj:
                if not isolation["isolated"]:
                    lead_q = lead_q.filter(Lead.assigned_rep_id == rep_obj.id)
                elif rep_obj.id in isolation["team_rep_uuids"]:
                    lead_q = lead_q.filter(Lead.assigned_rep_id == rep_obj.id)

        # Campaign filter
        if campaign_id:
            camp = db.query(Campaign).filter(
                (Campaign.campaign_id == campaign_id) | (func.cast(Campaign.id, String) == campaign_id)
            ).first()
            if camp:
                lead_q = lead_q.filter(Lead.campaign_id == camp.id)

        # Pipeline stage filter
        if stage:
            lead_q = lead_q.filter(Lead.pipeline_stage == stage)

        # Source filter (via campaign.source)
        if source:
            source_camp_ids = [c.id for c in db.query(Campaign).filter(Campaign.source == source).all()]
            if source_camp_ids:
                lead_q = lead_q.filter(Lead.campaign_id.in_(source_camp_ids))
            else:
                lead_q = lead_q.filter(False)

        all_leads = lead_q.all()

        # Aging bucket filter (computed post-query: not-attempted leads only)
        if aging_bucket:
            today = date.today()
            filtered = []
            for l in all_leads:
                if l.last_contacted_at is not None or l.status in ("converted", "lost"):
                    continue
                days_since = (today - l.created_at.date()).days if l.created_at else 0
                if aging_bucket == "0_7d" and days_since <= 7:
                    filtered.append(l)
                elif aging_bucket == "7_14d" and 7 < days_since <= 14:
                    filtered.append(l)
                elif aging_bucket == "14d_plus" and days_since > 14:
                    filtered.append(l)
            all_leads = filtered

        # Pending follow-ups filter: leads that have interactions with overdue follow-ups
        if pending_followups:
            today = date.today()
            pending_interaction_lead_ids = set(
                row[0] for row in db.query(Interaction.lead_id).filter(
                    Interaction.next_follow_up.isnot(None),
                    Interaction.next_follow_up <= today,
                    Interaction.lead_id.isnot(None)
                ).all()
            )
            all_leads = [l for l in all_leads if l.id in pending_interaction_lead_ids]

        # Show interactions mode: return interaction list instead of leads
        if show_interactions:
            rep_obj = None
            if rep_id:
                rep_obj = db.query(CompanyRep).filter(
                    (CompanyRep.rep_id == rep_id) | (func.cast(CompanyRep.id, String) == rep_id)
                ).first()
            int_q = db.query(Interaction)
            if rep_obj:
                int_q = int_q.filter(Interaction.company_rep_id == rep_obj.id)
            iso = get_rep_isolation_filter(current_user, db)
            if iso["isolated"]:
                int_q = int_q.filter(Interaction.company_rep_id.in_(iso["team_rep_uuids"]))
            ints = int_q.order_by(Interaction.created_at.desc()).limit(200).all()
            # Pre-fetch
            int_rep_ids = set(x.company_rep_id for x in ints if x.company_rep_id)
            int_lead_ids = set(x.lead_id for x in ints if x.lead_id)
            int_cust_ids = set(x.customer_id for x in ints if x.customer_id)
            int_brk_ids = set(x.broker_id for x in ints if x.broker_id)
            int_reps = {r.id: r for r in db.query(CompanyRep).filter(CompanyRep.id.in_(int_rep_ids)).all()} if int_rep_ids else {}
            int_leads = {l.id: l for l in db.query(Lead).filter(Lead.id.in_(int_lead_ids)).all()} if int_lead_ids else {}
            int_custs = {c.id: c for c in db.query(Customer).filter(Customer.id.in_(int_cust_ids)).all()} if int_cust_ids else {}
            int_brks = {b.id: b for b in db.query(Broker).filter(Broker.id.in_(int_brk_ids)).all()} if int_brk_ids else {}
            int_rows = []
            for x in ints:
                r = int_reps.get(x.company_rep_id)
                ld = int_leads.get(x.lead_id)
                cu = int_custs.get(x.customer_id)
                br = int_brks.get(x.broker_id)
                contact_name = ld.name if ld else cu.name if cu else br.name if br else "-"
                contact_id = ld.lead_id if ld else cu.customer_id if cu else br.broker_id if br else ""
                contact_mobile = ld.mobile if ld else cu.mobile if cu else br.mobile if br else ""
                int_rows.append({
                    "interaction_id": x.interaction_id,
                    "rep_name": r.name if r else "",
                    "contact_name": contact_name,
                    "contact_id": contact_id,
                    "contact_mobile": contact_mobile,
                    "interaction_type": x.interaction_type,
                    "status": x.status or "",
                    "notes": x.notes or "",
                    "next_follow_up": str(x.next_follow_up) if x.next_follow_up else "",
                    "created_at": str(x.created_at) if x.created_at else "",
                })
            return {"leads": int_rows, "total": len(int_rows), "mode": "interactions"}

        # Pre-fetch rep and campaign names
        rep_ids = set(l.assigned_rep_id for l in all_leads if l.assigned_rep_id)
        camp_ids = set(l.campaign_id for l in all_leads if l.campaign_id)
        rep_name_map = {}
        camp_name_map = {}
        if rep_ids:
            for r in db.query(CompanyRep).filter(CompanyRep.id.in_(rep_ids)).all():
                rep_name_map[r.id] = {"name": r.name, "rep_id": r.rep_id}
        if camp_ids:
            for c in db.query(Campaign).filter(Campaign.id.in_(camp_ids)).all():
                camp_name_map[c.id] = {"name": c.name, "campaign_id": c.campaign_id}

        today = date.today()
        rows = []
        for l in all_leads:
            rep_info = rep_name_map.get(l.assigned_rep_id, {})
            camp_info = camp_name_map.get(l.campaign_id, {})
            days_since = (today - l.created_at.date()).days if l.created_at else None
            rows.append({
                "lead_id": l.lead_id,
                "name": l.name,
                "mobile": l.mobile or "",
                "email": l.email or "",
                "pipeline_stage": l.pipeline_stage or "",
                "status": l.status or "",
                "source": l.source or "",
                "temperature": l.temperature or "",
                "assigned_rep": rep_info.get("name", "Unassigned"),
                "assigned_rep_id": rep_info.get("rep_id", ""),
                "campaign": camp_info.get("name", ""),
                "campaign_id": camp_info.get("campaign_id", ""),
                "days_since_creation": days_since,
                "last_contacted": l.last_contacted_at.isoformat() if l.last_contacted_at else "",
                "created_at": l.created_at.isoformat() if l.created_at else "",
                "city": l.city or "",
                "area": l.area or "",
            })

        # CSV export
        if export_format == "csv":
            import csv as csv_mod
            from io import StringIO
            output = StringIO()
            if rows:
                writer = csv_mod.DictWriter(output, fieldnames=list(rows[0].keys()))
                writer.writeheader()
                writer.writerows(rows)
            content = output.getvalue()
            return StreamingResponse(
                iter([content]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=leads_drilldown.csv"}
            )

        # Excel export
        if export_format == "excel":
            import openpyxl
            from io import BytesIO
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Leads Drilldown"
            if rows:
                headers = list(rows[0].keys())
                ws.append(headers)
                for row in rows:
                    ws.append([row.get(h, "") for h in headers])
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=leads_drilldown.xlsx"}
            )

        return {
            "leads": rows,
            "total": len(rows),
            "filters": {
                "stage": stage, "aging_bucket": aging_bucket,
                "rep_id": rep_id, "campaign_id": campaign_id, "source": source
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Drilldown error: {str(e)}")


# ============================================
# TASK MANAGEMENT ENDPOINTS
# ============================================
from app.services.task_service import task_service as _task_service
from app.services.voice_query_service import voice_query_service as _voice_service

VALID_TASK_TYPES = {"general", "follow_up", "documentation", "meeting", "inventory",
    "transaction", "customer", "report", "approval", "sales", "collection",
    "site_visit", "legal", "recovery", "reconciliation"}
VALID_PRIORITIES = {"low", "medium", "high", "urgent"}

def _build_rep_name_cache(db, tasks):
    """Batch-fetch rep names to avoid N+1 queries in _task_to_dict loops."""
    ids = set()
    for t in tasks:
        if t.assignee_id: ids.add(t.assignee_id)
        if t.created_by: ids.add(t.created_by)
        for cid in (t.collaborator_ids or []):
            try: ids.add(uuid.UUID(cid) if isinstance(cid, str) else cid)
            except (ValueError, AttributeError): pass
    if not ids:
        return {}
    reps = db.query(CompanyRep.id, CompanyRep.name).filter(CompanyRep.id.in_(list(ids))).all()
    return {r.id: r.name for r in reps}

def _task_to_dict(task, db, name_cache=None):
    """Convert Task model to dict with assignee/creator/collaborator names.
    Pass name_cache (from _build_rep_name_cache) to avoid N+1 queries in loops."""
    if name_cache is not None:
        assignee_name = name_cache.get(task.assignee_id) if task.assignee_id else None
        creator_name = name_cache.get(task.created_by) if task.created_by else None
    else:
        assignee_name = None
        creator_name = None
        if task.assignee_id:
            assignee = db.query(CompanyRep).filter(CompanyRep.id == task.assignee_id).first()
            assignee_name = assignee.name if assignee else None
        if task.created_by:
            creator = db.query(CompanyRep).filter(CompanyRep.id == task.created_by).first()
            creator_name = creator.name if creator else None

    # Resolve collaborator names
    collaborator_ids_raw = task.collaborator_ids or []
    collaborators = []
    for cid in collaborator_ids_raw:
        cid_str = str(cid)
        try:
            cid_uuid = uuid.UUID(cid_str)
        except (ValueError, AttributeError):
            continue
        if name_cache is not None:
            cname = name_cache.get(cid_uuid, "Unknown")
        else:
            rep = db.query(CompanyRep).filter(CompanyRep.id == cid_uuid).first()
            cname = rep.name if rep else "Unknown"
        collaborators.append({"id": cid_str, "name": cname})

    return {
        "id": str(task.id),
        "task_id": task.task_id,
        "title": task.title,
        "description": task.description,
        "task_type": task.task_type,
        "department": task.department,
        "priority": task.priority,
        "status": task.status,
        "assignee_id": str(task.assignee_id) if task.assignee_id else None,
        "assignee_name": assignee_name,
        "created_by": str(task.created_by) if task.created_by else None,
        "creator_name": creator_name,
        "delegated_by": str(task.delegated_by) if task.delegated_by else None,
        "parent_task_id": str(task.parent_task_id) if task.parent_task_id else None,
        "due_date": task.due_date.isoformat() if task.due_date else None,
        "completed_at": task.completed_at.isoformat() if task.completed_at else None,
        "completion_notes": task.completion_notes,
        "pending_assignment": task.pending_assignment,
        "original_assignee_text": task.original_assignee_text,
        "linked_inventory_id": str(task.linked_inventory_id) if task.linked_inventory_id else None,
        "linked_transaction_id": str(task.linked_transaction_id) if task.linked_transaction_id else None,
        "linked_customer_id": str(task.linked_customer_id) if task.linked_customer_id else None,
        "linked_project_id": str(task.linked_project_id) if task.linked_project_id else None,
        "collaborator_ids": [str(c) for c in collaborator_ids_raw],
        "collaborators": collaborators,
        "created_at": task.created_at.isoformat() if task.created_at else None,
        "updated_at": task.updated_at.isoformat() if task.updated_at else None,
    }

def _is_valid_uuid(val):
    try:
        uuid.UUID(str(val))
        return True
    except (ValueError, AttributeError):
        return False

def _check_task_access(task, current_user):
    """Verify user has access to a task. Admin/CCO see all, others see own/collaborated."""
    if current_user.role in ("admin", "cco"):
        return True
    if task.assignee_id == current_user.id or task.created_by == current_user.id:
        return True
    return str(current_user.id) in [str(c) for c in (task.collaborator_ids or [])]

@app.post("/api/tasks")
async def create_task(
    title: str = Form(...),
    description: Optional[str] = Form(None),
    task_type: str = Form("general", alias="type"),
    priority: str = Form("medium"),
    assignee_id: Optional[str] = Form(None, alias="assigned_to"),
    due_date: Optional[str] = Form(None),
    department: Optional[str] = Form(None),
    crm_entity_type: Optional[str] = Form(None),
    crm_entity_id: Optional[str] = Form(None),
    collaborator_ids: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    if current_user.role == "viewer":
        raise HTTPException(status_code=403, detail="Viewers cannot create tasks")
    # Input validation (#13, #14)
    if len(title) > 500:
        raise HTTPException(status_code=400, detail="Title too long (max 500 chars)")
    if description and len(description) > 5000:
        raise HTTPException(status_code=400, detail="Description too long (max 5000 chars)")
    task_type_lower = task_type.lower()
    priority_lower = priority.lower()
    if task_type_lower not in VALID_TASK_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid task type. Valid: {', '.join(sorted(VALID_TASK_TYPES))}")
    if priority_lower not in VALID_PRIORITIES:
        raise HTTPException(status_code=400, detail=f"Invalid priority. Valid: {', '.join(sorted(VALID_PRIORITIES))}")
    try:
        parsed_due = date.fromisoformat(due_date) if due_date else None
        parsed_assignee = uuid.UUID(assignee_id) if assignee_id else None
        # Parse collaborators (JSON array string)
        parsed_collabs = []
        if collaborator_ids:
            import json as json_mod
            try:
                parsed_collabs = json_mod.loads(collaborator_ids)
            except (json_mod.JSONDecodeError, TypeError):
                parsed_collabs = [c.strip() for c in collaborator_ids.split(',') if c.strip()]
            parsed_collabs = [str(c) for c in parsed_collabs if _is_valid_uuid(str(c))]
        # Map generic crm_entity_type/crm_entity_id to specific linked fields
        linked_inventory_id = crm_entity_id if crm_entity_type == "inventory" else None
        linked_transaction_id = crm_entity_id if crm_entity_type == "transaction" else None
        linked_customer_id = crm_entity_id if crm_entity_type in ("customer", "lead") else None
        linked_project_id = crm_entity_id if crm_entity_type == "project" else None
        task = _task_service.create_task(
            db, current_user.id, title, description, task_type_lower, priority_lower,
            parsed_assignee, parsed_due, department,
            linked_inventory_id, linked_transaction_id,
            linked_customer_id, linked_project_id
        )
        if parsed_collabs:
            task.collaborator_ids = parsed_collabs
            from sqlalchemy.orm.attributes import flag_modified
            flag_modified(task, 'collaborator_ids')
            db.commit()
            db.refresh(task)
        return _task_to_dict(task, db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/tasks/from-text")
async def create_task_from_text(
    text: str = Form(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    if current_user.role == "viewer":
        raise HTTPException(status_code=403, detail="Viewers cannot create tasks")
    if len(text) > 2000:
        raise HTTPException(status_code=400, detail="Text too long (max 2000 chars)")
    # Rate limit voice-driven task creation same as voice queries
    user_key = current_user.rep_id or str(current_user.id)
    if not voice_limiter.check(user_key, VOICE_RATE_LIMIT, VOICE_RATE_WINDOW):
        raise HTTPException(status_code=429, detail="Rate limit exceeded for voice task creation")
    try:
        task = _task_service.create_task_from_text(db, text, current_user.id)
        return _task_to_dict(task, db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/tasks")
async def list_tasks(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    department: Optional[str] = None,
    task_type: Optional[str] = None,
    search: Optional[str] = None,
    assignee_id: Optional[str] = None,
    skip: int = 0, limit: int = 50,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    role = current_user.role
    user_id = current_user.id

    # Fix #6: Only admin/cco can filter by arbitrary assignee_id
    if assignee_id and role in ("admin", "cco"):
        tasks = _task_service.get_tasks(db, user_id=assignee_id, role=role,
                                         status=status, priority=priority, department=department,
                                         task_type=task_type, search=search, assignee_only=True,
                                         limit=limit, offset=skip, current_rep_id=current_user.rep_id)
    else:
        tasks = _task_service.get_tasks(db, user_id=user_id, role=role,
                                         status=status, priority=priority, department=department,
                                         task_type=task_type, search=search, limit=limit, offset=skip,
                                         current_rep_id=current_user.rep_id)
    # Fix #7: Batch-fetch rep names to avoid N+1 queries
    name_cache = _build_rep_name_cache(db, tasks)
    return [_task_to_dict(t, db, name_cache) for t in tasks]

@app.get("/api/tasks/my")
async def my_tasks(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    tasks = _task_service.get_my_tasks(db, current_user.id)
    name_cache = _build_rep_name_cache(db, tasks)
    return [_task_to_dict(t, db, name_cache) for t in tasks]

@app.get("/api/tasks/reports/summary")
async def task_summary(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    return _task_service.get_task_summary(db, current_user.id, current_user.role, current_rep_id=current_user.rep_id)

@app.get("/api/tasks/executive-summary")
async def executive_summary(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """C-suite executive summary — tasks grouped by direct reports with stats and recent updates."""
    return _task_service.get_executive_summary(db, current_rep_id=current_user.rep_id)

@app.get("/api/tasks/departments/config")
async def departments_config(
    current_user = Depends(get_current_user)
):
    return _task_service.get_departments_config()

@app.post("/api/tasks/daily-report")
async def generate_daily_report(
    rep_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Generate daily task summary reports. Admin/CCO only."""
    if current_user.role not in ("admin", "cco"):
        raise HTTPException(status_code=403, detail="Admin/CCO only")
    try:
        results = _task_service.generate_daily_reports(db, target_rep_id=rep_id)
        # Also generate org report when generating for all users (no specific rep)
        org_result = None
        if not rep_id:
            try:
                org_result = _task_service.generate_org_report(db)
            except Exception as e:
                logger.error(f"Org report generation failed: {e}", exc_info=True)
        response = {"message": f"Generated {len(results)} report(s)", "reports": results}
        if org_result:
            response["org_report"] = org_result
        return response
    except Exception as e:
        logger.error(f"Daily report generation failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/tasks/daily-report/org")
async def get_org_daily_report(
    report_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get the organizational task overview report. Admin/CCO only."""
    from fastapi.responses import HTMLResponse
    if current_user.role not in ("admin", "cco"):
        raise HTTPException(403, "Admin/CCO only")
    target_date = date.fromisoformat(report_date) if report_date else date.today()
    file_path = MEDIA_ROOT / "reports" / "daily" / target_date.isoformat() / f"ORG_{target_date.isoformat()}.html"
    if not file_path.exists():
        raise HTTPException(404, f"No org report found for {target_date}")
    return HTMLResponse(content=file_path.read_text(encoding='utf-8'))

@app.get("/api/tasks/daily-report/{rep_id}")
async def get_daily_report(
    rep_id: str,
    report_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get the latest daily report for a rep. Returns the HTML file."""
    from fastapi.responses import HTMLResponse
    # Admin/CCO can view any, others can only view their own
    if current_user.role not in ("admin", "cco") and current_user.rep_id != rep_id:
        raise HTTPException(403, "Access denied")

    target_date = date.fromisoformat(report_date) if report_date else date.today()
    file_path = MEDIA_ROOT / "reports" / "daily" / target_date.isoformat() / f"{rep_id}_{target_date.isoformat()}.html"

    if not file_path.exists():
        raise HTTPException(404, f"No report found for {rep_id} on {target_date}")

    return HTMLResponse(content=file_path.read_text(encoding='utf-8'))

@app.get("/api/tasks/status-options/{task_type_name}")
async def status_options(
    task_type_name: str,
    current_user = Depends(get_current_user)
):
    return {"task_type": task_type_name, "statuses": _task_service.get_valid_statuses(task_type_name)}

@app.get("/api/tasks/{task_id}")
async def get_task(
    task_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    task = _task_service._find_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    # Fix #5: IDOR check
    if not _check_task_access(task, current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    result = _task_to_dict(task, db)
    # Include comments and activities (Fix #8: batch-fetch names)
    comments = _task_service.get_comments(db, task_id)
    activities = _task_service.get_activities(db, task_id)
    # Batch-fetch all author/actor names
    people_ids = set()
    for c in comments:
        if c.author_id: people_ids.add(c.author_id)
    for a in activities:
        if a.actor_id: people_ids.add(a.actor_id)
    people_names = {}
    if people_ids:
        reps = db.query(CompanyRep.id, CompanyRep.name).filter(CompanyRep.id.in_(list(people_ids))).all()
        people_names = {r.id: r.name for r in reps}
    result["comments"] = [{
        "id": str(c.id), "content": c.content,
        "author_id": str(c.author_id),
        "author_name": people_names.get(c.author_id, "Unknown"),
        "created_at": c.created_at.isoformat() if c.created_at else None,
    } for c in comments]
    result["activities"] = [{
        "id": str(a.id), "action": a.action,
        "old_value": a.old_value, "new_value": a.new_value,
        "actor_id": str(a.actor_id),
        "actor_name": people_names.get(a.actor_id, "Unknown"),
        "created_at": a.created_at.isoformat() if a.created_at else None,
    } for a in activities]
    # Include subtasks with micro-tasks
    subtasks = db.query(Task).filter(Task.parent_task_id == task.id).order_by(Task.created_at).all()
    sub_cache = _build_rep_name_cache(db, subtasks)
    # Batch-load micro-tasks for all subtasks
    sub_ids = [s.id for s in subtasks]
    all_micro_tasks = []
    if sub_ids:
        all_micro_tasks = db.query(MicroTask).filter(MicroTask.task_id.in_(sub_ids)).order_by(MicroTask.sort_order, MicroTask.created_at).all()
    # Build people name cache for micro-tasks
    mt_people_ids = set()
    for mt in all_micro_tasks:
        if mt.assignee_id:
            mt_people_ids.add(mt.assignee_id)
        if mt.created_by:
            mt_people_ids.add(mt.created_by)
        if mt.completed_by:
            mt_people_ids.add(mt.completed_by)
    mt_name_cache = {}
    if mt_people_ids:
        mt_reps = db.query(CompanyRep.id, CompanyRep.name).filter(CompanyRep.id.in_(list(mt_people_ids))).all()
        mt_name_cache = {r.id: r.name for r in mt_reps}
    # Group micro-tasks by subtask
    mt_by_sub = {}
    for mt in all_micro_tasks:
        mt_by_sub.setdefault(mt.task_id, []).append(mt)
    # Batch count comments per subtask and per micro-task
    sub_comment_counts = {}
    if sub_ids:
        counts = db.query(TaskComment.task_id, func.count(TaskComment.id)).filter(
            TaskComment.task_id.in_(sub_ids), TaskComment.micro_task_id == None
        ).group_by(TaskComment.task_id).all()
        sub_comment_counts = {tid: cnt for tid, cnt in counts}
    # Latest activity per subtask => last modified metadata
    sub_last_activity = {}
    if sub_ids:
        sub_activities = db.query(TaskActivity).filter(
            TaskActivity.task_id.in_(sub_ids)
        ).order_by(TaskActivity.created_at.desc()).all()
        for a in sub_activities:
            if a.task_id not in sub_last_activity:
                sub_last_activity[a.task_id] = a
    sub_last_actor_names = {}
    sub_actor_ids = set(a.actor_id for a in sub_last_activity.values() if a and a.actor_id)
    if sub_actor_ids:
        sub_reps = db.query(CompanyRep.id, CompanyRep.name).filter(CompanyRep.id.in_(list(sub_actor_ids))).all()
        sub_last_actor_names = {r.id: r.name for r in sub_reps}
    enriched_subtasks = []
    for s in subtasks:
        sd = _task_to_dict(s, db, sub_cache)
        s_mts = mt_by_sub.get(s.id, [])
        sd["micro_tasks"] = [_micro_task_to_dict(mt, db, mt_name_cache) for mt in s_mts]
        sd["micro_task_progress"] = {
            "total": len(s_mts),
            "completed": sum(1 for mt in s_mts if mt.is_completed),
        }
        sd["comment_count"] = sub_comment_counts.get(s.id, 0)
        sd["is_completed"] = s.status == "completed"
        last_activity = sub_last_activity.get(s.id)
        sd["last_modified_at"] = (
            last_activity.created_at.isoformat() if last_activity and last_activity.created_at
            else (s.updated_at.isoformat() if s.updated_at else None)
        )
        sd["last_modified_by_name"] = (
            sub_last_actor_names.get(last_activity.actor_id) if last_activity and last_activity.actor_id
            else sd.get("creator_name")
        )
        enriched_subtasks.append(sd)
    result["subtasks"] = enriched_subtasks
    return result

@app.put("/api/tasks/{task_id}")
async def update_task(
    task_id: str,
    title: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    priority: Optional[str] = Form(None),
    status_val: Optional[str] = Form(None, alias="status"),
    department: Optional[str] = Form(None),
    assignee_id: Optional[str] = Form(None),
    due_date: Optional[str] = Form(None),
    collaborator_ids: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    task = _task_service._find_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if not _check_task_access(task, current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    # Input validation
    if title and len(title) > 500:
        raise HTTPException(status_code=400, detail="Title too long (max 500 chars)")
    if description is not None and len(description) > 5000:
        raise HTTPException(status_code=400, detail="Description too long (max 5000 chars)")
    if priority and priority.lower() not in VALID_PRIORITIES:
        raise HTTPException(status_code=400, detail=f"Invalid priority. Valid: {', '.join(sorted(VALID_PRIORITIES))}")
    changes = []
    if title:
        task.title = title
        changes.append(f"title updated")
    if description is not None:
        task.description = description
        changes.append("description updated")
    if priority:
        old_pri = task.priority
        task.priority = priority
        changes.append(f"priority: {old_pri} -> {priority}")
    if department:
        task.department = department
        changes.append(f"department: {department}")
    if assignee_id:
        task.assignee_id = uuid.UUID(assignee_id)
        rep = db.query(CompanyRep).filter(CompanyRep.id == uuid.UUID(assignee_id)).first()
        changes.append(f"assigned to {rep.name if rep else assignee_id}")
    if due_date:
        task.due_date = date.fromisoformat(due_date)
        changes.append(f"due date: {due_date}")
    if collaborator_ids is not None:
        import json as json_mod
        if collaborator_ids == '' or collaborator_ids == '[]':
            new_collabs = []
        else:
            try:
                new_collabs = json_mod.loads(collaborator_ids)
            except (json_mod.JSONDecodeError, TypeError):
                new_collabs = [c.strip() for c in collaborator_ids.split(',') if c.strip()]
        new_collabs = [str(c) for c in new_collabs if _is_valid_uuid(str(c))]
        old_set = set(str(c) for c in (task.collaborator_ids or []))
        new_set = set(new_collabs)
        task.collaborator_ids = new_collabs
        from sqlalchemy.orm.attributes import flag_modified
        flag_modified(task, 'collaborator_ids')
        # Notify newly added collaborators
        added = new_set - old_set
        for cid in added:
            collab_rep = db.query(CompanyRep).filter(CompanyRep.id == uuid.UUID(cid)).first()
            if collab_rep:
                create_notification(
                    db, collab_rep.rep_id, "task_assigned",
                    f"Added as collaborator: {task.title}",
                    f"Added by {current_user.name}",
                    category="task", entity_type="task", entity_id=task.task_id
                )
        collab_names = []
        for cid in new_collabs:
            rep = db.query(CompanyRep).filter(CompanyRep.id == uuid.UUID(cid)).first()
            if rep: collab_names.append(rep.name)
        changes.append(f"collaborators: {', '.join(collab_names) if collab_names else 'none'}")
    if status_val:
        updated = _task_service.update_task_status(db, task_id, status_val, current_user.id)
        return _task_to_dict(updated, db)

    task.updated_at = datetime.utcnow()
    # Log activity and notify stakeholders for field updates
    if changes:
        _task_service._log_activity(db, task.id, current_user.id, "updated", "", "; ".join(changes))
        _task_service._notify_task_stakeholders(
            db, task, current_user.id,
            f"Task updated: {task.title}",
            "; ".join(changes)
        )
    db.commit()
    return _task_to_dict(task, db)

@app.delete("/api/tasks/{task_id}")
async def delete_task(
    task_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Delete a task. Admin/CCO can delete any, others can delete tasks they created."""
    task = _task_service._find_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if current_user.role not in ("admin", "cco") and task.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Only the creator or admin can delete tasks")
    # Delete subtasks first to avoid FK constraint violation
    subtasks = db.query(Task).filter(Task.parent_task_id == task.id).all()
    for sub in subtasks:
        db.delete(sub)
    db.delete(task)
    db.commit()
    return {"detail": "Task deleted", "task_id": task_id}

@app.post("/api/tasks/{task_id}/complete")
async def complete_task(
    task_id: str,
    notes: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    try:
        task = _task_service.complete_task(db, task_id, current_user.id, notes)
        return _task_to_dict(task, db)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/tasks/{task_id}/delegate")
async def delegate_task(
    task_id: str,
    new_assignee_id: str = Form(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    try:
        task = _task_service.delegate_task(db, task_id, new_assignee_id, current_user.id)
        return _task_to_dict(task, db)
    except (ValueError, PermissionError) as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/tasks/{task_id}/comments")
async def add_task_comment(
    task_id: str,
    content: str = Form(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    if len(content) > 5000:
        raise HTTPException(status_code=400, detail="Comment too long (max 5000 chars)")
    try:
        comment = _task_service.add_comment(db, task_id, current_user.id, content)
        author = db.query(CompanyRep).filter(CompanyRep.id == comment.author_id).first()
        return {
            "id": str(comment.id), "content": comment.content,
            "author_id": str(comment.author_id),
            "author_name": author.name if author else "Unknown",
            "created_at": comment.created_at.isoformat() if comment.created_at else None,
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

@app.get("/api/tasks/{task_id}/comments")
async def get_task_comments(
    task_id: str,
    scope: str = Query("task", regex="^(task|all)$"),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    try:
        task = _task_service._find_task(db, task_id)
        if not task:
            raise HTTPException(status_code=404, detail="Task not found")
        if scope == "all":
            # Aggregate comments: this task + subtasks + micro-tasks
            subtasks = db.query(Task).filter(Task.parent_task_id == task.id).all()
            all_task_ids = [task.id] + [s.id for s in subtasks]
            # Build title lookup for subtasks
            subtask_titles = {s.id: s.title for s in subtasks}
            # Get all micro-tasks under subtasks for title lookup
            sub_ids = [s.id for s in subtasks]
            micro_tasks_all = []
            mt_titles = {}
            mt_parent_map = {}  # micro_task_id -> subtask_id
            if sub_ids:
                micro_tasks_all = db.query(MicroTask).filter(MicroTask.task_id.in_(sub_ids)).all()
                mt_titles = {mt.id: mt.title for mt in micro_tasks_all}
                mt_parent_map = {mt.id: mt.task_id for mt in micro_tasks_all}
            comments = db.query(TaskComment).filter(
                TaskComment.task_id.in_(all_task_ids)
            ).order_by(TaskComment.created_at.desc()).limit(200).all()
        else:
            comments = db.query(TaskComment).filter(
                TaskComment.task_id == task.id, TaskComment.micro_task_id == None
            ).order_by(TaskComment.created_at.desc()).limit(50).all()
        author_ids = set(c.author_id for c in comments if c.author_id)
        author_names = {}
        if author_ids:
            reps = db.query(CompanyRep.id, CompanyRep.name).filter(CompanyRep.id.in_(list(author_ids))).all()
            author_names = {r.id: r.name for r in reps}
        result = []
        for c in comments:
            entry = {
                "id": str(c.id), "content": c.content,
                "author_id": str(c.author_id),
                "author_name": author_names.get(c.author_id, "Unknown"),
                "created_at": c.created_at.isoformat() if c.created_at else None,
            }
            if scope == "all":
                # Tag with context
                if c.micro_task_id:
                    entry["context"] = {
                        "type": "micro_task",
                        "title": mt_titles.get(c.micro_task_id, ""),
                        "parent_subtask_title": subtask_titles.get(mt_parent_map.get(c.micro_task_id), ""),
                    }
                elif c.task_id != task.id:
                    entry["context"] = {
                        "type": "subtask",
                        "title": subtask_titles.get(c.task_id, ""),
                        "parent_subtask_title": None,
                    }
                else:
                    entry["context"] = {
                        "type": "task",
                        "title": task.title,
                        "parent_subtask_title": None,
                    }
            result.append(entry)
        return result
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

@app.get("/api/tasks/{task_id}/activities")
async def get_task_activities(
    task_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    try:
        activities = _task_service.get_activities(db, task_id)
        actor_ids = set(a.actor_id for a in activities if a.actor_id)
        actor_names = {}
        if actor_ids:
            reps = db.query(CompanyRep.id, CompanyRep.name).filter(CompanyRep.id.in_(list(actor_ids))).all()
            actor_names = {r.id: r.name for r in reps}
        return [{
            "id": str(a.id), "action": a.action,
            "old_value": a.old_value, "new_value": a.new_value,
            "actor_id": str(a.actor_id),
            "actor_name": actor_names.get(a.actor_id, "Unknown"),
            "created_at": a.created_at.isoformat() if a.created_at else None,
        } for a in activities]
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

@app.post("/api/tasks/{task_id}/subtasks")
async def create_subtask(
    task_id: str,
    title: str = Form(...),
    description: Optional[str] = Form(None),
    assignee_id: Optional[str] = Form(None),
    priority: str = Form("medium"),
    due_date: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    parent = _task_service._find_task(db, task_id)
    if not parent:
        raise HTTPException(status_code=404, detail="Parent task not found")
    try:
        parsed_due = date.fromisoformat(due_date) if due_date else None
        parsed_assignee = uuid.UUID(assignee_id) if assignee_id else None
        subtask = _task_service.create_task(
            db, current_user.id, title, description,
            parent.task_type, priority, parsed_assignee, parsed_due,
            parent.department, str(parent.linked_inventory_id) if parent.linked_inventory_id else None,
            str(parent.linked_transaction_id) if parent.linked_transaction_id else None,
            str(parent.linked_customer_id) if parent.linked_customer_id else None,
            str(parent.linked_project_id) if parent.linked_project_id else None,
            parent_task_id=parent.id,
        )
        return _task_to_dict(subtask, db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ============================================
# MICRO-TASK ENDPOINTS
# ============================================

def _micro_task_to_dict(mt, db, name_cache=None):
    """Convert MicroTask model to dict with assignee/creator/completer names."""
    assignee_name = None
    creator_name = None
    completed_by_name = None
    if mt.assignee_id:
        if name_cache is not None:
            assignee_name = name_cache.get(mt.assignee_id)
        else:
            rep = db.query(CompanyRep).filter(CompanyRep.id == mt.assignee_id).first()
            assignee_name = rep.name if rep else None
    if mt.created_by:
        if name_cache is not None:
            creator_name = name_cache.get(mt.created_by)
        else:
            rep = db.query(CompanyRep).filter(CompanyRep.id == mt.created_by).first()
            creator_name = rep.name if rep else None
    if mt.completed_by:
        if name_cache is not None:
            completed_by_name = name_cache.get(mt.completed_by)
        else:
            rep = db.query(CompanyRep).filter(CompanyRep.id == mt.completed_by).first()
            completed_by_name = rep.name if rep else None
    comment_count = db.query(func.count(TaskComment.id)).filter(
        TaskComment.micro_task_id == mt.id
    ).scalar() or 0
    return {
        "id": str(mt.id),
        "task_id": str(mt.task_id),
        "title": mt.title,
        "is_completed": mt.is_completed,
        "assignee_id": str(mt.assignee_id) if mt.assignee_id else None,
        "assignee_name": assignee_name,
        "due_date": mt.due_date.isoformat() if mt.due_date else None,
        "sort_order": mt.sort_order,
        "created_by": str(mt.created_by),
        "creator_name": creator_name,
        "completed_at": mt.completed_at.isoformat() if mt.completed_at else None,
        "completed_by": str(mt.completed_by) if mt.completed_by else None,
        "completed_by_name": completed_by_name,
        "comment_count": comment_count,
        "created_at": mt.created_at.isoformat() if mt.created_at else None,
        "updated_at": mt.updated_at.isoformat() if mt.updated_at else None,
        "last_modified_at": mt.updated_at.isoformat() if mt.updated_at else (mt.created_at.isoformat() if mt.created_at else None),
        "last_modified_by_name": completed_by_name or creator_name,
    }

@app.post("/api/tasks/{task_id}/micro-tasks")
async def create_micro_task(
    task_id: str,
    title: str = Form(...),
    assignee_id: Optional[str] = Form(None),
    due_date: Optional[str] = Form(None),
    sort_order: int = Form(0),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Create a micro-task on a subtask."""
    task = _task_service._find_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if not task.parent_task_id:
        raise HTTPException(status_code=400, detail="Micro-tasks can only be added to subtasks (tasks with a parent)")
    if not _check_task_access(task, current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    parsed_due = date.fromisoformat(due_date) if due_date else None
    parsed_assignee = uuid.UUID(assignee_id) if assignee_id else None
    mt = MicroTask(
        task_id=task.id,
        title=title,
        assignee_id=parsed_assignee,
        due_date=parsed_due,
        sort_order=sort_order,
        created_by=current_user.id,
    )
    db.add(mt)
    _task_service._log_activity(db, task.id, current_user.id, "micro_task_created", None, title[:100])
    db.commit()
    db.refresh(mt)
    return _micro_task_to_dict(mt, db)

@app.get("/api/tasks/{task_id}/micro-tasks")
async def list_micro_tasks(
    task_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """List micro-tasks for a subtask."""
    task = _task_service._find_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if not _check_task_access(task, current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    mts = db.query(MicroTask).filter(MicroTask.task_id == task.id).order_by(MicroTask.sort_order, MicroTask.created_at).all()
    # Batch-fetch assignee/creator/completer names
    people_ids = set()
    for mt in mts:
        if mt.assignee_id:
            people_ids.add(mt.assignee_id)
        if mt.created_by:
            people_ids.add(mt.created_by)
        if mt.completed_by:
            people_ids.add(mt.completed_by)
    name_cache = {}
    if people_ids:
        reps = db.query(CompanyRep.id, CompanyRep.name).filter(CompanyRep.id.in_(list(people_ids))).all()
        name_cache = {r.id: r.name for r in reps}
    return [_micro_task_to_dict(mt, db, name_cache) for mt in mts]

@app.put("/api/micro-tasks/reorder")
async def reorder_micro_tasks(
    items: str = Form(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Batch reorder micro-tasks. Expects items as JSON: [{"id": "uuid", "sort_order": 0}, ...]"""
    import json as json_mod
    try:
        parsed = json_mod.loads(items)
    except (json_mod.JSONDecodeError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid JSON in items field")
    if not isinstance(parsed, list):
        raise HTTPException(status_code=400, detail="items must be a JSON array")
    for item in parsed:
        try:
            mt_uuid = uuid.UUID(item["id"])
        except (ValueError, KeyError):
            continue
        mt = db.query(MicroTask).filter(MicroTask.id == mt_uuid).first()
        if mt:
            mt.sort_order = int(item.get("sort_order", 0))
            mt.updated_at = datetime.utcnow()
    db.commit()
    return {"detail": f"Reordered {len(parsed)} micro-tasks"}

@app.put("/api/micro-tasks/{micro_task_id}")
async def update_micro_task(
    micro_task_id: str,
    title: Optional[str] = Form(None),
    is_completed: Optional[str] = Form(None),
    assignee_id: Optional[str] = Form(None),
    due_date: Optional[str] = Form(None),
    sort_order: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Update a micro-task (title, toggle complete, assignee, due date, sort order)."""
    try:
        mt_uuid = uuid.UUID(micro_task_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid micro-task ID")
    mt = db.query(MicroTask).filter(MicroTask.id == mt_uuid).first()
    if not mt:
        raise HTTPException(status_code=404, detail="Micro-task not found")
    # Access check via parent task
    parent_task = db.query(Task).filter(Task.id == mt.task_id).first()
    if not parent_task or not _check_task_access(parent_task, current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    if title is not None:
        mt.title = title
    if is_completed is not None:
        was_completed = mt.is_completed
        mt.is_completed = is_completed.lower() in ("true", "1", "yes")
        if mt.is_completed and not was_completed:
            mt.completed_at = datetime.utcnow()
            mt.completed_by = current_user.id
            _task_service._log_activity(db, mt.task_id, current_user.id, "micro_task_completed", None, mt.title[:100])
        elif not mt.is_completed and was_completed:
            mt.completed_at = None
            mt.completed_by = None
            _task_service._log_activity(db, mt.task_id, current_user.id, "micro_task_reopened", None, mt.title[:100])
    if assignee_id is not None:
        mt.assignee_id = uuid.UUID(assignee_id) if assignee_id else None
    if due_date is not None:
        mt.due_date = date.fromisoformat(due_date) if due_date else None
    if sort_order is not None:
        mt.sort_order = sort_order
    mt.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(mt)
    return _micro_task_to_dict(mt, db)

@app.delete("/api/micro-tasks/{micro_task_id}")
async def delete_micro_task(
    micro_task_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Delete a micro-task."""
    try:
        mt_uuid = uuid.UUID(micro_task_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid micro-task ID")
    mt = db.query(MicroTask).filter(MicroTask.id == mt_uuid).first()
    if not mt:
        raise HTTPException(status_code=404, detail="Micro-task not found")
    parent_task = db.query(Task).filter(Task.id == mt.task_id).first()
    if not parent_task or not _check_task_access(parent_task, current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    _task_service._log_activity(db, mt.task_id, current_user.id, "micro_task_deleted", mt.title[:100], None)
    db.delete(mt)
    db.commit()
    return {"detail": "Micro-task deleted"}

@app.post("/api/micro-tasks/{micro_task_id}/comments")
async def add_micro_task_comment(
    micro_task_id: str,
    content: str = Form(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Add a comment to a micro-task."""
    if len(content) > 5000:
        raise HTTPException(status_code=400, detail="Comment too long (max 5000 chars)")
    try:
        mt_uuid = uuid.UUID(micro_task_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid micro-task ID")
    mt = db.query(MicroTask).filter(MicroTask.id == mt_uuid).first()
    if not mt:
        raise HTTPException(status_code=404, detail="Micro-task not found")
    parent_task = db.query(Task).filter(Task.id == mt.task_id).first()
    if not parent_task or not _check_task_access(parent_task, current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    comment = TaskComment(
        task_id=mt.task_id,
        micro_task_id=mt.id,
        author_id=current_user.id,
        content=content,
    )
    db.add(comment)
    _task_service._log_activity(db, mt.task_id, current_user.id, "commented", None, f"[micro-task] {content[:80]}")
    db.commit()
    db.refresh(comment)
    author = db.query(CompanyRep).filter(CompanyRep.id == comment.author_id).first()
    return {
        "id": str(comment.id), "content": comment.content,
        "task_id": str(comment.task_id),
        "micro_task_id": str(comment.micro_task_id),
        "author_id": str(comment.author_id),
        "author_name": author.name if author else "Unknown",
        "created_at": comment.created_at.isoformat() if comment.created_at else None,
    }

@app.get("/api/micro-tasks/{micro_task_id}/comments")
async def get_micro_task_comments(
    micro_task_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get comments for a micro-task."""
    try:
        mt_uuid = uuid.UUID(micro_task_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid micro-task ID")
    mt = db.query(MicroTask).filter(MicroTask.id == mt_uuid).first()
    if not mt:
        raise HTTPException(status_code=404, detail="Micro-task not found")
    parent_task = db.query(Task).filter(Task.id == mt.task_id).first()
    if not parent_task or not _check_task_access(parent_task, current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    comments = db.query(TaskComment).filter(
        TaskComment.micro_task_id == mt.id
    ).order_by(TaskComment.created_at.desc()).limit(50).all()
    author_ids = set(c.author_id for c in comments if c.author_id)
    author_names = {}
    if author_ids:
        reps = db.query(CompanyRep.id, CompanyRep.name).filter(CompanyRep.id.in_(list(author_ids))).all()
        author_names = {r.id: r.name for r in reps}
    return [{
        "id": str(c.id), "content": c.content,
        "task_id": str(c.task_id),
        "micro_task_id": str(c.micro_task_id) if c.micro_task_id else None,
        "author_id": str(c.author_id),
        "author_name": author_names.get(c.author_id, "Unknown"),
        "created_at": c.created_at.isoformat() if c.created_at else None,
    } for c in comments]

# ============================================
# CROSS-TAB ENTITY TASK ENDPOINTS
# ============================================

def _entity_tasks_response(db, tasks, current_user):
    """Build cross-tab task list with subtask/micro-task counts."""
    name_cache = _build_rep_name_cache(db, tasks)
    result = []
    for t in tasks:
        d = _task_to_dict(t, db, name_cache)
        # Subtask counts
        subtasks = db.query(Task).filter(Task.parent_task_id == t.id).all()
        d["subtask_count"] = len(subtasks)
        d["subtask_completed"] = sum(1 for s in subtasks if s.status == "completed")
        # Micro-task counts across all subtasks
        sub_ids = [s.id for s in subtasks]
        if sub_ids:
            mt_total = db.query(func.count(MicroTask.id)).filter(MicroTask.task_id.in_(sub_ids)).scalar() or 0
            mt_done = db.query(func.count(MicroTask.id)).filter(MicroTask.task_id.in_(sub_ids), MicroTask.is_completed == True).scalar() or 0
        else:
            mt_total = 0
            mt_done = 0
        d["micro_task_count"] = mt_total
        d["micro_task_completed"] = mt_done
        # Comment count (all levels)
        comment_count = db.query(func.count(TaskComment.id)).filter(TaskComment.task_id == t.id).scalar() or 0
        if sub_ids:
            comment_count += db.query(func.count(TaskComment.id)).filter(TaskComment.task_id.in_(sub_ids)).scalar() or 0
        d["comment_count"] = comment_count
        result.append(d)
    return {"tasks": result, "total": len(result)}

@app.get("/api/customers/{customer_id}/tasks")
async def get_customer_tasks(
    customer_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get tasks linked to a customer."""
    try:
        cust_uuid = uuid.UUID(customer_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid customer ID")
    tasks = db.query(Task).filter(
        Task.linked_customer_id == cust_uuid,
        Task.parent_task_id == None
    ).order_by(Task.created_at.desc()).all()
    return _entity_tasks_response(db, tasks, current_user)

@app.get("/api/projects/{project_id}/tasks")
async def get_project_tasks(
    project_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get tasks linked to a project."""
    try:
        proj_uuid = uuid.UUID(project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid project ID")
    tasks = db.query(Task).filter(
        Task.linked_project_id == proj_uuid,
        Task.parent_task_id == None
    ).order_by(Task.created_at.desc()).all()
    return _entity_tasks_response(db, tasks, current_user)

@app.get("/api/inventory/{inventory_id}/tasks")
async def get_inventory_tasks(
    inventory_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get tasks linked to an inventory item."""
    try:
        inv_uuid = uuid.UUID(inventory_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid inventory ID")
    tasks = db.query(Task).filter(
        Task.linked_inventory_id == inv_uuid,
        Task.parent_task_id == None
    ).order_by(Task.created_at.desc()).all()
    return _entity_tasks_response(db, tasks, current_user)

@app.get("/api/transactions/{transaction_id}/tasks")
async def get_transaction_tasks(
    transaction_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get tasks linked to a transaction."""
    try:
        txn_uuid = uuid.UUID(transaction_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid transaction ID")
    tasks = db.query(Task).filter(
        Task.linked_transaction_id == txn_uuid,
        Task.parent_task_id == None
    ).order_by(Task.created_at.desc()).all()
    return _entity_tasks_response(db, tasks, current_user)

# ============================================
# VOICE QUERY ENDPOINTS
# ============================================

@app.post("/api/voice/query")
async def voice_query(
    query: str = Form(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    # Rate limit: voice queries
    user_key = current_user.rep_id or str(current_user.id)
    if not voice_limiter.check(user_key, VOICE_RATE_LIMIT, VOICE_RATE_WINDOW):
        raise HTTPException(status_code=429, detail=f"Voice query rate limit exceeded ({VOICE_RATE_LIMIT}/{VOICE_RATE_WINDOW}s). Please wait.")
    if not _check_daily_limit(user_key, "voice_queries", VOICE_DAILY_LIMIT):
        raise HTTPException(status_code=429, detail=f"Daily voice query limit reached ({VOICE_DAILY_LIMIT}/day). Contact admin.")
    _track_daily_usage(user_key, "voice_queries")
    # Enforce max query length
    if len(query) > 1000:
        raise HTTPException(status_code=400, detail="Query too long (max 1000 chars)")
    user_rep_id = current_user.rep_id
    return _voice_service.process_query(db, query, user_rep_id)

DEEPGRAM_API_KEY = os.getenv("DEEPGRAM_API_KEY", "")
DEEPGRAM_API_URL = "https://api.deepgram.com/v1/listen"

@app.post("/api/voice/upload")
async def voice_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Accept audio file, transcribe via Deepgram, then run voice query pipeline."""
    if not DEEPGRAM_API_KEY:
        raise HTTPException(status_code=503, detail="Voice transcription not configured (DEEPGRAM_API_KEY missing)")
    # Rate limit same as text queries
    user_key = current_user.rep_id or str(current_user.id)
    if not voice_limiter.check(user_key, VOICE_RATE_LIMIT, VOICE_RATE_WINDOW):
        raise HTTPException(status_code=429, detail=f"Voice query rate limit exceeded ({VOICE_RATE_LIMIT}/{VOICE_RATE_WINDOW}s). Please wait.")
    if not _check_daily_limit(user_key, "voice_queries", VOICE_DAILY_LIMIT):
        raise HTTPException(status_code=429, detail=f"Daily voice query limit reached ({VOICE_DAILY_LIMIT}/day). Contact admin.")

    audio_bytes = await file.read()
    if len(audio_bytes) < 1000:
        raise HTTPException(status_code=400, detail="Audio too short. Please record for at least 1 second.")
    if len(audio_bytes) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Audio too large (max 25MB)")

    # Transcribe via Deepgram
    content_type = file.content_type or "audio/webm"
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(
                DEEPGRAM_API_URL,
                content=audio_bytes,
                headers={
                    "Authorization": f"Token {DEEPGRAM_API_KEY}",
                    "Content-Type": content_type,
                },
                params={
                    "model": "nova-3",
                    "smart_format": "true",
                    "punctuate": "true",
                    "language": "en",
                },
            )
        if resp.status_code != 200:
            logger.error(f"Deepgram error {resp.status_code}: {resp.text[:200]}")
            raise HTTPException(status_code=502, detail="Speech transcription failed. Please try again.")
        dg_data = resp.json()
        transcript = dg_data.get("results", {}).get("channels", [{}])[0].get("alternatives", [{}])[0].get("transcript", "")
        confidence = dg_data.get("results", {}).get("channels", [{}])[0].get("alternatives", [{}])[0].get("confidence", 0)
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Speech transcription timed out. Please try again.")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Deepgram transcription error: {e}")
        raise HTTPException(status_code=502, detail="Speech transcription failed.")

    if not transcript or not transcript.strip():
        return {"response_text": "I couldn't understand the audio. Please try again or type your question.", "transcript": "", "confidence": 0}

    _track_daily_usage(user_key, "voice_queries")
    # Run through existing query pipeline
    result = _voice_service.process_query(db, transcript.strip(), current_user.rep_id)
    result["transcript"] = transcript.strip()
    result["stt_confidence"] = round(confidence, 3)
    return result

@app.get("/api/voice/history")
async def voice_history(
    skip: int = 0, limit: int = 50,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    user_rep_id = current_user.rep_id
    query = db.query(QueryHistory)
    if user_rep_id:
        query = query.filter(QueryHistory.user_rep_id == user_rep_id)
    items = query.order_by(QueryHistory.created_at.desc()).offset(skip).limit(limit).all()
    return [{
        "id": str(h.id), "query_text": h.query_text,
        "intent": h.intent, "domain": h.domain,
        "confidence": float(h.confidence) if h.confidence else 0,
        "response_text": h.response_text,
        "success": h.success, "processing_time_ms": float(h.processing_time_ms) if h.processing_time_ms else 0,
        "created_at": h.created_at.isoformat() if h.created_at else None,
    } for h in items]

@app.post("/api/voice/feedback")
async def voice_feedback(
    query_id: str = Form(...),
    feedback_type: str = Form(...),
    rating: Optional[int] = Form(None),
    correct_intent: Optional[str] = Form(None),
    correct_domain: Optional[str] = Form(None),
    correct_response: Optional[str] = Form(None),
    user_comment: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    feedback = QueryFeedback(
        query_history_id=uuid.UUID(query_id),
        feedback_type=feedback_type,
        rating=rating,
        correct_intent=correct_intent,
        correct_domain=correct_domain,
        correct_response=correct_response,
        user_comment=user_comment,
    )
    db.add(feedback)
    db.commit()
    return {"status": "ok", "id": str(feedback.id)}

@app.get("/api/voice/stats")
async def voice_stats(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    if current_user.role not in ("admin", "cco"):
        raise HTTPException(status_code=403, detail="Admin access required")
    total = db.query(QueryHistory).count()
    successful = db.query(QueryHistory).filter(QueryHistory.success == True).count()
    from sqlalchemy import func as sqlfunc
    intent_dist = dict(db.query(QueryHistory.intent, sqlfunc.count()).group_by(QueryHistory.intent).all())
    domain_dist = dict(db.query(QueryHistory.domain, sqlfunc.count()).group_by(QueryHistory.domain).all())
    return {
        "total_queries": total,
        "successful_queries": successful,
        "success_rate": round(successful / total * 100, 1) if total > 0 else 0,
        "intent_distribution": intent_dist,
        "domain_distribution": domain_dist,
    }

# ============================================
# API USAGE MONITORING & SERVER HEALTH
# ============================================

@app.get("/api/admin/usage")
async def get_api_usage(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get API usage stats for all users. Admin/CCO only."""
    if current_user.role not in ("admin", "cco"):
        raise HTTPException(status_code=403, detail="Admin access required")

    today = date.today().isoformat()
    usage_data = []
    for user_key, data in _daily_usage.items():
        if data.get("date") == today:
            usage_data.append({
                "user": user_key,
                "voice_queries": data.get("voice_queries", 0),
                "task_ops": data.get("task_ops", 0),
                "date": data["date"],
            })

    # Voice usage from DB (last 24h)
    from sqlalchemy import func as sqlfunc
    yesterday = datetime.utcnow() - timedelta(hours=24)
    db_voice_24h = db.query(QueryHistory).filter(QueryHistory.created_at >= yesterday).count()
    db_voice_by_user = dict(
        db.query(QueryHistory.user_rep_id, sqlfunc.count())
        .filter(QueryHistory.created_at >= yesterday)
        .group_by(QueryHistory.user_rep_id).all()
    )

    return {
        "today": today,
        "limits": {
            "voice_per_minute": VOICE_RATE_LIMIT,
            "voice_daily": VOICE_DAILY_LIMIT,
            "task_per_minute": TASK_RATE_LIMIT,
        },
        "daily_usage": usage_data,
        "voice_24h_total": db_voice_24h,
        "voice_24h_by_user": db_voice_by_user,
    }

@app.get("/api/admin/health")
async def server_health(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Server health check with resource usage. Admin/CCO only."""
    if current_user.role not in ("admin", "cco"):
        raise HTTPException(status_code=403, detail="Admin access required")

    import platform
    health = {
        "status": "healthy",
        "timestamp": datetime.utcnow().isoformat(),
        "platform": platform.system(),
        "python_version": platform.python_version(),
    }

    # DB connection check
    try:
        db.execute(text("SELECT 1"))
        health["database"] = "connected"
    except Exception as e:
        health["database"] = f"error: {str(e)}"
        health["status"] = "degraded"

    # DB size
    try:
        result = db.execute(text("SELECT pg_database_size(current_database())")).scalar()
        health["db_size_mb"] = round(result / (1024 * 1024), 1) if result else 0
    except Exception:
        health["db_size_mb"] = None

    # Table row counts (key tables only)
    try:
        counts = {}
        for table in ["customers", "brokers", "inventory", "transactions", "tasks", "query_history"]:
            try:
                cnt = db.execute(text(f"SELECT COUNT(*) FROM {table}")).scalar()
                counts[table] = cnt
            except Exception:
                counts[table] = -1
        health["table_counts"] = counts
    except Exception:
        pass

    # Memory usage (if psutil available)
    try:
        import psutil
        proc = psutil.Process()
        mem = proc.memory_info()
        health["memory"] = {
            "rss_mb": round(mem.rss / (1024 * 1024), 1),
            "vms_mb": round(mem.vms / (1024 * 1024), 1),
        }
        health["cpu_percent"] = proc.cpu_percent(interval=0.1)
        disk = psutil.disk_usage("/")
        health["disk"] = {
            "total_gb": round(disk.total / (1024**3), 1),
            "used_gb": round(disk.used / (1024**3), 1),
            "free_gb": round(disk.free / (1024**3), 1),
            "percent": disk.percent,
        }
    except ImportError:
        health["memory"] = "psutil not installed"

    # Rate limiter stats
    voice_limiter.cleanup()
    task_limiter.cleanup()
    health["active_rate_limit_users"] = {
        "voice": len(voice_limiter._requests),
        "task": len(task_limiter._requests),
    }

    return health


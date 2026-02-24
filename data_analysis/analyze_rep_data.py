"""
Comprehensive Sales Rep Lead Data Analysis
============================================
Parses 4 rep Excel files + 1 overlaps file, normalizes mobile numbers,
detects duplicates (intra-file and cross-file), and outputs a JSON summary.

Author: Claude Code analysis script
Date: 2026-02-24
"""

import sys
import os
import json
import re
from collections import defaultdict, Counter

# Force UTF-8 stdout on Windows
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl

# ─── Configuration ───────────────────────────────────────────────────────────

REP_FILES = [
    {
        "path": "C:/Users/Malik/Downloads/CRM Data Ali Zaidi.xlsx",
        "rep_id": "REP-0018",
        "rep_name": "Syed Ali Zaib Zaidi",
    },
    {
        "path": "C:/Users/Malik/Downloads/CRM Data - Iram Aslam.xlsx",
        "rep_id": "REP-0019",
        "rep_name": "Iram Aslam",
    },
    {
        "path": "C:/Users/Malik/Downloads/CRM Data Naeem Zaidi.xlsx",
        "rep_id": "REP-0017",
        "rep_name": "Syed Naeem Abbass Zaidi",
    },
    {
        "path": "C:/Users/Malik/Downloads/CRM Data Samia Rashid.xlsx",
        "rep_id": "REP-0016",
        "rep_name": "Samia Rashid",
    },
]

OVERLAPS_FILE = "C:/Users/Malik/Downloads/To be uploaded in CRM.xlsx"

OUTPUT_JSON = "C:/Users/Malik/desktop/radius2-analytics/data_analysis/rep_data_analysis.json"

# ─── Mobile Normalization ────────────────────────────────────────────────────


def normalize_mobile(raw_value):
    """
    Normalize a mobile number value from Excel.
    Returns (normalized_string, is_valid, raw_for_debug).

    Rules:
    - Strip whitespace, dashes, dots, parens
    - Float → int → str
    - 10-digit starting with 3 → prepend "0"
    - Starts with "92" + 10 digits → strip "92", prepend "0"
    - Starts with "920" + 10 digits → strip "920", prepend "0"
    - Skip None/empty/0
    - Valid = 11 digits starting with "0"
    """
    if raw_value is None:
        return None, False, None

    # Convert numeric types
    if isinstance(raw_value, (int, float)):
        if raw_value == 0 or raw_value == 0.0:
            return None, False, str(raw_value)
        # Convert float to int to avoid ".0" suffix
        val = str(int(raw_value))
    else:
        val = str(raw_value)
        # Handle scientific notation strings like "9.72E+11" → convert to full integer string
        if re.match(r'^[\d\.]+[eE][+\-]?\d+$', val.strip()):
            try:
                val = str(int(float(val.strip())))
            except (ValueError, OverflowError):
                pass
        # Handle string-encoded floats like "3055430688.0" — strip trailing ".0" before
        # general dot-stripping, so "3055430688.0" becomes "3055430688" not "30554306880"
        elif re.match(r'^\d+\.0$', val.strip()):
            val = val.strip()[:-2]

    # Strip whitespace, dashes, dots, parens, plus sign
    val = re.sub(r'[\s\-\.\(\)\+]', '', val)

    # Skip empty or zero-like
    if not val or val == '0' or val.lower() in ('none', 'nan', 'null', 'n/a', '-'):
        return None, False, str(raw_value)

    # Remove any non-digit characters remaining
    digits = re.sub(r'[^\d]', '', val)

    if not digits:
        return None, False, str(raw_value)

    # Rule: "920" prefix + 10 digits (like 9203065553246 → 13 digits)
    # 920 + 10 = 13 digits, strip "920" → 10 digits → prepend "0"
    if len(digits) == 13 and digits.startswith('920'):
        digits = '0' + digits[3:]  # strip "920", prepend "0"

    # Rule: "92" prefix + 10 digits (like 923008669975 → 12 digits)
    # 92 + 10 = 12 digits, strip "92" → 10 digits → prepend "0"
    elif len(digits) == 12 and digits.startswith('92'):
        digits = '0' + digits[2:]  # strip "92", prepend "0"

    # Rule: 10 digits starting with "3" → prepend "0"
    elif len(digits) == 10 and digits.startswith('3'):
        digits = '0' + digits

    # Check validity: standard Pakistani mobile = 11 digits starting with "0"
    is_valid = len(digits) == 11 and digits.startswith('0')

    return digits, is_valid, str(raw_value)


# ─── Column Mapping ──────────────────────────────────────────────────────────

def find_column_index(headers, target_patterns):
    """Find column index by matching against multiple possible header patterns."""
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_clean = str(h).strip().lower().rstrip(';').rstrip('.').rstrip()
        for pattern in target_patterns:
            if pattern in h_clean:
                return i
    return None


def map_columns(headers):
    """Map standard field names to column indices."""
    mapping = {}
    mapping['sr_no'] = find_column_index(headers, ['sr', 'sr.', 'sr no'])
    mapping['name'] = find_column_index(headers, ['name'])
    mapping['mobile'] = find_column_index(headers, ['mobile no'])
    mapping['add_mobile_1'] = find_column_index(headers, ['additional mobile 1'])
    mapping['add_mobile_2'] = find_column_index(headers, ['additional mobile 2'])
    mapping['add_mobile_3'] = find_column_index(headers, ['additional mobile 3'])
    mapping['add_mobile_4'] = find_column_index(headers, ['additional mobile 4'])
    mapping['email'] = find_column_index(headers, ['email'])
    mapping['source'] = find_column_index(headers, ['source'])
    mapping['occupation'] = find_column_index(headers, ['occupation'])
    mapping['area'] = find_column_index(headers, ['area'])
    mapping['city'] = find_column_index(headers, ['city'])
    return mapping


# ─── Parse Rep File ──────────────────────────────────────────────────────────

def parse_rep_file(file_info):
    """Parse a single rep's Excel file into a list of lead dicts."""
    path = file_info['path']
    rep_id = file_info['rep_id']
    rep_name = file_info['rep_name']

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = rows[0]
    col_map = map_columns(headers)

    leads = []
    for row_idx, row in enumerate(rows[1:], start=2):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue

        # Extract fields
        def get(field):
            idx = col_map.get(field)
            if idx is not None and idx < len(row):
                return row[idx]
            return None

        raw_name = get('name')
        name = str(raw_name).strip() if raw_name is not None else None

        # Normalize primary mobile
        raw_mobile = get('mobile')
        primary_mobile, primary_valid, primary_raw = normalize_mobile(raw_mobile)

        # Normalize additional mobiles
        additional_mobiles = []
        additional_mobiles_raw = []
        for field in ['add_mobile_1', 'add_mobile_2', 'add_mobile_3', 'add_mobile_4']:
            raw = get(field)
            norm, valid, raw_str = normalize_mobile(raw)
            if norm is not None:
                additional_mobiles.append({
                    'number': norm,
                    'valid': valid,
                    'raw': raw_str
                })

        email = get('email')
        if email is not None:
            email = str(email).strip()
            if email.lower() in ('none', '', 'nan'):
                email = None

        source = get('source')
        if source is not None:
            source = str(source).strip()
            if source.lower() in ('none', '', 'nan'):
                source = None

        occupation = get('occupation')
        if occupation is not None:
            occupation = str(occupation).strip()
            if occupation.lower() in ('none', '', 'nan'):
                occupation = None

        area = get('area')
        if area is not None:
            area = str(area).strip()
            if area.lower() in ('none', '', 'nan'):
                area = None

        city = get('city')
        if city is not None:
            city = str(city).strip()
            if city.lower() in ('none', '', 'nan'):
                city = None

        lead = {
            'row': row_idx,
            'name': name,
            'primary_mobile': primary_mobile,
            'primary_mobile_valid': primary_valid,
            'primary_mobile_raw': primary_raw,
            'additional_mobiles': additional_mobiles,
            'email': email,
            'source': source,
            'occupation': occupation,
            'area': area,
            'city': city,
            'rep_id': rep_id,
            'rep_name': rep_name,
        }
        leads.append(lead)

    return leads


# ─── Parse Overlaps File ─────────────────────────────────────────────────────

def parse_overlaps_file(path):
    """Parse the overlaps file and return a set of normalized mobile numbers and allocation info."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    overlap_mobiles = set()
    overlap_details = []  # list of {mobile, overlap_type, allocation, sheet}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = rows[0]

        # Find the Mobile column (might be at index 0 or 1 depending on sheet)
        mobile_col = None
        overlap_col = None
        allocation_col = None
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_lower = str(h).strip().lower()
            if h_lower == 'mobile':
                mobile_col = i
            elif h_lower == 'overlap':
                overlap_col = i
            elif 'allocation' in h_lower:
                allocation_col = i

        if mobile_col is None:
            print(f"  WARNING: No 'Mobile' column found in sheet '{sheet_name}'")
            continue

        for row in rows[1:]:
            if all(v is None for v in row):
                continue

            raw_mobile = row[mobile_col] if mobile_col < len(row) else None
            norm, valid, raw_str = normalize_mobile(raw_mobile)

            if norm is not None:
                overlap_mobiles.add(norm)

                overlap_type = None
                allocation = None
                if overlap_col is not None and overlap_col < len(row):
                    overlap_type = str(row[overlap_col]).strip() if row[overlap_col] else None
                if allocation_col is not None and allocation_col < len(row):
                    allocation = str(row[allocation_col]).strip() if row[allocation_col] else None

                overlap_details.append({
                    'mobile': norm,
                    'overlap_type': overlap_type,
                    'allocation': allocation,
                    'sheet': sheet_name,
                })

    wb.close()
    return overlap_mobiles, overlap_details


# ─── Analysis ────────────────────────────────────────────────────────────────

def analyze_leads(all_rep_leads, overlap_mobiles, overlap_details):
    """Run all analyses and return the summary dict."""
    result = {
        'per_file': {},
        'cross_file_duplicates': {},
        'overlap_file_coverage': {},
        'new_cross_duplicates': [],
        'total_unique_leads': 0,
        'weird_mobiles': [],
        'name_issues': [],
    }

    # Global structures
    mobile_to_reps = defaultdict(set)       # mobile → set of rep_ids
    mobile_to_leads = defaultdict(list)     # mobile → list of lead dicts
    all_unique_mobiles = set()
    all_leads_count = 0

    # ─── Per-file analysis ──────────────────────────────────────────
    for rep_info, leads in all_rep_leads:
        rep_id = rep_info['rep_id']
        rep_name = rep_info['rep_name']
        file_name = os.path.basename(rep_info['path'])

        total_rows = len(leads)
        all_leads_count += total_rows

        valid_mobile_count = 0
        no_mobile_count = 0
        internal_dup_mobiles = defaultdict(list)
        additional_mobiles_count = 0
        no_name_count = 0
        name_issue_list = []
        weird_mobile_list = []

        for lead in leads:
            pm = lead['primary_mobile']
            pm_valid = lead['primary_mobile_valid']
            name = lead['name']

            # Valid primary mobile
            if pm is not None and pm_valid:
                valid_mobile_count += 1
                internal_dup_mobiles[pm].append(lead['row'])
                mobile_to_reps[pm].add(rep_id)
                mobile_to_leads[pm].append({
                    'rep_id': rep_id,
                    'rep_name': rep_name,
                    'name': name,
                    'row': lead['row'],
                })
                all_unique_mobiles.add(pm)
            elif pm is not None and not pm_valid:
                # Has a number but it's not valid format
                valid_mobile_count += 1  # still counts as "has mobile"
                internal_dup_mobiles[pm].append(lead['row'])
                mobile_to_reps[pm].add(rep_id)
                mobile_to_leads[pm].append({
                    'rep_id': rep_id,
                    'rep_name': rep_name,
                    'name': name,
                    'row': lead['row'],
                })
                all_unique_mobiles.add(pm)
                weird_mobile_list.append({
                    'row': lead['row'],
                    'name': name,
                    'normalized': pm,
                    'raw': lead['primary_mobile_raw'],
                    'rep_id': rep_id,
                })
            else:
                no_mobile_count += 1

            # Additional mobiles
            for am in lead['additional_mobiles']:
                additional_mobiles_count += 1
                if not am['valid']:
                    weird_mobile_list.append({
                        'row': lead['row'],
                        'name': name,
                        'normalized': am['number'],
                        'raw': am['raw'],
                        'rep_id': rep_id,
                        'field': 'additional',
                    })

            # Name issues
            if name is None or name.strip() == '' or name.strip().lower() in (
                'none', 'not mentioned', 'n/a', '-', 'unknown'
            ):
                no_name_count += 1
                name_issue_list.append({
                    'row': lead['row'],
                    'name': name,
                    'mobile': pm,
                    'rep_id': rep_id,
                })

        # Find actual internal duplicates (mobile appearing more than once in same file)
        internal_dups = {
            mob: rows for mob, rows in internal_dup_mobiles.items() if len(rows) > 1
        }

        result['per_file'][rep_id] = {
            'rep_name': rep_name,
            'file_name': file_name,
            'total_rows': total_rows,
            'rows_with_mobile': valid_mobile_count,
            'rows_without_mobile': no_mobile_count,
            'internal_duplicate_mobiles': len(internal_dups),
            'internal_duplicate_details': {
                mob: {'count': len(rows), 'rows': rows}
                for mob, rows in sorted(internal_dups.items())
            },
            'rows_with_no_name': no_name_count,
            'additional_mobiles_count': additional_mobiles_count,
        }

        result['weird_mobiles'].extend(weird_mobile_list)
        result['name_issues'].extend(name_issue_list)

    # ─── Cross-file duplicate detection ─────────────────────────────
    cross_dups = {}
    for mobile, reps in mobile_to_reps.items():
        if len(reps) > 1:
            cross_dups[mobile] = {
                'rep_ids': sorted(list(reps)),
                'leads': [
                    {'rep_id': l['rep_id'], 'rep_name': l['rep_name'], 'name': l['name'], 'row': l['row']}
                    for l in mobile_to_leads[mobile]
                ]
            }

    result['cross_file_duplicates'] = {
        'count': len(cross_dups),
        'details': dict(sorted(cross_dups.items())),
    }

    # ─── Overlap file coverage ──────────────────────────────────────
    overlap_in_individual = set()
    overlap_not_found = set()

    for mob in overlap_mobiles:
        if mob in all_unique_mobiles:
            overlap_in_individual.add(mob)
        else:
            overlap_not_found.add(mob)

    # Cross-rep dupes NOT in overlaps file
    cross_dup_mobiles = set(cross_dups.keys())
    new_cross_dups = cross_dup_mobiles - overlap_mobiles

    result['overlap_file_coverage'] = {
        'total_overlap_mobiles': len(overlap_mobiles),
        'found_in_individual_files': len(overlap_in_individual),
        'not_found_in_individual_files': len(overlap_not_found),
        'not_found_list': sorted(list(overlap_not_found)),
        'coverage_percent': round(len(overlap_in_individual) / max(len(overlap_mobiles), 1) * 100, 1),
    }

    result['new_cross_duplicates'] = {
        'count': len(new_cross_dups),
        'description': 'Cross-rep duplicate mobiles that are NOT in the overlaps file (new/unresolved overlaps)',
        'mobiles': {
            mob: cross_dups[mob] for mob in sorted(new_cross_dups)
        }
    }

    result['total_unique_leads'] = len(all_unique_mobiles)
    result['total_leads_all_files'] = all_leads_count
    result['leads_without_any_mobile'] = sum(
        stats['rows_without_mobile'] for stats in result['per_file'].values()
    )

    # ─── Overlap file allocation breakdown ──────────────────────────
    allocation_counts = Counter()
    for d in overlap_details:
        alloc = d.get('allocation')
        if alloc:
            allocation_counts[alloc.strip()] += 1

    result['overlap_allocation_breakdown'] = dict(allocation_counts.most_common())

    return result


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 80)
    print("SALES REP LEAD DATA — COMPREHENSIVE ANALYSIS")
    print("=" * 80)
    print()

    # ─── Parse rep files ────────────────────────────────────────────
    all_rep_leads = []
    for rep_info in REP_FILES:
        print(f"Parsing: {os.path.basename(rep_info['path'])} ({rep_info['rep_id']})...")
        leads = parse_rep_file(rep_info)
        all_rep_leads.append((rep_info, leads))
        print(f"  → {len(leads)} rows loaded")
    print()

    # ─── Parse overlaps file ────────────────────────────────────────
    print(f"Parsing overlaps file: {os.path.basename(OVERLAPS_FILE)}...")
    overlap_mobiles, overlap_details = parse_overlaps_file(OVERLAPS_FILE)
    print(f"  → {len(overlap_mobiles)} unique overlap mobiles from {len(overlap_details)} entries")
    print()

    # ─── Run analysis ───────────────────────────────────────────────
    print("Running analysis...")
    result = analyze_leads(all_rep_leads, overlap_mobiles, overlap_details)
    print()

    # ─── Print Summary ──────────────────────────────────────────────
    print("=" * 80)
    print("PER-FILE STATISTICS")
    print("=" * 80)
    for rep_id, stats in result['per_file'].items():
        print(f"\n  {rep_id} — {stats['rep_name']} ({stats['file_name']})")
        print(f"    Total rows:             {stats['total_rows']}")
        print(f"    Rows with mobile:       {stats['rows_with_mobile']}")
        print(f"    Rows without mobile:    {stats['rows_without_mobile']}")
        print(f"    Internal dup mobiles:    {stats['internal_duplicate_mobiles']}")
        print(f"    Rows with no name:      {stats['rows_with_no_name']}")
        print(f"    Additional mobiles:      {stats['additional_mobiles_count']}")

        if stats['internal_duplicate_mobiles'] > 0:
            print(f"    Internal duplicates breakdown:")
            for mob, info in list(stats['internal_duplicate_details'].items())[:10]:
                print(f"      {mob}: {info['count']}x (rows {info['rows']})")
            if stats['internal_duplicate_mobiles'] > 10:
                print(f"      ... and {stats['internal_duplicate_mobiles'] - 10} more")

    print()
    print("=" * 80)
    print("CROSS-FILE DUPLICATES")
    print("=" * 80)
    cross = result['cross_file_duplicates']
    print(f"  Total cross-file duplicate mobiles: {cross['count']}")
    if cross['count'] > 0:
        # Show first 20
        for i, (mob, info) in enumerate(cross['details'].items()):
            if i >= 20:
                print(f"  ... and {cross['count'] - 20} more")
                break
            reps = ', '.join(info['rep_ids'])
            names = ' / '.join(f"{l['rep_name']}:{l['name']}" for l in info['leads'])
            print(f"  {mob} → [{reps}] — {names}")

    print()
    print("=" * 80)
    print("OVERLAP FILE COVERAGE")
    print("=" * 80)
    ov = result['overlap_file_coverage']
    print(f"  Total unique overlap mobiles:   {ov['total_overlap_mobiles']}")
    print(f"  Found in individual files:      {ov['found_in_individual_files']}")
    print(f"  NOT found in individual files:  {ov['not_found_in_individual_files']}")
    print(f"  Coverage:                       {ov['coverage_percent']}%")
    if ov['not_found_in_individual_files'] > 0:
        print(f"  Not-found mobiles (first 20):  {ov['not_found_list'][:20]}")

    print()
    print("=" * 80)
    print("NEW CROSS-REP DUPLICATES (NOT in overlaps file)")
    print("=" * 80)
    new_cd = result['new_cross_duplicates']
    print(f"  Count: {new_cd['count']}")
    if new_cd['count'] > 0:
        for i, (mob, info) in enumerate(new_cd['mobiles'].items()):
            if i >= 30:
                print(f"  ... and {new_cd['count'] - 30} more")
                break
            reps = ', '.join(info['rep_ids'])
            names = ' / '.join(f"{l['rep_name']}:{l['name']}" for l in info['leads'])
            print(f"  {mob} → [{reps}] — {names}")

    print()
    print("=" * 80)
    print("OVERLAP ALLOCATION BREAKDOWN")
    print("=" * 80)
    for alloc, count in result['overlap_allocation_breakdown'].items():
        print(f"  {alloc}: {count}")

    print()
    print("=" * 80)
    print("TOTALS")
    print("=" * 80)
    print(f"  Total leads across all files:      {result['total_leads_all_files']}")
    print(f"  Leads without any mobile:          {result['leads_without_any_mobile']}")
    print(f"  Total unique mobiles (all files):   {result['total_unique_leads']}")
    print(f"  Cross-file duplicate mobiles:       {result['cross_file_duplicates']['count']}")
    print(f"  Weird/non-standard mobiles:         {len(result['weird_mobiles'])}")
    print(f"  Name issues (empty/missing):        {len(result['name_issues'])}")

    if result['weird_mobiles']:
        print()
        print("=" * 80)
        print("WEIRD MOBILES (non-standard format after normalization)")
        print("=" * 80)
        for i, w in enumerate(result['weird_mobiles']):
            if i >= 30:
                print(f"  ... and {len(result['weird_mobiles']) - 30} more")
                break
            field = w.get('field', 'primary')
            print(f"  Row {w['row']} ({w['rep_id']}): '{w['raw']}' → '{w['normalized']}' [{field}] — {w['name']}")

    if result['name_issues']:
        print()
        print("=" * 80)
        print("NAME ISSUES (first 30)")
        print("=" * 80)
        for i, n in enumerate(result['name_issues'][:30]):
            print(f"  Row {n['row']} ({n['rep_id']}): name='{n['name']}', mobile={n['mobile']}")
        if len(result['name_issues']) > 30:
            print(f"  ... and {len(result['name_issues']) - 30} more")

    # ─── Write JSON output ──────────────────────────────────────────
    # Convert sets to lists for JSON serialization
    output = json.loads(json.dumps(result, default=str, ensure_ascii=False, indent=2))

    os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print()
    print(f"JSON output written to: {OUTPUT_JSON}")
    print()
    print("Analysis complete.")


if __name__ == '__main__':
    main()

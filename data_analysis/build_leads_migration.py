# -*- coding: utf-8 -*-
"""
Lead Data Migration Builder
Processes 4 individual rep files + 1 overlap resolution file
Generates SQL migration for Orbit CRM leads table
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import json
import re
import uuid
from collections import Counter, defaultdict
from datetime import datetime

# ─── CONFIG ───────────────────────────────────────────────────────────────────
REP_FILES = {
    'REP-0018': {
        'name': 'Syed Ali Zaib Zaidi',
        'uuid': 'd8d0b91f-9753-4c86-8149-30ebf312ad21',
        'file': 'C:/Users/Malik/Downloads/CRM Data Ali Zaidi.xlsx',
    },
    'REP-0019': {
        'name': 'Iram Aslam',
        'uuid': 'af723dce-2620-416f-bc29-fa7294a01cc9',
        'file': 'C:/Users/Malik/Downloads/CRM Data - Iram Aslam.xlsx',
    },
    'REP-0017': {
        'name': 'Syed Naeem Abbass Zaidi',
        'uuid': '0979ceba-e38a-47cc-b50b-cc0d684f2795',
        'file': 'C:/Users/Malik/Downloads/CRM Data Naeem Zaidi.xlsx',
    },
    'REP-0016': {
        'name': 'Samia Rashid',
        'uuid': '254c4e3d-15fe-44ed-9d01-2111824ff464',
        'file': 'C:/Users/Malik/Downloads/CRM Data Samia Rashid.xlsx',
    },
}

OVERLAP_FILE = 'C:/Users/Malik/Downloads/To be uploaded in CRM.xlsx'

# Allocation name → rep_id mapping (from overlap file)
ALLOC_TO_REP = {
    'Iram Aslam': 'REP-0019', 'Iram Aslam ': 'REP-0019',
    'Syed Ali Zaidi': 'REP-0018', 'Syed Ali Zaib Zaidi': 'REP-0018',
    'Samia Rashid': 'REP-0016', 'Samia Rasheed': 'REP-0016', 'Samia Rashid ': 'REP-0016',
    'Naeem': 'REP-0017', 'Naeem Abbas': 'REP-0017', 'Naeem Abbass': 'REP-0017',
    'Naeem Zaidi': 'REP-0017', 'Syed Naeem Abbas': 'REP-0017',
}
ALLOC_TO_COLNAME = {
    'Iram Aslam': 'Iram', 'Iram Aslam ': 'Iram',
    'Syed Ali Zaidi': 'Ali', 'Syed Ali Zaib Zaidi': 'Ali',
    'Samia Rashid': 'Samia', 'Samia Rasheed': 'Samia', 'Samia Rashid ': 'Samia',
    'Naeem': 'Naeem', 'Naeem Abbas': 'Naeem', 'Naeem Abbass': 'Naeem',
    'Naeem Zaidi': 'Naeem', 'Syed Naeem Abbas': 'Naeem',
}

# New duplicate resolution
NEW_DUP_OWNER = {'03077674577': 'REP-0019'}  # Muhammad Ahmed Shahid → Iram

# Specific mobile fixes
MOBILE_FIXES = {'9710508285799': '971508285799'}

# Junk occupation values to clean
JUNK_OCCUPATIONS = {
    'nathing': '', 'nothing': '', 'n/a': '', 'na': '', 'nil': '',
    'faisalabad': '', 'faislabad': '', 'lahore': '',
    'wapda city faisalabad': '', 'ibrahim': '',
    'out of reach': '', 'no answer': '', 'not interested': '',
    'wrong number': '', 'switched off': '',
}

# ─── MOBILE NORMALIZATION ─────────────────────────────────────────────────────
def normalize_mobile_multi(raw_value):
    """Parse a raw mobile value that may contain multiple numbers.
    Returns list of (normalized, is_standard_pk) tuples."""
    if raw_value is None:
        return []
    val = str(raw_value)
    if 'E+' in val or 'e+' in val:
        try:
            val = str(int(float(val)))
        except (ValueError, OverflowError):
            pass
    if val.endswith('.0'):
        val = val[:-2]
    # Clean unicode, prefixes, emails
    val = _clean_mobile_str(val)
    if val is None:
        return []
    # Split on common separators
    parts = re.split(r'\s*[-/&,]\s*', val)
    results = []
    for part in parts:
        part = re.sub(r'[\s\.\(\)\+\u200e\u200f\u200b\u202a-\u202e\ufeff\u2066-\u2069]', '', part)
        if not part or part == '0' or part == 'None' or not part.isdigit():
            continue
        # Apply specific fixes
        if part in MOBILE_FIXES:
            part = MOBILE_FIXES[part]
        # Pakistani normalization
        if len(part) == 13 and part.startswith('920'):
            part = '0' + part[3:]
        elif len(part) == 12 and part.startswith('92'):
            part = '0' + part[2:]
        elif len(part) == 10 and part[0] == '3':
            part = '0' + part
        if len(part) > 20:
            part = part[:20]
        is_pk = len(part) == 11 and part.startswith('03')
        results.append((part, is_pk))
    return results


def _clean_mobile_str(val):
    """Common mobile string cleaning steps."""
    # Strip Unicode invisible chars (LTR/RTL marks, zero-width spaces, etc.)
    val = re.sub(r'[\u200e\u200f\u200b\u200c\u200d\u202a\u202b\u202c\u202d\u202e\ufeff\u2066\u2067\u2068\u2069]', '', val)
    # Strip surrounding double quotes
    val = val.strip('"').strip('"').strip('"')
    # Strip p: prefix (phonebook notation)
    if val.lower().startswith('p:'):
        val = val[2:]
    # Skip if contains @ (email) or all-alpha (name/text)
    if '@' in val or (val.isalpha() and len(val) > 2):
        return None
    return val


def normalize_mobile(raw_value):
    """Normalize a mobile number (first number only). Returns (normalized, is_standard_pk)"""
    if raw_value is None:
        return None, False

    # Convert numeric to string
    val = str(raw_value)

    # Handle scientific notation from Excel (e.g., 9.72E+11)
    if 'E+' in val or 'e+' in val:
        try:
            val = str(int(float(val)))
        except (ValueError, OverflowError):
            return val, False

    # Remove .0 suffix from float conversion
    if val.endswith('.0'):
        val = val[:-2]

    # Clean unicode, prefixes, emails
    val = _clean_mobile_str(val)
    if val is None:
        return None, False

    # SPLIT multi-number cells BEFORE stripping delimiters
    # Common separators: " - ", " / ", " & ", "/", ","
    for sep in [' - ', ' / ', ' & ', '/', ',']:
        if sep in val:
            val = val.split(sep)[0]  # Take first number

    # Strip whitespace, dashes, dots, parens, plus sign
    val = re.sub(r'[\s\-\.\(\)\+]', '', val)

    # Apply specific fixes
    if val in MOBILE_FIXES:
        val = MOBILE_FIXES[val]

    # Skip empty/zero
    if not val or val == '0' or val == 'None':
        return None, False

    # Skip non-numeric (text/names/garbage in mobile field)
    if not val.isdigit():
        return None, False

    # Pakistani number normalization
    # 920XXXXXXXXXX (13 digits, starts with 920) → 0XXXXXXXXXX
    if len(val) == 13 and val.startswith('920'):
        val = '0' + val[3:]
    # 92XXXXXXXXXX (12 digits, starts with 92) → 0XXXXXXXXXX
    elif len(val) == 12 and val.startswith('92'):
        val = '0' + val[2:]
    # 3XXXXXXXXX (10 digits, starts with 3) → 03XXXXXXXXX
    elif len(val) == 10 and val[0] == '3':
        val = '0' + val
    # 03XXXXXXXXX (11 digits, starts with 03) → already standard
    elif len(val) == 11 and val.startswith('03'):
        pass  # Already standard Pakistani format

    # Safety: truncate to 20 chars max (varchar(20) in DB)
    if len(val) > 20:
        val = val[:20]

    is_pk = len(val) == 11 and val.startswith('03')
    return val, is_pk


def is_junk_name(name):
    """Check if a name is junk/missing"""
    if not name:
        return True
    n = str(name).strip().lower()
    return n in ('', 'none', 'not mentioned', 'not mentionedd', 'n/a', 'na',
                 'nil', 'unknown', 'test', '-', '--', '---')


def clean_occupation(occ):
    """Clean occupation field"""
    if not occ:
        return ''
    occ = str(occ).strip()
    if occ.lower() in JUNK_OCCUPATIONS:
        return JUNK_OCCUPATIONS[occ.lower()]
    return occ


def sql_escape(val):
    """Escape a string for SQL insertion"""
    if val is None or val == '':
        return 'NULL'
    s = str(val).replace("'", "''")
    return f"'{s}'"


# ─── STEP 1: Parse individual rep files ───────────────────────────────────────
print("=" * 60)
print("STEP 1: Parsing individual rep files")
print("=" * 60)

all_rep_leads = {}  # rep_id → [lead_dict]

for rep_id, config in REP_FILES.items():
    wb = openpyxl.load_workbook(config['file'])
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    # Find column indices (flexible header matching)
    name_idx = mobile_idx = None
    addl_mobile_idxs = []
    email_idx = source_idx = occ_idx = area_idx = city_idx = None

    for i, h in enumerate(header):
        if h is None:
            continue
        h_str = str(h).strip().lower().rstrip(';').rstrip('.')
        if h_str in ('name', 'name '):
            name_idx = i
        elif h_str in ('mobile no', 'mobile no '):
            mobile_idx = i
        elif 'additional mobile' in h_str:
            addl_mobile_idxs.append(i)
        elif h_str == 'email':
            email_idx = i
        elif h_str == 'source':
            source_idx = i
        elif h_str in ('occupation', 'occupation '):
            occ_idx = i
        elif h_str in ('area', 'area '):
            area_idx = i
        elif h_str == 'city':
            city_idx = i

    leads = []
    for row in data_rows:
        # Get primary mobile
        raw_mobile = row[mobile_idx] if mobile_idx is not None else None
        mobile, is_pk = normalize_mobile(raw_mobile)

        if mobile is None:
            continue  # Skip no-mobile rows

        # Get name
        name = str(row[name_idx]).strip() if name_idx is not None and row[name_idx] else None

        if is_junk_name(name):
            continue  # Skip junk name rows

        # Get additional mobiles (handle multi-number cells)
        addl_mobiles = []
        for idx in addl_mobile_idxs:
            if idx < len(row) and row[idx]:
                parsed = normalize_mobile_multi(row[idx])
                for addl_m, _ in parsed:
                    if addl_m and addl_m != mobile and addl_m not in addl_mobiles:
                        addl_mobiles.append(addl_m)

        # Get other fields
        email = str(row[email_idx]).strip() if email_idx is not None and row[email_idx] else ''
        source = str(row[source_idx]).strip() if source_idx is not None and row[source_idx] else 'Personal'
        occ = clean_occupation(row[occ_idx] if occ_idx is not None else None)
        area = str(row[area_idx]).strip() if area_idx is not None and row[area_idx] else ''
        city = str(row[city_idx]).strip() if city_idx is not None and row[city_idx] else ''

        leads.append({
            'name': name,
            'mobile': mobile,
            'additional_mobiles': addl_mobiles,
            'email': email if email.lower() != 'none' else '',
            'source': source if source.lower() != 'none' else 'Personal',
            'occupation': occ,
            'area': area if area.lower() != 'none' else '',
            'city': city if city.lower() != 'none' else '',
            'rep_id': rep_id,
        })

    all_rep_leads[rep_id] = leads
    print(f"  {rep_id} ({config['name']}): {len(leads)} leads after removing no-mobile and junk-name rows")


# ─── STEP 2: Deduplicate within each file ────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 2: Internal deduplication (keep first occurrence)")
print("=" * 60)

for rep_id in all_rep_leads:
    before = len(all_rep_leads[rep_id])
    seen = set()
    deduped = []
    for lead in all_rep_leads[rep_id]:
        if lead['mobile'] not in seen:
            seen.add(lead['mobile'])
            deduped.append(lead)
    all_rep_leads[rep_id] = deduped
    removed = before - len(deduped)
    print(f"  {rep_id}: {before} → {len(deduped)} (removed {removed} internal dups)")


# ─── STEP 3: Parse overlap resolution file ───────────────────────────────────
print("\n" + "=" * 60)
print("STEP 3: Parsing overlap resolution file (386 leads)")
print("=" * 60)

wb_overlap = openpyxl.load_workbook(OVERLAP_FILE)
overlap_leads = []  # {mobile, name, occupation, city, source, rep_id}
overlap_mobiles = set()

for sheet_name in wb_overlap.sheetnames:
    ws = wb_overlap[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    mobile_idx = alloc_idx = None
    rep_columns = {}

    for i, h in enumerate(header):
        if h is None:
            continue
        h_str = str(h).strip()
        if h_str == 'Mobile':
            mobile_idx = i
        elif 'Allocation' in h_str:
            alloc_idx = i
        elif '- Name' in h_str:
            rep = h_str.split(' - ')[0].strip()
            rep_columns.setdefault(rep, {})['name_idx'] = i
        elif '- Occupation' in h_str:
            rep = h_str.split(' - ')[0].strip()
            rep_columns.setdefault(rep, {})['occ_idx'] = i
        elif '- City' in h_str:
            rep = h_str.split(' - ')[0].strip()
            rep_columns.setdefault(rep, {})['city_idx'] = i
        elif '- Source' in h_str:
            rep = h_str.split(' - ')[0].strip()
            rep_columns.setdefault(rep, {})['source_idx'] = i

    for row in data_rows:
        mobile_raw = row[mobile_idx] if mobile_idx is not None else None
        alloc = str(row[alloc_idx]).strip() if alloc_idx is not None and row[alloc_idx] else None
        if not mobile_raw or str(mobile_raw).strip() in ('None', ''):
            continue

        mobile_str = str(mobile_raw).replace('.0', '').strip()
        mobile, _ = normalize_mobile(mobile_str)
        if not mobile:
            continue

        rep_id = ALLOC_TO_REP.get(alloc)
        col_name = ALLOC_TO_COLNAME.get(alloc)

        # Get allocated rep's data
        name = occ = city = source = ''
        if col_name and col_name in rep_columns:
            cols = rep_columns[col_name]
            name = str(row[cols['name_idx']]).strip() if cols.get('name_idx') is not None and row[cols['name_idx']] else ''
            occ = str(row[cols.get('occ_idx', 0)]).strip() if cols.get('occ_idx') is not None and row[cols.get('occ_idx')] else ''
            city = str(row[cols.get('city_idx', 0)]).strip() if cols.get('city_idx') is not None and row[cols.get('city_idx')] else ''
            source = str(row[cols.get('source_idx', 0)]).strip() if cols.get('source_idx') is not None and row[cols.get('source_idx')] else ''

            # Fallback name from other reps
            if is_junk_name(name):
                for other_rep, other_cols in rep_columns.items():
                    if other_rep == col_name:
                        continue
                    alt = str(row[other_cols['name_idx']]).strip() if other_cols.get('name_idx') is not None and row[other_cols['name_idx']] else ''
                    if not is_junk_name(alt):
                        name = alt
                        break

        overlap_mobiles.add(mobile)
        overlap_leads.append({
            'mobile': mobile,
            'name': name,
            'occupation': clean_occupation(occ),
            'city': city if city.lower() not in ('none', '') else '',
            'source': source if source.lower() not in ('none', '') else 'Personal',
            'area': '',
            'email': '',
            'additional_mobiles': [],
            'rep_id': rep_id,
        })

print(f"  Parsed {len(overlap_leads)} overlap leads ({len(overlap_mobiles)} unique mobiles)")


# ─── STEP 4: Remove overlap mobiles from individual files ─────────────────────
print("\n" + "=" * 60)
print("STEP 4: Removing overlap mobiles from individual files")
print("=" * 60)

# Also add the new duplicate
new_dup_mobile = '03077674577'
overlap_mobiles.add(new_dup_mobile)

for rep_id in all_rep_leads:
    before = len(all_rep_leads[rep_id])
    all_rep_leads[rep_id] = [l for l in all_rep_leads[rep_id] if l['mobile'] not in overlap_mobiles]
    removed = before - len(all_rep_leads[rep_id])
    print(f"  {rep_id}: removed {removed} overlap leads ({before} → {len(all_rep_leads[rep_id])})")


# ─── STEP 5: Add resolved overlaps back ──────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 5: Adding resolved overlaps back to allocated reps")
print("=" * 60)

# Add the new duplicate to Iram
# Find the lead data from Iram's original file for this mobile
new_dup_lead = {
    'mobile': new_dup_mobile,
    'name': 'Muhammad Ahmed Shahid',
    'occupation': '',
    'city': 'Faisalabad',
    'source': 'Personal',
    'area': '',
    'email': '',
    'additional_mobiles': [],
    'rep_id': 'REP-0019',
}

overlap_counts = Counter()
for lead in overlap_leads:
    all_rep_leads[lead['rep_id']].append(lead)
    overlap_counts[lead['rep_id']] += 1

# Add the new dup
all_rep_leads['REP-0019'].append(new_dup_lead)
overlap_counts['REP-0019'] += 1

for rep_id, count in sorted(overlap_counts.items()):
    print(f"  {rep_id}: +{count} resolved overlaps")


# ─── STEP 6: Final dedup pass (safety check) ─────────────────────────────────
print("\n" + "=" * 60)
print("STEP 6: Final deduplication pass")
print("=" * 60)

total_final = 0
for rep_id in all_rep_leads:
    before = len(all_rep_leads[rep_id])
    seen = set()
    deduped = []
    for lead in all_rep_leads[rep_id]:
        if lead['mobile'] not in seen:
            seen.add(lead['mobile'])
            deduped.append(lead)
    all_rep_leads[rep_id] = deduped
    removed = before - len(deduped)
    total_final += len(deduped)
    if removed > 0:
        print(f"  {rep_id}: removed {removed} more dups ({before} → {len(deduped)})")
    else:
        print(f"  {rep_id}: clean ({len(deduped)} leads)")

print(f"\n  TOTAL FINAL LEADS: {total_final}")

# Cross-rep final check
all_mobiles_final = []
for rep_id, leads in all_rep_leads.items():
    for l in leads:
        all_mobiles_final.append((l['mobile'], rep_id))

mobile_rep_map = defaultdict(set)
for m, r in all_mobiles_final:
    mobile_rep_map[m].add(r)

cross_dups = {m: reps for m, reps in mobile_rep_map.items() if len(reps) > 1}
if cross_dups:
    print(f"\n  WARNING: {len(cross_dups)} cross-rep duplicates remain!")
    for m, reps in list(cross_dups.items())[:10]:
        print(f"    {m}: {reps}")
else:
    print(f"\n  VERIFIED: No cross-rep duplicates in final dataset")


# ─── STEP 7: Generate SQL migration ──────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 7: Generating SQL migration")
print("=" * 60)

sql_lines = []
sql_lines.append("-- ============================================================")
sql_lines.append("-- Lead Data Migration: Sales Rep Raw Leads Upload")
sql_lines.append(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
sql_lines.append(f"-- Total leads: {total_final}")
sql_lines.append("-- Sources: 4 individual rep files + 1 overlap resolution file")
sql_lines.append("-- ============================================================")
sql_lines.append("")
sql_lines.append("BEGIN;")
sql_lines.append("")

# Get max lead_id number from existing leads
# We know existing leads go up to LEAD-00029
sql_lines.append("-- Get next lead_id sequence number")
sql_lines.append("-- Existing leads: LEAD-00001 to LEAD-00030 (17 total)")
sql_lines.append("")

lead_counter = 31  # Start from LEAD-00031 (LEAD-00030 exists in prod)

for rep_id, leads in sorted(all_rep_leads.items()):
    config = REP_FILES[rep_id]
    rep_uuid = config['uuid']

    sql_lines.append(f"-- ── {rep_id}: {config['name']} ({len(leads)} leads) ──")
    sql_lines.append("")

    for lead in leads:
        lead_id = f"LEAD-{lead_counter:05d}"
        lead_counter += 1

        name_esc = sql_escape(lead['name'])
        mobile_esc = sql_escape(lead['mobile'])
        email_esc = sql_escape(lead['email']) if lead['email'] else 'NULL'
        source_esc = sql_escape(lead.get('source', 'Personal') or 'Personal')
        occ_esc = sql_escape(lead['occupation']) if lead['occupation'] else 'NULL'
        area_esc = sql_escape(lead['area']) if lead['area'] else 'NULL'
        city_esc = sql_escape(lead['city']) if lead['city'] else 'NULL'

        # Additional mobiles as JSONB array
        addl = lead.get('additional_mobiles', [])
        if addl:
            addl_json = json.dumps(addl)
            addl_esc = f"'{addl_json}'::jsonb"
        else:
            addl_esc = "'[]'::jsonb"

        sql_lines.append(
            f"INSERT INTO leads (id, lead_id, name, mobile, email, source, occupation, "
            f"area, city, assigned_rep_id, status, pipeline_stage, lead_type, "
            f"additional_mobiles, country_code, metadata, created_at, updated_at) "
            f"VALUES ("
            f"uuid_generate_v4(), '{lead_id}', {name_esc}, {mobile_esc}, {email_esc}, "
            f"{source_esc}, {occ_esc}, {area_esc}, {city_esc}, "
            f"'{rep_uuid}', 'new', 'New', 'prospect', "
            f"{addl_esc}, '+92', '{{}}', NOW(), NOW());"
        )

    sql_lines.append("")

sql_lines.append("-- ── Verification Queries ──")
sql_lines.append("")
sql_lines.append("-- Check total leads uploaded")
sql_lines.append(f"-- Expected: {total_final} new + 16 existing = {total_final + 16} total")
sql_lines.append("SELECT COUNT(*) as total_leads FROM leads;")
sql_lines.append("")
sql_lines.append("-- Check per-rep breakdown")
sql_lines.append("SELECT cr.rep_id, cr.name, COUNT(l.id) as lead_count")
sql_lines.append("FROM leads l")
sql_lines.append("JOIN company_reps cr ON l.assigned_rep_id = cr.id")
sql_lines.append("GROUP BY cr.rep_id, cr.name")
sql_lines.append("ORDER BY cr.rep_id;")
sql_lines.append("")
sql_lines.append("-- Check for duplicate mobiles")
sql_lines.append("SELECT mobile, COUNT(*) as cnt FROM leads")
sql_lines.append("WHERE mobile IS NOT NULL AND mobile != ''")
sql_lines.append("GROUP BY mobile HAVING COUNT(*) > 1;")
sql_lines.append("")
sql_lines.append("COMMIT;")

# Write SQL file
sql_content = '\n'.join(sql_lines)
output_path = 'C:/Users/Malik/desktop/radius2-analytics/database/raw_leads_migration.sql'
with open(output_path, 'w', encoding='utf-8') as f:
    f.write(sql_content)

print(f"  SQL migration written to: {output_path}")
print(f"  Total INSERT statements: {lead_counter - 30}")
print(f"  Lead IDs: LEAD-00030 to LEAD-{lead_counter-1:05d}")

# ─── STEP 8: Summary stats ───────────────────────────────────────────────────
print("\n" + "=" * 60)
print("FINAL SUMMARY")
print("=" * 60)

for rep_id in sorted(all_rep_leads):
    config = REP_FILES[rep_id]
    leads = all_rep_leads[rep_id]
    pk_count = sum(1 for l in leads if len(l['mobile']) == 11 and l['mobile'].startswith('03'))
    intl_count = len(leads) - pk_count
    with_addl = sum(1 for l in leads if l.get('additional_mobiles'))
    with_occ = sum(1 for l in leads if l['occupation'])
    with_email = sum(1 for l in leads if l['email'])
    print(f"\n{rep_id} ({config['name']}):")
    print(f"  Total: {len(leads)}")
    print(f"  Pakistani mobiles: {pk_count}, International: {intl_count}")
    print(f"  With additional mobiles: {with_addl}")
    print(f"  With occupation: {with_occ}")
    print(f"  With email: {with_email}")

print(f"\nGRAND TOTAL: {total_final} leads")
print(f"Lead IDs: LEAD-00030 → LEAD-{lead_counter-1:05d}")

# Save summary JSON
summary = {
    'generated_at': datetime.now().isoformat(),
    'total_leads': total_final,
    'lead_id_range': f'LEAD-00030 to LEAD-{lead_counter-1:05d}',
    'per_rep': {},
    'sql_file': output_path,
}
for rep_id in sorted(all_rep_leads):
    leads = all_rep_leads[rep_id]
    summary['per_rep'][rep_id] = {
        'name': REP_FILES[rep_id]['name'],
        'count': len(leads),
        'pk_mobiles': sum(1 for l in leads if len(l['mobile']) == 11 and l['mobile'].startswith('03')),
        'intl_mobiles': sum(1 for l in leads if not (len(l['mobile']) == 11 and l['mobile'].startswith('03'))),
    }

with open('C:/Users/Malik/desktop/radius2-analytics/data_analysis/migration_summary.json', 'w') as f:
    json.dump(summary, f, indent=2)

print("\nDone! Migration SQL ready at: database/raw_leads_migration.sql")
